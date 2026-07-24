"""Native API 实现：文件/Excel/系统通知"""
import os
import uuid
import openpyxl
from PySide6.QtWidgets import QFileDialog, QMessageBox
from PySide6.QtCore import QObject, QRunnable, QThreadPool, Signal


class _ReadExcelTask(QRunnable):
    """异步读取 Excel：避免阻塞 QWebChannel 线程。"""

    def __init__(self, task_id: str, file_path: str, callback):
        super().__init__()
        self.task_id = task_id
        self.file_path = file_path
        self._callback = callback

    def run(self):
        result = {"task_id": self.task_id, "ok": False, "data": [], "error": ""}
        try:
            if not self.file_path or not os.path.exists(self.file_path):
                result["error"] = "文件不存在"
            else:
                wb = openpyxl.load_workbook(self.file_path, data_only=True)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(cell is not None for cell in row):
                        rows.append(dict(zip(headers, row)))
                result["ok"] = True
                result["data"] = rows
        except Exception as e:
            result["error"] = str(e)
            print(f"[NativeAPI] read_excel error: {e}")
        # 必须在 QThread 主线程中通过 Signal 通知前端
        self._callback(result)


class NativeAPI:
    """PyQt 本地能力封装"""

    def __init__(self, parent: QObject = None):
        self._bridge = None
        self._pool = QThreadPool.globalInstance()

    def set_bridge(self, bridge):
        self._bridge = bridge

    def select_file(self, filter_str: str = "All Files (*)") -> str:
        """打开文件选择对话框，返回文件路径"""
        dialog = QFileDialog()
        dialog.setNameFilter(filter_str)
        dialog.setFileMode(QFileDialog.ExistingFile)
        if dialog.exec():
            files = dialog.selectedFiles()
            return files[0] if files else ""
        return ""

    def save_file(self, default_name: str = "") -> str:
        """打开保存文件对话框，返回保存路径"""
        dialog = QFileDialog()
        dialog.setAcceptMode(QFileDialog.AcceptSave)
        dialog.selectFile(default_name)
        if dialog.exec():
            files = dialog.selectedFiles()
            return files[0] if files else ""
        return ""

    def read_excel_async(self, file_path: str) -> str:
        """异步读取 Excel，返回 task_id，前端通过监听 excelReadComplete 信号获取结果。

        大 Excel 文件（10w+ 行）下不会阻塞 WebChannel JS 线程。
        """
        task_id = uuid.uuid4().hex
        task = _ReadExcelTask(
            task_id,
            file_path,
            callback=lambda result: self._bridge.emit_excel_read_complete(result)
            if self._bridge else None,
        )
        self._pool.start(task)
        return task_id

    def write_excel(self, file_path: str, data: list) -> bool:
        """将数据写入 Excel 文件"""
        if not data:
            return False
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = list(data[0].keys())
            ws.append(headers)
            for row in data:
                ws.append([row.get(h, "") for h in headers])
            wb.save(file_path)
            return True
        except Exception as e:
            print(f"[NativeAPI] write_excel error: {e}")
            return False

    def show_notification(self, message: str):
        """显示系统通知"""
        QMessageBox.information(None, "通知", message)

    def get_app_version(self) -> str:
        """获取桌面壳版本"""
        return "1.0.0"

    def get_app_name(self) -> str:
        """获取应用名称"""
        return "PI Manager"

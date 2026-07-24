# -*- coding: utf-8 -*-
"""
PI Manager - 客户端主窗口
"""
import sys
import os
import threading
import time
import traceback
import concurrent.futures
import urllib.request
import ctypes
import logging
from datetime import datetime, timedelta
from functools import lru_cache

# 配置日志
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger('PI_Client')

# 数据处理 - 使用 openpyxl 替代 pandas
import openpyxl
from openpyxl.utils import get_column_letter

# PySide6 导入
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QLabel, QStackedWidget, QMessageBox, QTableWidget, QTableWidgetItem, QDialog,
    QFormLayout, QLineEdit, QTextEdit, QComboBox, QHeaderView, QAbstractItemView,
    QGridLayout, QCheckBox, QGroupBox, QFileDialog, QProgressDialog, QTabWidget,
    QScrollArea, QDateEdit, QMenu, QFrame, QStatusBar, QSpinBox, QDoubleSpinBox,
    QSizePolicy, QProgressBar
)
from PySide6.QtCore import Qt, QTimer, QDate, QEvent, QThread, QSettings, Signal, QObject
from PySide6.QtGui import (
    QIcon, QPalette, QColor, QFont, QFontDatabase, QBrush, QPixmap, QImage,
    QAction, QPainter
)

# 本地模块（添加到path）
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from cache_manager import cache_manager, CACHE_KEYS, set_user, invalidate_cache
from api.client import ApiClient
from api.cached_client import CachedApiClient
from config import Config
from product_categories import get_parent_category_options, get_child_category_options, get_category_full_path, get_category_name, get_category_options

from dialogs import (
    LoginWindow,
    InvoiceUploadDialog,
    FieldEditDialog,
    OrderEditDialog,
)
from widgets.action_bar import ActionBarFactory
from widgets.order_summary_edit_dialog import OrderSummaryEditDialog
from widgets.order_summary_dialogs import (
    CustomerRequirementDialog,
    CustomerModelDialog,
    CustomerReplyDialog
)
from widgets.export_preview_dialog import ExportPreviewDialog
from widgets.order_summary.order_summary_tab import OrderSummaryTab
from widgets.pi_management.pi_history_tab import PiHistoryTab
from widgets.settings_dialog import SettingsDialog
from services import OrderService

# Web 容器模块（QWebEngineView + QWebChannel）
try:
    from web_container import WebContainerView
    HAS_WEB_CONTAINER = True
except ImportError:
    HAS_WEB_CONTAINER = False
    WebContainerView = None

# TAB_ROUTES 使用延迟导入，避免与 dialogs/pi.py → main 的循环导入冲突

# 测试模块（可选）
try:
    from test_customer_reply import CustomerReplyTester
    HAS_CUSTOMER_REPLY_TEST = True
except ImportError:
    HAS_CUSTOMER_REPLY_TEST = False
    CustomerReplyTester = None

# ============================================
# 全局常量和配置
# ============================================

# 全局线程池（复用，避免重复创建）
_global_thread_pool = concurrent.futures.ThreadPoolExecutor(max_workers=8, thread_name_prefix="pi_manager")

# 图片内存缓存（LRU缓存，最多100张）
_image_cache = {}
_image_cache_lock = threading.Lock()
_MAX_IMAGE_CACHE_SIZE = 100

# 省份编码映射（静态数据）
PROVINCE_CODE_MAP = {
    "北京": "11", "天津": "12", "河北": "13", "山西": "14", "内蒙古": "15",
    "辽宁": "21", "吉林": "22", "黑龙江": "23",
    "上海": "31", "江苏": "32", "浙江": "33", "安徽": "34", "福建": "35", "江西": "36", "山东": "37",
    "河南": "41", "湖北": "42", "湖南": "43", "广东": "44", "广西": "45", "海南": "46",
    "重庆": "50", "四川": "51", "贵州": "52", "云南": "53", "西藏": "54",
    "陕西": "61", "甘肃": "62", "青海": "63", "宁夏": "64", "新疆": "65",
    "台湾": "71", "香港": "81", "澳门": "82"
}

# 城市编码映射（静态数据，模块级别只创建一次）
CITY_CODE_MAP = {
    "11": {"北京": "1"}, "12": {"天津": "1"}, "31": {"上海": "1"}, "50": {"重庆": "1"},
    "13": {"石家庄": "1", "唐山": "2", "秦皇岛": "3", "邯郸": "4", "邢台": "5", "保定": "6", "张家口": "7", "承德": "8", "沧州": "9", "廊坊": "A", "衡水": "B"},
    "14": {"太原": "1", "大同": "2", "阳泉": "3", "长治": "4", "晋城": "5", "朔州": "6", "晋中": "7", "运城": "8", "忻州": "9", "临汾": "A", "吕梁": "B"},
    "15": {"呼和浩特": "1", "包头": "2", "乌海": "3", "赤峰": "4", "通辽": "5", "鄂尔多斯": "6", "呼伦贝尔": "7", "巴彦淖尔": "8", "乌兰察布": "9", "兴安": "A", "锡林郭勒": "B", "阿拉善": "C"},
    "21": {"沈阳": "1", "大连": "2", "鞍山": "3", "抚顺": "4", "本溪": "5", "丹东": "6", "锦州": "7", "营口": "8", "阜新": "9", "辽阳": "A", "盘锦": "B", "铁岭": "C", "朝阳": "D", "葫芦岛": "E"},
    "22": {"长春": "1", "吉林": "2", "四平": "3", "辽源": "4", "通化": "5", "白山": "6", "松原": "7", "白城": "8", "延边": "9"},
    "23": {"哈尔滨": "1", "齐齐哈尔": "2", "鸡西": "3", "鹤岗": "4", "双鸭山": "5", "大庆": "6", "伊春": "7", "佳木斯": "8", "七台河": "9", "牡丹江": "A", "黑河": "B", "绥化": "C", "大兴安岭": "D"},
    "32": {"南京": "1", "无锡": "2", "徐州": "3", "常州": "4", "苏州": "5", "南通": "6", "连云港": "7", "淮安": "8", "盐城": "9", "扬州": "A", "镇江": "B", "泰州": "C", "宿迁": "D"},
    "33": {"杭州": "1", "宁波": "2", "温州": "3", "嘉兴": "4", "湖州": "5", "绍兴": "6", "金华": "7", "衢州": "8", "舟山": "9", "台州": "A", "丽水": "B"},
    "34": {"合肥": "1", "芜湖": "2", "蚌埠": "3", "淮南": "4", "马鞍山": "5", "淮北": "6", "铜陵": "7", "安庆": "8", "黄山": "9", "阜阳": "A", "宿州": "B", "滁州": "C", "六安": "D", "宣城": "E", "池州": "F", "亳州": "G"},
    "35": {"福州": "1", "厦门": "2", "莆田": "3", "三明": "4", "泉州": "5", "漳州": "6", "南平": "7", "龙岩": "8", "宁德": "9"},
    "36": {"南昌": "1", "景德镇": "2", "萍乡": "3", "九江": "4", "新余": "5", "鹰潭": "6", "赣州": "7", "吉安": "8", "宜春": "9", "抚州": "A", "上饶": "B"},
    "37": {"济南": "1", "青岛": "2", "淄博": "3", "枣庄": "4", "东营": "5", "烟台": "6", "潍坊": "7", "济宁": "8", "泰安": "9", "威海": "A", "日照": "B", "临沂": "C", "德州": "D", "滨州": "E", "菏泽": "F"},
    "41": {"郑州": "1", "开封": "2", "洛阳": "3", "平顶山": "4", "安阳": "5", "鹤壁": "6", "新乡": "7", "焦作": "8", "濮阳": "9", "许昌": "A", "漯河": "B", "三门峡": "C", "南阳": "D", "商丘": "E", "信阳": "F", "周口": "G", "驻马店": "H", "济源": "I"},
    "42": {"武汉": "1", "黄石": "2", "十堰": "3", "宜昌": "4", "襄阳": "5", "鄂州": "6", "荆门": "7", "孝感": "8", "荆州": "9", "黄冈": "A", "咸宁": "B", "随州": "C", "恩施": "D", "仙桃": "E", "潜江": "F", "天门": "G"},
    "43": {"长沙": "1", "株洲": "2", "湘潭": "3", "衡阳": "4", "邵阳": "5", "岳阳": "6", "常德": "7", "张家界": "8", "益阳": "9", "郴州": "A", "永州": "B", "怀化": "C", "娄底": "D", "湘西": "E"},
    "44": {"广州": "1", "韶关": "2", "深圳": "3", "珠海": "4", "汕头": "5", "佛山": "6", "江门": "7", "湛江": "8", "茂名": "9", "肇庆": "A", "惠州": "B", "梅州": "C", "汕尾": "D", "河源": "E", "阳江": "F", "清远": "G", "东莞": "H", "中山": "I", "潮州": "J", "揭阳": "K", "云浮": "L"},
    "45": {"南宁": "1", "柳州": "2", "桂林": "3", "梧州": "4", "北海": "5", "防城港": "6", "钦州": "7", "贵港": "8", "玉林": "9", "百色": "A", "贺州": "B", "河池": "C", "来宾": "D", "崇左": "E"},
    "46": {"海口": "1", "三亚": "2", "三沙": "3", "儋州": "4"},
    "51": {"成都": "1", "自贡": "2", "攀枝花": "3", "泸州": "4", "德阳": "5", "绵阳": "6", "广元": "7", "遂宁": "8", "内江": "9", "乐山": "A", "南充": "B", "眉山": "C", "宜宾": "D", "广安": "E", "达州": "F", "雅安": "G", "巴中": "H", "资阳": "I", "阿坝": "J", "甘孜": "K", "凉山": "L"},
    "52": {"贵阳": "1", "六盘水": "2", "遵义": "3", "安顺": "4", "毕节": "5", "铜仁": "6", "黔西南": "7", "黔东南": "8", "黔南": "9"},
    "53": {"昆明": "1", "曲靖": "2", "玉溪": "3", "保山": "4", "昭通": "5", "丽江": "6", "普洱": "7", "临沧": "8", "楚雄": "9", "红河": "A", "文山": "B", "西双版纳": "C", "大理": "D", "怒江": "E", "迪庆": "F"},
    "54": {"拉萨": "1", "日喀则": "2", "昌都": "3", "林芝": "4", "山南": "5", "那曲": "6", "阿里": "7"},
    "61": {"西安": "1", "铜川": "2", "宝鸡": "3", "咸阳": "4", "渭南": "5", "延安": "6", "汉中": "7", "榆林": "8", "安康": "9", "商洛": "A"},
    "62": {"兰州": "1", "嘉峪关": "2", "金昌": "3", "白银": "4", "天水": "5", "武威": "6", "张掖": "7", "平凉": "8", "酒泉": "9", "庆阳": "A", "定西": "B", "陇南": "C", "临夏": "D", "甘南": "E"},
    "63": {"西宁": "1", "海东": "2", "海北": "3", "黄南": "4", "海南": "5", "果洛": "6", "玉树": "7", "海西": "8"},
    "64": {"银川": "1", "石嘴山": "2", "吴忠": "3", "固原": "4", "中卫": "5"},
    "65": {"乌鲁木齐": "1", "克拉玛依": "2", "吐鲁番": "3", "哈密": "4", "昌吉": "5", "博尔塔拉": "6", "巴音郭楞": "7", "阿克苏": "8", "克孜勒苏": "9", "喀什": "A", "和田": "B", "伊犁": "C", "塔城": "D", "阿勒泰": "E", "石河子": "F", "阿拉尔": "G", "图木舒克": "H", "五家渠": "I", "北屯": "J", "铁门关": "K", "双河": "L", "可克达拉": "M", "昆玉": "N", "胡杨河": "O", "新星": "P"},
    "71": {"台北": "1", "高雄": "2", "台中": "3", "台南": "4", "新北": "5", "桃园": "6"},
    "81": {"香港": "1"}, "82": {"澳门": "1"}
}

def get_font(size=10, weight=QFont.Weight.Normal):
    font = QFont()
    font.setPointSize(size)
    font.setWeight(weight)
    # 尝试多种中文字体
    available = QFontDatabase.families()
    for family in ["Microsoft YaHei", "SimHei", "SimSun", "KaiTi", "WenQuanYi Micro Hei", "Heiti SC"]:
        if family in available:
            font.setFamily(family)
            return font
    # 如果没有中文字体，使用默认字体
    font.setFamily(QFont().defaultFamily())
    return font

def set_global_font(app):
    font = get_font(10)
    app.setFont(font)

# 部门配置映射（从配置文件获取）
DEPARTMENT_CONFIG = Config.DEPARTMENT_DB_CONFIG

class ProductItemEditDialog(QDialog):
    """产品项编辑对话框"""
    def __init__(self, item, products, parent=None):
        super().__init__(parent)
        self.item = item.copy()
        self.products = products
        self.setWindowTitle("编辑产品")
        self.setMinimumWidth(500)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # 产品下拉选择（从产品列表读取）
        product_layout = QHBoxLayout()
        product_layout.addWidget(QLabel("选择产品:"))
        self.product_combo = QComboBox()
        self.product_combo.addItem("-- 请选择 --", None)
        for p in self.products:
            product_name = p.get('detail_desc', '') or p.get('name', '')
            self.product_combo.addItem(product_name, p.get('id'))
        product_layout.addWidget(self.product_combo)
        layout.addLayout(product_layout)
        
        # 可编辑字段
        fields = [
            ("客户产品编号", "customer_product_code", ""),
            ("OE号", "oe_number", ""),
            ("产品名称", "product_name", ""),
            ("客户型号", "customer_model", ""),
            ("数量", "quantity", "number"),
            ("单价", "unit_price", "number"),
        ]
        
        self.editors = {}
        form_layout = QFormLayout()
        for label, key, field_type in fields:
            if field_type == "number":
                editor = QDoubleSpinBox()
                editor.setRange(0, 99999999)
                editor.setDecimals(2)
                try:
                    val = float(self.item.get(key, 0) or 0)
                    editor.setValue(val)
                except:
                    editor.setValue(0)
            else:
                editor = QLineEdit(str(self.item.get(key, '')))
                editor.setFixedHeight(30)
            self.editors[key] = editor
            form_layout.addRow(f"{label}:", editor)
        
        layout.addLayout(form_layout)
        
        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        
        save_btn = QPushButton("保存")
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #10b981;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
            }
        """)
        save_btn.clicked.connect(self._on_save)
        btn_layout.addWidget(save_btn)
        
        layout.addLayout(btn_layout)
        
        # 产品选择变化时自动填充
        self.product_combo.currentIndexChanged.connect(self._on_product_selected)
    
    def _on_product_selected(self, index):
        """选择产品后自动填充字段"""
        product_id = self.product_combo.currentData()
        if not product_id:
            return
        
        for p in self.products:
            if p.get('id') == product_id:
                # 自动填充可用的字段
                if not self.editors['customer_product_code'].text():
                    self.editors['customer_product_code'].setText(p.get('product_code', ''))
                if not self.editors['oe_number'].text():
                    # 尝试获取主要OE号
                    oe_list = self.parent().parent().api_client.get_product_oes(product_id) or []
                    primary_oe = next((oe for oe in oe_list if oe.get('is_primary')), None)
                    if primary_oe:
                        self.editors['oe_number'].setText(primary_oe.get('oe_number', ''))
                if not self.editors['product_name'].text():
                    self.editors['product_name'].setText(p.get('detail_desc', '') or p.get('name', ''))
                break
    
    def _on_save(self):
        print("[DEBUG] ProductItemEditDialog._on_save: 开始保存")
        for key, editor in self.editors.items():
            if isinstance(editor, QLineEdit):
                self.item[key] = editor.text()
                print(f"[DEBUG] 字段 {key}: {editor.text()}")
            elif isinstance(editor, QDoubleSpinBox):
                self.item[key] = editor.value()
                print(f"[DEBUG] 字段 {key}: {editor.value()}")
        print(f"[DEBUG] ProductItemEditDialog._on_save: 保存完成, item={self.item}")
        self.accept()
    
    def get_item(self):
        return self.item


class ReplyHistoryDialog(QDialog):
    """客户回复历史记录对话框"""
    def __init__(self, replies, parent=None):
        super().__init__(parent)
        self.replies = replies or []
        self.setWindowTitle("客户回复历史")
        self.setMinimumWidth(600)
        self.setMinimumHeight(400)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        self.reply_table = QTableWidget()
        self.reply_table.setColumnCount(4)
        self.reply_table.setHorizontalHeaderLabels(["时间", "类型", "内容", "状态"])
        
        for reply in self.replies:
            row = self.reply_table.rowCount()
            self.reply_table.insertRow(row)
            self.reply_table.setItem(row, 0, QTableWidgetItem(reply.get('created_at', '')[:19]))
            self.reply_table.setItem(row, 1, QTableWidgetItem(reply.get('reply_type', '')))
            self.reply_table.setItem(row, 2, QTableWidgetItem(reply.get('reply_content', '')))
            self.reply_table.setItem(row, 3, QTableWidgetItem(reply.get('status', '')))
        
        layout.addWidget(self.reply_table)
        
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn)


class SupplierSchemeDialog(QDialog):
    """供应商方案编辑弹窗"""
    def __init__(self, api_client, suppliers, customers, scheme=None, parent=None):
        super().__init__(parent)
        self.api_client = api_client
        self.suppliers = suppliers
        self.customers = customers
        self.scheme = scheme or {}
        self.is_edit = bool(scheme)
        self.setWindowTitle("编辑供应商方案" if self.is_edit else "添加供应商方案")
        self.setFixedSize(650, 700)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        form_layout = QFormLayout()
        form_layout.setSpacing(12)

        # 供应商
        self.supplier_combo = QComboBox()
        self.supplier_combo.setFixedHeight(35)
        self.supplier_combo.addItem("请选择供应商", None)
        for s in self.suppliers:
            self.supplier_combo.addItem(f"{s.get('supplier_code')} - {s.get('supplier_name')}", s.get('id'))
        if self.scheme.get('supplier_id'):
            for i in range(self.supplier_combo.count()):
                if self.supplier_combo.itemData(i) == self.scheme.get('supplier_id'):
                    self.supplier_combo.setCurrentIndex(i)
                    break
        form_layout.addRow("供应商 *:", self.supplier_combo)

        # 方案类型选择
        self.scheme_type_combo = QComboBox()
        self.scheme_type_combo.setFixedHeight(35)
        self.scheme_type_combo.addItem("🏷️ 默认方案（不指定客户）", {'type': 'default', 'customer_id': None})
        self.scheme_type_combo.addItem("👤 指定客户专属方案", {'type': 'customer', 'customer_id': None})
        self.scheme_type_combo.currentIndexChanged.connect(self.on_scheme_type_changed)
        form_layout.addRow("方案类型 *:", self.scheme_type_combo)

        # 客户选择（仅在指定客户时显示）
        self.customer_combo = QComboBox()
        self.customer_combo.setFixedHeight(35)
        self.customer_combo.addItem("请选择客户", None)
        for c in self.customers:
            self.customer_combo.addItem(f"{c.get('customer_code')} - {c.get('customer_name')}", c.get('id'))
        self.customer_combo.setVisible(False)
        
        # 根据编辑数据设置初始值
        if self.is_edit:
            if self.scheme.get('customer_id'):
                self.scheme_type_combo.setCurrentIndex(1)
                self.customer_combo.setVisible(True)
                for i in range(self.customer_combo.count()):
                    if self.customer_combo.itemData(i) == self.scheme.get('customer_id'):
                        self.customer_combo.setCurrentIndex(i)
                        break
            else:
                self.scheme_type_combo.setCurrentIndex(0)
        
        form_layout.addRow("选择客户:", self.customer_combo)

        # 客户产品编号（新建时默认使用父窗口的OE号）
        self.customer_code_input = QLineEdit()
        self.customer_code_input.setPlaceholderText("客户在对方系统中的产品编号")
        # 兼容两种字段名：factory_code（API返回）和 customer_product_code（前端使用）
        default_code = self.scheme.get('factory_code') or self.scheme.get('customer_product_code', '')
        if not default_code and not self.is_edit and self.parent() and hasattr(self.parent(), 'oe_input'):
            default_code = self.parent().oe_input.text().strip()
        self.customer_code_input.setText(default_code)
        form_layout.addRow("客户产品编号:", self.customer_code_input)

        layout.addLayout(form_layout)

        # 价格信息
        price_group = QGroupBox("价格信息")
        price_layout = QGridLayout()
        price_layout.setSpacing(10)

        self.exw_incl_input = QLineEdit()
        self.exw_incl_input.setPlaceholderText("EXW含税价")
        # 兼容两种字段名
        exw_val = self.scheme.get('exw_price_incl') or self.scheme.get('purchase_price', '')
        self.exw_incl_input.setText(str(exw_val or ''))
        price_layout.addWidget(QLabel("EXW含税价:"), 0, 0)
        price_layout.addWidget(self.exw_incl_input, 0, 1)

        self.exw_excl_input = QLineEdit()
        self.exw_excl_input.setPlaceholderText("EXW不含税价")
        self.exw_excl_input.setText(str(self.scheme.get('exw_price_excl', '') or ''))
        price_layout.addWidget(QLabel("EXW不含税价:"), 0, 2)
        price_layout.addWidget(self.exw_excl_input, 0, 3)

        self.fob_incl_input = QLineEdit()
        self.fob_incl_input.setPlaceholderText("FOB含税价")
        self.fob_incl_input.setText(str(self.scheme.get('fob_price_incl', '') or ''))
        price_layout.addWidget(QLabel("FOB含税价:"), 1, 0)
        price_layout.addWidget(self.fob_incl_input, 1, 1)

        self.fob_excl_input = QLineEdit()
        self.fob_excl_input.setPlaceholderText("FOB不含税价")
        self.fob_excl_input.setText(str(self.scheme.get('fob_price_excl', '') or ''))
        price_layout.addWidget(QLabel("FOB不含税价:"), 1, 2)
        price_layout.addWidget(self.fob_excl_input, 1, 3)

        self.freight_input = QLineEdit()
        self.freight_input.setPlaceholderText("运费")
        self.freight_input.setText(str(self.scheme.get('freight', '') or ''))
        price_layout.addWidget(QLabel("运费:"), 2, 0)
        price_layout.addWidget(self.freight_input, 2, 1)

        self.packing_fee_input = QLineEdit()
        self.packing_fee_input.setPlaceholderText("包装费")
        self.packing_fee_input.setText(str(self.scheme.get('packing_fee', '') or ''))
        price_layout.addWidget(QLabel("包装费:"), 2, 2)
        price_layout.addWidget(self.packing_fee_input, 2, 3)

        price_group.setLayout(price_layout)
        layout.addWidget(price_group)

        # 包装尺寸
        size_group = QGroupBox("包装尺寸")
        size_layout = QGridLayout()
        self.carton_length_input = QLineEdit()
        self.carton_length_input.setPlaceholderText("长(cm)")
        self.carton_length_input.setText(str(self.scheme.get('carton_length_cm', '') or ''))
        size_layout.addWidget(QLabel("纸箱长(cm):"), 0, 0)
        size_layout.addWidget(self.carton_length_input, 0, 1)

        self.carton_width_input = QLineEdit()
        self.carton_width_input.setPlaceholderText("宽(cm)")
        self.carton_width_input.setText(str(self.scheme.get('carton_width_cm', '') or ''))
        size_layout.addWidget(QLabel("纸箱宽(cm):"), 0, 2)
        size_layout.addWidget(self.carton_width_input, 0, 3)

        self.carton_height_input = QLineEdit()
        self.carton_height_input.setPlaceholderText("高(cm)")
        self.carton_height_input.setText(str(self.scheme.get('carton_height_cm', '') or ''))
        size_layout.addWidget(QLabel("纸箱高(cm):"), 1, 0)
        size_layout.addWidget(self.carton_height_input, 1, 1)

        self.units_input = QLineEdit()
        self.units_input.setPlaceholderText("每箱数量")
        self.units_input.setText(str(self.scheme.get('units_per_carton', '') or ''))
        size_layout.addWidget(QLabel("每箱数量:"), 1, 2)
        size_layout.addWidget(self.units_input, 1, 3)
        size_group.setLayout(size_layout)
        layout.addWidget(size_group)

        # 重量信息
        weight_group = QGroupBox("重量信息")
        weight_layout = QHBoxLayout()
        self.gross_weight_input = QLineEdit()
        self.gross_weight_input.setPlaceholderText("毛重(kg)")
        self.gross_weight_input.setText(str(self.scheme.get('gross_weight_kg', '') or ''))
        weight_layout.addWidget(QLabel("毛重(kg):"))
        weight_layout.addWidget(self.gross_weight_input)

        self.weight_input = QLineEdit()
        self.weight_input.setPlaceholderText("净重(kg)")
        self.weight_input.setText(str(self.scheme.get('weight_kg', '') or ''))
        weight_layout.addWidget(QLabel("净重(kg):"))
        weight_layout.addWidget(self.weight_input)
        weight_group.setLayout(weight_layout)
        layout.addWidget(weight_group)

        # 备注
        self.remark_input = QTextEdit()
        self.remark_input.setPlaceholderText("备注信息")
        self.remark_input.setText(self.scheme.get('remark', ''))
        self.remark_input.setMaximumHeight(60)
        layout.addWidget(QLabel("备注:"))
        layout.addWidget(self.remark_input)
        
        # 设为默认方案
        self.is_default_checkbox = QCheckBox("设为默认供应商方案（优先使用）")
        self.is_default_checkbox.setStyleSheet("color: #2563eb; font-weight: 500;")
        if self.scheme.get('is_default'):
            self.is_default_checkbox.setChecked(True)
        layout.addWidget(self.is_default_checkbox)

        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        save_btn = QPushButton("保存")
        save_btn.setFixedWidth(100)
        save_btn.clicked.connect(self.save_scheme)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        btn_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.setFixedWidth(100)
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)
        self.setLayout(layout)

    def on_scheme_type_changed(self):
        """方案类型变化时显示/隐藏客户选择"""
        scheme_data = self.scheme_type_combo.currentData()
        if scheme_data and scheme_data.get('type') == 'customer':
            self.customer_combo.setVisible(True)
        else:
            self.customer_combo.setVisible(False)

    def save_scheme(self):
        supplier_id = self.supplier_combo.currentData()
        
        if not supplier_id:
            QMessageBox.warning(self, "警告", "请选择供应商")
            return
        
        # 根据方案类型获取客户ID
        scheme_data = self.scheme_type_combo.currentData()
        customer_id = None
        if scheme_data and scheme_data.get('type') == 'customer':
            customer_id = self.customer_combo.currentData()
            if not customer_id:
                QMessageBox.warning(self, "警告", "请选择客户")
                return

        def try_float(value):
            try:
                return float(value) if value.strip() else None
            except ValueError:
                return None

        supplier_name = self.supplier_combo.currentText()
        customer_name = self.customer_combo.currentText()

        self.scheme_data = {
            "id": self.scheme.get('id') if self.is_edit else None,
            "supplier_id": supplier_id,
            "supplier_name": supplier_name,
            "customer_id": customer_id,
            "customer_name": customer_name,
            "customer_product_code": self.customer_code_input.text().strip(),
            "is_default": self.is_default_checkbox.isChecked(),
            "exw_price_incl": try_float(self.exw_incl_input.text()),
            "exw_price_excl": try_float(self.exw_excl_input.text()),
            "fob_price_incl": try_float(self.fob_incl_input.text()),
            "fob_price_excl": try_float(self.fob_excl_input.text()),
            "freight": try_float(self.freight_input.text()),
            "packing_fee": try_float(self.packing_fee_input.text()),
            "carton_length_cm": try_float(self.carton_length_input.text()),
            "carton_width_cm": try_float(self.carton_width_input.text()),
            "carton_height_cm": try_float(self.carton_height_input.text()),
            "units_per_carton": int(self.units_input.text()) if self.units_input.text().strip() else None,
            "gross_weight_kg": try_float(self.gross_weight_input.text()),
            "weight_kg": try_float(self.weight_input.text()),
            "remark": self.remark_input.toPlainText()
        }
        print(f"DEBUG - save_scheme: scheme_data = {self.scheme_data}")
        self.accept()

    def get_scheme_data(self):
        return getattr(self, 'scheme_data', None)


class CustomerDialog(QDialog):
    def __init__(self, api_client: ApiClient, customer=None):
        super().__init__()
        self.api_client = api_client
        self.customer = customer
        self.is_edit = customer is not None
        self.init_ui()
        if self.is_edit:
            self.load_contacts()

    def init_ui(self):
        self.setWindowTitle("编辑客户" if self.is_edit else "新增客户")
        self.setMinimumSize(750, 600)

        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        form_layout = QFormLayout()
        form_layout.setSpacing(12)

        self.dept_combo = QComboBox()
        self.dept_combo.addItems([
            "S - 索英普",
            "W - 维那",
            "M - 马迪那",
            "D - 银达"
        ])
        if self.customer:
            dept_text_map = {"S": "S - 索英普", "W": "W - 维那", "M": "M - 马迪那", "D": "D - 银达"}
            saved_dept = self.customer.get('dept_id', 'S')
            self.dept_combo.setCurrentText(dept_text_map.get(saved_dept, "S - 索英普"))
        form_layout.addRow("部门:", self.dept_combo)

        if self.is_edit:
            self.code_label = QLabel(self.customer.get('customer_code', ''))
            form_layout.addRow("客户编号:", self.code_label)
        
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("请输入客户名称")
        if self.customer:
            self.name_input.setText(self.customer.get('customer_name', ''))
        form_layout.addRow("客户名称 *:", self.name_input)

        self.country_input = QLineEdit()
        self.country_input.setPlaceholderText("请输入所在国家")
        if self.customer:
            self.country_input.setText(self.customer.get('country', ''))
        form_layout.addRow("所在国家 *:", self.country_input)

        self.basic_require_input = QTextEdit()
        self.basic_require_input.setPlaceholderText("请输入通用交易条款")
        self.basic_require_input.setMaximumHeight(60)
        if self.customer:
            self.basic_require_input.setText(self.customer.get('basic_require', ''))
        form_layout.addRow("基本要求:", self.basic_require_input)

        self.special_input = QTextEdit()
        self.special_input.setPlaceholderText("请输入特殊要求，如特定包装、标签等")
        self.special_input.setMaximumHeight(60)
        if self.customer:
            self.special_input.setText(self.customer.get('special_require', ''))
        form_layout.addRow("特殊要求:", self.special_input)

        self.payment_input = QLineEdit()
        self.payment_input.setPlaceholderText("如 T/T 30天")
        if self.customer:
            self.payment_input.setText(self.customer.get('payment_terms', ''))
        form_layout.addRow("付款条款:", self.payment_input)

        layout.addLayout(form_layout)

        contacts_group = QGroupBox("联系人信息")
        contacts_layout = QVBoxLayout()
        contacts_layout.setSpacing(5)

        self.contacts_table = QTableWidget()
        self.contacts_table.setColumnCount(4)
        self.contacts_table.setHorizontalHeaderLabels(["姓名", "电话", "邮箱", "职位"])
        self.contacts_table.setMaximumHeight(150)
        self.contacts_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.contacts_table.setAlternatingRowColors(True)
        header = self.contacts_table.horizontalHeader()
        for i in range(4):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
        contacts_layout.addWidget(self.contacts_table)

        btn_row = QHBoxLayout()
        add_btn = QPushButton("+ 添加联系人")
        add_btn.setFixedWidth(100)
        add_btn.clicked.connect(self.add_contact_row)
        remove_btn = QPushButton("- 删除选中")
        remove_btn.setFixedWidth(100)
        remove_btn.clicked.connect(self.remove_selected_contact)
        btn_row.addWidget(add_btn)
        btn_row.addWidget(remove_btn)
        btn_row.addStretch()
        contacts_layout.addLayout(btn_row)

        contacts_group.setLayout(contacts_layout)
        layout.addWidget(contacts_group)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        save_btn = QPushButton("保存")
        save_btn.setFixedWidth(100)
        save_btn.clicked.connect(self.save_customer)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        btn_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.setFixedWidth(100)
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)
        self.setLayout(layout)

    def load_contacts(self):
        if not self.customer:
            return
        try:
            contacts = self.api_client.get_customer_contacts(self.customer['id'])
            self.populate_contacts_table(contacts or [])
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载联系人失败: {str(e)}")

    def populate_contacts_table(self, contacts):
        self.contacts_table.setRowCount(len(contacts))
        for row, contact in enumerate(contacts):
            self.contacts_table.setItem(row, 0, QTableWidgetItem(contact.get('name', '')))
            self.contacts_table.setItem(row, 1, QTableWidgetItem(contact.get('phone', '')))
            self.contacts_table.setItem(row, 2, QTableWidgetItem(contact.get('email', '')))
            self.contacts_table.setItem(row, 3, QTableWidgetItem(contact.get('position', '')))

    def add_contact_row(self):
        row = self.contacts_table.rowCount()
        self.contacts_table.insertRow(row)
        for col in range(4):
            self.contacts_table.setItem(row, col, QTableWidgetItem(""))

    def remove_selected_contact(self):
        current_row = self.contacts_table.currentRow()
        if current_row >= 0:
            self.contacts_table.removeRow(current_row)

    def save_customer(self):
        if not self.name_input.text().strip():
            QMessageBox.warning(self, "提示", "请输入客户名称")
            return
        if not self.country_input.text().strip():
            QMessageBox.warning(self, "提示", "请输入所在国家")
            return

        dept_map = {"S - 索英普": "S", "W - 维那": "W", "M - 马迪那": "M", "D - 银达": "D"}
        dept_id = dept_map.get(self.dept_combo.currentText(), "S")

        data = {
            "dept_id": dept_id,
            "customer_name": self.name_input.text().strip(),
            "country": self.country_input.text().strip(),
            "basic_require": self.basic_require_input.toPlainText().strip(),
            "special_require": self.special_input.toPlainText().strip(),
            "payment_terms": self.payment_input.text().strip()
        }

        try:
            if self.is_edit:
                result = self.api_client.update_customer(self.customer['id'], data)
                self.save_contacts(self.customer['id'])
            else:
                result = self.api_client.create_customer(data)
                if result and 'id' in result:
                    self.save_contacts(result['id'])
            invalidate_cache("customers")
            self.accept()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存失败: {str(e)}")

    def save_contacts(self, customer_id):
        contacts_to_save = []
        for row in range(self.contacts_table.rowCount()):
            name = self.contacts_table.item(row, 0).text().strip()
            phone = self.contacts_table.item(row, 1).text().strip()
            email = self.contacts_table.item(row, 2).text().strip()
            position = self.contacts_table.item(row, 3).text().strip()

            if name or phone or email:
                contacts_to_save.append({
                    "name": name,
                    "phone": phone,
                    "email": email,
                    "position": position,
                    "is_primary": 1 if row == 0 else 0
                })

        try:
            old_contacts = self.api_client.get_customer_contacts(customer_id)
            for contact in old_contacts:
                self.api_client.delete_customer_contact(customer_id, contact['id'])
        except Exception:
            pass

        for contact_data in contacts_to_save:
            try:
                self.api_client.create_customer_contact(customer_id, contact_data)
            except Exception as e:
                print(f"创建联系人失败: {e}")


class CustomerDetailDialog(QDialog):
    def __init__(self, api_client: ApiClient, customer):
        super().__init__()
        self.api_client = api_client
        self.customer = customer
        self.addresses = []
        self.contacts = []
        self.pi_orders = []
        self.init_ui()
        self.load_data()

    def init_ui(self):
        self.setWindowTitle(f"客户详情 - {self.customer.get('customer_name', '')}")
        self.setMinimumSize(800, 600)

        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        # 标签页
        self.tab_widget = QTabWidget()
        
        # 基本信息页
        self.basic_tab = QWidget()
        self.setup_basic_tab()
        
        # 收货地址页
        self.address_tab = QWidget()
        self.setup_address_tab()
        
        # 联系人页
        self.contact_tab = QWidget()
        self.setup_contact_tab()
        
        # PI订单历史页
        self.pi_tab = QWidget()
        self.setup_pi_tab()

        self.tab_widget.addTab(self.basic_tab, "基本信息")
        self.tab_widget.addTab(self.address_tab, "收货地址")
        self.tab_widget.addTab(self.contact_tab, "联系人")
        self.tab_widget.addTab(self.pi_tab, "交易历史")

        layout.addWidget(self.tab_widget)

        # 关闭按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        close_btn = QPushButton("关闭")
        close_btn.setFixedWidth(100)
        close_btn.clicked.connect(self.close)
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)

        self.setLayout(layout)

    def setup_basic_tab(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(12)

        form_layout = QFormLayout()
        form_layout.setSpacing(10)

        form_layout.addRow(QLabel("<b>客户编号:</b>"), QLabel(self.customer.get('customer_code', '')))
        form_layout.addRow(QLabel("<b>客户名称:</b>"), QLabel(self.customer.get('customer_name', '')))
        form_layout.addRow(QLabel("<b>所属部门:</b>"), QLabel(self.customer.get('dept_id', '')))
        form_layout.addRow(QLabel("<b>所在国家:</b>"), QLabel(self.customer.get('country', '')))
        
        basic_require = self.customer.get('basic_require', '')
        form_layout.addRow(QLabel("<b>基本要求:</b>"), QLabel(basic_require if basic_require else "-"))
        
        form_layout.addRow(QLabel("<b>付款条款:</b>"), QLabel(self.customer.get('payment_terms', '') or "-"))
        
        status = self.customer.get('status', 1)
        status_text = "启用" if status == 1 else "禁用"
        status_color = "#10b981" if status == 1 else "#ef4444"
        status_label = QLabel(status_text)
        status_label.setStyleSheet(f"color: {status_color}; font-weight: bold;")
        form_layout.addRow(QLabel("<b>状态:</b>"), status_label)

        layout.addLayout(form_layout)
        layout.addStretch()

        special_require = self.customer.get('special_require', '')
        if special_require:
            special_group = QGroupBox("特殊要求")
            special_group.setStyleSheet("QGroupBox { font-weight: bold; border: 2px solid #dc2626; border-radius: 5px; }")
            special_layout = QVBoxLayout()
            special_label = QLabel(special_require)
            special_label.setWordWrap(True)
            special_label.setStyleSheet("color: #dc2626; padding: 5px;")
            special_layout.addWidget(special_label)
            special_group.setLayout(special_layout)
            layout.addWidget(special_group)

        self.basic_tab.setLayout(layout)

    def setup_address_tab(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(10, 10, 10, 10)

        # 工具栏
        toolbar = QHBoxLayout()
        add_btn = QPushButton("+ 添加地址")
        add_btn.clicked.connect(self.add_address)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 6px 12px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        toolbar.addWidget(add_btn)
        toolbar.addStretch()
        layout.addLayout(toolbar)

        # 地址列表
        self.addresses_table = QTableWidget()
        self.addresses_table.setColumnCount(6)
        self.addresses_table.setHorizontalHeaderLabels(["国家", "港口", "详细地址", "默认地址", "编辑", "删除"])
        self.addresses_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.addresses_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        layout.addWidget(self.addresses_table)

        self.address_tab.setLayout(layout)

    def setup_contact_tab(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(10, 10, 10, 10)

        # 工具栏
        toolbar = QHBoxLayout()
        toolbar.addStretch()
        
        add_btn = QPushButton("+ 新增联系人")
        add_btn.clicked.connect(self.add_contact)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        toolbar.addWidget(add_btn)
        
        layout.addLayout(toolbar)

        self.contacts_table = QTableWidget()
        self.contacts_table.setColumnCount(7)
        self.contacts_table.setHorizontalHeaderLabels(["姓名", "职位", "电话", "邮箱", "是否主要", "编辑", "删除"])
        self.contacts_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.contacts_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        layout.addWidget(self.contacts_table)

        self.contact_tab.setLayout(layout)

    def setup_pi_tab(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(10, 10, 10, 10)

        self.pi_table = QTableWidget()
        self.pi_table.setColumnCount(10)
        self.pi_table.setHorizontalHeaderLabels(["", "ID", "PI号", "金额", "币种", "状态", "创建时间", "操作", "完成", "导出"])
        self.pi_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.pi_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.pi_table.setColumnWidth(0, 40)
        self.pi_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.pi_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.pi_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self.pi_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        self.pi_table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        self.pi_table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        self.pi_table.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeMode.ResizeToContents)
        self.pi_table.horizontalHeader().setSectionResizeMode(8, QHeaderView.ResizeMode.ResizeToContents)
        self.pi_table.horizontalHeader().setSectionResizeMode(9, QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.pi_table)

        self.pi_tab.setLayout(layout)

    def load_data(self):
        try:
            # 加载地址
            self.addresses = self.api_client.get_customer_addresses(self.customer['id'])
            self.load_addresses_table()

            # 加载联系人
            self.contacts = self.api_client.get_customer_contacts(self.customer['id'])
            self.load_contacts_table()

            # 加载PI订单
            self.pi_orders = self.api_client.get_customer_pi_list(self.customer['id'])
            self.load_pi_table()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载数据失败: {str(e)}")

    def load_addresses_table(self):
        self.addresses_table.setRowCount(len(self.addresses))
        for row, addr in enumerate(self.addresses):
            self.addresses_table.setItem(row, 0, QTableWidgetItem(addr.get('country', '')))
            self.addresses_table.setItem(row, 1, QTableWidgetItem(addr.get('port', '')))
            self.addresses_table.setItem(row, 2, QTableWidgetItem(addr.get('address_detail', '')))
            
            is_default = addr.get('is_default', 0)
            default_text = "是" if is_default == 1 else "否"
            self.addresses_table.setItem(row, 3, QTableWidgetItem(default_text))

            # 编辑按钮
            edit_btn = QPushButton("编辑")
            edit_btn.setFixedWidth(50)
            edit_btn.setStyleSheet("""
                QPushButton {
                    background-color: #3b82f6;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 4px 8px;
                }
                QPushButton:hover { background-color: #2563eb; }
            """)
            edit_btn.clicked.connect(lambda _, addr=addr: self.edit_address(addr))
            self.addresses_table.setCellWidget(row, 4, edit_btn)
            
            # 删除按钮
            delete_btn = QPushButton("删除")
            delete_btn.setFixedWidth(50)
            delete_btn.setStyleSheet("""
                QPushButton {
                    background-color: #ef4444;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 4px 8px;
                }
                QPushButton:hover { background-color: #dc2626; }
            """)
            delete_btn.clicked.connect(lambda _, addr=addr: self.delete_address(addr))
            self.addresses_table.setCellWidget(row, 5, delete_btn)

    def load_contacts_table(self):
        self.contacts_table.setRowCount(len(self.contacts))
        for row, contact in enumerate(self.contacts):
            self.contacts_table.setItem(row, 0, QTableWidgetItem(contact.get('name', '')))
            self.contacts_table.setItem(row, 1, QTableWidgetItem(contact.get('position', '')))
            self.contacts_table.setItem(row, 2, QTableWidgetItem(contact.get('phone', '')))
            self.contacts_table.setItem(row, 3, QTableWidgetItem(contact.get('email', '')))
            
            is_primary = contact.get('is_primary', 0)
            primary_text = "是" if is_primary == 1 else "否"
            self.contacts_table.setItem(row, 4, QTableWidgetItem(primary_text))

            # 操作按钮
            btn_layout = QHBoxLayout()
            
            # 编辑按钮
            edit_btn = QPushButton("编辑")
            edit_btn.setFixedWidth(50)
            edit_btn.setStyleSheet("""
                QPushButton {
                    background-color: #3b82f6;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 4px 8px;
                }
                QPushButton:hover { background-color: #2563eb; }
            """)
            edit_btn.clicked.connect(lambda _, c=contact: self.edit_contact(c))
            self.contacts_table.setCellWidget(row, 5, edit_btn)
            
            # 删除按钮
            delete_btn = QPushButton("删除")
            delete_btn.setFixedWidth(50)
            delete_btn.setStyleSheet("""
                QPushButton {
                    background-color: #ef4444;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 4px 8px;
                }
                QPushButton:hover { background-color: #dc2626; }
            """)
            delete_btn.clicked.connect(lambda _, c=contact: self.delete_contact(c))
            self.contacts_table.setCellWidget(row, 6, delete_btn)

    def load_pi_table(self):
        self.pi_table.setRowCount(len(self.pi_orders))
        for row, pi in enumerate(self.pi_orders):
            self.pi_table.setItem(row, 0, QTableWidgetItem(pi.get('pi_number', '')))
            self.pi_table.setItem(row, 1, QTableWidgetItem(str(pi.get('total_amount', ''))))
            
            status = pi.get('status', '')
            self.pi_table.setItem(row, 2, QTableWidgetItem(status))
            
            created_at = pi.get('created_at', '')
            if created_at:
                created_at = created_at[:19] if isinstance(created_at, str) else str(created_at)
            self.pi_table.setItem(row, 3, QTableWidgetItem(created_at))

    def add_address(self):
        dialog = AddressDialog(self.api_client, customer_id=self.customer['id'])
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_data()

    def edit_address(self, address):
        dialog = AddressDialog(self.api_client, customer_id=self.customer['id'], address=address)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_data()

    def delete_address(self, address):
        reply = QMessageBox.question(self, "确认删除", "确定要删除这个地址吗？",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.api_client.delete_customer_address(self.customer['id'], address['id'])
                self.load_data()
            except Exception as e:
                QMessageBox.warning(self, "错误", f"删除失败：{str(e)}")

    def add_contact(self):
        dialog = ContactDialog(self.api_client, customer_id=self.customer['id'])
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_data()

    def edit_contact(self, contact):
        dialog = ContactDialog(self.api_client, customer_id=self.customer['id'], contact=contact)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_data()

    def delete_contact(self, contact):
        reply = QMessageBox.question(self, "确认删除", "确定要删除这个联系人吗？",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.api_client.delete_customer_contact(self.customer['id'], contact['id'])
                self.load_data()
            except Exception as e:
                QMessageBox.warning(self, "错误", f"删除失败：{str(e)}")


class AddressDialog(QDialog):
    def __init__(self, api_client: ApiClient, customer_id, address=None):
        super().__init__()
        self.api_client = api_client
        self.customer_id = customer_id
        self.address = address
        self.is_edit = address is not None
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("编辑地址" if self.is_edit else "添加地址")
        self.setFixedSize(400, 300)

        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        form_layout = QFormLayout()
        form_layout.setSpacing(12)

        self.country_input = QLineEdit()
        if self.address:
            self.country_input.setText(self.address.get('country', ''))
        form_layout.addRow("国家:", self.country_input)

        self.port_input = QLineEdit()
        if self.address:
            self.port_input.setText(self.address.get('port', ''))
        form_layout.addRow("港口:", self.port_input)

        self.detail_input = QTextEdit()
        if self.address:
            self.detail_input.setText(self.address.get('address_detail', ''))
        self.detail_input.setMaximumHeight(80)
        form_layout.addRow("详细地址:", self.detail_input)

        self.default_checkbox = QCheckBox("设为默认地址")
        if self.address and self.address.get('is_default', 0) == 1:
            self.default_checkbox.setChecked(True)
        form_layout.addRow("", self.default_checkbox)

        layout.addLayout(form_layout)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        save_btn = QPushButton("保存")
        save_btn.setFixedWidth(100)
        save_btn.clicked.connect(self.save_address)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        btn_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.setFixedWidth(100)
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)
        self.setLayout(layout)

    def save_address(self):
        data = {
            "country": self.country_input.text().strip(),
            "port": self.port_input.text().strip(),
            "address_detail": self.detail_input.toPlainText().strip(),
            "is_default": 1 if self.default_checkbox.isChecked() else 0
        }

        try:
            if self.is_edit:
                self.api_client.update_customer_address(self.customer_id, self.address['id'], data)
            else:
                self.api_client.create_customer_address(self.customer_id, data)
            self.accept()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存失败：{str(e)}")


class ContactDialog(QDialog):
    def __init__(self, api_client: ApiClient, customer_id, contact=None):
        super().__init__()
        self.api_client = api_client
        self.customer_id = customer_id
        self.contact = contact
        self.is_edit = contact is not None
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("编辑联系人" if self.is_edit else "添加联系人")
        self.setFixedSize(400, 350)

        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        form_layout = QFormLayout()
        form_layout.setSpacing(12)

        self.name_input = QLineEdit()
        if self.contact:
            self.name_input.setText(self.contact.get('name', ''))
        form_layout.addRow("姓名:", self.name_input)

        self.position_input = QLineEdit()
        if self.contact:
            self.position_input.setText(self.contact.get('position', ''))
        form_layout.addRow("职位:", self.position_input)

        self.phone_input = QLineEdit()
        if self.contact:
            self.phone_input.setText(self.contact.get('phone', ''))
        form_layout.addRow("电话:", self.phone_input)

        self.email_input = QLineEdit()
        if self.contact:
            self.email_input.setText(self.contact.get('email', ''))
        form_layout.addRow("邮箱:", self.email_input)

        self.primary_checkbox = QCheckBox("设为主要联系人")
        if self.contact and self.contact.get('is_primary', 0) == 1:
            self.primary_checkbox.setChecked(True)
        form_layout.addRow("", self.primary_checkbox)

        layout.addLayout(form_layout)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        save_btn = QPushButton("保存")
        save_btn.setFixedWidth(100)
        save_btn.clicked.connect(self.save_contact)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        btn_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.setFixedWidth(100)
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)
        self.setLayout(layout)

    def save_contact(self):
        if not self.name_input.text().strip():
            QMessageBox.warning(self, "提示", "请输入联系人姓名")
            return

        data = {
            "name": self.name_input.text().strip(),
            "position": self.position_input.text().strip(),
            "phone": self.phone_input.text().strip(),
            "email": self.email_input.text().strip(),
            "is_primary": 1 if self.primary_checkbox.isChecked() else 0
        }

        try:
            if self.is_edit:
                self.api_client.update_customer_contact(self.customer_id, self.contact['id'], data)
            else:
                self.api_client.create_customer_contact(self.customer_id, data)
            self.accept()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存失败：{str(e)}")


class SupplierDialog(QDialog):
    def __init__(self, api_client: ApiClient, supplier=None):
        super().__init__()
        self.api_client = api_client
        self.supplier = supplier
        self.is_edit = supplier is not None
        self.provinces = []
        self.cities = []
        self.selected_city_code = ""
        self.init_ui()
        QTimer.singleShot(0, self.load_provinces)

    def load_provinces(self):
        try:
            self.provinces = self.api_client.get_provinces()
            self.province_combo.clear()
            self.province_combo.addItems(self.provinces)
            if self.supplier and self.supplier.get('region'):
                region = self.supplier.get('region', '')
                for prov in self.provinces:
                    if region.startswith(prov):
                        self.province_combo.setCurrentText(prov)
                        self.load_cities(prov)
                        city_name = region[len(prov):].strip()
                        if city_name and city_name in self.cities:
                            self.city_combo.setCurrentText(city_name)
                        break
        except Exception as e:
            print(f"加载省份失败: {e}")

    def load_cities(self, province):
        try:
            self.cities = self.api_client.get_cities(province)
            self.city_combo.clear()
            self.city_combo.addItems(self.cities)
        except Exception as e:
            print(f"加载城市失败: {e}")

    def on_province_changed(self, province):
        self.load_cities(province)

    def on_city_changed(self, city):
        province = self.province_combo.currentText()
        try:
            # 使用模块级别的静态映射（只创建一次）
            p_code = PROVINCE_CODE_MAP.get(province, "")
            c_map = CITY_CODE_MAP.get(p_code, {})
            self.selected_city_code = p_code + c_map.get(city, "0")
        except Exception as e:
            print(f"获取城市编码失败: {e}")

    def init_ui(self):
        self.setWindowTitle("编辑供应商" if self.is_edit else "新增供应商")
        self.setFixedSize(500, 480)

        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        form_layout = QFormLayout()
        form_layout.setSpacing(12)

        if self.is_edit:
            self.code_label = QLabel(self.supplier.get('supplier_code', ''))
            form_layout.addRow("供应商编号:", self.code_label)

        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("请输入供应商名称")
        if self.supplier:
            self.name_input.setText(self.supplier.get('supplier_name', ''))
        form_layout.addRow("供应商名称:", self.name_input)

        province_layout = QHBoxLayout()
        self.province_combo = QComboBox()
        self.province_combo.setFixedHeight(35)
        self.province_combo.currentTextChanged.connect(self.on_province_changed)
        province_layout.addWidget(self.province_combo)

        self.city_combo = QComboBox()
        self.city_combo.setFixedHeight(35)
        self.city_combo.currentTextChanged.connect(self.on_city_changed)
        province_layout.addWidget(self.city_combo)
        form_layout.addRow("省份/城市:", province_layout)

        self.contact_input = QLineEdit()
        self.contact_input.setPlaceholderText("请输入联系人")
        if self.supplier:
            self.contact_input.setText(self.supplier.get('contact_person', ''))
        form_layout.addRow("联系人:", self.contact_input)

        self.phone_input = QLineEdit()
        self.phone_input.setPlaceholderText("请输入联系电话")
        if self.supplier:
            self.phone_input.setText(self.supplier.get('phone', ''))
        form_layout.addRow("联系电话:", self.phone_input)

        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("请输入邮箱地址")
        if self.supplier:
            self.email_input.setText(self.supplier.get('email', ''))
        form_layout.addRow("邮箱:", self.email_input)

        self.address_input = QTextEdit()
        self.address_input.setPlaceholderText("请输入详细地址")
        if self.supplier:
            self.address_input.setText(self.supplier.get('address', ''))
        self.address_input.setMaximumHeight(80)
        form_layout.addRow("详细地址:", self.address_input)

        layout.addLayout(form_layout)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        save_btn = QPushButton("保存")
        save_btn.setFixedWidth(100)
        save_btn.clicked.connect(self.save_supplier)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        btn_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.setFixedWidth(100)
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)
        self.setLayout(layout)

    def save_supplier(self):
        if not self.name_input.text().strip():
            QMessageBox.warning(self, "提示", "请输入供应商名称")
            return

        province = self.province_combo.currentText()
        city = self.city_combo.currentText()
        region = f"{province} {city}" if province and city else ""

        data = {
            "supplier_name": self.name_input.text().strip(),
            "province": province,
            "city": city,
            "city_code": self.selected_city_code,
            "region": region,
            "contact_person": self.contact_input.text().strip(),
            "phone": self.phone_input.text().strip(),
            "email": self.email_input.text().strip(),
            "address": self.address_input.toPlainText().strip()
        }

        try:
            if self.is_edit:
                self.api_client.update_supplier(self.supplier['id'], data)
            else:
                self.api_client.create_supplier(data)
            self.accept()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存失败: {str(e)}")


class MainWindow(QMainWindow):

    def __init__(self, api_client: ApiClient, dept_id: str, index_path: str = None):
        super().__init__()
        self.api_client = api_client
        self.dept_id = dept_id
        self.index_path = index_path
        # 全局变量
        self.default_profit_margin = 25.0  # 默认毛利率
        self.exchange_rate = 7.24          # 默认汇率
        self.server_version = self._fetch_server_version()
        self.init_ui()
        self.load_globals()
        self.load_data()

    def _fetch_server_version(self) -> str:
        """获取服务端版本号
        🔧 2026-06-29 修复：api_client.get() 会自动在 endpoint 前添加 /api/ 前缀，
        原调用 self.api_client.get("/api/version") 会导致请求 URL 变成 /api/api/version → 404。
        改为不带 /api/ 前缀的端点（"version"），由 _build_url 拼成 /api/version。
        """
        try:
            resp = self.api_client.get("version")
            if resp:
                return resp.get("version", "未知")
        except Exception:
            pass
        return "未知"

    def _check_update_async(self):
        """
        🔧 2026-06-29 修复：
        - A4: 新增 min_compatible 和 is_blocked 返回值处理
          当 is_blocked=True 时，显示阻塞消息并强制引导用户升级
        """
        from config import Config

        result = check_for_updates()
        has_update, latest_ver, changelog, force_update, download_url, sha256_url, min_compatible, is_blocked = result

        if is_blocked:
            # A4 修复：低于最低兼容版本，强制阻塞并引导升级
            QApplication.instance().processEvents()
            block_msg = (
                f"⚠️ 版本兼容检查失败\n\n"
                f"当前版本: {Config.APP_VERSION}\n"
                f"最低兼容版本: {min_compatible}\n\n"
                f"{changelog or '请下载最新版本后再启动程序。'}"
            )
            reply = QMessageBox.critical(
                self,
                "版本过低，无法启动",
                block_msg,
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )
            if reply == QMessageBox.Yes and download_url:
                success = download_and_install_update(
                    download_url, sha256_url, sys.executable, latest_ver
                )
                if success:
                    QMessageBox.information(self, "提示", "正在安装更新，请稍候...")
                    QApplication.instance().quit()
                else:
                    QMessageBox.warning(self, "提示", "下载失败，请手动从更新服务下载最新版本。")
            else:
                QMessageBox.warning(self, "提示", "程序将退出，请升级后再启动。")
                QApplication.instance().quit()
            return

        if has_update:
            QApplication.instance().processEvents()
            update_dialog = UpdateDialog(
                Config.APP_VERSION,
                latest_ver,
                changelog,
                force_update
            )
            update_dialog.exec()
            if update_dialog.was_update_accepted() and download_url:
                # A6 修复：URL 拼接已在 check_for_updates() 返回时统一处理，此处不再重复拼接
                success = download_and_install_update(
                    download_url,
                    sha256_url,
                    sys.executable,
                    latest_ver
                )
                if success:
                    QMessageBox.information(self, "提示", "正在安装更新，请稍候...")
                    QApplication.instance().quit()
                else:
                    QMessageBox.warning(self, "更新失败", "下载或安装更新包失败，请手动下载安装。")

    @staticmethod
    def _is_not_none(value):
        """
        检查值是否不为None或NaN（替代 pd.notna）

        Args:
            value: 要检查的值

        Returns:
            bool: 如果值不是 None、空字符串或 NaN，返回 True
        """
        if value is None:
            return False
        if isinstance(value, float) and str(value) == 'nan':
            return False
        if isinstance(value, str) and value.strip() == '':
            return False
        return True
    
    def load_globals(self):
        """加载全局变量（使用本地配置，无网络延迟）

        🔧 2026-06-29 优化：同步预加载客户列表缓存。
        """
        try:
            from config.local_settings_manager import load_local_settings
            settings = load_local_settings()
            self.default_profit_margin = settings.get('default_profit_margin', 25.0)
            self.exchange_rate = settings.get('exchange_rate', 7.24)
            print(f"[INFO] 全局变量加载: 毛利率={self.default_profit_margin}%, 汇率={self.exchange_rate}")
        except Exception as e:
            print(f"[WARN] 加载全局变量失败，使用默认值: {e}")
            self.default_profit_margin = 25.0
            self.exchange_rate = 7.24

        # 🔧 2026-06-29 同步预加载客户列表缓存
        try:
            from api.cached_client import CachedApiClient
            if isinstance(self.api_client, CachedApiClient):
                # 触发缓存预加载（会立即返回缓存或发起 API 请求）
                self.api_client.get_customers()
                print(f"[INFO] 客户列表缓存预加载完成")
        except Exception as e:
            print(f"[WARN] 客户列表预加载失败: {e}")
    
    def calculate_estimated_usd_price(self, factory_rmb_price):
        """计算预估美金报价
        公式: 预估美金报价 = 工厂人民币价格 × (1 + 毛利率) / 汇率
        """
        if not factory_rmb_price or factory_rmb_price <= 0:
            return 0
        margin_factor = 1 + (self.default_profit_margin / 100)
        return factory_rmb_price * margin_factor / self.exchange_rate
    
    def calculate_order_profit_margin(self, customer_usd_price, total_rmb_amount, exchange_rate=None):
        """计算订单预估毛利率
        公式: 预估毛利率 = 客户美金报价 × 汇率 / 总金额
        注意: 需要将总金额转为美金计算毛利率
        """
        if not exchange_rate:
            exchange_rate = self.exchange_rate
        if not customer_usd_price or not total_rmb_amount or total_rmb_amount <= 0:
            return 0
        # 客户总美金 = 客户美金报价 × 汇率（折算成人民币）
        customer_total_rmb = customer_usd_price * exchange_rate
        # 毛利率 = (客户人民币 - 成本人民币) / 客户人民币
        if customer_total_rmb <= 0:
            return 0
        profit_margin = (customer_total_rmb - total_rmb_amount) / customer_total_rmb * 100
        return max(0, profit_margin)  # 不返回负数

    def init_ui(self):
        from config import Config
        client_ver = Config.APP_VERSION.lstrip('v')
        self.setWindowTitle(f"PI订单管理系统 [{client_ver}] [Server {self.server_version}] - {DEPARTMENT_CONFIG[self.dept_id]['name']}")
        self.setMinimumSize(1200, 800)
        # 默认全屏显示
        self.showMaximized()

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)

        header = QWidget()
        header.setFixedHeight(70)
        header.setStyleSheet("background-color: #2563eb;")
        header_layout = QHBoxLayout()
        header_layout.setContentsMargins(20, 0, 20, 0)

        title = QLabel(f"📦 PI订单管理系统 [{client_ver}] [Server {self.server_version}] - {DEPARTMENT_CONFIG[self.dept_id]['name']}")
        title.setFont(get_font(16, QFont.Weight.Bold))
        title.setStyleSheet("color: white;")
        header_layout.addWidget(title)

        header_layout.addStretch()

        # 用户信息和管理员模式切换
        self.user_info_label = QLabel()
        self.user_info_label.setStyleSheet("color: white; font-size: 14px;")
        header_layout.addWidget(self.user_info_label)
        
        self.admin_mode_label = QLabel()
        self.admin_mode_label.setStyleSheet("color: #fbbf24; font-size: 12px; font-weight: bold;")
        header_layout.addWidget(self.admin_mode_label)
        
        # 退出登录按钮
        logout_btn = QPushButton("退出")
        logout_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc2626;
                color: white;
                border: none;
                padding: 5px 15px;
                border-radius: 4px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #b91c1c;
            }
        """)
        logout_btn.clicked.connect(self.logout)
        header_layout.addWidget(logout_btn)

        header.setLayout(header_layout)
        main_layout.addWidget(header)

        content = QWidget()
        content_layout = QHBoxLayout()
        content_layout.setContentsMargins(0, 0, 0, 0)

        sidebar = QWidget()
        sidebar.setFixedWidth(200)
        sidebar.setStyleSheet("background-color: #1e293b;")
        sidebar_layout = QVBoxLayout()
        sidebar_layout.setContentsMargins(0, 20, 0, 0)
        sidebar_layout.setSpacing(5)

        self.tab_buttons = {}
        tabs = [
            ("产品管理", "products"),
            ("客户管理", "customers"),
            ("供应商管理", "suppliers"),
            ("报价管理", "quotes"),
            ("PI管理", "pi"),
            ("采购管理", "purchase"),
            ("出货管理", "shipment"),
            ("客户付款", "customer_payment"),
            ("供应商付款", "supplier_payment"),
            ("库存管理", "inventory"),
            ("订单总表", "order_summary"),
        ]

        for name, key in tabs:
            btn = QPushButton(name)
            btn.setFixedHeight(45)
            btn.setFont(get_font(10))
            btn.setStyleSheet("""
                QPushButton {
                    color: white;
                    background-color: transparent;
                    border: none;
                    text-align: left;
                    padding-left: 20px;
                }
                QPushButton:hover {
                    background-color: #334155;
                }
                QPushButton:checked {
                    background-color: #2563eb;
                }
            """)
            btn.setCheckable(True)
            btn.clicked.connect(lambda checked, k=key: self.switch_tab(k))
            sidebar_layout.addWidget(btn)
            self.tab_buttons[key] = btn

        self.tab_buttons["products"].setChecked(True)

        sidebar_layout.addStretch()

        # 帮助区域
        help_label = QLabel("帮助")
        help_label.setStyleSheet("color: #64748b; font-size: 11px; padding-left: 20px; margin-top: 10px;")
        sidebar_layout.addWidget(help_label)

        # 反馈 Bug 按钮
        bug_btn = QPushButton("🐛 反馈 Bug")
        bug_btn.setFixedHeight(35)
        bug_btn.setFont(get_font(10))
        bug_btn.setStyleSheet("""
            QPushButton {
                color: white;
                background-color: transparent;
                border: none;
                text-align: left;
                padding-left: 20px;
            }
            QPushButton:hover {
                background-color: #334155;
            }
        """)
        bug_btn.clicked.connect(self.open_bug_report)
        sidebar_layout.addWidget(bug_btn)

        # 设置按钮
        settings_btn = QPushButton("⚙ 设置")
        settings_btn.setFixedHeight(40)
        settings_btn.setFont(get_font(10))
        settings_btn.setStyleSheet("""
            QPushButton {
                color: #94a3b8;
                background-color: transparent;
                border: none;
                text-align: left;
                padding-left: 20px;
            }
            QPushButton:hover {
                background-color: #334155;
                color: white;
            }
        """)
        settings_btn.clicked.connect(self.open_settings)
        sidebar_layout.addWidget(settings_btn)

        sidebar.setLayout(sidebar_layout)
        content_layout.addWidget(sidebar)

        self._web_view = self._create_web_content()
        content_layout.addWidget(self._web_view)
        from web_container.routes import TAB_ROUTES
        self._web_view.navigate_to(TAB_ROUTES['products'])

        content.setLayout(content_layout)
        main_layout.addWidget(content)

        central_widget.setLayout(main_layout)

    def _create_web_content(self):
        """创建 Web 容器内容区（单一 QWebEngineView）"""
        from config.local_settings_manager import get_frontend_url
        self._web_container_widget = QWidget()
        layout = QVBoxLayout(self._web_container_widget)
        layout.setContentsMargins(0, 0, 0, 0)
        if HAS_WEB_CONTAINER:
            web_view = WebContainerView(remote_url=None, parent=self, index_path=self.index_path)
        else:
            web_view = QLabel("Web 容器模块不可用（缺少 PySide6 QtWebEngine 组件）")
            web_view.setAlignment(Qt.AlignmentFlag.AlignCenter)
            web_view.setStyleSheet("color: #666; font-size: 14px;")
        layout.addWidget(web_view)
        return web_view

    def switch_tab(self, key):
        from web_container.routes import TAB_ROUTES
        path = TAB_ROUTES.get(key)
        if path and self._web_view is not None and hasattr(self._web_view, 'navigate_to'):
            self._web_view.navigate_to(path)
        for name, button in self.tab_buttons.items():
            button.setChecked(name == key)

    def _load_async(self, api_method, update_method, error_msg="加载失败", loading_msg=None):
        """通用异步加载方法，使用QThread确保UI在主线程更新"""
        from PySide6.QtCore import QThread
        
        # 显示加载提示
        if loading_msg:
            self._show_loading_tip(loading_msg)
        
        class LoaderThread(QThread):
            def __init__(self, api_method, parent=None):
                super().__init__(parent)
                self.api_method = api_method
                self.result_data = []
                self.error_occurred = False
            
            def run(self):
                try:
                    data = self.api_method()
                    self.result_data = data if data else []
                except Exception as e:
                    print(f"{error_msg}: {e}")
                    self.error_occurred = True
                    self.result_data = []
        
        def on_finished(result):
            self._hide_loading_tip()
            update_method(result)
        
        thread = LoaderThread(api_method, self)
        thread.finished.connect(lambda: on_finished(thread.result_data))
        thread.start()
        return thread
    
    def load_customers_async(self):
        """异步加载客户数据"""
        self._load_async(
            self.api_client.get_customers,
            self._update_customers_table,
            "加载客户失败",
            loading_msg="正在加载客户..."
        )
    
    def _update_customers_table(self, customers):
        """在主线程更新客户表格"""
        if not customers:
            customers = []
        self.customers_table.setRowCount(len(customers))
        for row, c in enumerate(customers):
            checkbox = QCheckBox()
            checkbox.setStyleSheet("margin-left: 50%;")
            self.customers_table.setCellWidget(row, 0, checkbox)
            self.customers_table.setItem(row, 1, QTableWidgetItem(str(c.get('id', ''))))
            self.customers_table.setItem(row, 2, QTableWidgetItem(c.get('customer_code', '')))
            self.customers_table.setItem(row, 3, QTableWidgetItem(c.get('customer_name', '')))
            self.customers_table.setItem(row, 4, QTableWidgetItem(c.get('country', '')))
            self.customers_table.setItem(row, 5, QTableWidgetItem(""))
            self.customers_table.setItem(row, 6, QTableWidgetItem(""))
            self.customers_table.setItem(row, 7, QTableWidgetItem(c.get('payment_terms', '')))
            status = c.get('status', 1)
            status_text = "启用" if status == 1 else "禁用"
            status_color = "#10b981" if status == 1 else "#ef4444"
            status_item = QTableWidgetItem(status_text)
            status_item.setForeground(QBrush(QColor(status_color)))
            self.customers_table.setItem(row, 8, status_item)
            action_bar = ActionBarFactory.create_customer_action_bar(
                edit_callback=lambda _, c=c: self.edit_customer(c),
                toggle_callback=lambda _, c=c: self.toggle_customer_status(c),
                status=status
            )
            self.customers_table.setCellWidget(row, 9, action_bar)
        self.customer_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
        # 异步加载联系人和地址
        QTimer.singleShot(0, lambda: self._load_customer_extra_info(customers))
    
    def load_suppliers_async(self):
        """异步加载供应商数据"""
        self._load_async(
            self.api_client.get_suppliers,
            self._update_suppliers_table,
            "加载供应商失败",
            loading_msg="正在加载供应商..."
        )
    
    def _update_suppliers_table(self, suppliers):
        """在主线程更新供应商表格"""
        if not suppliers:
            suppliers = []
        self.suppliers_table.setRowCount(len(suppliers))
        for row, s in enumerate(suppliers):
            checkbox = QCheckBox()
            checkbox.setStyleSheet("margin-left: 15px;")
            self.suppliers_table.setCellWidget(row, 0, checkbox)
            self.suppliers_table.setItem(row, 1, QTableWidgetItem(str(s.get('id', ''))))
            self.suppliers_table.setItem(row, 2, QTableWidgetItem(s.get('supplier_code', '')))
            self.suppliers_table.setItem(row, 3, QTableWidgetItem(s.get('supplier_name', '')))
            self.suppliers_table.setItem(row, 4, QTableWidgetItem(s.get('region', '')))
            self.suppliers_table.setItem(row, 5, QTableWidgetItem(s.get('contact_person', '')))
            self.suppliers_table.setItem(row, 6, QTableWidgetItem(s.get('phone', '')))
            edit_btn = QPushButton("编辑")
            edit_btn.setFixedWidth(60)
            edit_btn.clicked.connect(lambda _, s=s: self.edit_supplier(s))
            self.suppliers_table.setCellWidget(row, 7, edit_btn)

    def create_products_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        toolbar = QHBoxLayout()
        
        title = QLabel("产品列表")
        title.setFont(get_font(14, QFont.Weight.Bold))
        toolbar.addWidget(title)
        
        toolbar.addStretch()

        # 搜索框
        search_layout = QHBoxLayout()
        search_layout.setSpacing(10)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("搜索（OE号/产品编号/工厂编号/品牌/描述）")
        self.search_input.setFixedWidth(250)
        search_layout.addWidget(self.search_input)

        self.category_filter_level1 = QComboBox()
        self.category_filter_level1.setFixedWidth(120)
        self.category_filter_level1.addItem("全部大类", 0)
        for code, name in get_parent_category_options():
            self.category_filter_level1.addItem(name, code)
        self.category_filter_level1.currentIndexChanged.connect(self.on_category_level1_changed)
        search_layout.addWidget(self.category_filter_level1)
        
        self.category_filter_level2 = QComboBox()
        self.category_filter_level2.setFixedWidth(120)
        self.category_filter_level2.addItem("全部子类", 0)
        search_layout.addWidget(self.category_filter_level2)
        
        # 日志
        print("[产品管理] 类别筛选器初始化完成")

        search_btn = QPushButton("搜索")
        search_btn.clicked.connect(self.search_products)
        search_btn.setStyleSheet("""
            QPushButton {
                background-color: #3b82f6;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 6px 12px;
            }
            QPushButton:hover { background-color: #2563eb; }
        """)
        search_layout.addWidget(search_btn)

        reset_btn = QPushButton("重置")
        reset_btn.clicked.connect(self.reset_search)
        reset_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 6px 12px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        search_layout.addWidget(reset_btn)

        toolbar.addLayout(search_layout)

        add_btn = QPushButton("+ 新增产品")
        add_btn.clicked.connect(self.add_product)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        toolbar.addWidget(add_btn)

        # 筛选栏 - 保留客户筛选
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)

        filter_layout.addWidget(QLabel("客户:"))
        self.product_customer_filter = QComboBox()
        self.product_customer_filter.setFixedWidth(150)
        self.product_customer_filter.addItem("全部客户", 0)
        filter_layout.addWidget(self.product_customer_filter)

        filter_btn = QPushButton("筛选")
        filter_btn.clicked.connect(self.filter_by_customer)
        filter_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        filter_layout.addWidget(filter_btn)

        filter_layout.addStretch()

        toolbar.addLayout(filter_layout)

        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.load_products)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        toolbar.addWidget(refresh_btn)

        import_btn = QPushButton("批量导入")
        import_btn.clicked.connect(self.import_products)
        import_btn.setStyleSheet("""
            QPushButton {
                background-color: #10b981;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #059669; }
        """)
        toolbar.addWidget(import_btn)

        # 批量操作按钮
        batch_layout = QHBoxLayout()
        batch_layout.setSpacing(10)
        
        self.select_all_checkbox = QCheckBox("全选")
        self.select_all_checkbox.stateChanged.connect(self.toggle_select_all_products)
        batch_layout.addWidget(self.select_all_checkbox)
        
        batch_disable_btn = QPushButton("批量禁用")
        batch_disable_btn.clicked.connect(self.batch_toggle_product_status)
        batch_disable_btn.setStyleSheet("""
            QPushButton {
                background-color: #f59e0b;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 6px 12px;
            }
            QPushButton:hover { background-color: #d97706; }
        """)
        batch_layout.addWidget(batch_disable_btn)
        
        batch_delete_btn = QPushButton("批量删除")
        batch_delete_btn.clicked.connect(self.batch_delete_products)
        batch_delete_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc2626;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 6px 12px;
            }
            QPushButton:hover { background-color: #b91c1c; }
        """)
        batch_layout.addWidget(batch_delete_btn)
        
        toolbar.addLayout(batch_layout)

        layout.addLayout(toolbar)

        self.products_table = QTableWidget()
        self.products_table.setColumnCount(15)
        self.products_table.setHorizontalHeaderLabels([
            "", "客户产品编号", "系统编号", "OE号", "图片", "产品名称",
            "客户型号", "客户", "类别", "颜色", "品牌", "USD", "RMB", "规格", "操作"
        ])
        # 设置列宽
        self.products_table.setColumnWidth(0, 30)   # 复选框
        self.products_table.setColumnWidth(1, 120)  # 客户产品编号
        self.products_table.setColumnWidth(2, 120)  # 系统编号
        self.products_table.setColumnWidth(3, 100)  # OE号
        self.products_table.setColumnWidth(4, 70)   # 图片
        self.products_table.setColumnWidth(5, 150)  # 产品名称
        self.products_table.setColumnWidth(6, 100)  # 客户型号
        self.products_table.setColumnWidth(7, 100)  # 客户
        self.products_table.setColumnWidth(8, 80)   # 类别
        self.products_table.setColumnWidth(9, 60)   # 颜色
        self.products_table.setColumnWidth(10, 80)  # 品牌
        self.products_table.setColumnWidth(11, 80)  # USD
        self.products_table.setColumnWidth(12, 80)  # RMB
        self.products_table.setColumnWidth(13, 150)  # 规格
        self.products_table.setColumnWidth(14, 70)  # 操作
        self.products_table.verticalHeader().setDefaultSectionSize(70)
        self.products_table.doubleClicked.connect(self.on_product_double_click)
        self.setup_table_context_menu(self.products_table, ["", "客户产品编号", "系统编号", "OE号", "图片", "产品名称", "客户型号", "客户", "类别", "颜色", "品牌", "USD", "RMB", "规格", "操作"])
        layout.addWidget(self.products_table)

        widget.setLayout(layout)
        return widget

    def create_customers_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        toolbar = QHBoxLayout()
        title = QLabel("客户列表")
        title.setFont(get_font(14, QFont.Weight.Bold))
        toolbar.addWidget(title)
        toolbar.addStretch()

        self.customer_select_all_checkbox = QCheckBox("全选")
        self.customer_select_all_checkbox.clicked.connect(self.toggle_select_all_customers)
        toolbar.addWidget(self.customer_select_all_checkbox)

        self.customer_search_input = QLineEdit()
        self.customer_search_input.setPlaceholderText("搜索客户名称/编号...")
        self.customer_search_input.setFixedWidth(200)
        self.customer_search_input.returnPressed.connect(self.search_customers)
        toolbar.addWidget(self.customer_search_input)

        self.customer_country_filter = QComboBox()
        self.customer_country_filter.addItem("全部国家", 0)
        self.customer_country_filter.setFixedWidth(150)
        toolbar.addWidget(self.customer_country_filter)

        search_btn = QPushButton("搜索")
        search_btn.clicked.connect(self.search_customers)
        search_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        toolbar.addWidget(search_btn)

        add_btn = QPushButton("+ 新增客户")
        add_btn.clicked.connect(self.add_customer)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        toolbar.addWidget(add_btn)

        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.load_customers)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        toolbar.addWidget(refresh_btn)

        layout.addLayout(toolbar)

        self.customers_table = QTableWidget()
        self.customers_table.setColumnCount(10)
        self.customers_table.setHorizontalHeaderLabels(["选择", "ID", "客户编号", "客户名称", "国家", "默认联系人", "默认地址", "付款条款", "状态", "操作"])
        self.customers_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.customers_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.customers_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.customers_table.doubleClicked.connect(self.on_customer_double_click)
        self.setup_table_context_menu(self.customers_table, ["选择", "ID", "客户编号", "客户名称", "国家", "默认联系人", "默认地址", "付款条款", "状态", "操作"])
        layout.addWidget(self.customers_table)

        widget.setLayout(layout)
        return widget

    def create_suppliers_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        toolbar = QHBoxLayout()
        title = QLabel("供应商列表")
        title.setFont(get_font(14, QFont.Weight.Bold))
        toolbar.addWidget(title)
        toolbar.addStretch()

        self.select_all_checkbox = QCheckBox("全选")
        self.select_all_checkbox.clicked.connect(self.toggle_select_all_suppliers)
        toolbar.addWidget(self.select_all_checkbox)

        delete_btn = QPushButton("删除选中")
        delete_btn.clicked.connect(self.delete_selected_suppliers)
        delete_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc2626;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #b91c1c; }
        """)
        toolbar.addWidget(delete_btn)

        import_btn = QPushButton("批量导入")
        import_btn.clicked.connect(self.import_suppliers)
        import_btn.setStyleSheet("""
            QPushButton {
                background-color: #16a34a;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #15803d; }
        """)
        toolbar.addWidget(import_btn)

        add_btn = QPushButton("+ 新增供应商")
        add_btn.clicked.connect(self.add_supplier)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        toolbar.addWidget(add_btn)

        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.load_suppliers)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        toolbar.addWidget(refresh_btn)

        layout.addLayout(toolbar)

        self.suppliers_table = QTableWidget()
        self.suppliers_table.setColumnCount(8)
        self.suppliers_table.setHorizontalHeaderLabels(["", "ID", "供应商编号", "供应商名称", "地区", "联系人", "电话", "操作"])
        self.suppliers_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.suppliers_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.suppliers_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.suppliers_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.suppliers_table.setColumnWidth(0, 40)
        self.suppliers_table.doubleClicked.connect(self.on_supplier_double_click)
        self.setup_table_context_menu(self.suppliers_table, ["", "ID", "供应商编号", "供应商名称", "地区", "联系人", "电话", "操作"])
        layout.addWidget(self.suppliers_table)

        widget.setLayout(layout)
        return widget

    def toggle_select_all_suppliers(self, checked):
        for row in range(self.suppliers_table.rowCount()):
            checkbox = self.suppliers_table.cellWidget(row, 0)
            if checkbox:
                checkbox.setChecked(checked)

    def delete_selected_suppliers(self):
        selected_ids = []
        for row in range(self.suppliers_table.rowCount()):
            checkbox = self.suppliers_table.cellWidget(row, 0)
            if checkbox and checkbox.isChecked():
                supplier_id = int(self.suppliers_table.item(row, 1).text())
                selected_ids.append(supplier_id)

        if not selected_ids:
            QMessageBox.warning(self, "提示", "请先选择要删除的供应商")
            return

        reply = QMessageBox.question(self, "确认删除", f"确定要删除选中的 {len(selected_ids)} 个供应商吗？",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.No:
            return

        for supplier_id in selected_ids:
            try:
                self.api_client.delete_supplier(supplier_id)
            except Exception as e:
                QMessageBox.warning(self, "错误", f"删除供应商失败: {str(e)}")

        self.load_suppliers()
        self.select_all_checkbox.setChecked(False)

    def create_pi_tab(self):
        # 2026-06-12 需求#42：使用 QTabWidget 包裹，新增"历史与正式纪录"子 Tab
        outer = QWidget()
        outer_layout = QVBoxLayout(outer)
        outer_layout.setContentsMargins(0, 0, 0, 0)

        self._pi_tab_widget = QTabWidget()

        # === Tab 1: PI 订单列表（原有内容）===
        list_widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        toolbar = QHBoxLayout()
        title = QLabel("PI订单列表")
        title.setFont(get_font(14, QFont.Weight.Bold))
        toolbar.addWidget(title)
        toolbar.addStretch()

        add_btn = QPushButton("+ 新建PI")
        add_btn.clicked.connect(self.add_pi)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        toolbar.addWidget(add_btn)

        self.pi_batch_delete_btn = QPushButton("批量删除")
        self.pi_batch_delete_btn.clicked.connect(self.batch_delete_pi)
        self.pi_batch_delete_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc2626;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #b91c1c; }
        """)
        toolbar.addWidget(self.pi_batch_delete_btn)

        self.pi_batch_export_btn = QPushButton("批量导出")
        self.pi_batch_export_btn.clicked.connect(self.batch_export_pi)
        self.pi_batch_export_btn.setStyleSheet("""
            QPushButton {
                background-color: #059669;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #047857; }
        """)
        toolbar.addWidget(self.pi_batch_export_btn)

        # 预览导出按钮
        self.pi_preview_export_btn = QPushButton("预览导出")
        self.pi_preview_export_btn.clicked.connect(self._on_preview_pi_export)
        self.pi_preview_export_btn.setStyleSheet("""
            QPushButton {
                background-color: #7c3aed;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #6d28d9; }
        """)
        toolbar.addWidget(self.pi_preview_export_btn)

        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.load_pi_orders_async)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        toolbar.addWidget(refresh_btn)

        layout.addLayout(toolbar)

        self.pi_table = QTableWidget()
        self.pi_table.setColumnCount(10)
        self.pi_table.setHorizontalHeaderLabels(["", "ID", "PI号", "金额", "币种", "状态", "创建时间", "操作", "完成", "导出"])
        self.pi_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.pi_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.pi_table.setColumnWidth(0, 40)
        self.pi_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.pi_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.pi_table.doubleClicked.connect(self.on_pi_double_click)
        self.setup_table_context_menu(self.pi_table, ["", "ID", "PI号", "金额", "币种", "状态", "创建时间", "操作", "完成", "导出"])
        layout.addWidget(self.pi_table)

        list_widget.setLayout(layout)
        self._pi_tab_widget.addTab(list_widget, "PI 列表")

        # === Tab 2: 历史与正式纪录（需求#42）===
        self._pi_history_tab = PiHistoryTab(self.api_client, self)
        self._pi_tab_widget.addTab(self._pi_history_tab, "历史与正式纪录")

        outer_layout.addWidget(self._pi_tab_widget)
        return outer

    def create_purchase_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        toolbar = QHBoxLayout()
        title = QLabel("采购订单列表")
        title.setFont(get_font(14, QFont.Weight.Bold))
        toolbar.addWidget(title)
        toolbar.addStretch()

        add_btn = QPushButton("+ 新建采购单")
        add_btn.clicked.connect(self.add_purchase)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        toolbar.addWidget(add_btn)

        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.load_purchase_orders)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        toolbar.addWidget(refresh_btn)

        export_contract_btn = QPushButton("📄 导出合同")
        export_contract_btn.clicked.connect(self._on_export_purchase_contract)
        export_contract_btn.setStyleSheet("""
            QPushButton {
                background-color: #8b5cf6;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #7c3aed; }
        """)
        toolbar.addWidget(export_contract_btn)

        layout.addLayout(toolbar)

        self.purchase_table = QTableWidget()
        self.purchase_table.setColumnCount(10)
        self.purchase_table.setHorizontalHeaderLabels(["ID", "采购单号", "PI号", "供应商", "金额", "状态", "操作", "确认", "入库", "合同"])

        # 2026-06-23：所有列允许手动拖动宽度（之前只设了 col 1 Stretch，其他列默认 Interactive 但 Stretch 列导致视觉错位）
        # 默认列宽（按需显示）
        _DEFAULT_WIDTHS = [50, 200, 150, 180, 100, 80, 80, 80, 80, 80]
        for c_idx, w in enumerate(_DEFAULT_WIDTHS):
            self.purchase_table.horizontalHeader().setSectionResizeMode(
                c_idx, QHeaderView.ResizeMode.Interactive
            )
            self.purchase_table.setColumnWidth(c_idx, w)

        # 加载用户上次保存的列宽
        self._load_purchase_table_column_widths()
        # 监听拖动，节流保存
        self.purchase_table.horizontalHeader().sectionResized.connect(
            self._on_purchase_table_column_resized
        )

        self.purchase_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.purchase_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.setup_table_context_menu(self.purchase_table, ["ID", "采购单号", "PI号", "供应商", "金额", "状态", "操作"])
        layout.addWidget(self.purchase_table)

        widget.setLayout(layout)
        return widget

    # ==================== 采购订单列表表格列宽持久化（2026-06-23） ====================
    _PURCHASE_TABLE_COL_WIDTHS_KEY = "main_window/purchase_table_column_widths"

    def _load_purchase_table_column_widths(self):
        """从 QSettings 加载采购订单表格列宽"""
        if not hasattr(self, 'purchase_table'):
            return
        try:
            settings = QSettings("PI-Manager", "MainWindow")
            saved = settings.value(self._PURCHASE_TABLE_COL_WIDTHS_KEY, None)
            if not saved:
                return
            widths = None
            if isinstance(saved, dict):
                widths = {int(k): int(v) for k, v in saved.items()}
            elif isinstance(saved, str) and saved.startswith('{'):
                import ast
                parsed = ast.literal_eval(saved)
                if isinstance(parsed, dict):
                    widths = {int(k): int(v) for k, v in parsed.items()}
            if not widths:
                return
            for col, w in widths.items():
                if 0 <= col < self.purchase_table.columnCount() and w > 10:
                    self.purchase_table.setColumnWidth(col, w)
        except Exception as e:
            print(f"[MainWindow] 加载采购订单列表列宽失败: {e}")

    def _on_purchase_table_column_resized(self, logical_index: int, old_size: int, new_size: int):
        """列宽拖动后节流保存"""
        if not hasattr(self, '_save_purchase_col_width_timer'):
            self._save_purchase_col_width_timer = QTimer(self)
            self._save_purchase_col_width_timer.setSingleShot(True)
            self._save_purchase_col_width_timer.setInterval(500)
            self._save_purchase_col_width_timer.timeout.connect(self._save_purchase_table_column_widths)
        self._save_purchase_col_width_timer.start()

    def _save_purchase_table_column_widths(self):
        """保存采购订单表格列宽到 QSettings"""
        if not hasattr(self, 'purchase_table'):
            return
        try:
            widths = {
                col: self.purchase_table.columnWidth(col)
                for col in range(self.purchase_table.columnCount())
            }
            settings = QSettings("PI-Manager", "MainWindow")
            settings.setValue(self._PURCHASE_TABLE_COL_WIDTHS_KEY, widths)
        except Exception as e:
            print(f"[MainWindow] 保存采购订单列表列宽失败: {e}")

    def create_quotes_tab(self):
        """报价管理标签页"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        toolbar = QHBoxLayout()
        title = QLabel("报价管理")
        title.setFont(get_font(14, QFont.Weight.Bold))
        toolbar.addWidget(title)
        toolbar.addStretch()

        add_btn = QPushButton("+ 新建报价单")
        add_btn.clicked.connect(self.add_quote)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        toolbar.addWidget(add_btn)
        
        batch_delete_btn = QPushButton("批量删除")
        batch_delete_btn.clicked.connect(self.batch_delete_quotes)
        batch_delete_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc2626;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #b91c1c; }
        """)
        toolbar.addWidget(batch_delete_btn)
        
        batch_export_btn = QPushButton("批量导出")
        batch_export_btn.clicked.connect(self.batch_export_quotes)
        batch_export_btn.setStyleSheet("""
            QPushButton {
                background-color: #059669;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #047857; }
        """)
        toolbar.addWidget(batch_export_btn)
        
        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.load_quotes_async)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #6b7280;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #4b5563; }
        """)
        toolbar.addWidget(refresh_btn)

        layout.addLayout(toolbar)

        # 报价单列表
        self.quote_table = QTableWidget()
        self.quote_table.setColumnCount(10)
        self.quote_table.setHorizontalHeaderLabels(["", "ID", "报价单号", "客户", "金额", "币种", "状态", "有效期", "备注", "操作"])
        self.quote_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.quote_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.quote_table.setColumnHidden(1, True)  # 隐藏ID列
        self.quote_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.quote_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.quote_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.quote_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.quote_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        self.quote_table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        self.quote_table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        self.quote_table.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeMode.ResizeToContents)
        self.quote_table.horizontalHeader().setSectionResizeMode(8, QHeaderView.ResizeMode.Stretch)
        self.quote_table.horizontalHeader().setSectionResizeMode(9, QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.quote_table)

        widget.setLayout(layout)
        return widget

    def load_quotes_async(self):
        """异步加载报价单"""
        self._load_async(
            self.api_client.get_quotes,
            self._update_quote_table,
            "加载报价单失败",
            loading_msg="正在加载报价单..."
        )

    def _update_quote_table(self, quotes):
        """更新报价单表格"""
        try:
            self.quote_table.setRowCount(len(quotes))
            status_map = {1: "草稿", 2: "已发送", 3: "已接受", 4: "已拒绝"}
            
            for row, q in enumerate(quotes):
                # 复选框
                checkbox = QTableWidgetItem()
                checkbox.setFlags(checkbox.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                checkbox.setCheckState(Qt.CheckState.Unchecked)
                checkbox.setData(Qt.ItemDataRole.UserRole, q.get('id'))
                self.quote_table.setItem(row, 0, checkbox)
                
                self.quote_table.setItem(row, 1, QTableWidgetItem(str(q.get('id', ''))))
                self.quote_table.setItem(row, 2, QTableWidgetItem(q.get('quote_no', '')))
                self.quote_table.setItem(row, 3, QTableWidgetItem(q.get('customer_name', '')))
                self.quote_table.setItem(row, 4, QTableWidgetItem(f"${q.get('total_amount', 0):,.2f}"))
                self.quote_table.setItem(row, 5, QTableWidgetItem(q.get('currency', 'USD')))
                self.quote_table.setItem(row, 6, QTableWidgetItem(status_map.get(q.get('status', 1), '草稿')))
                valid_until = q.get('valid_until')
                if valid_until:
                    valid_until = str(valid_until)[:10]
                self.quote_table.setItem(row, 7, QTableWidgetItem(valid_until or '-'))
                self.quote_table.setItem(row, 8, QTableWidgetItem(q.get('remark', '') or '-'))
                
                # 操作按钮
                btn_widget = QWidget()
                btn_layout = QHBoxLayout()
                btn_layout.setContentsMargins(0, 0, 0, 0)
                
                edit_btn = QPushButton("编辑")
                edit_btn.setFixedWidth(50)
                edit_btn.clicked.connect(lambda _, qid=q.get('id'): self.edit_quote(qid))
                btn_layout.addWidget(edit_btn)
                
                pi_btn = QPushButton("转PI")
                pi_btn.setFixedWidth(50)
                pi_btn.setStyleSheet("color: #10b981;")
                pi_btn.clicked.connect(lambda _, qid=q.get('id'): self.convert_quote_to_pi(qid))
                btn_layout.addWidget(pi_btn)
                
                btn_widget.setLayout(btn_layout)
                self.quote_table.setCellWidget(row, 9, btn_widget)
        except Exception as e:
            print(f"更新报价单表格失败: {e}")
        finally:
            self.quote_table.viewport().update()
    
    def get_selected_quote_ids(self):
        """获取选中的报价单ID列表"""
        ids = []
        for row in range(self.quote_table.rowCount()):
            item = self.quote_table.item(row, 0)
            if item and item.checkState() == Qt.CheckState.Checked:
                quote_id = item.data(Qt.ItemDataRole.UserRole)
                if quote_id:
                    ids.append(quote_id)
        return ids
    
    def toggle_all_quotes(self, state):
        """全选/取消全选"""
        check_state = Qt.CheckState.Checked if state else Qt.CheckState.Unchecked
        for row in range(self.quote_table.rowCount()):
            item = self.quote_table.item(row, 0)
            if item:
                item.setCheckState(check_state)
    
    def batch_delete_quotes(self):
        """批量删除报价单"""
        selected_ids = self.get_selected_quote_ids()
        if not selected_ids:
            QMessageBox.warning(self, "提示", "请先选择要删除的报价单")
            return
        
        reply = QMessageBox.question(
            self, "确认删除", 
            f"确定要删除选中的 {len(selected_ids)} 个报价单吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        try:
            result = self.api_client.batch_delete_quotes(selected_ids)
            deleted = result.get('deleted', 0)
            errors = result.get('errors', [])
            if errors:
                QMessageBox.warning(self, "部分删除失败", f"成功删除 {deleted} 个\n失败: {len(errors)} 个\n{errors}")
            else:
                QMessageBox.information(self, "成功", f"已删除 {deleted} 个报价单")
            self.load_quotes_async()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"删除失败: {str(e)}")
            self.load_quotes_async()
    
    def batch_export_quotes(self):
        """批量导出报价单"""
        selected_ids = self.get_selected_quote_ids()
        if not selected_ids:
            QMessageBox.warning(self, "提示", "请先选择要导出的报价单")
            return
        
        try:
            quotes_data = []
            for quote_id in selected_ids:
                quote = self.api_client.get_quote(quote_id)
                quotes_data.append(quote)
            
            if not quotes_data:
                QMessageBox.information(self, "提示", "没有可导出的数据")
                return
            
            # 构建导出数据
            export_rows = []
            for q in quotes_data:
                for item in q.get('items', []):
                    export_rows.append({
                        '报价单号': q.get('quote_no', ''),
                        '客户': q.get('customer_name', ''),
                        '币种': q.get('currency', 'USD'),
                        '总金额': q.get('total_amount', 0),
                        '有效期': q.get('valid_until', ''),
                        '状态': ['草稿', '已发送', '已接受', '已拒绝'][q.get('status', 1) - 1] if q.get('status', 1) <= 4 else '',
                        '备注': q.get('remark', ''),
                        '产品编号': item.get('product_id', ''),
                        'OE号': item.get('oe_number', ''),
                        '客户编号': item.get('customer_code', ''),
                        '产品描述': item.get('detail_desc', ''),
                        '数量': item.get('quantity', 0),
                        '单价': item.get('unit_price', 0),
                        '总价': item.get('total_price', 0),
                        '明细备注': item.get('remark', ''),
                    })
            
            # 使用 openpyxl 创建Excel并导出（替代 pandas）
            from PySide6.QtWidgets import QFileDialog
            file_path, _ = QFileDialog.getSaveFileName(
                self, "保存报价单",
                f"报价单导出_{len(selected_ids)}个.xlsx",
                "Excel Files (*.xlsx)"
            )
            if file_path:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "报价单"
                
                # 写入表头
                if export_rows:
                    headers = list(export_rows[0].keys())
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col_idx, value=header)
                    
                    # 写入数据
                    for row_idx, row_data in enumerate(export_rows, 2):
                        for col_idx, header in enumerate(headers, 1):
                            value = row_data.get(header)
                            # 处理None值
                            if value is None or value == '':
                                value = ''
                            ws.cell(row=row_idx, column=col_idx, value=value)
                
                wb.save(file_path)
                QMessageBox.information(self, "成功", f"已导出 {len(export_rows)} 条明细到:\n{file_path}")
        except Exception as e:
            QMessageBox.warning(self, "错误", f"导出失败: {str(e)}")

    def add_quote(self):
        """新建报价单"""
        dialog = QuoteDialog(self, self.api_client, self.dept_id)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            QMessageBox.information(self, "成功", "报价单已保存")
            self.load_quotes_async()

    def edit_quote(self, quote_id):
        """编辑报价单"""
        try:
            quote = self.api_client.get_quote(quote_id)
            dialog = QuoteDialog(self, self.api_client, self.dept_id, quote)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                self.load_quotes_async()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"获取报价单失败: {str(e)}")

    def convert_quote_to_pi(self, quote_id):
        """将报价单转为PI"""
        reply = QMessageBox.question(self, "确认", "确定要将此报价单转为PI吗？")
        if reply == QMessageBox.StandardButton.Yes:
            try:
                result = self.api_client.convert_quote_to_pi(quote_id)
                QMessageBox.information(self, "成功", f"报价单已转为PI\nPI单号: {result.get('pi_no', '')}")
                self.load_quotes_async()
            except Exception as e:
                QMessageBox.warning(self, "错误", f"转换失败: {str(e)}")

    def create_shipment_tab(self):
        """出货管理标签页 - 双模式设计（2026-06-05 重写）"""
        from widgets.shipment_tab import ShipmentTab
        
        # 使用新的双模式出货Tab组件
        self._shipment_tab = ShipmentTab(self.api_client, self)
        
        # 连接信号
        self._shipment_tab.mode_changed.connect(self._on_shipment_mode_changed)
        
        return self._shipment_tab
    
    def _on_shipment_mode_changed(self, mode: str):
        """出货Tab模式切换回调"""
        if mode == "detail":
            # 进入详情模式
            pass
        elif mode == "list":
            # 返回列表模式
            pass

    def create_customer_payment_tab(self):
        """客户付款列表 - 12 列 PI 聚合展示模式 (spec #45)"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        # ===== 工具栏：标题 + 筛选控件 + 操作按钮 =====
        toolbar = QHBoxLayout()

        # 标题
        title = QLabel("客户付款列表")
        title.setFont(get_font(14, QFont.Weight.Bold))
        toolbar.addWidget(title)

        # 分隔符
        toolbar.addSpacing(30)

        # 客户筛选下拉框
        toolbar.addWidget(QLabel("客户:"))
        self.cp_customer_filter = QComboBox()
        self.cp_customer_filter.setMinimumWidth(120)
        self.cp_customer_filter.addItem("全部客户")
        self.cp_customer_filter.currentIndexChanged.connect(self._on_cp_customer_filter_changed)
        toolbar.addWidget(self.cp_customer_filter)

        # PI号搜索框
        toolbar.addWidget(QLabel("PI号:"))
        self.cp_pi_search = QLineEdit()
        self.cp_pi_search.setPlaceholderText("PI号搜索...")
        self.cp_pi_search.setMaximumWidth(150)
        self.cp_pi_search.textChanged.connect(self._on_cp_pi_search_changed)
        toolbar.addWidget(self.cp_pi_search)

        # 仅未结清复选框
        self.cp_only_unpaid = QCheckBox("仅未结清")
        self.cp_only_unpaid.stateChanged.connect(self._on_cp_only_unpaid_changed)
        toolbar.addWidget(self.cp_only_unpaid)

        # 弹性空间
        toolbar.addStretch()

        # 新建付款记录按钮
        add_btn = QPushButton("+ 新建付款记录")
        add_btn.clicked.connect(self.add_customer_payment)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        toolbar.addWidget(add_btn)

        # 刷新按钮
        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.load_customer_payments)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        toolbar.addWidget(refresh_btn)

        layout.addLayout(toolbar)

        # ===== 表格：12 列结构 =====
        self.customer_payment_table = QTableWidget()
        self.customer_payment_table.setColumnCount(12)
        self.customer_payment_table.setHorizontalHeaderLabels([
            "PI号", "客户名称",
            "付1金额", "付1日期",
            "付2金额", "付2日期",
            "付3金额", "付3日期",
            "总应收", "未付款",
            "水单", "操作"
        ])

        # 设置列宽（基于规格文档 3.1 节）
        column_widths = {
            0: 140,   # PI号
            1: 120,   # 客户名称
            2: 100, 3: 100,  # 付款1
            4: 100, 5: 100,  # 付款2
            6: 100, 7: 100,  # 付款3
            8: 110,   # 总应收
            9: 110,   # 未付款
            10: 80,   # 水单
            11: 80,   # 操作
        }
        for col, width in column_widths.items():
            self.customer_payment_table.setColumnWidth(col, width)

        # PI号列可拉伸以填充剩余空间
        self.customer_payment_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Stretch
        )

        # 表格行为配置
        self.customer_payment_table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows
        )
        self.customer_payment_table.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers
        )

        # 右键菜单配置（更新为 12 列）
        self.setup_table_context_menu(
            self.customer_payment_table,
            ["PI号", "客户名称", "付1金额", "付1日期", "付2金额", "付2日期",
             "付3金额", "付3日期", "总应收", "未付款", "水单", "操作"]
        )

        layout.addWidget(self.customer_payment_table)
        widget.setLayout(layout)

        # 延迟填充客户下拉框（等待客户数据加载完成）
        from PySide6.QtCore import QTimer
        QTimer.singleShot(1000, self._populate_cp_customer_filter)

        return widget

    def _populate_cp_customer_filter(self):
        """
        填充客户付款 Tab 的客户筛选下拉框
        (任务 4 中实现完整逻辑)
        """
        if not hasattr(self, 'customers') or not self.customers:
            return
        # 保留第一项"全部客户"
        while self.cp_customer_filter.count() > 1:
            self.cp_customer_filter.removeItem(1)
        for cust in self.customers:
            cust_id = cust.get('id')
            cust_name = cust.get('customer_name') or cust.get('name', '未知')
            self.cp_customer_filter.addItem(cust_name, userData=cust_id)

    def create_supplier_payment_tab(self):
        """供应商付款标签页 - 主从表设计（参考库存管理）"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        toolbar = QHBoxLayout()
        title = QLabel("供应商付款管理")
        title.setFont(get_font(14, QFont.Weight.Bold))
        toolbar.addWidget(title)
        toolbar.addStretch()

        add_btn = QPushButton("+ 新建付款记录")
        add_btn.clicked.connect(self.add_supplier_payment)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        toolbar.addWidget(add_btn)

        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.load_supplier_payments_async)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        toolbar.addWidget(refresh_btn)

        layout.addLayout(toolbar)

        # 主表：付款汇总
        self.supplier_payment_table = QTableWidget()
        self.supplier_payment_table.setColumnCount(8)
        self.supplier_payment_table.setHorizontalHeaderLabels(["ID", "供应商", "采购单", "总金额", "已付金额", "未付金额", "状态", "操作"])
        self.supplier_payment_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.supplier_payment_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.supplier_payment_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.setup_table_context_menu(self.supplier_payment_table, ["ID", "供应商", "采购单", "总金额", "已付金额", "未付金额", "状态", "操作"])
        layout.addWidget(self.supplier_payment_table)

        # 详情标签
        detail_label = QLabel("📋 付款阶段详情（请点击上方记录查看）")
        detail_label.setFont(get_font(12, QFont.Weight.Bold))
        layout.addWidget(detail_label)

        # 从表：付款阶段明细
        self.supplier_payment_stage_table = QTableWidget()
        self.supplier_payment_stage_table.setColumnCount(7)
        self.supplier_payment_stage_table.setHorizontalHeaderLabels(["阶段", "应付金额", "已付金额", "状态", "付款日期", "凭证", "操作"])
        self.supplier_payment_stage_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.supplier_payment_stage_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.supplier_payment_stage_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.supplier_payment_stage_table.setMaximumHeight(200)
        layout.addWidget(self.supplier_payment_stage_table)

        # 绑定点击事件
        try:
            self.supplier_payment_table.cellClicked.disconnect()
        except RuntimeError:
            pass
        self.supplier_payment_table.cellClicked.connect(self.show_supplier_payment_stages)

        widget.setLayout(layout)
        return widget

    def create_inventory_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        toolbar = QHBoxLayout()
        title = QLabel("库存管理（按OE号分组）")
        title.setFont(get_font(14, QFont.Weight.Bold))
        toolbar.addWidget(title)
        toolbar.addStretch()

        # 2026-06-23: 库存状态筛选下拉框
        self.inventory_status_filter = QComboBox()
        self.inventory_status_filter.setFixedWidth(120)
        self.inventory_status_filter.addItem("全部状态", None)
        self.inventory_status_filter.addItem("采购在途", 1)
        self.inventory_status_filter.addItem("待入库", 2)
        self.inventory_status_filter.addItem("已入库", 3)
        self.inventory_status_filter.addItem("历史库存", 4)
        self.inventory_status_filter.currentIndexChanged.connect(self.on_inventory_status_filter_changed)
        toolbar.addWidget(self.inventory_status_filter)

        add_btn = QPushButton("+ 新建库存")
        add_btn.clicked.connect(self.add_inventory)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        toolbar.addWidget(add_btn)

        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.load_inventories)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        toolbar.addWidget(refresh_btn)

        layout.addLayout(toolbar)

        # 搜索框
        search_layout = QHBoxLayout()
        self.inventory_search_input = QLineEdit()
        self.inventory_search_input.setPlaceholderText("搜索OE号...")
        self.inventory_search_input.setFixedHeight(35)
        self.inventory_search_input.returnPressed.connect(self.search_inventory)
        search_layout.addWidget(self.inventory_search_input)
        
        search_btn = QPushButton("搜索")
        search_btn.clicked.connect(self.search_inventory)
        search_btn.setStyleSheet("""
            QPushButton {
                background-color: #3b82f6;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
        """)
        search_layout.addWidget(search_btn)
        layout.addLayout(search_layout)

        # 主表格：显示OE号（产品维度）
        self.inventory_product_table = QTableWidget()
        self.inventory_product_table.setColumnCount(10)
        self.inventory_product_table.setHorizontalHeaderLabels(["OE号", "产品编号", "总库存", "供应商", "客户", "状态分布", "最近入库供应商", "最近出库客户", "最近变更", "操作"])
        self.inventory_product_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        self.inventory_product_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        self.inventory_product_table.setColumnWidth(0, 150)
        self.inventory_product_table.setColumnWidth(1, 110)
        self.inventory_product_table.setColumnWidth(2, 60)
        self.inventory_product_table.setColumnWidth(3, 60)
        self.inventory_product_table.setColumnWidth(4, 60)
        self.inventory_product_table.setColumnWidth(5, 100)
        self.inventory_product_table.setColumnWidth(6, 100)
        self.inventory_product_table.setColumnWidth(7, 100)
        self.inventory_product_table.setColumnWidth(8, 120)
        self.inventory_product_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.inventory_product_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        layout.addWidget(QLabel("📦 产品库存汇总（双击展开查看详情）"))
        layout.addWidget(self.inventory_product_table)
        
        # 分隔线
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setStyleSheet("background-color: #e5e7eb;")
        line.setFixedHeight(2)
        layout.addWidget(line)
        
        # 子表格：显示选中OE号的详细库存记录
        self.inventory_detail_label = QLabel("📋 库存详情（请先选择上方产品）")
        layout.addWidget(self.inventory_detail_label)
        
        self.inventory_detail_table = QTableWidget()
        self.inventory_detail_table.setColumnCount(9)
        self.inventory_detail_table.setHorizontalHeaderLabels(["ID", "供应商", "客户", "数量", "库位", "状态", "备注", "创建时间", "操作"])
        self.inventory_detail_table.setColumnWidth(0, 40)
        self.inventory_detail_table.setColumnWidth(1, 100)
        self.inventory_detail_table.setColumnWidth(2, 100)
        self.inventory_detail_table.setColumnWidth(3, 60)
        self.inventory_detail_table.setColumnWidth(4, 80)
        self.inventory_detail_table.setColumnWidth(5, 50)
        self.inventory_detail_table.setColumnWidth(6, 120)
        self.inventory_detail_table.setColumnWidth(7, 130)
        self.inventory_detail_table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.Stretch)
        self.inventory_detail_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.inventory_detail_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        layout.addWidget(self.inventory_detail_table)

        status_label = QLabel("● 黄色: 采购在途 | ● 蓝色: 待入库 | ● 绿色: 已入库 | ● 黑色: 历史库存")
        status_label.setStyleSheet("color: #6b7280; padding: 5px;")
        layout.addWidget(status_label)

        widget.setLayout(layout)
        return widget

    def load_data(self):
        """异步加载所有数据，避免阻塞UI"""
        self.update_user_info()
        if hasattr(self, '_web_view') and self._web_view is not None:
            return
        
        # 显示预加载状态提示
        self._show_loading_tip("正在同步服务器数据...")
        
        # 所有模块都使用异步加载，同时触发
        QTimer.singleShot(0, self.load_products_async)
        QTimer.singleShot(0, self.load_customers_async)
        QTimer.singleShot(0, self.load_suppliers_async)
        QTimer.singleShot(0, self.load_pi_orders_async)
        QTimer.singleShot(0, self.load_purchase_orders_async)
        QTimer.singleShot(0, self.load_shipments_async)
        QTimer.singleShot(0, self.load_customer_payments_async)
        QTimer.singleShot(0, self.load_supplier_payments_async)
        QTimer.singleShot(0, self.load_inventories_async)
    
    def _show_loading_tip(self, message: str):
        """显示加载提示（状态栏或临时覆盖层）"""
        # 在主窗口底部状态栏显示提示
        if not hasattr(self, '_status_label'):
            self._status_label = QLabel()
            self._status_label.setStyleSheet("""
                QLabel {
                    background-color: #2563eb;
                    color: white;
                    padding: 8px 20px;
                    font-size: 12px;
                }
            """)
            self.statusBar().addPermanentWidget(self._status_label)
        self._status_label.setText(f"  ⏳ {message} ")
        # 3秒后自动隐藏
        QTimer.singleShot(3000, self._hide_loading_tip)
    
    def _hide_loading_tip(self):
        """隐藏加载提示"""
        if hasattr(self, '_status_label'):
            self._status_label.setText("")
    
    def _test_customer_reply(self):
        """测试客户回复API"""
        from PySide6.QtWidgets import QMessageBox
        
        if not HAS_CUSTOMER_REPLY_TEST or not CustomerReplyTester:
            QMessageBox.information(self, "提示", "测试模块未导入成功")
            return
        
        print("\n" + "="*60)
        print("开始测试客户回复API...")
        print("="*60)
        
        try:
            tester = CustomerReplyTester(self.api_client)
            results = tester.run_all_tests()
            tester.print_summary()
            
            # 显示结果
            msg = f"测试完成!\n\n通过: {results['passed']}/{results['total']}"
            if results['failed'] > 0:
                msg += f"\n失败: {results['failed']}"
            QMessageBox.information(self, "测试结果", msg)
        except Exception as e:
            QMessageBox.warning(self, "测试失败", f"测试过程中出错:\n{str(e)}")
            import traceback
            traceback.print_exc()
    
    def _update_work_status(self, task_id: str, text: str, progress: int = -1):
        """更新右下角工作状态"""
        if not hasattr(self, '_work_status_manager'):
            self._work_status_manager = {}
        
        self._work_status_manager[task_id] = {
            "text": text,
            "progress": progress,
            "timestamp": time.time()
        }
        
        self._refresh_work_status_display()
    
    def _remove_work_status(self, task_id: str):
        """移除工作状态"""
        if hasattr(self, '_work_status_manager') and task_id in self._work_status_manager:
            del self._work_status_manager[task_id]
            self._refresh_work_status_display()
    
    def _refresh_work_status_display(self):
        """刷新工作状态显示"""
        if not hasattr(self, '_work_status_label'):
            self._work_status_label = QLabel()
            self._work_status_label.setStyleSheet("""
                QLabel {
                    background-color: #059669;
                    color: white;
                    padding: 5px 15px;
                    font-size: 11px;
                }
            """)
            self.statusBar().insertPermanentWidget(1, self._work_status_label)
        
        if not hasattr(self, '_work_status_manager') or not self._work_status_manager:
            self._work_status_label.setText("  ✓ 就绪 ")
            return
        
        parts = []
        for item in self._work_status_manager.values():
            text = item["text"]
            progress = item["progress"]
            if progress >= 0:
                parts.append(f"{text} ({progress}%)")
            else:
                parts.append(text)
        
        # 限制显示长度
        display_text = " | ".join(parts[:3])
        if len(parts) > 3:
            display_text += f" (+{len(parts)-3} 更多)"
        
        self._work_status_label.setText(f"  🔄 {display_text} ")
        
        # 清理过期状态（超过5分钟）
        now = time.time()
        expired = [k for k, v in self._work_status_manager.items() 
                   if now - v["timestamp"] > 300]
        for k in expired:
            del self._work_status_manager[k]
        if expired:
            self._refresh_work_status_display()

    def update_user_info(self):
        """更新用户信息显示"""
        if hasattr(self.api_client, 'current_user') and self.api_client.current_user:
            user = self.api_client.current_user
            self.user_info_label.setText(f"👤 {user.get('real_name', '用户')}")
            
            if user.get('is_admin'):
                self.admin_mode_label.setText("🔑 管理员模式")
            else:
                self.admin_mode_label.setText("👤 普通用户")
        else:
            self.user_info_label.setText("👤 未登录")
            self.admin_mode_label.setText("")
    
    def logout(self):
        """退出登录"""
        reply = QMessageBox.question(
            self, 
            "确认退出", 
            "确定要退出登录吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            if hasattr(self.api_client, 'logout'):
                self.api_client.logout()
            self.close()

    def open_bug_report(self):
        """打开 Bug 报告对话框"""
        dialog = BugReportDialog(self.api_client, self)
        dialog.exec()

    def open_settings(self):
        """打开设置对话框"""
        dialog = SettingsDialog(self.api_client, self, web_view=getattr(self, '_web_view', None))
        if dialog.exec():
            # 刷新全局变量
            self.load_globals()
    
    def create_order_summary_tab(self):
        """创建订单总表Tab - 使用 OrderSummaryTab 组件
        
        2026-06-04: 重构为使用独立模块
        """
        # 使用新的 OrderSummaryTab 组件
        self._order_summary_tab = OrderSummaryTab(self.api_client, self)
        
        # 保持原有信号连接（兼容性）
        self._order_summary_tab.order_list_clicked.connect(self._on_order_list_click)
        self._order_summary_tab.order_list_double_clicked.connect(self._on_order_list_double_click)
        self._order_summary_tab.order_detail_double_clicked.connect(self._on_order_detail_double_click)
        self._order_summary_tab.piActionRequested.connect(self._on_pi_action_requested)
        # 详情面板操作按钮
        self._order_summary_tab.purchaseRequested.connect(self._on_purchase_clicked)
        # 2026-06-11 任务 9：tab 内部已处理『采购全部/采购该产品/重新采购』，主窗仅订阅完成信号用于刷新
        self._order_summary_tab.purchaseCompleted.connect(self._on_purchase_completed)
        # 2026-06-23 入库后自动刷新库存管理 Tab（无需手动按刷新）
        self._order_summary_tab.itemUpdated.connect(self._on_inventory_refresh_requested)
        self._order_summary_tab.supplementRequested.connect(self._on_supplement_products)
        self._order_summary_tab.shipmentRequested.connect(self._on_shipment_from_order_summary)
        self._order_summary_tab.replyExportRequested.connect(self._on_reply_export_clicked)
        self._order_summary_tab.backToOrderListRequested.connect(self._back_to_order_list)
        # 需求#41：添加付款按钮 → 打开收款对话框
        self._order_summary_tab.paymentAddRequested.connect(self._on_payment_add_from_order)
        
        # 向后兼容：暴露旧属性供遗留代码使用
        # 这些属性指向新模块的内部组件，使旧代码可以继续工作
        self.order_list_table = self._order_summary_tab.get_list_panel().get_table()
        self.order_detail_table = self._order_summary_tab.get_detail_panel().get_table()
        self.order_list_panel = self._order_summary_tab.get_list_panel()
        self.order_detail_panel = self._order_summary_tab.get_detail_panel()

        # 2026-06-23 修复：保存正式纪录成功后刷新产品管理列表
        if self.order_detail_panel:
            self.order_detail_panel.formalRecordSaved.connect(
                lambda: self.load_products(use_cache=False)
            )
            # 2026-07-02：连接右键菜单信号
            self.order_detail_panel.editProductRequested.connect(self._on_edit_product)
            self.order_detail_panel.changeSupplierRequested.connect(self._on_change_supplier_from_menu)
            self.order_detail_panel.purchaseSnapshotRequested.connect(self._on_purchase_snapshot)
            self.order_detail_panel.openShopUrlRequested.connect(self._on_open_shop_url)

        # 兼容：初始化遗留属性
        self._order_summary_orders = []
        self._order_summary_filtered = []
        self._selected_order_index = None
        self._order_summary_view_mode = "list"
        self._current_order_detail = None
        
        # 创建并返回组件
        self._order_summary_tab.create()
        
        # 连接列表加载完成回调以更新兼容属性
        self._order_summary_tab.get_service().data_loaded.connect(self._on_service_data_loaded)
        
        return self._order_summary_tab

    def _create_web_order_tab(self):
        """创建 Web 订单 Tab - 使用 QWebEngineView 加载 Vue 前端"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 0)

        if HAS_WEB_CONTAINER:
            # Web 容器：加载 Vue 前端
            from config import Config
            remote_url = getattr(Config, 'WEB_FRONTEND_URL', 'https://piapi.wakabashia.tj.cn')
            self._web_view = WebContainerView(remote_url, self)
            layout.addWidget(self._web_view)
        else:
            # 模块不可用，显示提示
            hint = QLabel("Web 容器模块不可用（缺少 PySide6 QtWebEngine 组件）")
            hint.setAlignment(Qt.AlignmentFlag.AlignCenter)
            hint.setStyleSheet("color: #666; font-size: 14px;")
            layout.addWidget(hint)

        return widget
    
    def _on_service_data_loaded(self, orders):
        """订单数据加载完成，更新兼容属性"""
        self._order_summary_orders = orders
        self._order_summary_filtered = orders

    def _update_order_summary_view(self):
        """更新订单总表视图显示 - 委托给 OrderSummaryTab"""
        if hasattr(self, '_order_summary_tab') and self._order_summary_tab is not None:
            self._order_summary_tab.refresh_data()

    # ------------------------------------------------------------------
    # 2026-07-02：订单产品右键菜单槽函数
    # ------------------------------------------------------------------
    def _on_edit_product(self, row, column):
        """打开编辑订单产品 Dialog"""
        if self.order_detail_panel:
            self.order_detail_panel.open_edit_dialog(
                row, focus_column=column if column >= 0 else None
            )

    def _on_change_supplier_from_menu(self, row):
        """右键菜单：更换供应商"""
        if not self.order_detail_panel:
            return
        item = self.order_detail_panel.get_item_at_row(row)
        if not item:
            return
        from widgets.supplier_change_dialog import SupplierChangeDialog
        dlg = SupplierChangeDialog(item, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            data = dlg.get_data()
            try:
                self.api_client.change_supplier(item["id"], data)
                QMessageBox.information(self, "成功", "供应商已更换，采购单已重新生成")
                self.order_detail_panel.show_order_detail(
                    self.order_detail_panel._current_order,
                    self.order_detail_panel._current_items,
                )
            except Exception as e:
                QMessageBox.warning(self, "失败", str(e))

    def _on_purchase_snapshot(self, row):
        """右键菜单：采购快照"""
        if not self.order_detail_panel:
            return
        item = self.order_detail_panel.get_item_at_row(row)
        if not item:
            return
        from widgets.purchase_snapshot_dialog import PurchaseSnapshotDialog
        dlg = PurchaseSnapshotDialog(item, parent=self)
        dlg.exec()

    def _on_open_shop_url(self, row):
        """右键菜单：访问店铺网站"""
        if not self.order_detail_panel:
            return
        item = self.order_detail_panel.get_item_at_row(row)
        if not item:
            return
        url = item.get("shop_url", "")
        if not url:
            QMessageBox.information(self, "提示", "店铺链接为空")
            return
        from PySide6.QtGui import QDesktopServices
        from PySide6.QtCore import QUrl
        QDesktopServices.openUrl(QUrl(url))

    def _on_shipment_from_order_summary(self):
        """订单总表模式一：出货按钮点击处理 - 跳转到出货Tab"""
        # 跳转到出货Tab
        if hasattr(self, 'content_stack'):
            for i in range(self.content_stack.count()):
                widget = self.content_stack.widget(i)
                if widget and hasattr(widget, '__class__') and 'shipment' in str(widget.__class__.__name__).lower():
                    self.content_stack.setCurrentIndex(i)
                    return
            # 如果没找到，尝试用固定索引
            try:
                self.content_stack.setCurrentIndex(6)  # 出货Tab
            except Exception:
                pass

    def _on_pi_action_requested(self, order: dict, mode: str):
        """PI操作按钮点击回调（从 OrderListPanel 转发）[6.0.2]"""
        from PySide6.QtWidgets import QMessageBox
        
        order_id = order.get('pi_id') or order.get('id')
        pi_no = order.get('pi_no') or order.get('order_no', '')
        
        if not order_id:
            QMessageBox.warning(self, "错误", "订单数据缺少ID")
            return
        
        dialog_map = {
            'order': ("确认完成下单", f"确认完成下单？\n临时 PI 号将正式化。\n当前 PI 号：{pi_no}"),
            'ship':  ("确认完成出货", f"确认完成出货？\n订单状态将切换为「出货完毕」。\n当前 PI 号：{pi_no}"),
            'draft': ("确认重新生成", f"订单已有 PI 号：{pi_no}\n\n是否重新生成？"),
        }
        title, msg = dialog_map.get(mode, ("确认操作", f"是否继续？\n当前 PI 号：{pi_no}"))
        
        reply = QMessageBox.question(self, title, msg,
                                      QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        try:
            result = self.api_client.generate_pi(order_id, force=True)
            if result.get('success'):
                new_pi_no = result.get('pi_no', '')
                msg_map = {
                    'order': f"PI 号已正式化：{new_pi_no}\n订单状态：进行中-待出货",
                    'ship':  "出货完成！\n订单状态：出货完毕",
                    'draft': f"PI 号已生成：{new_pi_no}",
                }
                QMessageBox.information(self, "成功", msg_map.get(mode, "操作成功"))
                # 刷新数据
                self._order_summary_tab.refresh_data()
            else:
                QMessageBox.warning(self, "失败", result.get('error', '未知错误'))
        except Exception as e:
            QMessageBox.warning(self, "错误", f"操作失败: {e}")
    
    def _on_order_list_click(self, row, column):
        """单击订单列表行，仅选中行（不显示详情）"""
        if row < 0 or row >= len(self._order_summary_filtered):
            return
        
        if column == 0:
            return

        self._selected_order_index = row
        order = self._order_summary_filtered[row]
        self.order_list_table.selectRow(row)
        logger.info(f"[订单总表] 单击选中: index={row}, order_no={order.get('order_no')}")

    def _back_to_order_list(self):
        """返回订单列表视图"""
        self._order_summary_view_mode = "list"
        self._update_order_summary_view()

    def _on_payment_add_from_order(self, order: dict):
        """需求#41：订单总表"添加付款"按钮 → 打开客户收款对话框"""
        try:
            from widgets.order_summary.customer_payment_dialog import CustomerPaymentDialog
            dialog = CustomerPaymentDialog(order, self.api_client, self)
            if dialog.exec() == dialog.DialogCode.Accepted:
                # 刷新订单列表以更新付款状态
                self._order_summary_tab.refresh_data()
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.warning(self, "错误", f"打开收款对话框失败:\n{e}")

    def _on_purchase_clicked(self):
        """[6.2.3] 采购按钮点击薄包装（2026-06-11 任务 9）
        实际逻辑已下沉到 OrderSummaryTab._on_purchase_all_clicked；
        此处保留方法以兼容其他可能的位置调用。
        """
        # 委托给 tab 处理（如果 tab 可用）
        if hasattr(self, '_order_summary_tab') and self._order_summary_tab:
            try:
                self._order_summary_tab._on_purchase_all_clicked()
                return
            except Exception as e:
                print(f"[WARN] OrderSummaryTab._on_purchase_all_clicked 失败: {e}")

    def _on_reply_export_clicked(self, items: list):
        """[导出回复记录] 按钮点击处理"""
        try:
            from widgets.reply_export_dialog import ReplyExportDialog

            dialog = ReplyExportDialog(items=items, api_client=self.api_client, parent=self)
            dialog.setWindowModality(Qt.WindowModality.ApplicationModal)
            dialog.exec()

        except ImportError as e:
            print(f"[ERROR] 导出回复记录模块未找到: {e}")
            QMessageBox.warning(self, "错误", f"导出功能模块未找到: {str(e)}")
        except Exception as e:
            print(f"[ERROR] 打开导出回复记录对话框失败: {e}")
    
    def _on_purchase_completed(self, purchase_data):
        """[6.2.3] 采购完成回调：刷新订单详情、采购管理Tab并更新状态灯

        2026-06-23 修复：必须重新调 API 拉最新 order detail
        原代码直接用缓存的 _current_order_detail（采购前抓的）→ show_order_detail
        优先用 order.get('items') 渲染 → 用户看不到采购后的最新数据
        （采购单价、供应商、打包规格、采购选项等都不刷新）
        必须 return 到订单总表再双击进入才会重新触发 _show_order_detail。
        """
        print(f"[6.2.3] 采购完成: {purchase_data}")
        # 2026-06-23：先调 API 拉最新的 order detail，覆盖缓存
        order_id = self._current_order_detail.get('id') if self._current_order_detail else None
        if order_id:
            try:
                latest = self.api_client.get_pi_detail(order_id)
                if latest:
                    self._current_order_detail = latest
                    # 同步 _order_summary_filtered 缓存里这一行（保持选中行引用一致）
                    if (self._selected_order_index is not None
                            and self._selected_order_index < len(self._order_summary_filtered)):
                        self._order_summary_filtered[self._selected_order_index] = latest
                    print(f"[6.2.3] 采购后重新拉取 order_id={order_id} 成功")
            except Exception as e:
                print(f"[ERROR] 采购后重新拉取订单详情失败: {e}")
        # 刷新订单详情面板
        self._reload_order_detail()
        # 2026-06-09 修复：刷新采购管理Tab，使新采购记录立即可见
        self.load_purchase_orders()
        # 刷新完成后，用当前选中行更新状态灯（采购状态变为已采/蓝色）
        current_row = self.order_detail_table.currentRow()
        if current_row >= 0:
            self._update_status_indicator(current_row)
    
    def _get_item_status(self, item):
        """[6.2.2] 获取产品4个状态的字典"""
        return {
            'is_temp': item.get('is_temp', False),
            'is_purchased': item.get('is_purchased', False),
            'has_stock': item.get('has_stock', True),
            'has_invoice': item.get('has_invoice', False)
        }
    
    def _calculate_status_colors(self, status):
        """[6.2.2] 计算4个状态灯的颜色和文字"""
        colors = []
        texts = []
        # 产品类型：正式绿色，临时黄色
        is_temp = status['is_temp']
        colors.append('#22c55e' if not is_temp else '#fbbf24')
        texts.append('正式' if not is_temp else '临时')
        # 采购状态：已采购蓝色，未采购灰色
        is_purchased = status['is_purchased']
        colors.append('#3b82f6' if is_purchased else '#6b7280')
        texts.append('已采' if is_purchased else '未采')
        # 库存状态：有库存绿色，缺货红色
        has_stock = status['has_stock']
        colors.append('#22c55e' if has_stock else '#ef4444')
        texts.append('有库' if has_stock else '缺库')
        # 发票状态：有发票绿色，无发票灰色
        has_invoice = status['has_invoice']
        colors.append('#22c55e' if has_invoice else '#6b7280')
        texts.append('有票' if has_invoice else '无票')
        return colors, texts
    
    def _update_status_indicator(self, row):
        """[6.2.2] 根据当前选中的行更新状态灯（圆点+文字）"""
        if not hasattr(self, '_current_order_detail'):
            self._set_status_dots_gray()
            return

        order = self._current_order_detail
        items = order.get('items', [])

        if row < 0 or row >= len(items):
            self._set_status_dots_gray()
            return

        item = items[row]
        status = self._get_item_status(item)
        colors, texts = self._calculate_status_colors(status)

        for i, (color, text) in enumerate(zip(colors, texts)):
            self._status_dots[i].setStyleSheet(f"background-color: {color}; border-radius: 6px;")
            self._status_labels[i].setText(text)
            self._status_labels[i].setStyleSheet(f"color: {color}; font-size: 12px;")

        # 如果缺货，应用特殊样式
        if not status['has_stock']:
            self._apply_out_of_stock_style(row)
    
    def _set_status_dots_gray(self):
        """[6.2.2] 设置所有状态灯为灰色（圆点+文字）"""
        for dot in self._status_dots:
            dot.setStyleSheet("background-color: #6b7280; border-radius: 6px;")
        default_texts = ["正式", "未采", "有库", "无票"]
        for i, label in enumerate(self._status_labels):
            label.setText(default_texts[i])
            label.setStyleSheet("color: #6b7280; font-size: 12px;")
    
    def _apply_out_of_stock_style(self, row):
        """[6.2.2] 应用缺货样式（整行灰色）"""
        for col in range(self.order_detail_table.columnCount()):
            item = self.order_detail_table.item(row, col)
            if item:
                item.setForeground(QBrush(QColor("#9ca3af")))
    
    def _on_detail_table_clicked(self, item):
        """[6.2.2] 详情表点击事件，更新状态灯"""
        row = item.row()
        self._update_status_indicator(row)
    
    def _on_inventory_refresh_requested(self):
        """2026-06-23 计划 A：入库后自动刷新库存管理 Tab

        触发场景：订单详情面板『入库』成功，发出 itemUpdated 信号
        行为：异步调 load_inventories_async，整张库存表（含『最近变更』列）重渲染
        错误处理：失败仅打印日志，不阻塞用户后续操作
        """
        import logging
        logger = logging.getLogger(__name__)
        try:
            self.load_inventories_async()
            logger.info("[🖥UI] ↻ 库存管理 Tab 已触发自动刷新")
        except Exception as e:
            logger.error(f"[🖥UI] ❌ 库存刷新失败: {e}")

    def _on_supplement_products(self):
        """[6.2.1] 补充商品按钮点击处理"""
        try:
            from widgets.order_import_dialog import OrderImportDialog

            # 获取当前订单 ID
            order_id = None
            if hasattr(self, '_current_order_detail'):
                order_id = self._current_order_detail.get('id') or self._current_order_detail.get('pi_id')
            elif self._selected_order_index is not None and self._selected_order_index < len(self._order_summary_filtered):
                order = self._order_summary_filtered[self._selected_order_index]
                order_id = order.get('id') or order.get('pi_id')

            if not order_id:
                QMessageBox.warning(self, "提示", "未找到当前订单信息")
                return

            # 创建补充商品对话框
            dialog = OrderImportDialog(
                self.api_client,
                self,
                is_supplement_mode=True,
                order_id=order_id
            )
            dialog.setWindowModality(Qt.WindowModality.ApplicationModal)
            dialog.import_completed.connect(self._on_supplement_completed)
            dialog.exec()

        except ImportError as e:
            print(f"[ERROR] 补充商品对话框模块未找到: {e}")
            QMessageBox.warning(self, "错误", f"补充商品功能模块未找到: {str(e)}")
    
    def _on_supplement_completed(self, success, success_count, failed_count, errors, created_order_ids=None):
        """[6.2.1] 补充商品完成回调"""
        if success:
            print(f"[6.2.1] 补充商品完成: {success_count} 个")
            # [问题 #28] 刷新订单详情前先清除缓存
            if hasattr(self, '_order_summary_tab'):
                self._order_summary_tab.get_service().refresh_cache()
            self._reload_order_detail()

    def _on_import_completed(self, success, success_count, failed_count, errors, created_order_ids=None):
        """导入订单完成回调：成功后刷新并打开第一个新订单详情"""
        if not success:
            return

        print(f"[INFO] 订单导入完成: {success_count} 个成功, 新订单 IDs={created_order_ids}")
        self._update_order_summary_view()

        order_id = (created_order_ids or [None])[0]
        if not order_id:
            return

        try:
            detail = self.api_client.get_pi_detail(order_id)
            if detail:
                self._show_order_detail(detail)
                # 切换到订单总表 Tab，让用户看到详情
                self.switch_tab("order_summary")
                print(f"[INFO] 已打开新订单详情: PI_ID={order_id}")
            else:
                print(f"[WARN] 无法获取新订单详情: PI_ID={order_id}")
        except Exception as e:
            print(f"[ERROR] 打开新订单详情失败: {e}")

    def _reload_order_detail(self):
        """[6.2.1] 重新加载订单详情"""
        if hasattr(self, '_current_order_detail'):
            # 重新显示当前订单
            self._show_order_detail(self._current_order_detail)
    
    def _update_order_summary_view(self):
        """更新订单总表视图显示 - 委托给 OrderSummaryTab"""
        # 委托给新组件处理
        if hasattr(self, '_order_summary_tab') and self._order_summary_tab is not None:
            # 2026-06-10 修复：先刷新数据再设置视图模式
            self._order_summary_tab.refresh_data()
            self._order_summary_tab.set_view_mode(self._order_summary_view_mode)
    
    def _on_order_list_double_click(self, row, column):
        """双击订单列表行，显示订单详情"""
        if row < 0 or row >= len(self._order_summary_filtered):
            return

        if column == 0:
            return

        order = self._order_summary_filtered[row]
        self._selected_order_index = row

        self._show_order_detail(order)
        self._order_summary_view_mode = "detail"
        self._update_order_summary_view()
        logger.info(f"[订单总表] 双击显示详情: order_no={order.get('order_no')}")
    
    def _on_order_detail_double_click(self, row, column):
        """双击订单详情单元格，打开编辑对话框 - v1.1支持权限控制"""
        # 获取当前订单（优先使用 _current_order_detail，否则从 _selected_order_index 获取）
        order = self._current_order_detail
        if order is None and self._selected_order_index is not None:
            order = self._order_summary_filtered[self._selected_order_index]
        
        if order is None:
            QMessageBox.information(self, "提示", "请先选择一个订单")
            return
        
        # 获取当前订单的 items（从 _order_summary_tab 获取最新数据）
        items = []
        order_id = order.get('id')
        if order_id and hasattr(self, '_order_summary_tab'):
            items = self._order_summary_tab.get_service().get_items_by_order_id(order_id) or []
        
        # 如果还是没有 items，使用订单内嵌的 items
        if not items:
            items = order.get('items', [])

        # 计算当前行对应的 item（考虑序号列和空行情况）
        item_index = row
        if items:
            # 如果有 items，行号对应 item 的索引
            if item_index >= len(items):
                item_index = 0
            item = items[item_index]
            # 从 items 中获取 pi_invoice_item.id（后端返回的 id 字段）
            pi_item_id = item.get('id')
        else:
            # 没有 items 时，使用订单本身
            pi_item_id = order.get('id')
            item = order

        # 列42（开票情况）特殊处理
        if column == 42:
            dialog = InvoiceUploadDialog(order, row, column, self)
            dialog.exec()
            return

        # 所有其他列：打开 OrderSummaryEditDialog
        from widgets.order_summary_edit_dialog import OrderSummaryEditDialog
        dialog = OrderSummaryEditDialog(item, self.api_client, self, order=order)
        result = dialog.exec()
        # 2026-06-09 修复：保存后重新从服务器获取最新数据，避免用旧内存数据刷新导致价格/金额显示为0
        # 2026-06-22 增强：同时更新 items 列表中的数据（包装方式、采购选项等字段）
        if result == QDialog.DialogCode.Accepted:
            pi_id = order.get('pi_id') or order.get('id')
            if pi_id:
                fresh_detail = self.api_client.get_pi_detail(pi_id)
                if fresh_detail:
                    order.update(fresh_detail)
                    # ✅ 关键修复：更新 items 列表中的数据
                    # 原问题：order.update() 只更新顶层字段，不会更新 items 列表
                    # 导致：包装方式、采购选项等字段在表格中不显示新值
                    fresh_items = fresh_detail.get('items', [])
                    if fresh_items and item_index < len(fresh_items):
                        # 用服务器返回的最新数据更新当前编辑的 item
                        updated_item = fresh_items[item_index]
                        # 更新 items 列表中的对应项
                        items_in_order = order.get('items', [])
                        if items_in_order and item_index < len(items_in_order):
                            items_in_order[item_index].update(updated_item)
                            print(f"[DEBUG] 已更新 items[{item_index}] 数据: packaging={updated_item.get('packaging')}, purchase_option_name={updated_item.get('purchase_option_name')}")
                    elif fresh_items:
                        # 如果无法按索引匹配，尝试按 pi_item_id 匹配
                        pi_item_id_for_match = item.get('id') or item.get('pi_item_id')
                        if pi_item_id_for_match:
                            for fresh_item in fresh_items:
                                if str(fresh_item.get('id')) == str(pi_item_id_for_match) or str(fresh_item.get('pi_item_id')) == str(pi_item_id_for_match):
                                    # 找到匹配的 item，更新内存中的数据
                                    items_in_order = order.get('items', [])
                                    for i, old_item in enumerate(items_in_order):
                                        if str(old_item.get('id')) == str(pi_item_id_for_match) or str(old_item.get('pi_item_id')) == str(pi_item_id_for_match):
                                            old_item.update(fresh_item)
                                            print(f"[DEBUG] 已通过 ID 匹配更新 items[{i}] 数据")
                                            break
                                    break

                    # 🔧 2026-06-22 关键修复：更新 OrderService 的 items 缓存
                    # 否则下次点击该订单时，会从缓存中读到旧的items
                    if hasattr(self, '_order_summary_tab') and self._order_summary_tab is not None:
                        try:
                            service = self._order_summary_tab.get_service()
                            if service and fresh_items:
                                service.update_items_cache(pi_id, fresh_items)
                                print(f"[DEBUG] 已更新 OrderService items 缓存: order_id={pi_id}")
                        except Exception as e:
                            print(f"[WARN] 更新 OrderService 缓存失败: {e}")

        self._show_order_detail(order)
        return
    
    def _show_product_feature_dialog(self, item, order):
        """显示产品特性编辑对话框 - 仅正式产品可编辑"""
        current_feature = item.get('product_feature', '')
        
        dialog = QDialog(self)
        dialog.setWindowTitle("编辑产品特性")
        dialog.setMinimumWidth(500)
        
        layout = QVBoxLayout(dialog)
        
        label = QLabel("请输入产品特性描述:")
        layout.addWidget(label)
        
        text_edit = QTextEdit()
        text_edit.setPlainText(current_feature or '')
        text_edit.setPlaceholderText("例如: 高温耐磨、防水防尘、静音设计等...")
        layout.addWidget(text_edit)
        
        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel
        )
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            new_feature = text_edit.toPlainText().strip()
            # TODO: 调用API保存产品特性
            print(f"[DEBUG] 产品特性已更新: {new_feature}")
    
    def _open_order_edit_dialog(self, order):
        """打开订单编辑对话框"""
        customers = self.api_client.get_customers() or []
        
        dialog = OrderEditDialog(order, self, customers)
        if dialog.exec():
            updated_order = dialog.get_order()
            
            pi_id = updated_order.get('pi_id')
            if pi_id:
                try:
                    api_data = {
                        'customer_id': updated_order.get('customer_id'),
                    }
                    if updated_order.get('currency'):
                        api_data['currency'] = updated_order.get('currency')
                    if updated_order.get('remark'):
                        api_data['remark'] = updated_order.get('remark')
                    
                    result = self.api_client.update_pi(pi_id, api_data)
                    QMessageBox.information(self, "成功", "订单已保存")
                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    QMessageBox.warning(self, "错误", f"保存失败: {e}")
            
            # 更新列表中的数据
            idx = self._selected_order_index
            if idx is not None and idx < len(self._order_summary_filtered):
                self._order_summary_filtered[idx] = updated_order
                # 如果在原始列表中也存在
                pi_id = updated_order.get('pi_id')
                for i, o in enumerate(self._order_summary_orders):
                    if o.get('pi_id') == pi_id:
                        self._order_summary_orders[i] = updated_order
                        break
                # 刷新显示
                self._populate_order_list_table(self._order_summary_filtered)
                self._show_order_detail(updated_order)
                self._order_summary_status.setText(f"已更新订单: {updated_order.get('order_no', '')}")
    
    def _open_field_edit_dialog(self, order, field_name, current_value, row, column):
        """打开字段编辑对话框"""
        dialog = FieldEditDialog(field_name, current_value, self)
        if dialog.exec():
            new_value = dialog.get_value()
            # 更新订单数据
            order[field_name] = new_value
            # 刷新详情显示
            self.order_detail_table.setItem(row, column, QTableWidgetItem(str(new_value)))
            # 更新列表
            self._order_summary_filtered[self._selected_order_index] = order
    
    def _show_order_detail(self, order):
        """显示订单详情 - 委托给 OrderSummaryTab"""
        logger.info(f"[订单总表] 显示订单详情, order_no={order.get('order_no')}")
        
        # 保存当前订单数据（向后兼容）
        self._current_order_detail = order
        self._order_summary_view_mode = "detail"
        
        # 保存当前订单到 _current_order_detail，确保双击详情单元格时能找到订单
        # 注意：不清空 _selected_order_index，让它继续指向当前订单
        if self._selected_order_index is not None:
            # 保存订单到对应索引位置
            if self._selected_order_index < len(self._order_summary_filtered):
                self._order_summary_filtered[self._selected_order_index] = order
        
        # 委托给新组件处理
        if hasattr(self, '_order_summary_tab') and self._order_summary_tab is not None:
            self._order_summary_tab.show_order_detail(order)
    
    # ============================================================
    # 以下方法已迁移到 widgets/order_summary/ 模块，标记为废弃
    # 这些方法由 OrderSummaryTab、OrderListPanel、OrderDetailPanel 接管
    # ============================================================
    def format_currency_display(self, amount, currency='USD'):
        """格式化金额显示 - 统一格式 {X.XX} {CUR}"""
        if amount is None:
            return f"0.00 {currency}"
        try:
            return f"{float(amount):.2f} {currency}"
        except (ValueError, TypeError):
            return f"0.00 {currency}"
    
    def _load_product_image(self, row, col, item, order):
        """加载产品图片 - 异步加载"""
        image_label = QLabel()
        image_label.setFixedSize(84, 84)
        image_label.setAlignment(Qt.AlignCenter)
        image_label.setStyleSheet("border: 1px solid #e5e7eb; background-color: #f9fafb;")
        image_label.setCursor(Qt.CursorShape.PointingHandCursor)
        
        image_url = (
            item.get('image_url') or 
            item.get('image') or 
            item.get('product_image') or 
            item.get('pic_url') or
            order.get('image_url') or 
            order.get('image') or
            order.get('product_image')
        )
        
        if image_url:
            self.load_image_async(image_label, str(image_url))
        else:
            image_label.setText("暂无图片")
        
        self.order_detail_table.setCellWidget(row, col, image_label)
    
    def _import_order_summary(self):
        """导入订单总表 - 使用OrderImportDialog"""
        print("[INFO] 导入订单总表功能")
        
        try:
            from widgets.order_import_dialog import OrderImportDialog
            
            dialog = OrderImportDialog(self.api_client, self)
            dialog.setWindowModality(Qt.WindowModality.ApplicationModal)
            dialog.import_completed.connect(self._on_import_completed)
            dialog.exec()
            
            print("[INFO] 订单导入对话框已关闭，刷新订单总表...")
            self._update_order_summary_view()
            
        except ImportError as e:
            print(f"[ERROR] 导入订单对话框模块未找到: {e}")
            QMessageBox.warning(
                self, 
                "导入功能", 
                "导入功能模块未找到，请联系管理员。\n\n错误详情：" + str(e)
            )
        except Exception as e:
            print(f"[ERROR] 导入订单失败: {e}")
            QMessageBox.warning(self, "错误", f"导入失败: {e}")
    
    # ============================================
    # 以下订单总表方法已迁移到 OrderService / OrderSummaryTab / OrderListPanel
    # - load_order_summary → OrderService.load_full_data_async
    # - _on_order_summary_data_ready → OrderSummaryTab.load_data
    # - _calculate_order_estimates → (已移除，不再需要)
    # - _populate_order_list_table → OrderListPanel.update_table
    # - _get_order_action_state → OrderListPanel._get_order_action_state
    # - _build_order_summary_row → (已移除，被 OrderDetailPanel 替代)
    # ============================================

    def _normalize_image_url(self, image_url: str) -> str:
        """把后端返回的相对路径（/images/xxx）补全为绝对 URL"""
        if not image_url:
            return image_url
        if image_url.startswith(("http://", "https://")):
            return image_url
        base = (Config.API_BASE_URL or "").rstrip("/")
        result = f"{base}{image_url}" if image_url.startswith("/") else f"{base}/{image_url}"
        return result

    def load_image_async(self, label, image_url):
        """异步加载图片 - 使用全局线程池和图片缓存"""
        global _image_cache, _image_cache_lock

        # 规范化 URL：相对路径 → 绝对 URL
        image_url = self._normalize_image_url(image_url)
        if not image_url:
            return

        # 先检查内存缓存
        with _image_cache_lock:
            if image_url in _image_cache:
                pixmap = _image_cache[image_url]
                label.setPixmap(pixmap)
                return
        
        # 显示加载中占位符
        label.setText("...")
        
        def fetch_image():
            try:
                # 使用缓存避免重复下载
                cache_key = f"img_{hash(image_url)}"
                cached_data = cache_manager.get(cache_key, max_age=3600)  # 图片缓存1小时
                if cached_data:
                    return cached_data
                
                image_data = urllib.request.urlopen(image_url, timeout=3).read()
                cache_manager.set(cache_key, image_data)
                return image_data
            except Exception as e:
                print(f"图片加载失败: {e}")
                return None
        
        def on_done(future):
            # 检查label是否仍然有效（防止widget被销毁后访问）
            try:
                from shiboken6 import isValid
                if not label or not isValid(label):
                    return
            except ImportError:
                pass

            image_data = future.result()
            if image_data:
                try:
                    image = QImage.fromData(image_data)
                    pixmap = QPixmap.fromImage(image).scaled(56, 56, Qt.KeepAspectRatio, Qt.SmoothTransformation)

                    # 存入内存缓存
                    with _image_cache_lock:
                        if len(_image_cache) >= _MAX_IMAGE_CACHE_SIZE:
                            # 简单LRU：删除最早的
                            oldest_key = next(iter(_image_cache))
                            del _image_cache[oldest_key]
                        _image_cache[image_url] = pixmap

                    label.setPixmap(pixmap)
                    return
                except Exception as e:
                    print(f"图片处理失败: {e}")
            try:
                label.setText("暂无图片")
            except RuntimeError:
                pass  # widget已被销毁，忽略错误
        
        # 使用全局线程池执行
        future = _global_thread_pool.submit(fetch_image)
        future.add_done_callback(on_done)

    def load_products(self, use_cache: bool = True):
        import logging
        logger = logging.getLogger(__name__)
        
        logger.info("=" * 60)
        logger.info("开始加载产品列表...")
        
        self._show_loading_tip("正在加载客户产品...")
        try:
            logger.info("步骤1: 调用客户产品API...")
            resp = self.api_client.get("/customer-products", params={"page_size": 500})
            logger.info(f"API响应类型: {type(resp).__name__}, 响应内容: {resp}")

            if resp and isinstance(resp, dict):
                products = resp.get('items', [])
                logger.info(f"从dict响应中获取items, 共 {len(products)} 个产品")
            elif isinstance(resp, list):
                products = resp
                logger.info(f"从list响应直接获取, 共 {len(products)} 个产品")
            else:
                products = resp or []
                logger.warning(f"响应格式异常, 使用空列表: {type(resp)}")

            # 清理缓存
            logger.info("步骤2: 清理产品缓存...")
            cache_manager.delete(CACHE_KEYS['PRODUCTS'])

            if products is None:
                products = []
                logger.warning("产品数据为None, 已转换为空列表")

            logger.info(f"步骤3: 产品数据准备完毕, 共 {len(products)} 个产品")

            # 保存用户当前的选择状态（恢复时使用）
            current_customer_id = self.product_customer_filter.currentData()
            logger.info(f"保存当前选择: 客户ID={current_customer_id}")

            # 先断开筛选信号，避免恢复选择时触发不必要的筛选
            logger.info("步骤4: 断开筛选事件信号...")
            signals_disconnected = False
            try:
                self.product_customer_filter.currentIndexChanged.disconnect()
                signals_disconnected = True
                logger.debug("筛选事件信号已全部断开")
            except RuntimeError as e:
                logger.debug(f"断开信号时无现有连接: {e}")
            except Exception as e:
                logger.warning(f"断开信号时发生异常: {e}")

            # 清空并填充客户下拉框
            logger.info("步骤5: 填充客户下拉框...")
            self.product_customer_filter.clear()
            self.product_customer_filter.addItem("全部客户", 0)
            
            customers = self.api_client.get_customers() or []
            self._customer_filter_cache = customers
            logger.info(f"加载客户数据: {len(customers)} 个客户")
            
            for customer in customers:
                customer_name = customer.get('customer_name') or customer.get('name') or '未知客户'
                customer_id = customer.get('id')
                logger.debug(f"添加客户: id={customer_id}, name={customer_name}")
                if customer_id is not None:
                    self.product_customer_filter.addItem(customer_name, customer_id)
                else:
                    logger.warning(f"客户数据缺少id字段: {customer}")
            
            # 恢复客户选择状态（信号已断开，不会触发筛选）
            if current_customer_id and current_customer_id != 0:
                customer_index = self.product_customer_filter.findData(current_customer_id)
                if customer_index >= 0:
                    self.product_customer_filter.setCurrentIndex(customer_index)
                    logger.info(f"已恢复客户选择: ID={current_customer_id}, Index={customer_index}")
                else:
                    logger.warning(f"客户ID {current_customer_id} 不存在，恢复到'全部客户'")
                    self.product_customer_filter.setCurrentIndex(0)
            else:
                self.product_customer_filter.setCurrentIndex(0)
                logger.info("恢复客户选择: 全部客户")

            # 缓存产品数据
            logger.info("步骤6: 缓存产品数据...")
            self._product_cache = products

            # 填充表格
            logger.info("步骤7: 填充产品表格...")
            self._populate_products_table(products)
            logger.info(f"表格填充完成, 显示 {len(products)} 条记录")

            # 重新绑定筛选事件
            logger.info("步骤8: 重新绑定筛选事件...")
            self.product_customer_filter.currentIndexChanged.connect(self.filter_products)
            logger.debug("客户筛选事件已绑定")

            # 手动触发一次筛选，确保显示正确的数据
            if signals_disconnected:
                logger.info("步骤10: 手动触发筛选以应用当前筛选条件...")
                self.filter_products()

            logger.info("步骤11: 隐藏加载提示...")
            self._hide_loading_tip()

            logger.info("=" * 60)
            logger.info("产品列表加载完成!")
            logger.info("=" * 60)
            
        except Exception as e:
            logger.error(f"加载产品列表时发生异常: {type(e).__name__}: {str(e)}")
            import traceback
            logger.error(f"异常堆栈: {traceback.format_exc()}")
            self._hide_loading_tip()
            QMessageBox.warning(self, "错误", f"加载产品数据失败: {str(e)}")

    def _populate_products_table(self, products):
        """填充产品表格数据"""
        self.products_table.setRowCount(len(products))

        for row, p in enumerate(products):
            product_id = p.get('id')

            # 0: 复选框
            checkbox = QCheckBox()
            checkbox.setCheckState(Qt.CheckState.Unchecked)
            checkbox.setStyleSheet("margin-left: 50%;")
            self.products_table.setCellWidget(row, 0, checkbox)

            codes = p.get('codes', []) or []
            oes = p.get('oes', []) or []

            # 1: 客户产品编号
            primary_code = p.get('primary_code', '') or ''
            self.products_table.setItem(row, 1, QTableWidgetItem(primary_code))

            # 2: 系统编号
            full_system_code = p.get('system_code', '') or ''
            display_code = full_system_code[-9:] if len(full_system_code) >= 9 else full_system_code
            system_code_item = QTableWidgetItem(display_code)
            system_code_item.setForeground(Qt.blue)
            # 存储完整系统编号到 userData，供双击时使用
            system_code_item.setData(Qt.UserRole, full_system_code)
            self.products_table.setItem(row, 2, system_code_item)

            # 3: OE号
            primary_oe = p.get('primary_oe', '') or ''
            if len(oes) > 1:
                btn = QPushButton("多OE号")
                btn.setStyleSheet("""
                    QPushButton {
                        background-color: #3b82f6;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        padding: 4px 8px;
                        font-size: 11px;
                    }
                    QPushButton:hover { background-color: #2563eb; }
                """)
                btn.clicked.connect(lambda checked, pid=product_id, oes_list=oes: self._show_product_oe_dialog(pid, oes_list))
                self.products_table.setCellWidget(row, 3, btn)
            else:
                self.products_table.setItem(row, 3, QTableWidgetItem(primary_oe))

            # 4: 图片
            image_label = QLabel()
            image_label.setFixedSize(60, 60)
            image_label.setStyleSheet("border: 1px solid #e5e7eb;")
            image_label.setAlignment(Qt.AlignCenter)

            image_url = self._normalize_image_url(p.get('image_url'))
            if image_url:
                self.load_image_async(image_label, image_url)
            else:
                image_label.setText("暂无图片")

            image_label.mousePressEvent = lambda event, prod=p: self.view_product_images(prod)
            image_label.setCursor(Qt.PointingHandCursor)
            self.products_table.setCellWidget(row, 4, image_label)

            # 5: 产品名称
            self.products_table.setItem(row, 5, QTableWidgetItem(p.get('product_name', '') or p.get('detail_desc', '')))

            # 6: 客户型号
            customer_model = p.get('customer_model', '') or ''
            self.products_table.setItem(row, 6, QTableWidgetItem(customer_model))

            # 7: 客户名称
            # 2026-06-23 修复：直接用后端 _build_response 返回的 customer_name
            # 原代码从 _customer_filter_cache 二次匹配，但 _customer_filter_cache 实际可能空 / id 类型不匹配 → 显示 "-"
            customer_display_name = p.get('customer_name') or '-'
            if customer_display_name == '-':
                # 兜底：再尝试从 cache 查一次
                customer_id = p.get('customer_id')
                if customer_id is not None and hasattr(self, '_customer_filter_cache'):
                    for customer in self._customer_filter_cache:
                        if customer.get('id') == customer_id or str(customer.get('id')) == str(customer_id):
                            customer_display_name = customer.get('customer_name', customer.get('name', '-'))
                            break
            self.products_table.setItem(row, 7, QTableWidgetItem(customer_display_name or '-'))

            # 8: 类别
            # 2026-06-23 修复：直接用后端 _build_response 返回的 category_name
            # 原代码查 _category_filter_cache，但该属性从未被赋值（hasattr 永远 False）→ 一直显示 "-"
            category_name = p.get('category_name') or '-'
            if category_name == '-':
                # 兜底：再尝试从 cache 查一次
                cat_code = p.get('category_id', '')
                if cat_code and hasattr(self, '_category_filter_cache'):
                    for category in self._category_filter_cache:
                        if category.get('code') == cat_code or str(category.get('id')) == str(cat_code):
                            category_name = category.get('name', '-')
                            break
            self.products_table.setItem(row, 8, QTableWidgetItem(category_name or '-'))

            # 9: 颜色
            self.products_table.setItem(row, 9, QTableWidgetItem(p.get('color', '') or '-'))

            # 10: 品牌
            self.products_table.setItem(row, 10, QTableWidgetItem(p.get('brand', '') or '-'))

            # 11: USD 价格
            price_usd = p.get('price_usd', 0) or 0
            price_usd_item = QTableWidgetItem(f"${price_usd:.2f}" if price_usd else "-")
            price_usd_item.setTextAlignment(Qt.AlignRight)
            self.products_table.setItem(row, 11, price_usd_item)

            # 12: RMB 价格
            price_rmb = p.get('price_rmb', 0) or 0
            price_rmb_item = QTableWidgetItem(f"¥{price_rmb:.2f}" if price_rmb else "-")
            price_rmb_item.setTextAlignment(Qt.AlignRight)
            self.products_table.setItem(row, 12, price_rmb_item)

            # 13: 规格描述
            specifications = p.get('specifications', '') or ''
            if specifications and len(specifications) > 20:
                display_spec = specifications[:20] + '...'
            else:
                display_spec = specifications or '-'
            self.products_table.setItem(row, 13, QTableWidgetItem(display_spec))

            # 14: 隐藏的产品ID（用于双击编辑）
            self.products_table.setItem(row, 14, QTableWidgetItem(str(p.get('id', ''))))

            # 14: 编辑按钮
            def create_edit_callback(product):
                return lambda: self.edit_product(product)

            action_widget = QWidget()
            action_widget.setProperty("product_id", product_id)
            action_layout = QHBoxLayout()
            action_layout.setContentsMargins(0, 0, 0, 0)

            edit_btn = QPushButton("编辑")
            edit_btn.setFixedWidth(50)
            edit_btn.clicked.connect(create_edit_callback(p))
            action_layout.addWidget(edit_btn)

            action_widget.setLayout(action_layout)
            self.products_table.setCellWidget(row, 14, action_widget)

        self.select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)

    def filter_by_customer(self):
        """根据客户筛选产品"""
        customer_id = self.product_customer_filter.currentData()
        if customer_id == 0:
            self.search_products()
            return

        keyword = self.search_input.text().strip()
        category_id = self.category_filter.currentData()
        category_id = category_id if category_id != 0 else None

        try:
            products = self.api_client.search_products(keyword, category_id, customer_id)
            self._populate_products_table(products)
        except Exception as e:
            if hasattr(self, '_logger'):
                self._logger.error(f"客户筛选失败: {str(e)}")
            print(f"客户筛选失败: {e}")

    def filter_products(self):
        """根据筛选条件过滤产品列表（本地缓存过滤）"""
        if not self._product_cache:
            return

        customer_id = self.product_customer_filter.currentData()

        filtered = []
        for product in self._product_cache:
            if customer_id and product.get('customer_id') != customer_id:
                continue

            filtered.append(product)

        self._populate_products_table(filtered)

    def view_product_images(self, product):
        """查看产品图片"""
        product_id = product.get('id')
        product_code = product.get('product_code', '')
        image_url = self._normalize_image_url(product.get('default_image_url'))
        
        if not image_url:
            QMessageBox.information(self, "图片查看", f"产品 {product_code} 暂无图片")
            return
        
        try:
            # 创建图片查看对话框
            dialog = QDialog(self)
            dialog.setWindowTitle(f"产品图片 - {product_code}")
            dialog.setMinimumSize(600, 600)
            
            layout = QVBoxLayout()
            
            # 图片标签
            image_label = QLabel()
            image_label.setAlignment(Qt.AlignCenter)
            
            # 加载图片
            image_data = urllib.request.urlopen(image_url).read()
            image = QImage.fromData(image_data)
            pixmap = QPixmap.fromImage(image).scaled(580, 580, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            image_label.setPixmap(pixmap)
            
            layout.addWidget(image_label)
            dialog.setLayout(layout)
            dialog.exec()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载图片失败: {str(e)}")
    
    def confirm_product_import(self, product):
        """确认产品导入"""
        print(f"DEBUG - confirm_product_import called for product: {product.get('product_code')}")

        product_id = product.get('id')
        product_code = product.get('product_code', '')

        # 检查产品是否已经导入
        if product.get('is_imported') == 1:
            QMessageBox.warning(self, "提示", f"产品 {product_code} 已经确认导入，无需重复操作")
            return

        reply = QMessageBox.question(
            self, "确认导入",
            f"确定要确认产品 {product_code} 已导入吗？确认后将无法修改（普通用户）",
            QMessageBox.Ok | QMessageBox.Cancel
        )

        if reply == QMessageBox.Ok:
            try:
                print(f"DEBUG - Calling API to confirm import for product_id: {product_id}")
                result = self.api_client.confirm_product_import(product_id)
                print(f"DEBUG - API response: {result}")
                QMessageBox.information(self, "成功", "产品导入已确认")
                self.load_products()
            except Exception as e:
                print(f"DEBUG - Confirm import failed: {str(e)}")
                QMessageBox.warning(self, "错误", f"确认导入失败: {str(e)}")
    
    def cancel_product_import(self, product):
        """取消产品导入确认（仅管理员）"""
        print(f"DEBUG - cancel_product_import called for product: {product.get('product_code')}")
        
        product_id = product.get('id')
        product_code = product.get('product_code', '')
        
        reply = QMessageBox.question(
            self, "取消导入", 
            f"确定要取消产品 {product_code} 的导入确认吗？",
            QMessageBox.Ok | QMessageBox.Cancel
        )
        
        if reply == QMessageBox.Ok:
            try:
                print(f"DEBUG - Calling API to cancel import for product_id: {product_id}")
                result = self.api_client.cancel_product_import(product_id)
                print(f"DEBUG - API response: {result}")
                QMessageBox.information(self, "成功", "产品导入确认已取消")
                self.load_products()
            except Exception as e:
                print(f"DEBUG - Cancel import failed: {str(e)}")
                QMessageBox.warning(self, "错误", f"取消导入失败: {str(e)}")

    def batch_confirm_import_products(self):
        """批量确认产品导入"""
        print("DEBUG - batch_confirm_import_products called")
        
        selected_products = []
        already_imported = []
        
        # 获取选中的产品
        for row in range(self.products_table.rowCount()):
            checkbox = self.products_table.cellWidget(row, 0)
            if checkbox and checkbox.isChecked():
                # 获取产品ID（从隐藏的第12列）
                id_item = self.products_table.item(row, 12)
                if not id_item:
                    continue
                product_id = int(id_item.text())
                product_code = self.products_table.item(row, 1)  # 客户产品编号在第2列（索引1）
                if not product_code:
                    product_code = self.products_table.item(row, 12)  # 或者从隐藏列获取
                else:
                    product_code = product_code.text()
                
                # 检查是否已导入（从数据源获取）
                # 这里需要重新获取产品数据来检查is_imported状态
                try:
                    product_detail = self.api_client.get_product_detail(product_id)
                    if product_detail.get('is_imported') == 1:
                        already_imported.append(product_code)
                    else:
                        selected_products.append({'id': product_id, 'code': product_code})
                except Exception as e:
                    print(f"DEBUG - Failed to get product detail for {product_id}: {str(e)}")
                    selected_products.append({'id': product_id, 'code': product_code})
        
        if not selected_products and not already_imported:
            QMessageBox.warning(self, "提示", "请先选择要确认导入的产品")
            return
        
        # 如果有已导入的产品，显示提示
        message = ""
        if already_imported:
            message = f"以下产品已确认导入，将跳过：\n{', '.join(already_imported)}\n\n"
        
        if not selected_products:
            QMessageBox.information(self, "提示", message + "没有需要确认导入的产品")
            return
        
        reply = QMessageBox.question(
            self, "批量确认导入",
            f"{message}确定要确认选中的 {len(selected_products)} 个产品已导入吗？\n确认后将无法修改（普通用户）",
            QMessageBox.Ok | QMessageBox.Cancel
        )
        
        if reply == QMessageBox.Ok:
            success_count = 0
            failed_count = 0
            failed_list = []
            
            for product in selected_products:
                try:
                    print(f"DEBUG - Calling API to confirm import for product_id: {product['id']}")
                    result = self.api_client.confirm_product_import(product['id'])
                    print(f"DEBUG - API response: {result}")
                    success_count += 1
                except Exception as e:
                    print(f"DEBUG - Confirm import failed for {product['code']}: {str(e)}")
                    failed_count += 1
                    failed_list.append(product['code'])
            
            # 显示结果
            result_msg = f"批量确认导入完成\n成功：{success_count} 个\n失败：{failed_count} 个"
            if failed_list:
                result_msg += f"\n失败产品：{', '.join(failed_list)}"
            
            QMessageBox.information(self, "批量确认导入结果", result_msg)
            self.load_products()

    def search_products(self):
        keyword = self.search_input.text().strip()
        
        # 两级类别筛选
        category_code = self.category_filter_level2.currentData()
        if category_code == 0 or category_code is None:
            category_code = self.category_filter_level1.currentData()
            if category_code == 0:
                category_code = None
        
        try:
            print(f"搜索参数 - 关键词: '{keyword}', 类别代码: {category_code}")
            # 2026-06-14 改用 category_code 参数（前端下拉框存 code 字符串）
            products = self.api_client.search_products(keyword, category_code=category_code)
            print(f"搜索结果数量: {len(products)}")
            self.products_table.setRowCount(len(products))
            for row, p in enumerate(products):
                product_id = p.get('id')

                # 0: 复选框
                checkbox = QCheckBox()
                checkbox.setCheckState(Qt.CheckState.Unchecked)
                checkbox.setStyleSheet("margin-left: 50%;")
                self.products_table.setCellWidget(row, 0, checkbox)

                # 获取OE和客户关联
                oe_list = []
                customer_product_list = []
                try:
                    oe_list = self.api_client.get_product_oes(product_id) or []
                    customer_product_list = self.api_client.get_product_customers(product_id) or []
                except:
                    pass

                # 1: 客户产品编号
                customer_product_code = ""
                if customer_product_list:
                    first_pc = customer_product_list[0]
                    full_code = first_pc.get('customer_product_code', '')
                    customer_code = first_pc.get('customer_code', '')
                    if full_code and customer_code:
                        customer_product_code = full_code.replace(customer_code, "", 1).lstrip("-")
                    else:
                        customer_product_code = full_code or ""
                self.products_table.setItem(row, 1, QTableWidgetItem(customer_product_code))

                # 2: 系统编号（与普通加载对齐 - 2026-06-14 修复错位）
                self.products_table.setItem(row, 2, QTableWidgetItem(p.get('product_code', '') or '-'))

                # 3: OE号
                primary_oe = next((oe for oe in oe_list if oe.get('is_primary')), None)
                if len(oe_list) > 1:
                    btn = QPushButton("多OE号")
                    btn.setStyleSheet("""
                        QPushButton {
                            background-color: #3b82f6;
                            color: white;
                            border: none;
                            border-radius: 4px;
                            padding: 4px 8px;
                            font-size: 11px;
                        }
                        QPushButton:hover { background-color: #2563eb; }
                    """)
                    btn.clicked.connect(lambda checked, pid=product_id, oes=oe_list: self._show_product_oe_dialog(pid, oes))
                    self.products_table.setCellWidget(row, 3, btn)
                elif primary_oe:
                    self.products_table.setItem(row, 3, QTableWidgetItem(primary_oe.get('oe_number', '')))
                else:
                    self.products_table.setItem(row, 3, QTableWidgetItem(p.get('oe_number', '') or '-'))

                # 4: 图片
                image_label = QLabel()
                image_label.setFixedSize(60, 60)
                image_label.setStyleSheet("border: 1px solid #e5e7eb;")
                image_label.setAlignment(Qt.AlignCenter)

                image_url = self._normalize_image_url(p.get('default_image_url'))
                if image_url:
                    try:
                        image_data = urllib.request.urlopen(image_url).read()
                        image = QImage.fromData(image_data)
                        pixmap = QPixmap.fromImage(image).scaled(56, 56, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                        image_label.setPixmap(pixmap)
                    except Exception as e:
                        image_label.setText("暂无图片")
                else:
                    image_label.setText("暂无图片")

                image_label.setCursor(Qt.PointingHandCursor)
                self.products_table.setCellWidget(row, 4, image_label)

                # 5: 产品名称
                self.products_table.setItem(row, 5, QTableWidgetItem(p.get('detail_desc', '')))

                # 6: 客户型号
                customer_model = ""
                if customer_product_list:
                    customer_model = customer_product_list[0].get('customer_model', '') or ""
                self.products_table.setItem(row, 6, QTableWidgetItem(customer_model))

                # 7: 客户名称
                # 2026-06-23 修复：直接用后端 _build_response 返回的 customer_name
                # 原代码用 customer_product_list[0].customer_code（P-00001）跟 self.customers 里的 customer_code（ME9）匹配 → 永远不命中
                customer_name = p.get('customer_name') or ''
                if not customer_name:
                    if customer_product_list:
                        customer_code = customer_product_list[0].get('customer_code', '')
                        for c in getattr(self, 'customers', []) or []:
                            if c.get('customer_code') == customer_code or str(c.get('id')) == str(customer_code):
                                customer_name = c.get('customer_name', c.get('name', '')) or ''
                                break
                self.products_table.setItem(row, 7, QTableWidgetItem(customer_name or '-'))

                # 8: 类别
                # 2026-06-23 修复：直接用后端 _build_response 返回的 category_name
                # 原代码查 _category_filter_cache，但该属性从未被赋值（getattr 默认 []）→ 一直显示 "-"
                category_name = p.get('category_name') or '-'
                if category_name == '-':
                    cat_code = p.get('category_id', '')
                    if cat_code:
                        for cat in getattr(self, '_category_filter_cache', []) or []:
                            if cat.get('code') == cat_code or str(cat.get('id')) == str(cat_code):
                                category_name = cat.get('name', '-')
                                break
                self.products_table.setItem(row, 8, QTableWidgetItem(category_name))

                # 9: 颜色
                self.products_table.setItem(row, 9, QTableWidgetItem(p.get('color', '') or '-'))

                # 10: 品牌
                self.products_table.setItem(row, 10, QTableWidgetItem(p.get('brand', '') or '-'))

                # 11: USD 价格
                price_usd = p.get('price_usd', 0) or 0
                if not price_usd:
                    exw_inc = p.get('exw_price_incl', 0) or 0
                    fob_exc = p.get('fob_price_excl', 0) or 0
                    exw_exc = p.get('exw_price_excl', 0) or 0
                    price_usd = exw_inc or fob_exc or exw_exc
                price_usd_item = QTableWidgetItem(f"${price_usd:.2f}" if price_usd else "-")
                price_usd_item.setTextAlignment(Qt.AlignRight)
                self.products_table.setItem(row, 11, price_usd_item)

                # 12: RMB 价格
                price_rmb = p.get('price_rmb', 0) or 0
                price_rmb_item = QTableWidgetItem(f"¥{price_rmb:.2f}" if price_rmb else "-")
                price_rmb_item.setTextAlignment(Qt.AlignRight)
                self.products_table.setItem(row, 12, price_rmb_item)

                # 13: 规格描述
                specifications = p.get('specifications', '') or ''
                display_spec = (specifications[:20] + '...') if len(specifications) > 20 else (specifications or '-')
                self.products_table.setItem(row, 13, QTableWidgetItem(display_spec))

                # 14: 操作区 - 编辑按钮（与 _populate_products_table 一致）
                edit_btn = QPushButton("编辑")
                edit_btn.setProperty("product_id", product_id)
                edit_btn.setFixedWidth(50)
                edit_btn.clicked.connect(lambda _, prod=p: self.edit_product(prod))
                self.products_table.setCellWidget(row, 14, edit_btn)

            self.select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
        except Exception as e:
            print(f"搜索错误: {str(e)}")
            QMessageBox.warning(self, "错误", f"搜索产品失败: {str(e)}")

    def on_category_level1_changed(self):
        """一级类别变化时，更新二级类别选项"""
        parent_code = self.category_filter_level1.currentData()
        self.category_filter_level2.clear()
        self.category_filter_level2.addItem("全部子类", 0)
        print(f"[产品管理] on_category_level1_changed 触发, parent_code={parent_code}")
        if parent_code and parent_code != 0:
            child_options = get_child_category_options(parent_code)
            for code, name in child_options:
                self.category_filter_level2.addItem(name, code)
    
    def reset_search(self):
        """重置搜索条件"""
        self.search_input.clear()
        if hasattr(self, 'category_filter_level1'):
            self.category_filter_level1.setCurrentIndex(0)
        if hasattr(self, 'category_filter_level2'):
            self.category_filter_level2.clear()
            self.category_filter_level2.addItem("全部子类", 0)
        self.load_products()

    def load_product_categories(self):
        """从数据库加载产品类别"""
        try:
            categories = self.api_client.get("/product-categories")
            # 清空现有选项（保留"全部分类"）
            self.category_filter.clear()
            self.category_filter.addItem("全部分类", 0)
            for cat in categories:
                self.category_filter.addItem(cat.get('name', ''), cat.get('id', 0))
        except Exception as e:
            print(f"加载产品类别失败: {str(e)}")

    def toggle_select_all_products(self, state):
        for row in range(self.products_table.rowCount()):
            checkbox = self.products_table.cellWidget(row, 0)
            if checkbox:
                checkbox.setCheckState(Qt.CheckState(state))

    def get_selected_product_ids(self):
        selected_ids = []
        for row in range(self.products_table.rowCount()):
            checkbox = self.products_table.cellWidget(row, 0)
            if checkbox and checkbox.isChecked():
                # 从第14列操作区获取产品ID
                action_widget = self.products_table.cellWidget(row, 14)
                product_id = None
                if action_widget:
                    product_id = action_widget.property("product_id")
                if not product_id:
                    # 兜底：从操作区按钮查找
                    btn = self.products_table.cellWidget(row, 14)
                    if isinstance(btn, QPushButton):
                        product_id = btn.property("product_id")
                if product_id:
                    selected_ids.append(int(product_id))
        return selected_ids

    def batch_toggle_product_status(self):
        selected_ids = self.get_selected_product_ids()
        if not selected_ids:
            QMessageBox.warning(self, "警告", "请先选择要操作的产品")
            return

        reply = QMessageBox.question(
            self, "确认操作", 
            f"确定要切换选中的 {len(selected_ids)} 个产品的状态吗？",
            QMessageBox.Ok | QMessageBox.Cancel
        )

        if reply == QMessageBox.Ok:
            try:
                for product_id in selected_ids:
                    self.api_client.update_product_status(product_id, None)
                QMessageBox.information(self, "成功", f"已成功切换 {len(selected_ids)} 个产品的状态")
                self.load_products()
            except Exception as e:
                QMessageBox.warning(self, "错误", f"批量操作失败: {str(e)}")

    def batch_delete_products(self):
        selected_ids = self.get_selected_product_ids()
        if not selected_ids:
            QMessageBox.warning(self, "警告", "请先选择要删除的产品")
            return

        reply = QMessageBox.question(
            self, "确认删除", 
            f"确定要删除选中的 {len(selected_ids)} 个产品吗？此操作不可恢复！",
            QMessageBox.Ok | QMessageBox.Cancel
        )

        if reply == QMessageBox.Ok:
            try:
                success_count = 0
                fail_count = 0
                for product_id in selected_ids:
                    print(f"DEBUG - 删除产品ID: {product_id}")
                    result = self.api_client.delete_product(product_id)
                    if result:
                        success_count += 1
                    else:
                        fail_count += 1
                
                if fail_count > 0:
                    QMessageBox.warning(self, "部分失败", f"成功删除 {success_count} 个，失败 {fail_count} 个")
                else:
                    QMessageBox.information(self, "成功", f"已成功删除 {success_count} 个产品（将在24小时后物理删除）")
                # 强制刷新，不使用缓存
                self.load_products(use_cache=False)
            except Exception as e:
                QMessageBox.warning(self, "错误", f"批量删除失败: {str(e)}")

    def view_product_images(self, product):
        """查看产品图片"""
        # 获取产品图片信息
        image_url = self._normalize_image_url(product.get('image_url'))
        sub_images = product.get('sub_images', [])
        
        # 合并主图和副图
        all_images = []
        if image_url:
            all_images.append(('主图', image_url))
        if sub_images:
            for i, url in enumerate(sub_images):
                all_images.append((f'副图{i+1}', url))
        
        if not all_images:
            QMessageBox.information(self, "图片查看", f"产品暂无图片")
            return
        
        try:
            # 创建图片查看对话框
            dialog = QDialog(self)
            dialog.setWindowTitle("产品图片")
            dialog.setMinimumSize(600, 600)
            
            layout = QVBoxLayout()
            
            scroll_area = QScrollArea()
            scroll_area.setWidgetResizable(True)
            scroll_content = QWidget()
            scroll_layout = QVBoxLayout(scroll_content)
            
            import requests
            for i, (name, url) in enumerate(all_images, 1):
                try:
                    response = requests.get(url, timeout=10)
                    pixmap = QPixmap()
                    pixmap.loadFromData(response.content)
                    
                    if not pixmap.isNull():
                        label = QLabel(f"{name}")
                        label.setFont(get_font(12, QFont.Weight.Bold))
                        scroll_layout.addWidget(label)
                        
                        image_label = QLabel()
                        image_label.setPixmap(pixmap.scaled(500, 400, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
                        image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                        scroll_layout.addWidget(image_label)
                except Exception as e:
                    scroll_layout.addWidget(QLabel(f"{name} 加载失败"))
            
            scroll_area.setWidget(scroll_content)
            layout.addWidget(scroll_area)
            
            close_btn = QPushButton("关闭")
            close_btn.clicked.connect(dialog.close)
            layout.addWidget(close_btn)
            
            dialog.setLayout(layout)
            dialog.exec()
            
        except Exception as e:
            QMessageBox.warning(self, "错误", f"获取产品图片失败: {str(e)}")

    def load_customers(self):
        try:
            customers = self.api_client.get_customers()
            if customers is None:
                customers = []
            self.customers_table.setRowCount(len(customers))
            
            countries = sorted(set([c.get('country', '') for c in customers if c.get('country')]))
            current_country = self.customer_country_filter.currentText()
            self.customer_country_filter.clear()
            self.customer_country_filter.addItem("全部国家", 0)
            for country in countries:
                self.customer_country_filter.addItem(country, country)
            index = self.customer_country_filter.findText(current_country)
            if index >= 0:
                self.customer_country_filter.setCurrentIndex(index)
            
            # 先显示基本信息（不等待联系人和地址）
            for row, c in enumerate(customers):
                checkbox = QCheckBox()
                checkbox.setStyleSheet("margin-left: 50%;")
                self.customers_table.setCellWidget(row, 0, checkbox)
                
                self.customers_table.setItem(row, 1, QTableWidgetItem(str(c.get('id', ''))))
                self.customers_table.setItem(row, 2, QTableWidgetItem(c.get('customer_code', '')))
                self.customers_table.setItem(row, 3, QTableWidgetItem(c.get('customer_name', '')))
                self.customers_table.setItem(row, 4, QTableWidgetItem(c.get('country', '')))
                self.customers_table.setItem(row, 5, QTableWidgetItem(""))  # 默认联系人（稍后填充）
                self.customers_table.setItem(row, 6, QTableWidgetItem(""))  # 默认地址（稍后填充）
                self.customers_table.setItem(row, 7, QTableWidgetItem(c.get('payment_terms', '')))
                
                status = c.get('status', 1)
                status_text = "启用" if status == 1 else "禁用"
                status_color = "#10b981" if status == 1 else "#ef4444"
                status_item = QTableWidgetItem(status_text)
                status_item.setForeground(QBrush(QColor(status_color)))
                self.customers_table.setItem(row, 8, status_item)

                action_bar = ActionBarFactory.create_customer_action_bar(
                    edit_callback=lambda _, c=c: self.edit_customer(c),
                    toggle_callback=lambda _, c=c: self.toggle_customer_status(c),
                    status=status
                )
                self.customers_table.setCellWidget(row, 9, action_bar)
            
            self.customer_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
            
            # 异步加载联系人和地址（不阻塞UI）
            QTimer.singleShot(0, lambda: self._load_customer_extra_info(customers))
            
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载客户数据失败: {str(e)}")
    
    def _load_customer_extra_info(self, customers=None):
        """异步加载客户的联系人和地址信息"""
        try:
            # 如果没有传入客户列表，则从API获取
            if customers is None:
                customers = self.api_client.get_customers()
            if not customers:
                return
            
            customer_contacts = {}
            customer_addresses = {}
            
            for c in customers:
                try:
                    contacts = self.api_client.get_customer_contacts(c['id'])
                    if contacts:
                        primary_contact = None
                        for contact in contacts:
                            if contact.get('is_primary', 0) == 1:
                                primary_contact = contact
                                break
                        if not primary_contact and contacts:
                            primary_contact = contacts[0]
                        if primary_contact:
                            customer_contacts[c['id']] = primary_contact.get('name', '')
                except Exception:
                    pass
                
                try:
                    addresses = self.api_client.get_customer_addresses(c['id'])
                    if addresses:
                        default_addr = None
                        for addr in addresses:
                            if addr.get('is_default', 0) == 1:
                                default_addr = addr
                                break
                        if not default_addr and addresses:
                            default_addr = addresses[0]
                        if default_addr:
                            addr_text = f"{default_addr.get('country', '')}-{default_addr.get('port', '')}"
                            customer_addresses[c['id']] = addr_text
                except Exception:
                    pass
            
            # 更新表格中的联系人和地址列
            for row in range(self.customers_table.rowCount()):
                id_item = self.customers_table.item(row, 1)
                if id_item:
                    cid = int(id_item.text())
                    contact = customer_contacts.get(cid, '')
                    address = customer_addresses.get(cid, '')
                    if contact:
                        self.customers_table.setItem(row, 5, QTableWidgetItem(contact))
                    if address:
                        self.customers_table.setItem(row, 6, QTableWidgetItem(address))
        except Exception as e:
            print(f"加载客户额外信息失败: {e}")

    def search_customers(self):
        keyword = self.customer_search_input.text().strip()
        country = self.customer_country_filter.currentData()
        country = country if country != 0 else None
        
        try:
            customers = self.api_client.search_customers(keyword, country)
            self.customers_table.setRowCount(len(customers))
            
            # 先显示基本信息
            for row, c in enumerate(customers):
                checkbox = QCheckBox()
                checkbox.setStyleSheet("margin-left: 50%;")
                self.customers_table.setCellWidget(row, 0, checkbox)
                
                self.customers_table.setItem(row, 1, QTableWidgetItem(str(c.get('id', ''))))
                self.customers_table.setItem(row, 2, QTableWidgetItem(c.get('customer_code', '')))
                self.customers_table.setItem(row, 3, QTableWidgetItem(c.get('customer_name', '')))
                self.customers_table.setItem(row, 4, QTableWidgetItem(c.get('country', '')))
                self.customers_table.setItem(row, 5, QTableWidgetItem(""))  # 默认联系人（稍后填充）
                self.customers_table.setItem(row, 6, QTableWidgetItem(""))  # 默认地址（稍后填充）
                self.customers_table.setItem(row, 7, QTableWidgetItem(c.get('payment_terms', '')))
                
                status = c.get('status', 1)
                status_text = "启用" if status == 1 else "禁用"
                status_color = "#10b981" if status == 1 else "#ef4444"
                status_item = QTableWidgetItem(status_text)
                status_item.setForeground(QBrush(QColor(status_color)))
                self.customers_table.setItem(row, 8, status_item)

                action_bar = ActionBarFactory.create_customer_action_bar(
                    edit_callback=lambda _, c=c: self.edit_customer(c),
                    toggle_callback=lambda _, c=c: self.toggle_customer_status(c),
                    status=status
                )
                self.customers_table.setCellWidget(row, 9, action_bar)
            
            self.customer_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
            
            # 异步加载联系人和地址
            QTimer.singleShot(0, lambda: self._load_customer_extra_info(customers))
        except Exception as e:
            QMessageBox.warning(self, "错误", f"搜索客户失败: {str(e)}")

    def toggle_customer_status(self, customer):
        try:
            result = self.api_client.toggle_customer_status(customer['id'])
            if result:
                self.load_customers()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"操作失败: {str(e)}")

    def on_customer_double_click(self, index):
        row = index.row()
        customer_id = self.customers_table.item(row, 1).text()
        try:
            customer = self.api_client.get_customer_detail(int(customer_id))
            self.view_customer_detail(customer)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载客户详情失败: {str(e)}")

    def on_pi_double_click(self, index):
        """双击PI行查看详情"""
        row = index.row()
        pi_id = self.pi_table.item(row, 1).text()
        try:
            pi_detail = self.api_client.get_pi_detail(int(pi_id))
            self.view_pi_detail(pi_detail)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载PI详情失败: {str(e)}")

    def view_pi_detail(self, pi):
        """查看PI详情（只读）"""
        try:
            pi_id = pi.get('id')
            pi_detail = self.api_client.get_pi_detail(pi_id)
            # 2026-06-23 修复：原来用 main.PIDialog（shim），但 main.PIDialog 和 dialogs.pi.PIDialog
            # 是互相 wrapper 的关系，会触发无限递归 (RecursionError: maximum recursion depth exceeded)
            # 直接用真正的实现 PIDialog
            from dialogs.pi import PIDialog as RealPIDialog
            dialog = RealPIDialog(self.api_client, self.dept_id, pi_detail, readonly=True)
            dialog.exec()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"查看PI详情失败: {str(e)}")

    def view_customer_detail(self, customer):
        dialog = CustomerDetailDialog(self.api_client, customer)
        dialog.exec()

    def setup_table_context_menu(self, table, headers):
        def show_context_menu(point):
            menu = QMenu()
            row = table.rowAt(point.y())
            if row < 0:
                return
            copy_cn = menu.addAction("复制中文信息")
            copy_en = menu.addAction("Copy English Info")
            menu.addSeparator()
            copy_row = menu.addAction("复制整行数据")
            action = menu.exec(table.mapToGlobal(point))
            if action == copy_cn or action == copy_en:
                texts = []
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    if item:
                        texts.append(str(item.text()))
                    else:
                        widget = table.cellWidget(row, col)
                        if widget:
                            if isinstance(widget, QPushButton):
                                texts.append(widget.text())
                            else:
                                texts.append("")
                        else:
                            texts.append("")
                text = " | ".join(texts) if action == copy_cn else " | ".join(texts)
                from PySide6.QtGui import QGuiApplication
                clipboard = QGuiApplication.clipboard()
                clipboard.setText(text)
            elif action == copy_row:
                row_data = []
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    if item:
                        row_data.append(str(item.text()))
                    else:
                        widget = table.cellWidget(row, col)
                        if widget and isinstance(widget, QPushButton):
                            row_data.append(widget.text())
                        else:
                            row_data.append("")
                from PySide6.QtGui import QGuiApplication
                clipboard = QGuiApplication.clipboard()
                clipboard.setText("\t".join(row_data))
        table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        table.customContextMenuRequested.connect(show_context_menu)

    def toggle_select_all_customers(self, state):
        for row in range(self.customers_table.rowCount()):
            checkbox = self.customers_table.cellWidget(row, 0)
            if checkbox:
                checkbox.setCheckState(Qt.CheckState(state))

    def load_suppliers(self):
        try:
            suppliers = self.api_client.get_suppliers()
            if suppliers is None:
                suppliers = []
            self.suppliers_table.setRowCount(len(suppliers))
            for row, s in enumerate(suppliers):
                checkbox = QCheckBox()
                checkbox.setStyleSheet("margin-left: 15px;")
                self.suppliers_table.setCellWidget(row, 0, checkbox)

                self.suppliers_table.setItem(row, 1, QTableWidgetItem(str(s.get('id', ''))))
                self.suppliers_table.setItem(row, 2, QTableWidgetItem(s.get('supplier_code', '')))
                self.suppliers_table.setItem(row, 3, QTableWidgetItem(s.get('supplier_name', '')))
                self.suppliers_table.setItem(row, 4, QTableWidgetItem(s.get('region', '')))
                self.suppliers_table.setItem(row, 5, QTableWidgetItem(s.get('contact_person', '')))
                self.suppliers_table.setItem(row, 6, QTableWidgetItem(s.get('phone', '')))

                edit_btn = QPushButton("编辑")
                edit_btn.setFixedWidth(60)
                edit_btn.clicked.connect(lambda _, s=s: self.edit_supplier(s))
                self.suppliers_table.setCellWidget(row, 7, edit_btn)
        except Exception as e:
            self._hide_loading_tip()
            QMessageBox.warning(self, "错误", f"加载供应商数据失败: {str(e)}")

    def load_pi_orders(self):
        self._show_loading_tip("正在加载PI订单...")
        try:
            pi_orders = self.api_client.get_pi_orders()
            if pi_orders is None:
                pi_orders = []
            self.pi_table.setRowCount(len(pi_orders))
            status_map = {1: "草稿", 2: "已确认", 3: "已发货", 4: "已完成"}
            for row, pi in enumerate(pi_orders):
                status = pi.get('status', 1)
                is_completed = status == 4  # 已完成状态不可操作
                
                # 选择框（已完成PI不可选）
                if is_completed:
                    checkbox = QTableWidgetItem("✓")
                    checkbox.setFlags(checkbox.flags() & ~Qt.ItemFlag.ItemIsSelectable)
                    checkbox.setForeground(QColor("#9ca3af"))  # 灰色
                    self.pi_table.setItem(row, 0, checkbox)
                else:
                    checkbox = QTableWidgetItem()
                    checkbox.setFlags(checkbox.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                    checkbox.setCheckState(Qt.CheckState.Unchecked)
                    checkbox.setData(Qt.ItemDataRole.UserRole, pi.get('id'))
                    self.pi_table.setItem(row, 0, checkbox)
                
                self.pi_table.setItem(row, 1, QTableWidgetItem(str(pi.get('id', ''))))
                self.pi_table.setItem(row, 2, QTableWidgetItem(pi.get('pi_no', '')))
                self.pi_table.setItem(row, 3, QTableWidgetItem(f"{pi.get('total_amount', 0):,.2f}"))
                self.pi_table.setItem(row, 4, QTableWidgetItem(pi.get('currency', 'USD')))
                status_text = status_map.get(status, "未知")
                status_item = QTableWidgetItem(status_text)
                if is_completed:
                    status_item.setForeground(QColor("#6b7280"))  # 灰色
                self.pi_table.setItem(row, 5, status_item)
                self.pi_table.setItem(row, 6, QTableWidgetItem(str(pi.get('created_at', ''))[:19] if pi.get('created_at') else ''))

                # 操作列(7)、完成列(8)、导出列(9)
                if is_completed:
                    # 已完成PI：操作列显示"-"，完成列显示"✓"，导出列显示导出按钮
                    self.pi_table.setItem(row, 7, QTableWidgetItem("-"))
                    self.pi_table.item(row, 7).setForeground(QColor("#9ca3af"))
                    self.pi_table.setItem(row, 8, QTableWidgetItem("✓"))
                    self.pi_table.item(row, 8).setForeground(QColor("#10b981"))
                    export_btn = QPushButton("导出")
                    export_btn.setFixedWidth(50)
                    export_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 4px; padding: 2px 4px;")
                    export_btn.clicked.connect(lambda _, p=pi: self.export_pi(p))
                    self.pi_table.setCellWidget(row, 9, export_btn)
                else:
                    # 未完成PI：操作列显示编辑按钮，完成列显示完成按钮，导出列显示导出按钮
                    edit_btn = QPushButton("编辑")
                    edit_btn.setFixedWidth(50)
                    edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 4px;")
                    edit_btn.clicked.connect(lambda _, p=pi: self.edit_pi(p))
                    self.pi_table.setCellWidget(row, 7, edit_btn)
                    
                    complete_btn = QPushButton("完成")
                    complete_btn.setFixedWidth(40)
                    complete_btn.setStyleSheet("background-color: #8b5cf6; color: white; border: none; border-radius: 4px; padding: 2px 4px;")
                    complete_btn.clicked.connect(lambda _, p=pi: self.complete_pi(p))
                    self.pi_table.setCellWidget(row, 8, complete_btn)
                    
                    export_btn = QPushButton("导出")
                    export_btn.setFixedWidth(40)
                    export_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 4px; padding: 2px 4px;")
                    export_btn.clicked.connect(lambda _, p=pi: self.export_pi(p))
                    self.pi_table.setCellWidget(row, 9, export_btn)
        except Exception as e:
            self._hide_loading_tip()
            QMessageBox.warning(self, "错误", f"加载PI订单数据失败: {str(e)}")

    def load_purchase_orders(self):
        self._show_loading_tip("正在加载采购订单...")
        try:
            purchase_orders = self.api_client.get_purchase_orders()
            if purchase_orders is None:
                purchase_orders = []
            self.purchase_table.setRowCount(len(purchase_orders))
            status_map = {1: "草稿", 2: "已确认", 3: "已入库"}
            for row, po in enumerate(purchase_orders):
                self.purchase_table.setItem(row, 0, QTableWidgetItem(str(po.get('id', ''))))
                self.purchase_table.setItem(row, 1, QTableWidgetItem(po.get('po_no', '')))
                self.purchase_table.setItem(row, 2, QTableWidgetItem(po.get('pi_no', '')))
                self.purchase_table.setItem(row, 3, QTableWidgetItem(po.get('supplier_name', '')))
                self.purchase_table.setItem(row, 4, QTableWidgetItem(str(po.get('total_amount', ''))))
                self.purchase_table.setItem(row, 5, QTableWidgetItem(status_map.get(po.get('status', 1), "未知")))

                edit_btn = QPushButton("编辑")
                edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                edit_btn.clicked.connect(lambda _, p=po: self.edit_purchase(p))
                self.purchase_table.setCellWidget(row, 6, edit_btn)
        except Exception as e:
            self._hide_loading_tip()
            QMessageBox.warning(self, "错误", f"加载采购订单数据失败: {str(e)}")

    def load_shipments(self):
        self._show_loading_tip("正在加载出货数据...")
        try:
            shipments = self.api_client.get_shipments()
            if shipments is None:
                shipments = []
            self.shipment_table.setRowCount(len(shipments))
            status_map = {1: "待出货", 2: "出货中", 3: "已出货", 4: "已到达"}
            for row, s in enumerate(shipments):
                self.shipment_table.setItem(row, 0, QTableWidgetItem(str(s.get('id', ''))))
                self.shipment_table.setItem(row, 1, QTableWidgetItem(s.get('pi_no', '')))
                self.shipment_table.setItem(row, 2, QTableWidgetItem(str(s.get('shipment_date', ''))[:10] if s.get('shipment_date') else ''))
                self.shipment_table.setItem(row, 3, QTableWidgetItem(s.get('container_no', '')))
                self.shipment_table.setItem(row, 4, QTableWidgetItem(s.get('bl_no', '')))
                self.shipment_table.setItem(row, 5, QTableWidgetItem(status_map.get(s.get('status', 1), "未知")))

                edit_btn = QPushButton("编辑")
                edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                edit_btn.clicked.connect(lambda _, s=s: self.edit_shipment(s))
                self.shipment_table.setCellWidget(row, 6, edit_btn)
        except Exception as e:
            self._hide_loading_tip()
            QMessageBox.warning(self, "错误", f"加载出货数据失败: {str(e)}")

    def load_customer_payments(self):
        self._show_loading_tip("正在加载客户付款...")
        try:
            payments = self.api_client.get_customer_payments()
            if payments is None:
                payments = []
            self.customer_payment_table.setRowCount(len(payments))
            for row, p in enumerate(payments):
                self.customer_payment_table.setItem(row, 0, QTableWidgetItem(str(p.get('id', ''))))
                self.customer_payment_table.setItem(row, 1, QTableWidgetItem(p.get('pi_no', '')))
                self.customer_payment_table.setItem(row, 2, QTableWidgetItem(str(p.get('payment_date', ''))[:10] if p.get('payment_date') else ''))
                self.customer_payment_table.setItem(row, 3, QTableWidgetItem(str(p.get('actual_amount', ''))))
                self.customer_payment_table.setItem(row, 4, QTableWidgetItem(p.get('payment_method', '')))

                edit_btn = QPushButton("编辑")
                edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                edit_btn.clicked.connect(lambda _, p=p: self.edit_customer_payment(p))
                self.customer_payment_table.setCellWidget(row, 5, edit_btn)
        except Exception as e:
            self._hide_loading_tip()
            QMessageBox.warning(self, "错误", f"加载客户付款数据失败: {str(e)}")

    def load_supplier_payments(self):
        self._show_loading_tip("正在加载供应商付款...")
        try:
            payments = self.api_client.get_supplier_payments()
            if payments is None:
                payments = []
            self.supplier_payment_table.setRowCount(len(payments))
            for row, p in enumerate(payments):
                self.supplier_payment_table.setItem(row, 0, QTableWidgetItem(str(p.get('id', ''))))
                self.supplier_payment_table.setItem(row, 1, QTableWidgetItem(p.get('supplier_name', '')))
                self.supplier_payment_table.setItem(row, 2, QTableWidgetItem(f"PO-{p.get('po_id', '')}"))
                self.supplier_payment_table.setItem(row, 3, QTableWidgetItem(p.get('payment_stage', '')))
                self.supplier_payment_table.setItem(row, 4, QTableWidgetItem(str(p.get('payment_date', ''))[:10] if p.get('payment_date') else ''))
                self.supplier_payment_table.setItem(row, 5, QTableWidgetItem(str(p.get('actual_amount', ''))))

                edit_btn = QPushButton("编辑")
                edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                edit_btn.clicked.connect(lambda _, p=p: self.edit_supplier_payment(p))
                self.supplier_payment_table.setCellWidget(row, 6, edit_btn)
        except Exception as e:
            self._hide_loading_tip()
            QMessageBox.warning(self, "错误", f"加载供应商付款数据失败: {str(e)}")

    # ========== 异步加载方法 ==========

    def _show_loading_indicator(self, table, message="加载中..."):
        """显示加载状态指示器"""
        table.setRowCount(1)
        table.setItem(0, 0, QTableWidgetItem(""))
        table.setItem(0, 1, QTableWidgetItem(message))
        for col in range(2, table.columnCount()):
            table.setItem(0, col, QTableWidgetItem(""))
        table.setEnabled(False)

    def _hide_loading_indicator(self, table):
        """隐藏加载状态指示器"""
        table.setEnabled(True)

    def load_pi_orders_async(self, force_refresh=False):
        """异步加载PI订单（带缓存和加载指示器）"""
        print("DEBUG - load_pi_orders_async started")
        # 显示加载状态（在主线程）
        self._show_loading_indicator(self.pi_table, "正在加载PI订单...")
        
        # 使用QThread来确保信号在主线程处理
        from PySide6.QtCore import QThread
        
        class PiLoaderThread(QThread):
            def __init__(self, api_client, force_refresh, parent=None):
                super().__init__(parent)
                self.api_client = api_client
                self.force_refresh = force_refresh
                self.result_data = []
                self.error_msg = None
            
            def run(self):
                try:
                    if self.force_refresh:
                        cache_manager.delete(CACHE_KEYS['PI_LIST'])
                    data = self.api_client.get_pi_orders()
                    if data:
                        cache_manager.set(CACHE_KEYS['PI_LIST'], data)
                    self.result_data = data if data else []
                except Exception as e:
                    print(f"加载PI订单失败: {e}")
                    self.error_msg = str(e)
                    self.result_data = []
        
        self._pi_loader_thread = PiLoaderThread(self.api_client, force_refresh, self)
        self._pi_loader_thread.finished.connect(
            lambda: self._on_pi_load_finished(self._pi_loader_thread.result_data)
        )
        self._pi_loader_thread.start()
        print("DEBUG - thread started")
    
    def _on_pi_load_finished(self, data):
        """PI加载完成回调"""
        print(f"DEBUG - _on_pi_load_finished with {len(data)} items")
        self._update_pi_table(data)

    def load_pi_orders(self):
        """同步加载PI订单（首次进入）"""
        try:
            # 先尝试从缓存加载
            cached = cache_manager.get(CACHE_KEYS['PI_LIST'], max_age=120)
            if cached is not None:
                self._update_pi_table(cached)
                return
            # 缓存不存在，从API加载
            pi_orders = self.api_client.get_pi_orders()
            if pi_orders:
                cache_manager.set(CACHE_KEYS['PI_LIST'], pi_orders)
            self._update_pi_table(pi_orders if pi_orders else [])
        except Exception as e:
            print(f"加载PI订单失败: {e}")
            self._update_pi_table([])

    def _update_pi_table(self, pi_orders):
        try:
            print(f"DEBUG - _update_pi_table called with {len(pi_orders) if pi_orders else 0} orders")
            print(f"DEBUG - pi_orders type: {type(pi_orders)}")
            if pi_orders and len(pi_orders) > 0:
                print(f"DEBUG - first order: {pi_orders[0]}")
            
            # 隐藏加载指示器
            self._hide_loading_indicator(self.pi_table)
            
            # 确保数据是列表
            if not pi_orders:
                pi_orders = []
            
            self.pi_table.setRowCount(len(pi_orders))
            status_map = {1: "草稿", 2: "已确认", 3: "已发货", 4: "已完成"}
            for row, pi in enumerate(pi_orders):
                status = pi.get('status', 1)
                is_completed = status == 4
                
                # 选择框（已完成PI不可选）
                if is_completed:
                    checkbox = QTableWidgetItem("✓")
                    checkbox.setFlags(checkbox.flags() & ~Qt.ItemFlag.ItemIsSelectable)
                    checkbox.setForeground(QColor("#9ca3af"))
                    self.pi_table.setItem(row, 0, checkbox)
                else:
                    checkbox = QTableWidgetItem()
                    checkbox.setFlags(checkbox.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                    checkbox.setCheckState(Qt.CheckState.Unchecked)
                    checkbox.setData(Qt.ItemDataRole.UserRole, pi.get('id'))
                    self.pi_table.setItem(row, 0, checkbox)
                
                self.pi_table.setItem(row, 1, QTableWidgetItem(str(pi.get('id', ''))))
                self.pi_table.setItem(row, 2, QTableWidgetItem(pi.get('pi_no', '')))
                self.pi_table.setItem(row, 3, QTableWidgetItem(f"{pi.get('total_amount', 0):,.2f}"))
                self.pi_table.setItem(row, 4, QTableWidgetItem(pi.get('currency', 'USD')))
                status_text = status_map.get(status, "未知")
                status_item = QTableWidgetItem(status_text)
                if is_completed:
                    status_item.setForeground(QColor("#6b7280"))
                self.pi_table.setItem(row, 5, status_item)
                self.pi_table.setItem(row, 6, QTableWidgetItem(str(pi.get('created_at', ''))[:19] if pi.get('created_at') else ''))
                
                # 操作列(7)、完成列(8)、导出列(9)
                if is_completed:
                    # 已完成PI：操作列显示"-"，完成列显示"✓"，导出列显示导出按钮
                    self.pi_table.setItem(row, 7, QTableWidgetItem("-"))
                    self.pi_table.item(row, 7).setForeground(QColor("#9ca3af"))
                    self.pi_table.setItem(row, 8, QTableWidgetItem("✓"))
                    self.pi_table.item(row, 8).setForeground(QColor("#10b981"))
                    export_btn = QPushButton("导出")
                    export_btn.setFixedWidth(50)
                    export_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 4px; padding: 2px 4px;")
                    export_btn.clicked.connect(lambda _, p=pi: self.export_pi(p))
                    self.pi_table.setCellWidget(row, 9, export_btn)
                else:
                    # 未完成PI：操作列显示编辑按钮，完成列显示完成按钮，导出列显示导出按钮
                    edit_btn = QPushButton("编辑")
                    edit_btn.setFixedWidth(50)
                    edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 4px;")
                    edit_btn.clicked.connect(lambda _, p=pi: self.edit_pi(p))
                    self.pi_table.setCellWidget(row, 7, edit_btn)
                    
                    complete_btn = QPushButton("完成")
                    complete_btn.setFixedWidth(40)
                    complete_btn.setStyleSheet("background-color: #8b5cf6; color: white; border: none; border-radius: 4px; padding: 2px 4px;")
                    complete_btn.clicked.connect(lambda _, p=pi: self.complete_pi(p))
                    self.pi_table.setCellWidget(row, 8, complete_btn)
                    
                    export_btn = QPushButton("导出")
                    export_btn.setFixedWidth(40)
                    export_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 4px; padding: 2px 4px;")
                    export_btn.clicked.connect(lambda _, p=pi: self.export_pi(p))
                    self.pi_table.setCellWidget(row, 9, export_btn)
        except Exception as e:
            print(f"更新PI订单表格失败: {e}")
            import traceback
            traceback.print_exc()

    def load_purchase_orders_async(self):
        """异步加载采购订单"""
        self._load_async(
            self.api_client.get_purchase_orders,
            self._update_purchase_table,
            "加载采购订单失败"
        )

    def _update_purchase_table(self, purchase_orders):
        try:
            self.purchase_table.setRowCount(len(purchase_orders))
            status_map = {1: "草稿", 2: "已确认", 3: "已入库"}
            for row, po in enumerate(purchase_orders):
                self.purchase_table.setItem(row, 0, QTableWidgetItem(str(po.get('id', ''))))
                self.purchase_table.setItem(row, 1, QTableWidgetItem(po.get('po_no', '')))
                self.purchase_table.setItem(row, 2, QTableWidgetItem(po.get('pi_no', '')))
                self.purchase_table.setItem(row, 3, QTableWidgetItem(po.get('supplier_name', '')))
                self.purchase_table.setItem(row, 4, QTableWidgetItem(str(po.get('total_amount', ''))))
                status = po.get('status', 1)
                status_item = QTableWidgetItem(status_map.get(status, "未知"))
                if status == 3:
                    status_item.setForeground(QBrush(QColor("#10b981")))
                elif status == 2:
                    status_item.setForeground(QBrush(QColor("#f59e0b")))
                self.purchase_table.setItem(row, 5, status_item)
                edit_btn = QPushButton("编辑")
                edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                edit_btn.clicked.connect(lambda _, p=po: self.edit_purchase(p))
                self.purchase_table.setCellWidget(row, 6, edit_btn)
                # 确认按钮
                confirm_btn = QPushButton("确认")
                confirm_btn.setStyleSheet("background-color: #f59e0b; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                confirm_btn.clicked.connect(lambda _, p=po: self.confirm_purchase_order(p))
                self.purchase_table.setCellWidget(row, 7, confirm_btn)
                # 入库按钮
                inbound_btn = QPushButton("入库")
                inbound_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                inbound_btn.clicked.connect(lambda _, p=po: self.inbound_purchase_order(p))
                self.purchase_table.setCellWidget(row, 8, inbound_btn)
                # 导出合同按钮
                contract_btn = QPushButton("📄")
                contract_btn.setStyleSheet("background-color: #8b5cf6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                contract_btn.setToolTip("导出国内采购合同")
                contract_btn.clicked.connect(lambda _, p=po: self._export_single_contract(p))
                self.purchase_table.setCellWidget(row, 9, contract_btn)
        except Exception as e:
            print(f"更新采购订单表格失败: {e}")

    def load_shipments_async(self):
        """异步加载出货数据 - 委托给新的ShipmentTab组件"""
        if hasattr(self, '_shipment_tab') and self._shipment_tab is not None:
            self._shipment_tab.refresh_data()
        else:
            # 降级：旧逻辑已移除，此处仅做兼容性占位
            pass

    def _update_shipment_table(self, shipments):
        """更新出货主表（汇总信息）"""
        try:
            self.shipment_table.setRowCount(len(shipments))
            status_map = {1: "待出货", 2: "出货中", 3: "已出货", 4: "已到达"}
            payment_status_map = {1: "未收款", 2: "部分收款", 3: "已收齐"}
            
            for row, s in enumerate(shipments):
                # ID
                self.shipment_table.setItem(row, 0, QTableWidgetItem(str(s.get('id', ''))))
                # PI号
                self.shipment_table.setItem(row, 1, QTableWidgetItem(s.get('pi_no', '')))
                # 总金额
                total_amount = s.get('total_amount', 0) or 0
                self.shipment_table.setItem(row, 2, QTableWidgetItem(f"{float(total_amount):,.2f}"))
                # 总箱数
                total_cartons = s.get('total_cartons', 0) or 0
                self.shipment_table.setItem(row, 3, QTableWidgetItem(str(total_cartons)))
                # 付款状态
                payment_status = s.get('payment_status', 1)
                payment_item = QTableWidgetItem(payment_status_map.get(payment_status, "未知"))
                if payment_status == 3:
                    payment_item.setForeground(QBrush(QColor("#10b981")))
                elif payment_status == 2:
                    payment_item.setForeground(QBrush(QColor("#f59e0b")))
                self.shipment_table.setItem(row, 4, payment_item)
                # 出货状态
                status = s.get('status', 1)
                status_item = QTableWidgetItem(status_map.get(status, "未知"))
                if status == 3:
                    status_item.setForeground(QBrush(QColor("#10b981")))
                elif status == 2:
                    status_item.setForeground(QBrush(QColor("#3b82f6")))
                self.shipment_table.setItem(row, 5, status_item)
                # 阶段数
                stages_count = s.get('stages_count', 0) or 0
                self.shipment_table.setItem(row, 6, QTableWidgetItem(str(stages_count)))
                # 操作按钮
                edit_btn = QPushButton("编辑")
                edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                edit_btn.clicked.connect(lambda _, s=s: self.edit_shipment(s))
                self.shipment_table.setCellWidget(row, 7, edit_btn)
                # 确认出货按钮
                confirm_btn = QPushButton("确认出货")
                confirm_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                confirm_btn.clicked.connect(lambda _, s=s: self.confirm_shipment_order(s))
                self.shipment_table.setCellWidget(row, 8, confirm_btn)
            
            # 清空阶段表
            self.shipment_stage_table.setRowCount(0)
        except Exception as e:
            print(f"更新出货表格失败: {e}")

    def show_shipment_stages(self, row, column):
        """显示选中出货记录的阶段明细"""
        try:
            shipment_id = int(self.shipment_table.item(row, 0).text())
            # 获取出货详情（包含stages）
            shipment = self.api_client.get_shipment(shipment_id)
            if not shipment:
                return
            stages = shipment.get('stages', [])
            self._update_shipment_stage_table(stages)
        except Exception as e:
            print(f"显示出货阶段失败: {e}")

    def _update_shipment_stage_table(self, stages):
        """更新出货阶段从表"""
        try:
            self.shipment_stage_table.setRowCount(len(stages))
            payment_status_map = {1: "未收款", 2: "部分收款", 3: "已收齐"}
            
            for row, s in enumerate(stages):
                # 阶段名称
                self.shipment_stage_table.setItem(row, 0, QTableWidgetItem(s.get('stage_name', '')))
                # 出货日期
                shipment_date = s.get('shipment_date', '')
                if shipment_date:
                    shipment_date = str(shipment_date)[:10]
                self.shipment_stage_table.setItem(row, 1, QTableWidgetItem(shipment_date))
                # 柜号
                self.shipment_stage_table.setItem(row, 2, QTableWidgetItem(s.get('container_no', '')))
                # 提单号
                self.shipment_stage_table.setItem(row, 3, QTableWidgetItem(s.get('bl_no', '')))
                # 数量
                quantity = s.get('quantity', 0) or 0
                self.shipment_stage_table.setItem(row, 4, QTableWidgetItem(f"{float(quantity):,.0f}"))
                # 库存
                inv_qty = s.get('inventory_quantity', 0) or 0
                inv_item = QTableWidgetItem(f"{float(inv_qty):,.0f}")
                if inv_qty > 0:
                    inv_item.setForeground(QBrush(QColor("#10b981")))
                self.shipment_stage_table.setItem(row, 5, inv_item)
                # 存放位置
                self.shipment_stage_table.setItem(row, 6, QTableWidgetItem(s.get('storage_location', '-')))
                # 付款状态
                payment_status = s.get('payment_status', 1)
                pay_item = QTableWidgetItem(payment_status_map.get(payment_status, "未知"))
                if payment_status == 3:
                    pay_item.setForeground(QBrush(QColor("#10b981")))
                elif payment_status == 2:
                    pay_item.setForeground(QBrush(QColor("#f59e0b")))
                self.shipment_stage_table.setItem(row, 7, pay_item)
                # 操作按钮
                btn_widget = QWidget()
                btn_layout = QHBoxLayout()
                btn_layout.setContentsMargins(0, 0, 0, 0)
                
                ci_btn = QPushButton("CI")
                ci_btn.setFixedWidth(40)
                ci_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 3px;")
                ci_btn.clicked.connect(lambda _, s=s: self.view_ci_document(s))
                btn_layout.addWidget(ci_btn)
                
                pl_btn = QPushButton("PL")
                pl_btn.setFixedWidth(40)
                pl_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 3px;")
                pl_btn.clicked.connect(lambda _, s=s: self.view_pl_document(s))
                btn_layout.addWidget(pl_btn)
                
                btn_widget.setLayout(btn_layout)
                self.shipment_stage_table.setCellWidget(row, 8, btn_widget)
        except Exception as e:
            print(f"更新出货阶段表格失败: {e}")

    def view_ci_document(self, stage):
        """查看CI文档"""
        ci_doc = stage.get('ci_document')
        if ci_doc:
            # TODO: 打开CI文档查看器
            QMessageBox.information(self, "CI文档", f"CI文档路径: {ci_doc}")
        else:
            QMessageBox.information(self, "CI文档", "该阶段暂无CI文档")

    def view_pl_document(self, stage):
        """查看PL文档"""
        pl_doc = stage.get('pl_document')
        if pl_doc:
            # TODO: 打开PL文档查看器
            QMessageBox.information(self, "PL文档", f"PL文档路径: {pl_doc}")
        else:
            QMessageBox.information(self, "PL文档", "该阶段暂无PL文档")

    def load_customer_payments_async(self):
        """异步加载客户付款"""
        self._load_async(
            self.api_client.get_customer_payments,
            self._update_customer_payment_table,
            "加载客户付款失败",
            loading_msg="正在加载客户付款..."
        )

    def _update_customer_payment_table(self, data):
        """
        更新客户付款表格（spec #45 实现 - 12列 PI 聚合模式）

        Args:
            data: dict with key 'items' containing list of PI-aggregated payment records
                  Each record has:
                    - pi_no: str
                    - customer_name: str
                    - total_amount: float
                    - paid_amount: float
                    - unpaid_amount: float
                    - payment1/2/3: dict or None (each has actual_amount, arrival_date)
                    - latest_water_image: str or None (base64)
                    - receipt_count: int
                    - pi_id: int
        """
        try:
            # 防御性检查：空数据或无效格式
            if not data:
                self._show_cp_empty_state()
                return

            items = data.get('items', []) if isinstance(data, dict) else []
            if not isinstance(items, list):
                logger.error(f"[CustomerPayment] 数据格式错误: {type(data)}")
                items = []

            if not items:
                self._show_cp_empty_state()
                return

            # 设置行数并清空表格
            self.customer_payment_table.setRowCount(len(items))

            for row, pi_data in enumerate(items):
                # ===== 列 0: PI号 (蓝色可点击) =====
                pi_no = pi_data.get('pi_no', '')
                pi_item = QTableWidgetItem(pi_no)
                pi_item.setForeground(QColor("#2563eb"))  # 蓝色
                pi_item.setData(Qt.UserRole, pi_data.get('pi_id'))  # 存储 ID 用于跳转
                self.customer_payment_table.setItem(row, 0, pi_item)

                # ===== 列 1: 客户名称 =====
                customer_name = pi_data.get('customer_name', '-')
                self.customer_payment_table.setItem(row, 1, QTableWidgetItem(customer_name))

                # ===== 列 2-7: 付款1/2/3 (金额 + 日期) =====
                for stage_idx in range(1, 4):
                    payment_key = f'payment{stage_idx}'
                    payment = pi_data.get(payment_key)

                    amount_col = 2 + (stage_idx - 1) * 2  # 2, 4, 6
                    date_col = amount_col + 1              # 3, 5, 7

                    if payment:
                        # 金额列（绿色）
                        amount = payment.get('actual_amount', 0) or 0
                        amount_item = QTableWidgetItem(f"{amount:.2f}")
                        amount_item.setForeground(QColor("#059669"))  # 绿色
                        amount_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                        self.customer_payment_table.setItem(row, amount_col, amount_item)

                        # 日期列
                        arrival_date = payment.get('arrival_date')
                        date_text = arrival_date[:10] if arrival_date else "-"
                        date_item = QTableWidgetItem(date_text)
                        if not arrival_date:
                            date_item.setForeground(QColor("#9ca3af"))  # 灰色
                        self.customer_payment_table.setItem(row, date_col, date_item)
                    else:
                        # 无此阶段付款 → 显示 "-"
                        for col in [amount_col, date_col]:
                            dash_item = QTableWidgetItem("-")
                            dash_item.setForeground(QColor("#9ca3af"))
                            dash_item.setTextAlignment(Qt.AlignCenter)
                            self.customer_payment_table.setItem(row, col, dash_item)

                # ===== 列 8: 总应收 =====
                total_amount = pi_data.get('total_amount', 0) or 0
                total_item = QTableWidgetItem(f"{total_amount:.2f}")
                total_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.customer_payment_table.setItem(row, 8, total_item)

                # ===== 列 9: 未付款 (红色高亮 if > 0) =====
                unpaid_amount = pi_data.get('unpaid_amount', 0) or 0
                unpaid_item = QTableWidgetItem(f"{unpaid_amount:.2f}")
                unpaid_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                if unpaid_amount > 0:
                    unpaid_item.setForeground(QColor("#dc2626"))  # 红色
                    font = unpaid_item.font()
                    font.setBold(True)
                    unpaid_item.setFont(font)
                else:
                    unpaid_item.setForeground(QColor("#059669"))  # 绿色 (已结清)
                self.customer_payment_table.setItem(row, 9, unpaid_item)

                # ===== 列 10: 水单 (缩略图 + 按钮) =====
                water_widget = QWidget()
                water_layout = QHBoxLayout(water_widget)
                water_layout.setContentsMargins(5, 2, 5, 2)
                water_layout.setSpacing(5)

                receipt_count = pi_data.get('receipt_count', 0)
                latest_water = pi_data.get('latest_water_image')

                if latest_water:
                    # 显示缩略图
                    try:
                        import base64 as b64
                        img_data = b64.b64decode(latest_water)
                        pixmap = QPixmap()
                        pixmap.loadFromData(img_data)
                        thumb = pixmap.scaled(30, 30, Qt.KeepAspectRatio)
                        thumb_label = QLabel()
                        thumb_label.setPixmap(thumb)
                        water_layout.addWidget(thumb_label)
                    except Exception:
                        pass  # 缩略图加载失败不影响功能

                # 查看水单按钮
                if receipt_count > 0:
                    view_btn = QPushButton(f"📄{receipt_count}")
                    view_btn.setToolTip("点击查看所有水单")
                    view_btn.setFixedSize(50, 28)
                    view_btn.setStyleSheet("""
                        QPushButton {
                            background-color: #3b82f6;
                            color: white;
                            border: none;
                            border-radius: 4px;
                            font-size: 11px;
                        }
                        QPushButton:hover { background-color: #2563eb; }
                    """)
                    view_btn.clicked.connect(
                        lambda checked, pid=pi_data.get('pi_id'), pno=pi_no:
                            self._on_view_water_bills(pid, pno)
                    )
                    water_layout.addWidget(view_btn)
                else:
                    no_water_label = QLabel("-")
                    no_water_label.setStyleSheet("color: #9ca3af;")
                    no_water_label.setAlignment(Qt.AlignCenter)
                    water_layout.addWidget(no_water_label)

                self.customer_payment_table.setCellWidget(row, 10, water_widget)

                # ===== 列 11: 操作按钮 =====
                action_widget = QWidget()
                action_layout = QHBoxLayout(action_widget)
                action_layout.setContentsMargins(5, 2, 5, 2)

                edit_btn = QPushButton("编辑")
                edit_btn.setFixedSize(50, 26)
                edit_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #f3f4f6;
                        color: #374151;
                        border: 1px solid #d1d5db;
                        border-radius: 4px;
                    }
                    QPushButton:hover { background-color: #e5e7eb; }
                """)
                # 注意: 编辑功能需要传入单个 payment 对象，
                # 当前为 PI 聚合视图，暂禁用（后续迭代实现按 PI 编辑）
                edit_btn.setEnabled(False)
                edit_btn.setToolTip("编辑功能将在后续版本支持")
                action_layout.addWidget(edit_btn)

                self.customer_payment_table.setCellWidget(row, 11, action_widget)

            logger.info(f"[CustomerPayment] 表格更新完成: {len(items)} 条 PI 记录")

        except Exception as e:
            logger.error(f"[CustomerPayment] 更新表格失败: {e}", exc_info=True)
            print(f"更新客户付款表格失败: {e}")

    def _show_cp_empty_state(self):
        """显示空状态占位符"""
        try:
            # 先清空所有 cell widget 和 item，避免残留的水单/编辑按钮
            self.customer_payment_table.clearContents()
            self.customer_payment_table.setRowCount(1)
            empty_item = QTableWidgetItem("(暂无付款数据)")
            empty_item.setTextAlignment(Qt.AlignCenter)
            empty_item.setForeground(QColor("#9ca3af"))
            font = empty_item.font()
            font.setPointSize(12)
            empty_item.setFont(font)
            self.customer_payment_table.setItem(0, 0, empty_item)
            self.customer_payment_table.setSpan(0, 0, 1, 12)  # 合并所有列
        except Exception as e:
            logger.error(f"[CustomerPayment] 显示空状态失败: {e}")

    # ===== 任务 4: 水单查看入口和筛选功能 =====

    def _on_view_water_bills(self, pi_id: int, pi_no: str):
        """
        查看指定 PI 的所有水单

        Args:
            pi_id: PI 的数据库 ID
            pi_no: PI 号码 (用于显示)
        """
        try:
            # 获取该 PI 的所有水单记录
            payments = self.api_client.get_customer_payments_by_pi(pi_id)

            if not payments:
                QMessageBox.information(
                    self, "水单查看",
                    f"PI {pi_no} 暂无水单记录"
                )
                return

            # 打开水单查看器弹窗
            from widgets.payment import WaterBillViewer
            viewer = WaterBillViewer(self, pi_no=pi_no, payments=payments)
            viewer.exec_()

        except Exception as e:
            logger.error(f"[CustomerPayment] 查看水单失败: {e}", exc_info=True)
            QMessageBox.critical(
                self, "错误",
                f"加载水单数据失败:\n{e}"
            )

    def _on_cp_customer_filter_changed(self, index: int):
        """
        客户筛选下拉框变更事件

        Args:
            index: 当前选中的索引 (0 = "全部客户")
        """
        customer_id = None
        if index > 0:  # 0 = "全部客户"
            # 从下拉框获取客户 ID
            customer_id = self.cp_customer_filter.itemData(index)

        self._refresh_customer_payment_table(
            customer_id=customer_id,
            pi_no=self.cp_pi_search.text().strip() or None,
            only_unpaid=self.cp_only_unpaid.isChecked()
        )

    def _on_cp_pi_search_changed(self, text: str):
        """
        PI号搜索框文本变更事件 (300ms 防抖)

        Args:
            text: 当前输入的文本
        """
        # 使用 QTimer 延迟执行，避免频繁请求
        if not hasattr(self, '_cp_search_timer'):
            self._cp_search_timer = QTimer()
            self._cp_search_timer.setSingleShot(True)
            self._cp_search_timer.timeout.connect(self._do_cp_pi_search)

        self._cp_search_timer.start(300)  # 300ms 防抖

    def _do_cp_pi_search(self):
        """执行 PI 号搜索"""
        self._refresh_customer_payment_table(
            customer_id=None,  # 保持当前客户筛选
            pi_no=self.cp_pi_search.text().strip() or None,
            only_unpaid=self.cp_only_unpaid.isChecked()
        )

    def _on_cp_only_unpaid_changed(self, state: int):
        """
        仅未结清复选框变更事件

        Args:
            state: Qt.Checked 或 Qt.Unchecked
        """
        self._refresh_customer_payment_table(
            customer_id=None,
            pi_no=self.cp_pi_search.text().strip() or None,
            only_unpaid=(state == Qt.Checked.value)
        )

    def _refresh_customer_payment_table(self, customer_id=None, pi_no=None, only_unpaid=False):
        """
        刷新客户付款表格（带筛选参数）

        Args:
            customer_id: 可选，客户 ID 过滤
            pi_no: 可选，PI 号模糊搜索（前端筛选，API 不支持）
            only_unpaid: 是否仅显示未结清 PI（前端筛选，API 不支持）
        """
        def fetch():
            """异步获取数据"""
            try:
                # 调用 API（仅支持 customer_id 参数）
                data = self.api_client.get_customer_payments(
                    customer_id=customer_id
                )

                # 确保 data 不为 None
                if data is None:
                    return {"items": []}

                # 如果是 list，包装为 dict
                if isinstance(data, list):
                    data = {"items": data}

                # 字段映射：API 返回的字段可能不同，需要适配
                # 期望字段: pi_no, customer_name, total_amount, paid_amount,
                #           unpaid_amount, payment1/2/3, latest_water_image, receipt_count, pi_id
                items = data.get('items', [])
                if not items and 'records' in data:
                    items = data['records']

                # 前端筛选: PI 号模糊搜索
                if pi_no:
                    items = [item for item in items if pi_no.lower() in str(item.get('pi_no', '')).lower()]

                # 前端筛选: 仅未结清
                if only_unpaid:
                    items = [item for item in items if (item.get('unpaid_amount', 0) or 0) > 0]

                return {"items": items}

            except ConnectionError as e:
                raise Exception(f"网络连接失败: 请检查后端服务是否启动")
            except TimeoutError as e:
                raise Exception(f"请求超时: 后端响应过慢，请稍后重试")
            except Exception as e:
                logger.error(f"[CustomerPayment] 加载失败: {e}")
                raise e

        def on_done(future):
            """异步完成回调"""
            try:
                data = future.result()
                self._update_customer_payment_table(data)
            except Exception as e:
                logger.error(f"[CustomerPayment] 刷新失败: {e}")
                QMessageBox.warning(
                    self, "错误",
                    f"加载客户付款数据失败:\n{e}"
                )

        # 异步加载（使用线程池）
        import concurrent.futures
        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(fetch)
            future.add_done_callback(on_done)

    def load_supplier_payments_async(self):
        """异步加载供应商付款"""
        self._load_async(
            self.api_client.get_supplier_payments,
            self._update_supplier_payment_table,
            "加载供应商付款失败",
            loading_msg="正在加载供应商付款..."
        )

    def _update_supplier_payment_table(self, payments):
        """更新供应商付款主表（汇总信息）"""
        try:
            self.supplier_payment_table.setRowCount(len(payments))
            for row, p in enumerate(payments):
                # ID
                self.supplier_payment_table.setItem(row, 0, QTableWidgetItem(str(p.get('id', ''))))
                # 供应商
                self.supplier_payment_table.setItem(row, 1, QTableWidgetItem(p.get('supplier_name', '')))
                # 采购单
                self.supplier_payment_table.setItem(row, 2, QTableWidgetItem(f"PO-{p.get('po_id', '')}"))
                # 总金额
                total = p.get('total_amount', 0) or 0
                self.supplier_payment_table.setItem(row, 3, QTableWidgetItem(f"{float(total):,.2f}"))
                # 已付金额
                paid = p.get('paid_amount', 0) or 0
                paid_item = QTableWidgetItem(f"{float(paid):,.2f}")
                if paid > 0:
                    paid_item.setForeground(QBrush(QColor("#10b981")))
                self.supplier_payment_table.setItem(row, 4, paid_item)
                # 未付金额
                unpaid = p.get('unpaid_amount', 0) or 0
                unpaid_item = QTableWidgetItem(f"{float(unpaid):,.2f}")
                if unpaid > 0:
                    unpaid_item.setForeground(QBrush(QColor("#ef4444")))
                self.supplier_payment_table.setItem(row, 5, unpaid_item)
                # 状态
                status_map = {1: "待付款", 2: "部分付款", 3: "已付清"}
                status = p.get('status', 1)
                status_text = status_map.get(status, "未知")
                status_item = QTableWidgetItem(status_text)
                if status == 3:
                    status_item.setForeground(QBrush(QColor("#10b981")))
                elif status == 2:
                    status_item.setForeground(QBrush(QColor("#f59e0b")))
                self.supplier_payment_table.setItem(row, 6, status_item)
                # 操作按钮
                edit_btn = QPushButton("编辑")
                edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                edit_btn.clicked.connect(lambda _, p=p: self.edit_supplier_payment(p))
                self.supplier_payment_table.setCellWidget(row, 7, edit_btn)
            # 清空阶段表
            self.supplier_payment_stage_table.setRowCount(0)
        except Exception as e:
            print(f"更新供应商付款表格失败: {e}")

    def show_supplier_payment_stages(self, row, column):
        """显示选中付款记录的阶段明细"""
        try:
            payment_id = int(self.supplier_payment_table.item(row, 0).text())
            # 获取付款详情（包含stages）
            payment = self.api_client.get_supplier_payment(payment_id)
            if not payment:
                return
            stages = payment.get('stages', [])
            self._update_supplier_payment_stage_table(stages)
        except Exception as e:
            print(f"显示付款阶段失败: {e}")

    def _update_supplier_payment_stage_table(self, stages):
        """更新付款阶段从表"""
        try:
            self.supplier_payment_stage_table.setRowCount(len(stages))
            status_map = {1: "待付", 2: "部分付", 3: "已付清"}
            for row, s in enumerate(stages):
                # 阶段名称
                self.supplier_payment_stage_table.setItem(row, 0, QTableWidgetItem(s.get('stage_name', '')))
                # 应付金额
                amount = s.get('amount', 0) or 0
                self.supplier_payment_stage_table.setItem(row, 1, QTableWidgetItem(f"{float(amount):,.2f}"))
                # 已付金额
                paid = s.get('paid_amount', 0) or 0
                paid_item = QTableWidgetItem(f"{float(paid):,.2f}")
                if paid > 0:
                    paid_item.setForeground(QBrush(QColor("#10b981")))
                self.supplier_payment_stage_table.setItem(row, 2, paid_item)
                # 状态
                status = s.get('status', 1)
                status_item = QTableWidgetItem(status_map.get(status, "未知"))
                if status == 3:
                    status_item.setForeground(QBrush(QColor("#10b981")))
                elif status == 2:
                    status_item.setForeground(QBrush(QColor("#f59e0b")))
                self.supplier_payment_stage_table.setItem(row, 3, status_item)
                # 付款日期
                payment_date = s.get('payment_date', '')
                if payment_date:
                    payment_date = str(payment_date)[:10]
                self.supplier_payment_stage_table.setItem(row, 4, QTableWidgetItem(payment_date))
                # 凭证
                has_proof = "有" if s.get('payment_proof') else "无"
                self.supplier_payment_stage_table.setItem(row, 5, QTableWidgetItem(has_proof))
                # 操作按钮
                pay_btn = QPushButton("付款")
                pay_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                pay_btn.clicked.connect(lambda _, s=s: self.pay_supplier_stage(s))
                self.supplier_payment_stage_table.setCellWidget(row, 6, pay_btn)
        except Exception as e:
            print(f"更新付款阶段表格失败: {e}")

    def load_inventories_async(self):
        """异步加载库存数据"""
        self._load_async(
            self.api_client.get_inventories,
            self._populate_inventories,
            "加载库存失败",
            loading_msg="正在加载库存数据..."
        )

    def load_inventories(self):
        """加载库存数据 - 按OE号分组显示（参考供应商方案模式）"""
        self._show_loading_tip("正在加载库存...")
        try:
            inventories = self.api_client.get_inventories()
            if inventories is None:
                inventories = []
            self._populate_inventories(inventories)
            self._hide_loading_tip()
        except Exception as e:
            self._hide_loading_tip()
            QMessageBox.warning(self, "错误", f"加载库存数据失败: {str(e)}")

    def _populate_inventories(self, inventories):
        """填充库存数据到表格"""
        try:
            # 按OE号分组统计
            self.inventory_data = {}
            for inv in inventories:
                oe = inv.get('oe_number') or inv.get('product_code', '未知')
                if oe not in self.inventory_data:
                    self.inventory_data[oe] = []
                self.inventory_data[oe].append(inv)
            
            self._load_inventory_product_table()
            self.inventory_detail_table.setRowCount(0)
            self.inventory_detail_label.setText("📋 库存详情（请点击上方产品查看）")
        except Exception as e:
            print(f"填充库存数据失败: {e}")

    def load_products_async(self):
        """异步加载产品数据 - 初次强制刷新"""
        # 初次加载时强制刷新，不使用缓存
        self.load_products(use_cache=False)

    def load_products_callback(self, products):
        """产品数据加载完成后更新UI"""
        try:
            self.load_products_with_data(products)
        except Exception as e:
            print(f"更新产品表格失败: {e}")

    def load_products_with_data(self, products):
        """用已有数据加载产品列表（异步回调使用）"""
        try:
            inventory_summary = {}
            try:
                inventory_summary = self.api_client.get_all_inventory_summary()
            except Exception:
                pass

            if products is None:
                products = []
            self.products_table.setRowCount(len(products))
            for row, p in enumerate(products):
                product_id = p.get('id')
                
                # 0: 复选框
                checkbox = QCheckBox()
                checkbox.setStyleSheet("margin-left: 50%;")
                self.products_table.setCellWidget(row, 0, checkbox)
                
                # 获取OE和客户关联
                oe_list = []
                customer_product_list = []
                try:
                    oe_list = self.api_client.get_product_oes(product_id) or []
                    customer_product_list = self.api_client.get_product_customers(product_id) or []
                except:
                    pass
                
                # 1: 客户产品编号
                customer_product_code = ""
                if customer_product_list:
                    first_pc = customer_product_list[0]
                    full_code = first_pc.get('customer_product_code', '')
                    customer_code = first_pc.get('customer_code', '')
                    if full_code and customer_code:
                        customer_product_code = full_code.replace(customer_code, "", 1).lstrip("-")
                    else:
                        customer_product_code = full_code or ""
                self.products_table.setItem(row, 1, QTableWidgetItem(customer_product_code))
                
                # 2: OE号
                primary_oe = next((oe for oe in oe_list if oe.get('is_primary')), None)
                if len(oe_list) > 1:
                    btn = QPushButton("多OE号")
                    btn.setStyleSheet("""
                        QPushButton {
                            background-color: #3b82f6;
                            color: white;
                            border: none;
                            border-radius: 4px;
                            padding: 4px 8px;
                            font-size: 11px;
                        }
                        QPushButton:hover { background-color: #2563eb; }
                    """)
                    btn.clicked.connect(lambda checked, pid=product_id, oes=oe_list: self._show_product_oe_dialog(pid, oes))
                    self.products_table.setCellWidget(row, 2, btn)
                elif primary_oe:
                    self.products_table.setItem(row, 2, QTableWidgetItem(primary_oe.get('oe_number', '')))
                else:
                    self.products_table.setItem(row, 2, QTableWidgetItem(p.get('oe_number', '') or '-'))
                
                # 3: 图片
                image_label = QLabel()
                image_label.setFixedSize(60, 60)
                image_label.setStyleSheet("border: 1px solid #e5e7eb;")
                image_label.setAlignment(Qt.AlignCenter)
                image_url = self._normalize_image_url(p.get('default_image_url'))
                if image_url:
                    self.load_image_async(image_label, image_url)
                else:
                    image_label.setText("暂无图片")
                image_label.setCursor(Qt.CursorShape.PointingHandCursor)
                self.products_table.setCellWidget(row, 3, image_label)
                
                # 4: 产品名称
                self.products_table.setItem(row, 4, QTableWidgetItem(p.get('detail_desc', '')))
                
                # 5: 客户型号
                customer_model = ""
                if customer_product_list:
                    customer_model = customer_product_list[0].get('customer_model', '') or ""
                self.products_table.setItem(row, 5, QTableWidgetItem(customer_model))
                
                # 6: 客户号（留空）
                self.products_table.setItem(row, 6, QTableWidgetItem(""))
                
                # 7: 产品特性（品牌）
                self.products_table.setItem(row, 7, QTableWidgetItem(p.get('brand', '') or '-'))
                
                # 8: 数量（库存）
                qty = inventory_summary.get(product_id, 0)
                qty_item = QTableWidgetItem(str(int(qty)) if qty else '0')
                if qty > 0:
                    qty_item.setForeground(QBrush(QColor("#10b981")))
                else:
                    qty_item.setForeground(QBrush(QColor("#6b7280")))
                qty_item.setTextAlignment(Qt.AlignCenter)
                self.products_table.setItem(row, 8, qty_item)
                # 9: 报价
                price = p.get('exw_price_incl', 0)
                price_text = f"{price} USD" if price else "-"
                price_item = QTableWidgetItem(price_text)
                price_item.setTextAlignment(Qt.AlignRight)
                self.products_table.setItem(row, 9, price_item)
                # 10: 操作
                action_widget = QWidget()
                action_layout = QHBoxLayout()
                action_layout.setContentsMargins(0, 0, 0, 0)
                
                edit_btn = QPushButton("编辑")
                edit_btn.setFixedWidth(50)
                edit_btn.clicked.connect(lambda _, prod=p: self.edit_product(prod))
                action_layout.addWidget(edit_btn)
                
                action_widget.setLayout(action_layout)
                self.products_table.setCellWidget(row, 10, action_widget)  # 修复：编辑按钮应该在列10
                
                # 11: 隐藏的ID列
                self.products_table.setItem(row, 11, QTableWidgetItem(str(product_id)))
        except Exception as e:
            print(f"更新产品表格失败: {e}")

    def _load_inventory_product_table(self, data=None):
        """加载产品库存汇总表"""
        if data is None:
            data = self.inventory_data
        
        # 获取最近变更日志
            product_logs = {}
            try:
                product_logs = self.api_client.get_product_logs() or {}
                print(f"DEBUG - 获取到 {len(product_logs)} 条产品日志")
            except Exception as e:
                print(f"DEBUG - 获取产品日志失败: {e}")
            
            self.inventory_product_table.setRowCount(len(data))
        
        for row, (oe_number, records) in enumerate(data.items()):
            # 获取产品ID
            product_id = records[0].get('product_id')
            print(f"DEBUG - 产品ID={product_id}, OE={oe_number}")
            
            # OE号（缩短显示）
            oe_display = str(oe_number)
            if len(oe_display) > 18:
                oe_display = oe_display[:15] + '...'
            oe_item = QTableWidgetItem(oe_display)
            oe_item.setToolTip(str(oe_number))  # 悬停显示完整OE号
            self.inventory_product_table.setItem(row, 0, oe_item)
            
            # 产品编号（取第一条记录的product_code）
            product_code = records[0].get('product_code', '') or ''
            if len(str(product_code)) > 14:
                product_code = str(product_code)[:11] + '...'
            self.inventory_product_table.setItem(row, 1, QTableWidgetItem(str(product_code)))
            
            # 总库存量
            total_qty = sum(float(r.get('total_quantity', 0) or 0) for r in records)
            qty_item = QTableWidgetItem(str(int(total_qty)))
            # 库存为0时标红
            if total_qty == 0:
                qty_item.setForeground(QBrush(QColor("#ef4444")))
            self.inventory_product_table.setItem(row, 2, qty_item)
            
            # 供应商数
            suppliers = set(r.get('supplier_name') for r in records if r.get('supplier_name'))
            self.inventory_product_table.setItem(row, 3, QTableWidgetItem(str(len(suppliers))))
            
            # 客户数
            customers = set(r.get('customer_name') for r in records if r.get('customer_name'))
            self.inventory_product_table.setItem(row, 4, QTableWidgetItem(str(len(customers))))
            
            # 状态分布（用颜色点表示）
            status_counts = {}
            for r in records:
                st = r.get('stock_type', 1)
                if isinstance(st, str):
                    try: st = int(st)
                    except ValueError: st = 1
                status_counts[st] = status_counts.get(st, 0) + 1
            
            status_parts = []
            if 1 in status_counts:
                status_parts.append(f"🟡{status_counts[1]}")
            if 2 in status_counts:
                status_parts.append(f"🔵{status_counts[2]}")
            if 3 in status_counts:
                status_parts.append(f"🟢{status_counts[3]}")
            if 4 in status_counts:
                status_parts.append(f"⚫{status_counts[4]}")
            self.inventory_product_table.setItem(row, 5, QTableWidgetItem(" ".join(status_parts) if status_parts else "-"))
            
            # 最近入库供应商
            # product_id 是整数，但 product_logs 的键是字符串
            log_info = product_logs.get(str(product_id), {}) or product_logs.get(product_id, {})
            print(f"DEBUG - 产品{product_id}的日志: {log_info}")
            supplier_name = log_info.get('supplier_name', '') or ''
            if len(supplier_name) > 10:
                supplier_name = supplier_name[:8] + '..'
            self.inventory_product_table.setItem(row, 6, QTableWidgetItem(supplier_name or '-'))
            
            # 最近出库客户
            customer_name = log_info.get('customer_name', '') or ''
            if len(customer_name) > 10:
                customer_name = customer_name[:8] + '..'
            self.inventory_product_table.setItem(row, 7, QTableWidgetItem(customer_name or '-'))
            
            # 最近变更时间
            last_change = log_info.get('last_change_time', '') or ''
            if last_change and len(str(last_change)) > 16:
                last_change = str(last_change)[:16].replace('T', ' ')
            self.inventory_product_table.setItem(row, 8, QTableWidgetItem(last_change))
            
            # 操作按钮（展开 + 添加）
            op_widget = QWidget()
            op_layout = QHBoxLayout(op_widget)
            op_layout.setContentsMargins(2, 2, 2, 2)
            op_layout.setSpacing(4)
            
            expand_btn = QPushButton("展开")
            expand_btn.setFixedWidth(50)
            expand_btn.setStyleSheet("""
                QPushButton {
                    background-color: #10b981; color: white; border: none;
                    border-radius: 4px; padding: 4px 8px; font-size: 12px;
                }
                QPushButton:hover { background-color: #059669; }
            """)
            expand_btn.clicked.connect(lambda _, oe=oe_number: self.show_inventory_detail(oe))
            op_layout.addWidget(expand_btn)
            
            add_btn = QPushButton("+添加")
            add_btn.setFixedWidth(50)
            add_btn.setStyleSheet("""
                QPushButton {
                    background-color: #3b82f6; color: white; border: none;
                    border-radius: 4px; padding: 4px 8px; font-size: 12px;
                }
                QPushButton:hover { background-color: #2563eb; }
            """)
            add_btn.clicked.connect(lambda _, oe=oe_number, recs=records: self.add_inventory_for_oe(oe, recs))
            op_layout.addWidget(add_btn)
            
            self.inventory_product_table.setCellWidget(row, 9, op_widget)
        
        # 双击展开详情
        try:
            self.inventory_product_table.cellDoubleClicked.disconnect()
        except TypeError:
            pass
        self.inventory_product_table.cellDoubleClicked.connect(
            lambda r, c: self.show_inventory_detail(list(data.keys())[r])
        )
    
    def show_inventory_detail(self, oe_number):
        """显示指定OE号的库存详情"""
        records = self.inventory_data.get(oe_number, [])
        self.inventory_detail_label.setText(f"📋 '{oe_number}' 的库存详情（共 {len(records)} 条记录）")
        
        self.inventory_detail_table.setRowCount(len(records))
        
        for row, inv in enumerate(records):
            # ID
            self.inventory_detail_table.setItem(row, 0, QTableWidgetItem(str(inv.get('id', ''))))
            # 供应商（缩短显示）
            supplier_name = inv.get('supplier_name', '-') or '-'
            if len(str(supplier_name)) > 10:
                supplier_name = str(supplier_name)[:8] + '..'
            supplier_item = QTableWidgetItem(supplier_name)
            supplier_item.setToolTip(inv.get('supplier_name', '-') or '-')
            self.inventory_detail_table.setItem(row, 1, supplier_item)
            # 客户（缩短显示）
            customer_name = inv.get('customer_name', '-') or '-'
            if len(str(customer_name)) > 10:
                customer_name = str(customer_name)[:8] + '..'
            customer_item = QTableWidgetItem(customer_name)
            customer_item.setToolTip(inv.get('customer_name', '-') or '-')
            self.inventory_detail_table.setItem(row, 2, customer_item)
            # 数量
            self.inventory_detail_table.setItem(row, 3, QTableWidgetItem(str(int(inv.get('total_quantity', 0) or 0))))
            # 库位
            self.inventory_detail_table.setItem(row, 4, QTableWidgetItem(inv.get('current_location', '') or ''))
            # 状态颜色 (2026-06-23: 4状态生命周期)
            stock_type = inv.get('stock_type', 1)
            if isinstance(stock_type, str):
                try:
                    stock_type = int(stock_type)
                except (ValueError, TypeError):
                    stock_type = 1
            # 新 4 状态颜色映射
            STOCK_COLOR_MAP = {
                1: "#FFD700",  # 黄: 采购在途
                2: "#4169E1",  # 蓝: 待入库
                3: "#32CD32",  # 绿: 已入库
                4: "#333333",  # 黑: 历史库存
            }
            STOCK_LABEL_MAP = {
                1: "采购在途",
                2: "待入库",
                3: "已入库",
                4: "历史库存",
            }
            color_hex = STOCK_COLOR_MAP.get(stock_type, "#cccccc")
            color_label = QLabel()
            color_label.setFixedSize(20, 20)
            color_label.setStyleSheet(f"background-color: {color_hex}; border: 1px solid #999; border-radius: 3px;")
            color_label.setToolTip(STOCK_LABEL_MAP.get(stock_type, f"状态{stock_type}"))
            self.inventory_detail_table.setCellWidget(row, 5, color_label)
            # 备注
            remark_item = QTableWidgetItem(inv.get('remark', '') or '')
            self.inventory_detail_table.setItem(row, 6, remark_item)
            # 创建时间
            created_at = inv.get('created_at', '') or ''
            if created_at and len(str(created_at)) > 16:
                created_at = str(created_at)[:16].replace('T', ' ')
            self.inventory_detail_table.setItem(row, 7, QTableWidgetItem(str(created_at)))
            # 编辑按钮
            edit_btn = QPushButton("编辑")
            edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
            edit_btn.clicked.connect(lambda _, inv=inv: self.edit_inventory(inv))
            self.inventory_detail_table.setCellWidget(row, 8, edit_btn)
    
    def add_inventory_for_oe(self, oe_number, existing_records):
        """为指定OE号添加库存记录"""
        # 获取产品信息
        product_id = existing_records[0].get('product_id') if existing_records else None
        if product_id:
            self.add_inventory_with_product(product_id, oe_number)
        else:
            self.add_inventory()
    
    def search_inventory(self):
        """搜索库存"""
        keyword = self.inventory_search_input.text().strip().lower()
        if not keyword:
            self._load_inventory_product_table()
            return
        
        # 过滤数据
        filtered_data = {}
        for oe_number, records in self.inventory_data.items():
            if keyword in str(oe_number).lower():
                filtered_data[oe_number] = records
            else:
                for r in records:
                    if (keyword in str(r.get('supplier_name', '')).lower() or 
                        keyword in str(r.get('customer_name', '')).lower()):
                        filtered_data[oe_number] = records
                        break
        
        self._load_inventory_product_table(filtered_data)

    def on_inventory_status_filter_changed(self, index):
        """2026-06-23: 库存状态筛选变更"""
        stock_type = self.inventory_status_filter.currentData()
        if stock_type is None:
            # 全部状态：重新加载
            self.load_inventories()
        else:
            # 按状态筛选
            self.load_inventories_with_filter(stock_type)

    def load_inventories_with_filter(self, stock_type: int):
        """按库存状态筛选加载"""
        try:
            inventories = self.api_client.get_inventories(stock_type=stock_type)
            if inventories is None:
                inventories = []
            self._populate_inventories(inventories)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载库存数据失败: {str(e)}")
    
    def add_inventory_with_product(self, product_id, oe_number):
        """为指定产品添加库存"""
        dialog = InventoryDialog(self.api_client, self.dept_id, product_id=product_id, oe_number=oe_number)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_inventories()

    def toggle_product_status(self, product):
        product_id = product.get('id')
        product_code = product.get('product_code', '')
        status = product.get('status', 1)
        status_text = "禁用" if status == 1 else "启用"
        
        reply = QMessageBox.question(
            self, "确认操作", 
            f"确定要{status_text}产品 {product_code} 吗？",
            QMessageBox.Ok | QMessageBox.Cancel
        )
        
        if reply == QMessageBox.Ok:
            try:
                self.api_client.toggle_product_status(product_id)
                QMessageBox.information(self, "成功", f"产品已{status_text}")
                self.load_products()
            except Exception as e:
                QMessageBox.warning(self, "错误", f"{status_text}产品失败: {str(e)}")

    def delete_product(self, product):
        product_id = product.get('id')
        product_code = product.get('product_code', '')
        
        reply = QMessageBox.question(
            self, "确认删除", 
            f"确定要删除产品 {product_code} 吗？此操作不可恢复！",
            QMessageBox.Ok | QMessageBox.Cancel
        )
        
        if reply == QMessageBox.Ok:
            try:
                self.api_client.delete_product(product_id)
                QMessageBox.information(self, "成功", "产品已删除")
                self.load_products()
            except Exception as e:
                QMessageBox.warning(self, "错误", f"删除产品失败: {str(e)}")

    def add_product(self):
        print("add_product called, dept_id:", self.dept_id)
        try:
            # 使用新的客户产品对话框
            from widgets.customer_product_dialog import CustomerProductDialog
            dialog = CustomerProductDialog(self.api_client)
            print("CustomerProductDialog created successfully")
            if dialog.exec() == QDialog.DialogCode.Accepted:
                self.load_products()
        except Exception as e:
            print("Error in add_product:", str(e))
            QMessageBox.warning(self, "错误", f"打开新增产品对话框失败: {str(e)}")

    def on_product_double_click(self, index):
        row = index.row()
        product = None

        # 方式1: 优先通过系统编号获取（第2列 userData 存储完整值）
        system_code_item = self.products_table.item(row, 2)
        if system_code_item:
            full_system_code = system_code_item.data(Qt.UserRole) or ''
            if full_system_code:
                try:
                    product = self.api_client.get(f"/customer-products/by-system-code/{full_system_code}")
                except Exception:
                    pass  # 系统编号查询失败，fallback 到 ID

        # 方式2: fallback 通过产品ID获取（第14列操作区存储 product_id）
        if not product:
            action_widget = self.products_table.cellWidget(row, 14)
            product_id = None
            if action_widget:
                product_id = action_widget.property("product_id")
                if not product_id and isinstance(action_widget, QPushButton):
                    product_id = action_widget.property("product_id")
            if product_id:
                try:
                    product = self.api_client.get(f"/customer-products/{product_id}")
                except Exception:
                    pass

        if not product:
            QMessageBox.warning(self, "错误", "无法获取产品信息")
            return

        self.edit_product(product)

    def edit_product(self, product):
        from widgets.customer_product_dialog import CustomerProductDialog
        dialog = CustomerProductDialog(self.api_client, product)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_products()

    def import_products(self):
        print("DEBUG - import_products called")
        
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择导入文件", "", "Excel文件 (*.xlsx *.xls)"
        )
        
        print(f"DEBUG - Selected file: {file_path}")
        
        if not file_path:
            return
        
        try:
            # 使用 openpyxl 读取 Excel（替代 pandas）
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active

            # 获取表头
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value) if cell.value else '')

            print(f"DEBUG - Excel file loaded, columns: {len(headers)}, rows: {ws.max_row}")

            products_data = []

            # 从第2行开始读取数据（跳过表头）
            for row_idx in range(2, ws.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers):
                    if col_idx < len(headers):
                        cell_value = ws.cell(row=row_idx, column=col_idx + 1).value
                        row_data[header] = cell_value

                product = {
                    "oe_number": str(row_data.get("OE号", "")),
                    "factory_code": str(row_data.get("工厂编号", "")),
                    "brand": str(row_data.get("品牌", "")),
                    "detail_desc": str(row_data.get("细节描述", "")),
                    "category_id": int(row_data.get("类别ID", 1)),
                    "supplier_id": int(row_data.get("供应商ID", 0)) if self._is_not_none(row_data.get("供应商ID")) else None,
                    "exw_price_incl": float(row_data.get("EXW含税价", 0)) if self._is_not_none(row_data.get("EXW含税价")) else None,
                    "exw_price_excl": float(row_data.get("EXW不含税价", 0)) if self._is_not_none(row_data.get("EXW不含税价")) else None,
                    "fob_price_incl": float(row_data.get("FOB含税价", 0)) if self._is_not_none(row_data.get("FOB含税价")) else None,
                    "fob_price_excl": float(row_data.get("FOB不含税价", 0)) if self._is_not_none(row_data.get("FOB不含税价")) else None,
                    "freight": float(row_data.get("运费", 0)) if self._is_not_none(row_data.get("运费")) else None,
                    "packing_fee": float(row_data.get("包装费", 0)) if self._is_not_none(row_data.get("包装费")) else None,
                    "purchase_channel": str(row_data.get("采购渠道", "")),
                    "carton_length_cm": float(row_data.get("纸箱长(cm)", 0)) if self._is_not_none(row_data.get("纸箱长(cm)")) else None,
                    "carton_width_cm": float(row_data.get("纸箱宽(cm)", 0)) if self._is_not_none(row_data.get("纸箱宽(cm)")) else None,
                    "carton_height_cm": float(row_data.get("纸箱高(cm)", 0)) if self._is_not_none(row_data.get("纸箱高(cm)")) else None,
                    "carton_volume_cbm": float(row_data.get("纸箱体积(CBM)", 0)) if self._is_not_none(row_data.get("纸箱体积(CBM)")) else None,
                    "carton_weight_kg": float(row_data.get("纸箱重量(KG)", 0)) if self._is_not_none(row_data.get("纸箱重量(KG)")) else None,
                    "pieces_per_carton": int(row_data.get("每箱数量", 0)) if self._is_not_none(row_data.get("每箱数量")) else None,
                    "unit": str(row_data.get("单位", "件")),
                    "moq": int(row_data.get("最小起订量", 0)) if self._is_not_none(row_data.get("最小起订量")) else None,
                }
                products_data.append(product)
            
            print(f"DEBUG - Prepared {len(products_data)} products for import")
            print(f"DEBUG - First product data: {products_data[0] if products_data else 'None'}")
            
            result = self.api_client.import_products(products_data)
            print(f"DEBUG - Import API result: {result}")
            
            if result.get("success"):
                QMessageBox.information(self, "成功", f"成功导入 {result.get('count', 0)} 个产品")
            else:
                QMessageBox.warning(self, "导入结果", f"导入完成，部分失败: {result.get('message', '')}")
            
            self.load_products()
            
        except Exception as e:
            print(f"DEBUG - Import failed: {str(e)}")
            QMessageBox.warning(self, "错误", f"导入失败: {str(e)}")

    def add_customer(self):
        dialog = CustomerDialog(self.api_client)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            cache_manager.delete(CACHE_KEYS['CUSTOMERS'])
            self.load_customers()

    def edit_customer(self, customer):
        dialog = CustomerDialog(self.api_client, customer)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            cache_manager.delete(CACHE_KEYS['CUSTOMERS'])
            self.load_customers()

    def import_suppliers(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择导入文件", "",
            "Excel文件 (*.xlsx *.xls);;CSV文件 (*.csv)"
        )

        if not file_path:
            return

        try:
            # 使用 openpyxl 读取 Excel/CSV（替代 pandas）
            if file_path.endswith('.csv'):
                import csv
                with open(file_path, 'r', encoding='utf-8', newline='') as f:
                    reader = csv.DictReader(f)
                    headers = reader.fieldnames or []
                    rows_data = [row for row in reader]
            else:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                ws = wb.active
                headers = []
                for cell in ws[1]:
                    headers.append(str(cell.value) if cell.value else '')
                rows_data = []
                for row_idx in range(2, ws.max_row + 1):
                    row_data = {}
                    for col_idx, header in enumerate(headers):
                        cell_value = ws.cell(row=row_idx, column=col_idx + 1).value
                        row_data[header] = cell_value
                    rows_data.append(row_data)

            supplier_list = []
            province_codes = self.PROVINCE_CODES
            city_data = self.CITY_DATA

            for row_data in rows_data:
                supplier_data = {}

                for col in headers:
                    col_lower = str(col).strip().lower()
                    value = row_data.get(col)

                    # 替代 pd.isna()
                    if value is None or (isinstance(value, float) and str(value) == 'nan'):
                        value = None
                    else:
                        value = str(value).strip()

                    if '供应商' in col_lower or '名称' in col_lower:
                        supplier_data['supplier_name'] = value
                    elif '省份' in col_lower:
                        supplier_data['province'] = value
                    elif '城市' in col_lower or '市' in col_lower:
                        supplier_data['city'] = value
                    elif '联系人' in col_lower:
                        supplier_data['contact_person'] = value
                    elif '电话' in col_lower or '手机' in col_lower:
                        supplier_data['phone'] = value
                    elif '邮箱' in col_lower or 'email' in col_lower:
                        supplier_data['email'] = value
                    elif '地址' in col_lower:
                        supplier_data['address'] = value
                
                if 'supplier_name' in supplier_data and supplier_data['supplier_name']:
                    province = supplier_data.get('province')
                    city = supplier_data.get('city')
                    
                    if province and city:
                        province_code = province_codes.get(province)
                        if province_code:
                            cities = city_data.get(province_code, {})
                            city_code = cities.get(city)
                            if city_code:
                                supplier_data['city_code'] = province_code + city_code
                    
                    supplier_list.append(supplier_data)
            
            if not supplier_list:
                QMessageBox.warning(self, "警告", "未找到有效的供应商数据")
                return
                
            progress = QProgressDialog("正在导入供应商...", "取消", 0, len(supplier_list), self)
            progress.setWindowModality(2)
            progress.show()
            
            def import_task():
                try:
                    result = self.api_client.post("/suppliers/batch", {"suppliers": supplier_list})
                    return result
                except Exception as e:
                    return {"error": str(e)}
            
            result = import_task()
            
            progress.close()
            
            if "error" in result:
                QMessageBox.critical(self, "导入失败", f"导入过程中发生错误: {result['error']}")
            else:
                success = result.get("success", 0)
                failed = result.get("failed", 0)
                msg = f"导入完成！\n成功: {success} 条\n失败: {failed} 条"
                if failed > 0:
                    failed_items = result.get("failed_items", [])
                    for item in failed_items[:5]:
                        msg += f"\n- {item['supplier_name']}: {item['error']}"
                    if len(failed_items) > 5:
                        msg += f"\n... 还有 {len(failed_items) - 5} 条失败记录"
                QMessageBox.information(self, "导入完成", msg)
                self.load_suppliers()
                
        except Exception as e:
            QMessageBox.critical(self, "导入失败", f"读取文件时发生错误: {str(e)}")

    def add_supplier(self):
        dialog = SupplierDialog(self.api_client)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            cache_manager.delete(CACHE_KEYS['SUPPLIERS'])
            self.load_suppliers()

    def on_supplier_double_click(self, index):
        row = index.row()
        supplier_id = self.suppliers_table.item(row, 1).text()
        try:
            supplier = self.api_client.get_supplier_detail(int(supplier_id))
            self.edit_supplier(supplier)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载供应商信息失败: {str(e)}")

    def edit_supplier(self, supplier):
        dialog = SupplierDialog(self.api_client, supplier)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            cache_manager.delete(CACHE_KEYS['SUPPLIERS'])
            self.load_suppliers()

    def add_pi(self):
        dialog = PIDialog(self.api_client, self.dept_id)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_pi_orders_async()

    def edit_pi(self, pi):
        """编辑PI订单"""
        try:
            # 获取完整PI详情
            pi_id = pi.get('id')
            if not pi_id:
                QMessageBox.warning(self, "错误", "无法获取PI ID")
                return
            pi_detail = self.api_client.get_pi_detail(pi_id)
            dialog = PIDialog(self.api_client, self.dept_id, pi_detail)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                self.load_pi_orders_async()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"打开PI失败: {str(e)}")

    def get_selected_pi_ids(self):
        """获取选中的PI ID列表"""
        ids = []
        for row in range(self.pi_table.rowCount()):
            item = self.pi_table.item(row, 0)
            if item and item.checkState() == Qt.CheckState.Checked:
                pi_id = item.data(Qt.ItemDataRole.UserRole)
                if pi_id:
                    ids.append(pi_id)
        return ids

    def batch_delete_pi(self):
        """批量删除PI订单"""
        selected_ids = self.get_selected_pi_ids()
        if not selected_ids:
            QMessageBox.warning(self, "提示", "请先选择要删除的PI订单")
            return
        
        reply = QMessageBox.question(
            self, "确认删除", 
            f"确定要删除选中的 {len(selected_ids)} 个PI订单吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        try:
            result = self.api_client.batch_delete_pi(selected_ids)
            deleted = result.get('deleted', 0)
            errors = result.get('errors', [])
            if errors:
                QMessageBox.warning(self, "部分删除失败", f"成功删除 {deleted} 个\n失败: {len(errors)} 个\n{errors}")
            else:
                QMessageBox.information(self, "成功", f"已删除 {deleted} 个PI订单")
            self.load_pi_orders_async()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"删除失败: {str(e)}")
            self.load_pi_orders_async()

    def _on_preview_pi_export(self):
        """预览PI导出"""
        selected_ids = self.get_selected_pi_ids()
        if not selected_ids:
            QMessageBox.warning(self, "提示", "请先选择要预览的PI订单")
            return

        if len(selected_ids) > 1:
            QMessageBox.warning(self, "提示", "请选择一个PI订单进行预览")
            return

        pi_id = selected_ids[0]
        try:
            # 获取预览数据
            data = self.api_client.get(f"/export/pi/{pi_id}/preview")
            # 显示预览对话框
            dialog = ExportPreviewDialog(self, "pi", data, self.api_client)
            if dialog.exec():
                # 用户确认导出
                edited_fields = dialog.editable_fields
                self._do_export_pi(pi_id, edited_fields)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"获取预览数据失败: {str(e)}")

    def _do_export_pi(self, pi_id: int, edited_fields: dict = None):
        """执行PI导出"""
        try:
            params = ""
            if edited_fields:
                params = "?" + "&".join([f"{k}={v}" for k, v in edited_fields.items()])

            response = self.api_client.get(f"/export/pi/{pi_id}{params}", raw=True)

            # 保存文件
            from PySide6.QtWidgets import QFileDialog
            file_path, _ = QFileDialog.getSaveFileName(
                self, "保存PI", f"PI_{pi_id}.xlsx",
                "Excel Files (*.xlsx)"
            )
            if file_path:
                with open(file_path, "wb") as f:
                    f.write(response.content)
                QMessageBox.information(self, "成功", f"已导出到:\n{file_path}")
        except Exception as e:
            QMessageBox.warning(self, "错误", f"导出失败: {str(e)}")

    def batch_export_pi(self):
        """批量导出PI订单"""
        selected_ids = self.get_selected_pi_ids()
        if not selected_ids:
            QMessageBox.warning(self, "提示", "请先选择要导出的PI订单")
            return
        
        try:
            pi_data_list = []
            for pi_id in selected_ids:
                pi_detail = self.api_client.get_pi_detail(pi_id)
                pi_data_list.append(pi_detail)
            
            if not pi_data_list:
                QMessageBox.information(self, "提示", "没有可导出的数据")
                return
            
            # 构建导出数据
            export_rows = []
            for pi in pi_data_list:
                for item in pi.get('items', []):
                    export_rows.append({
                        'PI号': pi.get('pi_no', ''),
                        '客户ID': pi.get('customer_id', ''),
                        '总金额': pi.get('total_amount', 0),
                        '币种': pi.get('currency', 'USD'),
                        '状态': ['草稿', '已确认', '已发货', '已完成'][pi.get('status', 1) - 1] if pi.get('status', 1) <= 4 else '',
                        '创建时间': pi.get('created_at', ''),
                        '产品编号': item.get('product_id', ''),
                        'OE号': item.get('oe_number', ''),
                        '客户编号': item.get('customer_code', ''),
                        '产品描述': item.get('detail_desc', ''),
                        '数量': item.get('quantity', 0),
                        '单价': item.get('unit_price', 0),
                        '总价': item.get('total_price', 0),
                        '备注': item.get('remark', ''),
                    })

            # 使用 openpyxl 创建Excel并导出（替代 pandas）
            from PySide6.QtWidgets import QFileDialog
            file_path, _ = QFileDialog.getSaveFileName(
                self, "保存PI订单",
                f"PI订单导出_{len(selected_ids)}个.xlsx",
                "Excel Files (*.xlsx)"
            )
            if file_path:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "PI订单"

                # 写入表头
                if export_rows:
                    headers = list(export_rows[0].keys())
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col_idx, value=header)

                    # 写入数据
                    for row_idx, row_data in enumerate(export_rows, 2):
                        for col_idx, header in enumerate(headers, 1):
                            value = row_data.get(header)
                            # 处理None值
                            if value is None or value == '':
                                value = ''
                            ws.cell(row=row_idx, column=col_idx, value=value)

                wb.save(file_path)
                QMessageBox.information(self, "成功", f"已导出 {len(export_rows)} 条明细到:\n{file_path}")
        except Exception as e:
            QMessageBox.warning(self, "错误", f"导出失败: {str(e)}")

    def export_pi(self, pi):
        """导出PI为Excel"""
        try:
            import tempfile
            content = self.api_client.export_pi_excel(pi.get('id'))
            # 保存到临时文件
            filename = f"PI_{pi.get('pi_no', pi.get('id'))}.xlsx"
            filepath = tempfile.gettempdir() + "\\" + filename
            with open(filepath, 'wb') as f:
                f.write(content)
            # 打开文件
            import os
            os.startfile(filepath)
            QMessageBox.information(self, "成功", f"PI已导出: {filename}")
        except Exception as e:
            QMessageBox.warning(self, "错误", f"导出失败: {str(e)}")

    def complete_pi(self, pi):
        """将PI标记为已完成"""
        reply = QMessageBox.question(
            self, "确认完成",
            f"确定要将PI单 {pi.get('pi_no', '')} 标记为已完成吗？\n完成后将不可编辑。",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.api_client.update_pi_status(pi.get('id'), 4)
                cache_manager.delete(CACHE_KEYS['PI_LIST'])
                self.load_pi_orders_async()
                QMessageBox.information(self, "成功", f"PI单 {pi.get('pi_no', '')} 已标记为完成")
            except Exception as e:
                QMessageBox.warning(self, "错误", f"操作失败: {str(e)}")

    def add_purchase(self):
        dialog = PurchaseDialog(self.api_client, self.dept_id)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_purchase_orders()

    def edit_purchase(self, purchase):
        dialog = PurchaseDialog(self.api_client, self.dept_id, purchase)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_purchase_orders_async()

    def confirm_purchase_order(self, purchase):
        """确认采购单"""
        try:
            self.api_client.confirm_purchase(purchase.get('id'))
            QMessageBox.information(self, "成功", "采购单已确认")
            self.load_purchase_orders_async()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"确认失败: {str(e)}")

    def inbound_purchase_order(self, purchase):
        """采购单入库"""
        try:
            self.api_client.inbound_purchase_order(purchase.get('id'))
            QMessageBox.information(self, "成功", "采购单已入库")
            self.load_purchase_orders_async()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"入库失败: {str(e)}")

    def _on_export_purchase_contract(self):
        """工具栏：导出选中行的采购合同"""
        selected_rows = set(item.row() for item in self.purchase_table.selectedItems())
        if not selected_rows:
            QMessageBox.warning(self, "提示", "请先选择要导出的采购单")
            return

        for row in selected_rows:
            po_id_item = self.purchase_table.item(row, 0)
            if po_id_item:
                po_id = int(po_id_item.text())
                self._export_single_contract({'id': po_id})

    def _export_single_contract(self, purchase):
        """导出单个采购合同"""
        try:
            import tempfile, os
            po_id = purchase.get('id')
            po_no = purchase.get('po_no', str(po_id))
            content = self.api_client.export_contract_excel(po_id)

            filename = f"Contract_{po_no}.xlsx"
            filepath = os.path.join(tempfile.gettempdir(), filename)
            with open(filepath, 'wb') as f:
                f.write(content)

            os.startfile(filepath)
            print(f"[采购合同] 导出成功: {filename}")
        except Exception as e:
            QMessageBox.warning(self, "错误", f"导出采购合同失败: {str(e)}")

    def add_shipment(self):
        dialog = ShipmentDialog(self.api_client, self.dept_id)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_shipments_async()

    def edit_shipment(self, shipment):
        dialog = ShipmentDialog(self.api_client, self.dept_id, shipment)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_shipments_async()

    def confirm_shipment_order(self, shipment):
        """确认出货"""
        try:
            self.api_client.confirm_shipment(shipment.get('id'))
            QMessageBox.information(self, "成功", "出货已确认")
            self.load_shipments_async()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"确认失败: {str(e)}")

    def add_customer_payment(self):
        dialog = CustomerPaymentDialog(self.api_client, self.dept_id)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_customer_payments()

    def edit_customer_payment(self, payment):
        dialog = CustomerPaymentDialog(self.api_client, self.dept_id, payment)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_customer_payments()

    def add_supplier_payment(self):
        dialog = SupplierPaymentDialog(self.api_client, self.dept_id)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_supplier_payments()

    def edit_supplier_payment(self, payment):
        dialog = SupplierPaymentDialog(self.api_client, self.dept_id, payment)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_supplier_payments()

    def add_inventory(self):
        dialog = InventoryDialog(self.api_client, self.dept_id)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_inventories()

    def edit_inventory(self, inventory):
        dialog = InventoryDialog(self.api_client, self.dept_id, inventory)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_inventories()


class PIDialog(QDialog):
    """PI 对话框（真实实现，与 client/dialogs/pi.py 完全等价）

    2026-06-23 修复循环递归：原 shim 用 `from dialogs.pi import PIDialog; self._impl = ...()`
    而 dialogs/pi.PIDialog.__init__ 又 `from main import PIDialog` 创建实例 → 无限递归
    （RecursionError: maximum recursion depth exceeded）。

    修复：dialogs/pi.py 改为 `from main import PIDialog`（class alias），
    本类直接继承 QDialog 并初始化 UI，**不再实例化**其他 PIDialog。
    """

    def __init__(self, api_client, dept_id, pi=None, readonly=False):
        super().__init__()
        self.api_client = api_client
        self.dept_id = dept_id
        self.pi = pi
        self.is_edit = pi is not None
        self.readonly = readonly
        self.customers = []
        self.products = []
        self.items = []
        # 真正 UI 初始化（在 load_data 前完成）
        self.init_ui()
        # 延迟加载数据
        from PySide6.QtCore import QTimer
        QTimer.singleShot(0, self.load_data)
    
    def load_data(self):
        """加载数据（客户同步加载保证回填，产品异步加载）"""
        # 客户列表需要同步加载，因为后面要回填选中项
        try:
            cached_customers = cache_manager.get(CACHE_KEYS['CUSTOMERS'], max_age=300)
            if cached_customers is not None:
                self.customers = cached_customers
            else:
                self.customers = self.api_client.get_customers()
                cache_manager.set(CACHE_KEYS['CUSTOMERS'], self.customers)
            
            self.customer_combo.clear()
            self.customer_combo.addItem("", "")
            for c in self.customers:
                self.customer_combo.addItem(f"{c.get('customer_code')} - {c.get('customer_name')}", c.get('id'))
            
            if self.is_edit and not self.readonly:
                self.customer_combo.setEnabled(True)
                self.currency_combo.setEnabled(True)
        except Exception as e:
            print(f"加载客户失败: {e}")
        
        # 产品列表异步加载（不需要回填）
        self.load_products()
        
        # 编辑模式：回填现有数据（在客户加载完成后执行）
        if self.is_edit and self.pi:
            self._fill_existing_data()
    
    def _fill_existing_data(self):
        """回填编辑模式下的现有PI数据"""
        # 回填客户
        customer_id = self.pi.get('customer_id')
        if customer_id:
            idx = self.customer_combo.findData(customer_id)
            if idx >= 0:
                self.customer_combo.setCurrentIndex(idx)
        
        # 回填币种
        currency = self.pi.get('currency', 'USD')
        idx = self.currency_combo.findText(currency)
        if idx >= 0:
            self.currency_combo.setCurrentIndex(idx)
        
        # 回填产品明细
        items = self.pi.get('items', [])
        if items:
            self.items = []
            for item in items:
                self.items.append({
                    "product_id": item.get('product_id'),
                    "product_code": item.get('product_code', ''),
                    "oe_number": item.get('oe_number', ''),
                    "quantity": item.get('quantity', 1),
                    "unit_price": item.get('unit_price', 0),
                    "customer_code": item.get('customer_code', ''),
                    "detail_desc": item.get('detail_desc', ''),
                    "remark": item.get('remark', '')
                })
            self.update_items_table()

    def load_customers(self):
        try:
            self.customers = self.api_client.get_customers()
            self.customer_combo.clear()
            self.customer_combo.addItem("", "")
            for c in self.customers:
                self.customer_combo.addItem(f"{c.get('customer_code')} - {c.get('customer_name')}", c.get('id'))
            
            # 编辑模式下启用客户和币种选择
            if self.is_edit and not self.readonly:
                self.customer_combo.setEnabled(True)
                self.currency_combo.setEnabled(True)
        except Exception as e:
            print(f"加载客户失败: {e}")

    def load_products(self):
        """异步加载产品列表"""
        # 先显示加载中
        self.product_combo.blockSignals(True)
        self.product_combo.clear()
        self.product_combo.addItem("加载中...", None)
        self.product_combo.blockSignals(False)
        
        def fetch():
            try:
                # 尝试从缓存加载
                products = cache_manager.get(CACHE_KEYS['PRODUCTS'], max_age=300)
                if products is None:
                    products = self.api_client.get_products()
                    cache_manager.set(CACHE_KEYS['PRODUCTS'], products)
                
                self.products = products
                self.all_products = products.copy() if products else []
                
                # 用QTimer切换到主线程更新UI
                QTimer.singleShot(0, lambda: self.update_product_combo(self.all_products))
            except Exception as e:
                print(f"加载产品失败: {e}")
                QTimer.singleShot(0, lambda: self.product_combo.setItemText(0, "加载失败"))
        
        # 使用全局线程池异步加载
        _global_thread_pool.submit(fetch)
    
    def update_product_combo(self, products):
        """更新产品下拉框（分批加载避免卡顿）"""
        self.product_combo.blockSignals(True)
        self.product_combo.clear()
        self.product_combo.addItem("", None)
        
        # 限制显示数量，避免过多产品导致卡顿
        max_display = 100
        display_products = products[:max_display] if len(products) > max_display else products
        
        for p in display_products:
            product_code = p.get('product_code', '')
            oe_number = p.get('oe_number', '')
            # 截断长文本
            oe_display = oe_number[:15] + "..." if len(oe_number) > 15 else oe_number
            self.product_combo.addItem(f"{product_code} - {oe_display}", p)
        
        if len(products) > max_display:
            self.product_combo.addItem(f"...还有 {len(products) - max_display} 个产品，请使用搜索", None)
        
        self.product_combo.blockSignals(False)
    
    def filter_products(self, text):
        """根据搜索关键词过滤产品"""
        if not text:
            self.update_product_combo(self.all_products)
            return
        text = text.lower()
        filtered = [
            p for p in self.all_products 
            if text in str(p.get('product_code', '')).lower() 
            or text in str(p.get('oe_number', '')).lower()
        ]
        self.update_product_combo(filtered)
    
    def on_product_selected(self, index):
        """产品选择变化时更新图片预览和显示供应商方案"""
        product = self.product_combo.currentData()
        if product:
            # 更新图片预览
            image_url = self._normalize_image_url(product.get('default_image_url')) or product.get('image_url')
            if image_url:
                try:
                    import urllib.request
                    image_data = urllib.request.urlopen(image_url).read()
                    image = QImage.fromData(image_data)
                    pixmap = QPixmap.fromImage(image).scaled(46, 46, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    self.product_image_preview.setPixmap(pixmap)
                except Exception:
                    self.product_image_preview.setText("❌")
            else:
                self.product_image_preview.setText("图")
            
            # 加载供应商方案
            self.load_product_schemes(product.get('id'))
        else:
            self.product_image_preview.setText("图")
            self.scheme_row.setEnabled(False)
    
    def load_product_schemes(self, product_id):
        """加载产品的供应商方案（从PrdProductSupplier表读取）"""
        try:
            schemes = self.api_client.get_product_schemes(product_id)
            self.scheme_combo.clear()
            self.scheme_combo.addItem("-- 选择供应商方案 --", None)
            if schemes:
                for s in schemes:
                    price = s.get('purchase_price', 0) or 0
                    customer = s.get('customer_name', '通用') or '通用'
                    label = (
                        f"【{s.get('supplier_name', '供应商')}】"
                        f" ({customer})"
                        f" 价格:{price:.2f}"
                    )
                    self.scheme_combo.addItem(label, s)
                self.scheme_row.setEnabled(True)
            else:
                self.scheme_row.setEnabled(False)
        except Exception as e:
            print(f"加载供应商方案失败: {e}")
            self.scheme_row.setEnabled(False)
    
    def on_scheme_selected(self, index):
        """供应商方案选择变化"""
        pass  # 可以在这里添加预览功能
    
    def apply_scheme(self):
        """使用选中的供应商方案填充单价"""
        scheme = self.scheme_combo.currentData()
        if not scheme:
            QMessageBox.warning(self, "提示", "请先选择一个供应商方案")
            return
        
        # 使用purchase_price作为单价
        price = scheme.get('purchase_price', 0) or 0
        if price:
            self.unit_price_input.setText(f"{price:.2f}")
            QMessageBox.information(self, "成功", f"已使用【{scheme.get('supplier_name', '供应商')}】的方案\n采购价: {price:.2f} 已填入")
        else:
            QMessageBox.warning(self, "提示", "该方案尚未设置价格，请在产品管理中设置")
    
    def create_new_scheme(self):
        """为当前产品新建供应商方案（直接弹出SupplierSchemeDialog）"""
        product = self.product_combo.currentData()
        if not product:
            QMessageBox.warning(self, "警告", "请先选择一个产品")
            return
        
        try:
            # 加载供应商和客户列表
            suppliers = self.api_client.get_suppliers()
            customers = self.api_client.get_customers()
            
            # 直接弹出供应商方案编辑弹窗
            dialog = SupplierSchemeDialog(self.api_client, suppliers, customers, parent=self)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                scheme_data = dialog.get_scheme_data()
                if scheme_data:
                    # 通过API创建供应商方案
                    scheme_data['product_id'] = product.get('id')
                    self.api_client.create_product_scheme(product.get('id'), scheme_data)
                    QMessageBox.information(self, "成功", "供应商方案已创建")
                    # 刷新方案列表
                    self.load_product_schemes(product.get('id'))
        except Exception as e:
            QMessageBox.warning(self, "错误", f"创建供应商方案失败: {str(e)}")

    def init_ui(self):
        self.setWindowTitle("查看PI单" if self.readonly else ("编辑PI单" if self.is_edit else "新建PI单"))
        self.setMinimumSize(850, 750)
        self.resize(850, 750)

        layout = QVBoxLayout()
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(10)

        # ===== 上部分：基本信息 =====
        basic_group = QGroupBox("基本信息")
        basic_layout = QFormLayout()
        basic_layout.setSpacing(10)

        self.customer_combo = QComboBox()
        self.customer_combo.setFixedHeight(32)
        if self.readonly:
            self.customer_combo.setEnabled(False)
        basic_layout.addRow("客户:", self.customer_combo)

        self.currency_combo = QComboBox()
        self.currency_combo.setFixedHeight(32)
        self.currency_combo.addItems(["USD", "CNY", "EUR"])
        if self.readonly:
            self.currency_combo.setEnabled(False)
        basic_layout.addRow("货币:", self.currency_combo)
        basic_group.setLayout(basic_layout)
        layout.addWidget(basic_group)

        # ===== 中部分：添加产品区域 =====
        add_product_group = QGroupBox("添加产品")
        add_product_layout = QVBoxLayout()
        add_product_layout.setSpacing(8)

        # 产品搜索和选择行
        product_row = QHBoxLayout()
        product_row.addWidget(QLabel("搜索产品:"))
        
        self.product_search = QLineEdit()
        self.product_search.setPlaceholderText("输入产品号/OE号...")
        self.product_search.setFixedWidth(150)
        self.product_search.textChanged.connect(self.filter_products)
        product_row.addWidget(self.product_search)
        
        self.product_combo = QComboBox()
        self.product_combo.setMinimumWidth(180)
        self.product_combo.currentIndexChanged.connect(self.on_product_selected)
        product_row.addWidget(self.product_combo)
        
        self.product_image_preview = QLabel()
        self.product_image_preview.setFixedSize(50, 50)
        self.product_image_preview.setStyleSheet("border: 1px solid #d1d5db; border-radius: 4px; background-color: #f9fafb;")
        self.product_image_preview.setAlignment(Qt.AlignCenter)
        self.product_image_preview.setText("图")
        product_row.addWidget(self.product_image_preview)
        
        product_row.addWidget(QLabel("数量:"))
        self.quantity_input = QLineEdit()
        self.quantity_input.setPlaceholderText("0")
        self.quantity_input.setFixedWidth(80)
        product_row.addWidget(self.quantity_input)
        
        product_row.addWidget(QLabel("单价:"))
        self.unit_price_input = QLineEdit()
        self.unit_price_input.setPlaceholderText("0.00")
        self.unit_price_input.setFixedWidth(80)
        product_row.addWidget(self.unit_price_input)
        
        if not self.readonly:
            add_item_btn = QPushButton("+ 添加")
            add_item_btn.setStyleSheet("background-color: #2563eb; color: white; border: none; border-radius: 4px; padding: 6px 16px;")
            add_item_btn.clicked.connect(self.add_item)
            product_row.addWidget(add_item_btn)
        
        product_row.addStretch()
        add_product_layout.addLayout(product_row)

        # 供应商方案选择行（选择产品后显示）
        self.scheme_row = QHBoxLayout()
        self.scheme_row.addWidget(QLabel("供应商方案:"))
        self.scheme_combo = QComboBox()
        self.scheme_combo.setMinimumWidth(200)
        self.scheme_combo.currentIndexChanged.connect(self.on_scheme_selected)
        self.scheme_row.addWidget(self.scheme_combo)
        
        use_scheme_btn = QPushButton("使用方案")
        use_scheme_btn.setStyleSheet("background-color: #059669; color: white; border: none; border-radius: 4px; padding: 6px 12px;")
        use_scheme_btn.clicked.connect(self.apply_scheme)
        self.scheme_row.addWidget(use_scheme_btn)
        
        new_scheme_btn = QPushButton("+ 新建方案")
        new_scheme_btn.setStyleSheet("background-color: #f59e0b; color: white; border: none; border-radius: 4px; padding: 6px 12px;")
        new_scheme_btn.clicked.connect(self.create_new_scheme)
        self.scheme_row.addWidget(new_scheme_btn)
        
        self.scheme_row.addStretch()
        add_product_layout.addLayout(self.scheme_row)
        
        self.scheme_row_widget = None  # 用于控制显示
        add_product_group.setLayout(add_product_layout)
        layout.addWidget(add_product_group)

        # ===== 下部分：产品明细列表 =====
        items_group = QGroupBox("产品明细列表")
        items_layout = QVBoxLayout()

        self.items_table = QTableWidget()
        self.items_table.setColumnCount(6)
        self.items_table.setHorizontalHeaderLabels(["产品编号", "OE号", "数量", "单价", "总价", "操作"])
        self.items_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.items_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.items_table.setMinimumHeight(200)
        if self.readonly:
            self.items_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        items_layout.addWidget(self.items_table)
        items_group.setLayout(items_layout)
        layout.addWidget(items_group)

        # 底部按钮
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        if not self.readonly:
            save_btn = QPushButton("保存")
            save_btn.setFixedWidth(100)
            save_btn.clicked.connect(self.save_pi)
            save_btn.setStyleSheet("""
                QPushButton { background-color: #2563eb; color: white; border: none; border-radius: 6px; padding: 8px 24px; }
                QPushButton:hover { background-color: #1d4ed8; }
            """)
            buttons_layout.addWidget(save_btn)

        close_btn = QPushButton("关闭" if self.readonly else "取消")
        close_btn.setFixedWidth(100)
        close_btn.clicked.connect(self.accept if self.readonly else self.reject)
        close_btn.setStyleSheet("""
            QPushButton { background-color: #e5e7eb; color: #374151; border: none; border-radius: 6px; padding: 8px 24px; }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        buttons_layout.addWidget(close_btn)

        layout.addLayout(buttons_layout)
        self.setLayout(layout)

    def add_item(self):
        product = self.product_combo.currentData()
        quantity = self.quantity_input.text().strip()
        unit_price = self.unit_price_input.text().strip()

        if not product or not quantity or not unit_price:
            QMessageBox.warning(self, "警告", "请填写完整信息")
            return

        try:
            quantity = int(quantity)
            unit_price = float(unit_price)
        except ValueError:
            QMessageBox.warning(self, "警告", "数量和单价必须是数字")
            return

        self.items.append({
            "product_id": product.get('id'),
            "product_code": product.get('product_code'),
            "oe_number": product.get('oe_number'),
            "quantity": quantity,
            "unit_price": unit_price,
            "customer_code": "",
            "detail_desc": "",
            "remark": ""
        })

        self.update_items_table()

    def update_items_table(self):
        self.items_table.setRowCount(len(self.items))
        for row, item in enumerate(self.items):
            quantity = item.get('quantity', 0)
            unit_price = item.get('unit_price', 0)
            total_price = quantity * unit_price
            
            self.items_table.setItem(row, 0, QTableWidgetItem(item.get('product_code', '')))
            self.items_table.setItem(row, 1, QTableWidgetItem(item.get('oe_number', '')))
            self.items_table.setItem(row, 2, QTableWidgetItem(str(quantity)))
            self.items_table.setItem(row, 3, QTableWidgetItem(f"{unit_price:.2f}"))
            self.items_table.setItem(row, 4, QTableWidgetItem(f"{total_price:.2f}"))

            delete_btn = QPushButton("删除")
            delete_btn.setFixedWidth(50)
            delete_btn.setStyleSheet("background-color: #dc2626; color: white; border: none; border-radius: 4px; padding: 2px;")
            delete_btn.clicked.connect(lambda _, r=row: self.remove_item(r))
            self.items_table.setCellWidget(row, 5, delete_btn)

    def remove_item(self, row):
        del self.items[row]
        self.update_items_table()

    def save_pi(self):
        """异步保存PI单"""
        customer_id = self.customer_combo.currentData()
        currency = self.currency_combo.currentText()

        if not customer_id:
            QMessageBox.warning(self, "警告", "请选择客户")
            return

        if not self.items:
            QMessageBox.warning(self, "警告", "请添加产品明细")
            return

        # 禁用保存按钮，防止重复提交
        self.setEnabled(False)
        
        pi_data = {
            "dept_id": self.dept_id,
            "customer_id": customer_id,
            "currency": currency,
            "items": self.items,
            "payment_stages": []
        }

        def do_save():
            try:
                if self.is_edit:
                    result = self.api_client.update_pi(self.pi.get('id'), pi_data)
                else:
                    result = self.api_client.create_pi(pi_data)
                
                # 清除PI列表缓存
                cache_manager.delete(CACHE_KEYS['PI_LIST'])
                
                # 在主线程更新UI
                from PySide6.QtCore import QMetaObject, Qt
                QMetaObject.invokeMethod(self, "_on_save_success",
                                        Qt.ConnectionType.QueuedConnection)
            except Exception as e:
                self._save_error_msg = str(e)
                from PySide6.QtCore import QMetaObject, Qt
                QMetaObject.invokeMethod(self, "_on_save_error",
                                        Qt.ConnectionType.QueuedConnection)
        
        # 使用全局线程池异步保存
        _global_thread_pool.submit(do_save)
    
    def _on_save_success(self):
        """保存成功回调"""
        self.setEnabled(True)
        QMessageBox.information(self, "成功", "PI单已保存")
        self.accept()
    
    def _on_save_error(self):
        """保存失败回调"""
        self.setEnabled(True)
        error_msg = getattr(self, '_save_error_msg', '未知错误')
        QMessageBox.warning(self, "错误", f"保存失败: {error_msg}")


class PurchaseDialog(QDialog):
    """[已废弃] 请使用 client/widgets/purchase_dialog.py 中的 PurchaseDialog
    新规格文档: docs/superpowers/specs/功能迭代/采购Dialog设计.md
    """
    def __init__(self, api_client, dept_id, purchase=None):
        super().__init__()
        self.api_client = api_client
        self.dept_id = dept_id
        self.purchase = purchase
        self.is_edit = purchase is not None
        self.pi_orders = []
        self.suppliers = []
        self.products = []
        self.items = []
        self.init_ui()
        QTimer.singleShot(0, self.load_data)
    
    def load_data(self):
        self.load_pi_orders()
        self.load_suppliers()
        self.load_products()

    def load_pi_orders(self):
        try:
            self.pi_orders = self.api_client.get_pi_orders()
            self.pi_combo.clear()
            self.pi_combo.addItem("", "")
            for pi in self.pi_orders:
                self.pi_combo.addItem(f"{pi.get('pi_no')} - {pi.get('total_amount')} {pi.get('currency')}", pi)
        except Exception as e:
            print(f"加载PI订单失败: {e}")

    def load_suppliers(self):
        try:
            self.suppliers = self.api_client.get_suppliers()
            self.supplier_combo.clear()
            self.supplier_combo.addItem("", "")
            for s in self.suppliers:
                self.supplier_combo.addItem(f"{s.get('supplier_code')} - {s.get('supplier_name')}", s.get('id'))
        except Exception as e:
            print(f"加载供应商失败: {e}")

    def load_products(self):
        try:
            self.products = self.api_client.get_products()
            self.product_combo.clear()
            self.product_combo.addItem("", "")
            for p in self.products:
                self.product_combo.addItem(f"{p.get('product_code')} - {p.get('oe_number')}", p)
        except Exception as e:
            print(f"加载产品失败: {e}")

    def init_ui(self):
        self.setWindowTitle("编辑采购单" if self.is_edit else "新建采购单")
        self.setFixedSize(700, 600)

        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        form_layout = QFormLayout()
        form_layout.setSpacing(10)

        self.pi_combo = QComboBox()
        self.pi_combo.setFixedHeight(35)
        form_layout.addRow("关联PI单:", self.pi_combo)

        self.supplier_combo = QComboBox()
        self.supplier_combo.setFixedHeight(35)
        form_layout.addRow("供应商:", self.supplier_combo)

        self.currency_combo = QComboBox()
        self.currency_combo.setFixedHeight(35)
        self.currency_combo.addItems(["CNY", "USD", "EUR"])
        form_layout.addRow("货币:", self.currency_combo)

        layout.addLayout(form_layout)

        items_group = QGroupBox("采购明细")
        items_layout = QVBoxLayout()

        toolbar = QHBoxLayout()
        self.product_combo = QComboBox()
        self.product_combo.setFixedWidth(150)
        toolbar.addWidget(self.product_combo)

        self.factory_code_input = QLineEdit()
        self.factory_code_input.setPlaceholderText("工厂编号")
        self.factory_code_input.setFixedWidth(100)
        toolbar.addWidget(self.factory_code_input)

        self.color_input = QLineEdit()
        self.color_input.setPlaceholderText("颜色")
        self.color_input.setFixedWidth(80)
        toolbar.addWidget(self.color_input)

        self.quantity_input = QLineEdit()
        self.quantity_input.setPlaceholderText("数量")
        self.quantity_input.setFixedWidth(80)
        toolbar.addWidget(self.quantity_input)

        self.unit_price_input = QLineEdit()
        self.unit_price_input.setPlaceholderText("单价")
        self.unit_price_input.setFixedWidth(80)
        toolbar.addWidget(self.unit_price_input)

        add_item_btn = QPushButton("+ 添加")
        add_item_btn.clicked.connect(self.add_item)
        toolbar.addWidget(add_item_btn)

        items_layout.addLayout(toolbar)

        self.items_table = QTableWidget()
        self.items_table.setColumnCount(10)
        self.items_table.setHorizontalHeaderLabels(["产品编号", "OE号", "工厂编号", "颜色", "数量", "单价", "出厂价", "出厂含税价", "FOB价", "操作"])
        self.items_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.items_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        self.items_table.setColumnWidth(1, 120)
        self.items_table.setFixedHeight(200)
        items_layout.addWidget(self.items_table)

        items_group.setLayout(items_layout)
        layout.addWidget(items_group)

        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        save_btn = QPushButton("保存")
        save_btn.clicked.connect(self.save_purchase)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        buttons_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)

        self.setLayout(layout)

    def add_item(self):
        product = self.product_combo.currentData()
        factory_code = self.factory_code_input.text().strip()
        color = self.color_input.text().strip()
        quantity = self.quantity_input.text().strip()
        unit_price = self.unit_price_input.text().strip()

        if not product or not quantity or not unit_price:
            QMessageBox.warning(self, "警告", "请填写完整信息")
            return

        try:
            quantity = int(quantity)
            unit_price = float(unit_price)
        except ValueError:
            QMessageBox.warning(self, "警告", "数量和单价必须是数字")
            return

        self.items.append({
            "product_id": product.get('id'),
            "product_code": product.get('product_code'),
            "oe_number": product.get('oe_number'),
            "factory_code": factory_code,
            "color": color,
            "detail_requirement": '',
            "quantity": quantity,
            "unit_price": unit_price,
            "price_ex_factory": product.get('exw_price_excl') or product.get('price_ex_factory'),
            "price_ex_factory_tax": product.get('exw_price_incl') or product.get('price_ex_factory_tax'),
            "price_fob": product.get('fob_price_excl') or product.get('price_fob'),
            "price_fob_tax": product.get('fob_price_incl') or product.get('price_fob_tax'),
            "remark": ""
        })

        self.factory_code_input.clear()
        self.color_input.clear()
        self.quantity_input.clear()
        self.unit_price_input.clear()

        self.update_items_table()

    def update_items_table(self):
        self.items_table.setRowCount(len(self.items))
        for row, item in enumerate(self.items):
            self.items_table.setItem(row, 0, QTableWidgetItem(item.get('product_code', '')))
            self.items_table.setItem(row, 1, QTableWidgetItem(item.get('oe_number', '')))
            self.items_table.setItem(row, 2, QTableWidgetItem(item.get('factory_code', '')))
            self.items_table.setItem(row, 3, QTableWidgetItem(item.get('color', '')))
            self.items_table.setItem(row, 4, QTableWidgetItem(str(item.get('quantity', ''))))
            self.items_table.setItem(row, 5, QTableWidgetItem(str(item.get('unit_price', ''))))
            self.items_table.setItem(row, 6, QTableWidgetItem(str(item.get('price_ex_factory', ''))))
            self.items_table.setItem(row, 7, QTableWidgetItem(str(item.get('price_ex_factory_tax', ''))))
            self.items_table.setItem(row, 8, QTableWidgetItem(str(item.get('price_fob', ''))))

            delete_btn = QPushButton("删除")
            delete_btn.setFixedWidth(50)
            delete_btn.setStyleSheet("background-color: #dc2626; color: white; border: none; border-radius: 4px; padding: 2px;")
            delete_btn.clicked.connect(lambda _, r=row: self.remove_item(r))
            self.items_table.setCellWidget(row, 9, delete_btn)

    def remove_item(self, row):
        del self.items[row]
        self.update_items_table()

    def save_purchase(self):
        pi = self.pi_combo.currentData()
        supplier_id = self.supplier_combo.currentData()
        currency = self.currency_combo.currentText()

        if not pi:
            QMessageBox.warning(self, "警告", "请选择关联的PI单")
            return

        if not supplier_id:
            QMessageBox.warning(self, "警告", "请选择供应商")
            return

        if not self.items:
            QMessageBox.warning(self, "警告", "请添加采购明细")
            return

        purchase_data = {
            "dept_id": self.dept_id,
            "pi_id": pi.get('id'),
            "supplier_id": supplier_id,
            "currency": currency,
            "items": self.items
        }

        try:
            if self.is_edit:
                self.api_client.update_purchase(self.purchase.get('id'), purchase_data)
            else:
                self.api_client.create_purchase(purchase_data)
            QMessageBox.information(self, "成功", "采购单已保存")
            self.accept()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存失败: {str(e)}")


class ShipmentDialog(QDialog):
    """出货对话框 - 支持多阶段管理"""
    def __init__(self, api_client, dept_id=None, shipment=None):
        super().__init__()
        self.api_client = api_client
        self.dept_id = dept_id or "S"
        self.shipment = shipment
        self.is_edit = shipment is not None
        self.pi_orders = []
        self.stages = []  # 出货阶段列表
        self.init_ui()
        QTimer.singleShot(0, self.load_data)

    def load_data(self):
        """加载PI订单数据"""
        try:
            self.pi_orders = self.api_client.get_pi_orders()
            self.pi_combo.clear()
            self.pi_combo.addItem("", "")
            for pi in self.pi_orders:
                # 显示PI号、客户名、金额
                customer_name = pi.get('customer_name', '') or ''
                display_text = f"{pi.get('pi_no')} - {customer_name} - ${pi.get('total_amount', 0)}"
                self.pi_combo.addItem(display_text, pi)
            
            # 编辑模式：回填数据
            if self.shipment:
                idx = self.pi_combo.findData(self.shipment.get('pi_id'))
                if idx >= 0:
                    self.pi_combo.setCurrentIndex(idx)
                    self.pi_combo.setEnabled(False)  # 编辑时不能修改PI
                # 加载已有的stages（如果有）
                if 'stages' in self.shipment and self.shipment['stages']:
                    self.stages = self.shipment['stages']
                    self.refresh_stages_table()
        except Exception as e:
            print(f"加载PI订单失败: {e}")

    def init_ui(self):
        """初始化UI"""
        self.setWindowTitle("编辑出货" if self.is_edit else "新建出货")
        self.setMinimumSize(800, 600)
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        # 基本信息区域
        basic_group = QGroupBox("基本信息")
        basic_layout = QFormLayout()
        basic_layout.setSpacing(10)

        self.pi_combo = QComboBox()
        self.pi_combo.setFixedHeight(35)
        basic_layout.addRow("PI单:", self.pi_combo)

        basic_group.setLayout(basic_layout)
        layout.addWidget(basic_group)

        # 出货阶段管理区域
        stages_group = QGroupBox("出货阶段管理")
        stages_layout = QVBoxLayout()

        # 阶段列表表格
        self.stages_table = QTableWidget()
        self.stages_table.setColumnCount(6)
        self.stages_table.setHorizontalHeaderLabels(["阶段名称", "出货日期", "柜号", "提单号", "数量", "操作"])
        self.stages_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.stages_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.stages_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.stages_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.stages_table.setMaximumHeight(250)
        stages_layout.addWidget(self.stages_table)

        # 添加阶段按钮
        add_stage_layout = QHBoxLayout()
        add_stage_layout.addStretch()

        add_stage_btn = QPushButton("+ 添加出货阶段")
        add_stage_btn.clicked.connect(self.add_stage)
        add_stage_btn.setStyleSheet("""
            QPushButton {
                background-color: #3b82f6;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #2563eb; }
        """)
        add_stage_layout.addWidget(add_stage_btn)
        stages_layout.addLayout(add_stage_layout)

        # 汇总信息
        summary_layout = QHBoxLayout()
        summary_layout.addStretch()
        self.total_stages_label = QLabel("阶段数: 0")
        self.total_stages_label.setStyleSheet("font-weight: bold; color: #374151;")
        summary_layout.addWidget(self.total_stages_label)
        summary_layout.addSpacing(20)
        self.total_qty_label = QLabel("总数量: 0")
        self.total_qty_label.setStyleSheet("font-weight: bold; color: #10b981;")
        summary_layout.addWidget(self.total_qty_label)
        stages_layout.addLayout(summary_layout)

        stages_group.setLayout(stages_layout)
        layout.addWidget(stages_group)

        # 按钮区域
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        save_btn = QPushButton("保存")
        save_btn.clicked.connect(self.save_shipment)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        buttons_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)
        self.setLayout(layout)

    def add_stage(self):
        """添加出货阶段"""
        dialog = ShipmentStageDialog(self, len(self.stages))
        if dialog.exec() == QDialog.DialogCode.Accepted:
            stage_data = dialog.get_stage_data()
            self.stages.append(stage_data)
            self.refresh_stages_table()

    def edit_stage(self, index):
        """编辑出货阶段"""
        if index < 0 or index >= len(self.stages):
            return
        dialog = ShipmentStageDialog(self, index, self.stages[index])
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.stages[index] = dialog.get_stage_data()
            self.refresh_stages_table()

    def delete_stage(self, index):
        """删除出货阶段"""
        if index < 0 or index >= len(self.stages):
            return
        reply = QMessageBox.question(self, "确认", f"确定要删除阶段 '{self.stages[index].get('stage_name')}' 吗？")
        if reply == QMessageBox.StandardButton.Yes:
            self.stages.pop(index)
            self.refresh_stages_table()

    def refresh_stages_table(self):
        """刷新阶段表格"""
        self.stages_table.setRowCount(len(self.stages))
        
        total_qty = 0
        
        for row, stage in enumerate(self.stages):
            # 阶段名称
            self.stages_table.setItem(row, 0, QTableWidgetItem(stage.get('stage_name', '')))
            # 出货日期
            self.stages_table.setItem(row, 1, QTableWidgetItem(str(stage.get('shipment_date', ''))))
            # 柜号
            self.stages_table.setItem(row, 2, QTableWidgetItem(stage.get('container_no', '')))
            # 提单号
            self.stages_table.setItem(row, 3, QTableWidgetItem(stage.get('bl_no', '')))
            # 数量
            qty = stage.get('quantity', 0) or 0
            total_qty += float(qty)
            self.stages_table.setItem(row, 4, QTableWidgetItem(str(qty)))
            
            # 操作按钮
            btn_widget = QWidget()
            btn_layout = QHBoxLayout()
            btn_layout.setContentsMargins(0, 0, 0, 0)
            
            edit_btn = QPushButton("编辑")
            edit_btn.setFixedWidth(50)
            edit_btn.clicked.connect(lambda _, r=row: self.edit_stage(r))
            btn_layout.addWidget(edit_btn)
            
            del_btn = QPushButton("删除")
            del_btn.setFixedWidth(50)
            del_btn.setStyleSheet("color: #ef4444;")
            del_btn.clicked.connect(lambda _, r=row: self.delete_stage(r))
            btn_layout.addWidget(del_btn)
            
            btn_widget.setLayout(btn_layout)
            self.stages_table.setCellWidget(row, 5, btn_widget)
        
        # 更新汇总
        self.total_stages_label.setText(f"阶段数: {len(self.stages)}")
        self.total_qty_label.setText(f"总数量: {total_qty}")

    def save_shipment(self):
        """保存出货记录"""
        pi = self.pi_combo.currentData()
        if not pi:
            QMessageBox.warning(self, "警告", "请选择PI单")
            return

        if not self.stages:
            QMessageBox.warning(self, "警告", "请至少添加一个出货阶段")
            return

        # 构建stages数据
        stages_data = []
        for stage in self.stages:
            stages_data.append({
                'id': stage.get('id'),  # 编辑时可能有id
                'stage_name': stage.get('stage_name'),
                'shipment_date': stage.get('shipment_date'),
                'container_no': stage.get('container_no'),
                'bl_no': stage.get('bl_no'),
                'quantity': stage.get('quantity'),
                'ci_document': stage.get('ci_document'),
                'pl_document': stage.get('pl_document'),
                'storage_location': stage.get('storage_location'),
                'remark': stage.get('remark')
            })

        shipment_data = {
            "dept_id": self.dept_id,
            "pi_id": pi.get('id'),
            "stages": stages_data,
            "items": []  # 出货明细，暂时为空
        }

        try:
            if self.is_edit:
                self.api_client.update_shipment(self.shipment.get('id'), shipment_data)
            else:
                self.api_client.create_shipment(shipment_data)
            QMessageBox.information(self, "成功", "出货记录已保存")
            self.accept()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存失败: {str(e)}")


class ShipmentStageDialog(QDialog):
    """出货阶段对话框"""
    def __init__(self, parent, stage_no, stage_data=None):
        super().__init__(parent)
        self.stage_no = stage_no
        self.stage_data = stage_data or {}
        self.is_edit = stage_data is not None
        self.init_ui()
        
        # 编辑模式回填数据
        if self.is_edit:
            self.load_stage_data()

    def init_ui(self):
        self.setWindowTitle(f"编辑出货阶段" if self.is_edit else f"添加出货阶段 #{self.stage_no + 1}")
        self.setFixedSize(500, 450)
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        form_layout = QFormLayout()
        form_layout.setSpacing(10)

        # 阶段名称
        self.stage_name_input = QLineEdit()
        self.stage_name_input.setFixedHeight(35)
        self.stage_name_input.setText(f"出货{self.stage_no + 1}")
        form_layout.addRow("阶段名称:", self.stage_name_input)

        # 出货日期
        self.shipment_date_input = QDateEdit()
        self.shipment_date_input.setCalendarPopup(True)
        self.shipment_date_input.setFixedHeight(35)
        self.shipment_date_input.setDate(QDate.currentDate())
        form_layout.addRow("出货日期:", self.shipment_date_input)

        # 柜号
        self.container_no_input = QLineEdit()
        self.container_no_input.setFixedHeight(35)
        self.container_no_input.setPlaceholderText("如: MSKU1234567")
        form_layout.addRow("柜号:", self.container_no_input)

        # 提单号
        self.bl_no_input = QLineEdit()
        self.bl_no_input.setFixedHeight(35)
        self.bl_no_input.setPlaceholderText("如: BL123456789")
        form_layout.addRow("提单号:", self.bl_no_input)

        # 数量
        self.quantity_input = QLineEdit()
        self.quantity_input.setFixedHeight(35)
        self.quantity_input.setPlaceholderText("出货数量")
        form_layout.addRow("数量:", self.quantity_input)

        # 存放位置
        self.storage_location_input = QLineEdit()
        self.storage_location_input.setFixedHeight(35)
        self.storage_location_input.setPlaceholderText("如: 上海港")
        form_layout.addRow("存放位置:", self.storage_location_input)

        # CI文件路径
        self.ci_document_input = QLineEdit()
        self.ci_document_input.setFixedHeight(35)
        self.ci_document_input.setPlaceholderText("CI文件路径或编号")
        form_layout.addRow("CI文件:", self.ci_document_input)

        # PL文件路径
        self.pl_document_input = QLineEdit()
        self.pl_document_input.setFixedHeight(35)
        self.pl_document_input.setPlaceholderText("PL文件路径或编号")
        form_layout.addRow("PL文件:", self.pl_document_input)

        # 备注
        self.remark_input = QLineEdit()
        self.remark_input.setFixedHeight(35)
        form_layout.addRow("备注:", self.remark_input)

        layout.addLayout(form_layout)

        # 按钮
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        save_btn = QPushButton("确定")
        save_btn.clicked.connect(self.validate_and_accept)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        buttons_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)
        self.setLayout(layout)

    def load_stage_data(self):
        """加载阶段数据（编辑模式）"""
        self.stage_name_input.setText(self.stage_data.get('stage_name', ''))
        if self.stage_data.get('shipment_date'):
            date = QDate.fromString(str(self.stage_data['shipment_date'])[:10], "yyyy-MM-dd")
            if date.isValid():
                self.shipment_date_input.setDate(date)
        self.container_no_input.setText(self.stage_data.get('container_no', ''))
        self.bl_no_input.setText(self.stage_data.get('bl_no', ''))
        self.quantity_input.setText(str(self.stage_data.get('quantity', '')))
        self.storage_location_input.setText(self.stage_data.get('storage_location', ''))
        self.ci_document_input.setText(self.stage_data.get('ci_document', ''))
        self.pl_document_input.setText(self.stage_data.get('pl_document', ''))
        self.remark_input.setText(self.stage_data.get('remark', ''))

    def validate_and_accept(self):
        """验证并确认"""
        if not self.stage_name_input.text().strip():
            QMessageBox.warning(self, "警告", "请输入阶段名称")
            return
        
        try:
            qty = float(self.quantity_input.text() or 0)
            if qty <= 0:
                QMessageBox.warning(self, "警告", "数量必须大于0")
                return
        except ValueError:
            QMessageBox.warning(self, "警告", "数量必须是数字")
            return

        self.accept()

    def get_stage_data(self):
        """获取阶段数据"""
        return {
            'id': self.stage_data.get('id') if self.is_edit else None,
            'stage_name': self.stage_name_input.text().strip(),
            'shipment_date': self.shipment_date_input.date().toString("yyyy-MM-dd"),
            'container_no': self.container_no_input.text().strip(),
            'bl_no': self.bl_no_input.text().strip(),
            'quantity': float(self.quantity_input.text() or 0),
            'storage_location': self.storage_location_input.text().strip(),
            'ci_document': self.ci_document_input.text().strip(),
            'pl_document': self.pl_document_input.text().strip(),
            'remark': self.remark_input.text().strip()
        }


class CustomerPaymentDialog(QDialog):
    def __init__(self, api_client, dept_id=None, payment=None):
        super().__init__()
        self.api_client = api_client
        self.dept_id = dept_id or "S"
        self.payment = payment
        self.is_edit = payment is not None
        self.pi_orders = []
        self.init_ui()
        QTimer.singleShot(0, self.load_data)

    def load_data(self):
        try:
            self.pi_orders = self.api_client.get_pi_orders()
            self.pi_combo.clear()
            self.pi_combo.addItem("", "")
            for pi in self.pi_orders:
                self.pi_combo.addItem(f"{pi.get('pi_no')} - {pi.get('total_amount')}", pi)
            if self.payment:
                idx = self.pi_combo.findData(self.payment.get('pi_id'))
                if idx >= 0:
                    self.pi_combo.setCurrentIndex(idx)
        except Exception as e:
            print(f"加载PI订单失败: {e}")

    def init_ui(self):
        self.setWindowTitle("编辑客户付款" if self.is_edit else "新建客户付款")
        self.setFixedSize(500, 350)
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        form_layout = QFormLayout()
        form_layout.setSpacing(10)

        self.pi_combo = QComboBox()
        self.pi_combo.setFixedHeight(35)
        form_layout.addRow("PI单:", self.pi_combo)

        self.payment_date_input = QDateEdit()
        self.payment_date_input.setCalendarPopup(True)
        self.payment_date_input.setFixedHeight(35)
        self.payment_date_input.setDate(QDate.currentDate())
        form_layout.addRow("付款日期:", self.payment_date_input)

        self.amount_input = QLineEdit()
        self.amount_input.setFixedHeight(35)
        form_layout.addRow("付款金额:", self.amount_input)

        self.payment_method_combo = QComboBox()
        self.payment_method_combo.setFixedHeight(35)
        self.payment_method_combo.addItems(["银行转账", "现金", "支票", "其他"])
        form_layout.addRow("付款方式:", self.payment_method_combo)

        self.remark_input = QTextEdit()
        self.remark_input.setFixedHeight(80)
        form_layout.addRow("备注:", self.remark_input)

        layout.addLayout(form_layout)

        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        save_btn = QPushButton("保存")
        save_btn.clicked.connect(self.save_payment)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        buttons_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)
        self.setLayout(layout)

        if self.payment:
            self.payment_date_input.setDate(QDate.fromString(self.payment.get('payment_date', ''), "yyyy-MM-dd"))
            self.amount_input.setText(str(self.payment.get('actual_amount', '')))
            method_map = {"银行转账": 0, "现金": 1, "支票": 2, "其他": 3}
            self.payment_method_combo.setCurrentIndex(method_map.get(self.payment.get('payment_method', '银行转账'), 0))
            self.remark_input.setPlainText(self.payment.get('remark', ''))

    def save_payment(self):
        pi = self.pi_combo.currentData()
        if not pi:
            QMessageBox.warning(self, "警告", "请选择PI单")
            return

        amount = self.amount_input.text().strip()
        if not amount:
            QMessageBox.warning(self, "警告", "请输入付款金额")
            return

        try:
            amount = float(amount)
        except ValueError:
            QMessageBox.warning(self, "警告", "付款金额必须是数字")
            return

        payment_data = {
            "dept_id": self.dept_id,
            "pi_id": pi.get('id'),
            "customer_id": pi.get('customer_id'),
            "amount": amount,
            "actual_amount": amount,
            "payment_date": self.payment_date_input.date().toString("yyyy-MM-dd"),
            "payment_method": self.payment_method_combo.currentText(),
            "remark": self.remark_input.toPlainText()
        }

        try:
            if self.is_edit:
                self.api_client.update_customer_payment(self.payment.get('id'), payment_data)
            else:
                self.api_client.create_customer_payment(payment_data)
            QMessageBox.information(self, "成功", "付款记录已保存")
            self.accept()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存失败: {str(e)}")


class SupplierPaymentDialog(QDialog):
    """供应商付款对话框 - 支持多阶段管理"""
    def __init__(self, api_client, dept_id=None, payment=None):
        super().__init__()
        self.api_client = api_client
        self.dept_id = dept_id or "S"
        self.payment = payment
        self.is_edit = payment is not None
        self.suppliers = []
        self.purchases = []
        self.stages = []  # 付款阶段列表
        self.init_ui()
        QTimer.singleShot(0, self.load_data)

    def load_data(self):
        """加载供应商和采购单数据"""
        try:
            self.suppliers = self.api_client.get_suppliers()
            self.supplier_combo.clear()
            self.supplier_combo.addItem("", "")
            for s in self.suppliers:
                self.supplier_combo.addItem(f"{s.get('supplier_code')} - {s.get('supplier_name')}", s.get('id'))
            
            # 编辑模式：回填数据
            if self.payment:
                idx = self.supplier_combo.findData(self.payment.get('supplier_id'))
                if idx >= 0:
                    self.supplier_combo.setCurrentIndex(idx)
                # 加载已有的stages
                if 'stages' in self.payment and self.payment['stages']:
                    self.stages = self.payment['stages']
                    self.refresh_stages_table()
        except Exception as e:
            print(f"加载供应商失败: {e}")

    def load_purchases(self):
        """加载供应商的采购单"""
        supplier_id = self.supplier_combo.currentData()
        if not supplier_id:
            return
        try:
            purchases = self.api_client.get_purchases_by_supplier(supplier_id)
            self.purchase_combo.clear()
            self.purchase_combo.addItem("", "")
            for p in purchases:
                self.purchase_combo.addItem(f"PO-{p.get('id')} - {p.get('total_amount', 0)}", p)
            # 编辑模式：回填采购单
            if self.payment and self.payment.get('po_id'):
                for i in range(self.purchase_combo.count()):
                    data = self.purchase_combo.itemData(i)
                    if data and data.get('id') == self.payment.get('po_id'):
                        self.purchase_combo.setCurrentIndex(i)
                        break
        except Exception as e:
            print(f"加载采购单失败: {e}")

    def init_ui(self):
        """初始化UI"""
        self.setWindowTitle("编辑供应商付款" if self.is_edit else "新建供应商付款")
        self.setMinimumSize(700, 600)
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        # 基本信息区域
        basic_group = QGroupBox("基本信息")
        basic_layout = QFormLayout()
        basic_layout.setSpacing(10)

        self.supplier_combo = QComboBox()
        self.supplier_combo.setFixedHeight(35)
        self.supplier_combo.currentIndexChanged.connect(self.load_purchases)
        basic_layout.addRow("供应商:", self.supplier_combo)

        self.purchase_combo = QComboBox()
        self.purchase_combo.setFixedHeight(35)
        basic_layout.addRow("采购单:", self.purchase_combo)

        self.payment_method_combo = QComboBox()
        self.payment_method_combo.setFixedHeight(35)
        self.payment_method_combo.addItems(["银行转账", "现金", "支票", "其他"])
        basic_layout.addRow("付款方式:", self.payment_method_combo)

        basic_group.setLayout(basic_layout)
        layout.addWidget(basic_group)

        # 付款阶段管理区域
        stages_group = QGroupBox("付款阶段管理")
        stages_layout = QVBoxLayout()

        # 阶段列表表格
        self.stages_table = QTableWidget()
        self.stages_table.setColumnCount(5)
        self.stages_table.setHorizontalHeaderLabels(["阶段名称", "应付金额", "已付金额", "状态", "操作"])
        self.stages_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.stages_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.stages_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.stages_table.setMaximumHeight(200)
        stages_layout.addWidget(self.stages_table)

        # 添加阶段按钮区域
        add_stage_layout = QHBoxLayout()
        add_stage_layout.addStretch()

        add_deposit_btn = QPushButton("+ 添加定金")
        add_deposit_btn.clicked.connect(lambda: self.add_stage('deposit'))
        add_deposit_btn.setStyleSheet("""
            QPushButton {
                background-color: #f59e0b;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
            }
            QPushButton:hover { background-color: #d97706; }
        """)
        add_stage_layout.addWidget(add_deposit_btn)

        add_balance_btn = QPushButton("+ 添加尾款")
        add_balance_btn.clicked.connect(lambda: self.add_stage('balance'))
        add_balance_btn.setStyleSheet("""
            QPushButton {
                background-color: #3b82f6;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
            }
            QPushButton:hover { background-color: #2563eb; }
        """)
        add_stage_layout.addWidget(add_balance_btn)

        stages_layout.addLayout(add_stage_layout)

        # 汇总信息
        summary_layout = QHBoxLayout()
        summary_layout.addStretch()
        self.total_label = QLabel("总金额: 0.00")
        self.total_label.setStyleSheet("font-weight: bold; color: #374151;")
        summary_layout.addWidget(self.total_label)
        summary_layout.addSpacing(20)
        self.paid_label = QLabel("已付: 0.00")
        self.paid_label.setStyleSheet("font-weight: bold; color: #10b981;")
        summary_layout.addWidget(self.paid_label)
        summary_layout.addSpacing(20)
        self.unpaid_label = QLabel("未付: 0.00")
        self.unpaid_label.setStyleSheet("font-weight: bold; color: #ef4444;")
        summary_layout.addWidget(self.unpaid_label)
        stages_layout.addLayout(summary_layout)

        stages_group.setLayout(stages_layout)
        layout.addWidget(stages_group)

        # 按钮区域
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        save_btn = QPushButton("保存")
        save_btn.clicked.connect(self.save_payment)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        buttons_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)
        self.setLayout(layout)

        # 编辑模式：回填付款方式
        if self.payment:
            method_map = {"银行转账": 0, "现金": 1, "支票": 2, "其他": 3}
            self.payment_method_combo.setCurrentIndex(method_map.get(self.payment.get('payment_method', '银行转账'), 0))

    def add_stage(self, stage_type):
        """添加付款阶段"""
        dialog = SupplierPaymentStageDialog(self, stage_type, len(self.stages))
        if dialog.exec() == QDialog.DialogCode.Accepted:
            stage_data = dialog.get_stage_data()
            self.stages.append(stage_data)
            self.refresh_stages_table()

    def edit_stage(self, index):
        """编辑付款阶段"""
        if index < 0 or index >= len(self.stages):
            return
        dialog = SupplierPaymentStageDialog(self, self.stages[index].get('stage_type', 'balance'), 
                                           index, self.stages[index])
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.stages[index] = dialog.get_stage_data()
            self.refresh_stages_table()

    def delete_stage(self, index):
        """删除付款阶段"""
        if index < 0 or index >= len(self.stages):
            return
        reply = QMessageBox.question(self, "确认", f"确定要删除阶段 '{self.stages[index].get('stage_name')}' 吗？")
        if reply == QMessageBox.StandardButton.Yes:
            self.stages.pop(index)
            self.refresh_stages_table()

    def refresh_stages_table(self):
        """刷新阶段表格"""
        self.stages_table.setRowCount(len(self.stages))
        status_map = {1: "待付", 2: "部分付", 3: "已付清"}
        
        total = 0
        paid = 0
        
        for row, stage in enumerate(self.stages):
            # 阶段名称
            self.stages_table.setItem(row, 0, QTableWidgetItem(stage.get('stage_name', '')))
            # 应付金额
            amount = stage.get('amount', 0) or 0
            total += float(amount)
            self.stages_table.setItem(row, 1, QTableWidgetItem(f"{float(amount):,.2f}"))
            # 已付金额
            stage_paid = stage.get('paid_amount', 0) or 0
            paid += float(stage_paid)
            self.stages_table.setItem(row, 2, QTableWidgetItem(f"{float(stage_paid):,.2f}"))
            # 状态
            status = stage.get('status', 1)
            self.stages_table.setItem(row, 3, QTableWidgetItem(status_map.get(status, "未知")))
            # 操作按钮
            btn_widget = QWidget()
            btn_layout = QHBoxLayout()
            btn_layout.setContentsMargins(0, 0, 0, 0)
            
            edit_btn = QPushButton("编辑")
            edit_btn.setFixedWidth(50)
            edit_btn.clicked.connect(lambda _, r=row: self.edit_stage(r))
            btn_layout.addWidget(edit_btn)
            
            del_btn = QPushButton("删除")
            del_btn.setFixedWidth(50)
            del_btn.setStyleSheet("color: #ef4444;")
            del_btn.clicked.connect(lambda _, r=row: self.delete_stage(r))
            btn_layout.addWidget(del_btn)
            
            btn_widget.setLayout(btn_layout)
            self.stages_table.setCellWidget(row, 4, btn_widget)
        
        # 更新汇总
        unpaid = total - paid
        self.total_label.setText(f"总金额: {total:,.2f}")
        self.paid_label.setText(f"已付: {paid:,.2f}")
        self.unpaid_label.setText(f"未付: {unpaid:,.2f}")

    def save_payment(self):
        """保存付款记录"""
        supplier_id = self.supplier_combo.currentData()
        purchase = self.purchase_combo.currentData()

        if not supplier_id:
            QMessageBox.warning(self, "警告", "请选择供应商")
            return

        if not purchase:
            QMessageBox.warning(self, "警告", "请选择采购单")
            return

        if not self.stages:
            QMessageBox.warning(self, "警告", "请至少添加一个付款阶段")
            return

        # 构建stages数据
        stages_data = []
        for stage in self.stages:
            stages_data.append({
                'id': stage.get('id'),  # 编辑时可能有id
                'stage_type': stage.get('stage_type'),
                'stage_name': stage.get('stage_name'),
                'amount': stage.get('amount'),
                'paid_amount': stage.get('paid_amount', 0),
                'status': stage.get('status', 1),
                'payment_date': stage.get('payment_date'),
                'payment_proof': stage.get('payment_proof'),
                'remark': stage.get('remark')
            })

        payment_data = {
            "dept_id": self.dept_id,
            "supplier_id": supplier_id,
            "po_id": purchase.get('id'),
            "payment_method": self.payment_method_combo.currentText(),
            "stages": stages_data,
            "remark": ""
        }

        try:
            print(f"DEBUG - 保存供应商付款: {payment_data}")
            if self.is_edit:
                result = self.api_client.update_supplier_payment(self.payment.get('id'), payment_data)
            else:
                result = self.api_client.create_supplier_payment(payment_data)
            print(f"DEBUG - 保存成功: {result}")
            QMessageBox.information(self, "成功", "付款记录已保存")
            self.accept()
        except Exception as e:
            print(f"DEBUG - 保存失败: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.warning(self, "错误", f"保存失败: {str(e)}")


class SupplierPaymentStageDialog(QDialog):
    """付款阶段编辑对话框"""
    def __init__(self, parent, stage_type, index, stage_data=None):
        super().__init__(parent)
        self.stage_type = stage_type
        self.index = index
        self.stage_data = stage_data or {}
        self.is_edit = stage_data is not None
        self.init_ui()

    def init_ui(self):
        stage_type_name = "定金" if self.stage_type == 'deposit' else f"尾款{self.index}"
        self.setWindowTitle(f"编辑{stage_type_name}" if self.is_edit else f"添加{stage_type_name}")
        self.setFixedSize(400, 350)
        
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        form_layout = QFormLayout()
        form_layout.setSpacing(10)

        # 阶段名称
        self.name_input = QLineEdit()
        self.name_input.setFixedHeight(35)
        default_name = self.stage_data.get('stage_name', '')
        if not default_name:
            default_name = "定金" if self.stage_type == 'deposit' else f"尾款{self.index + 1}"
        self.name_input.setText(default_name)
        form_layout.addRow("阶段名称:", self.name_input)

        # 应付金额
        self.amount_input = QLineEdit()
        self.amount_input.setFixedHeight(35)
        self.amount_input.setText(str(self.stage_data.get('amount', '')))
        form_layout.addRow("应付金额:", self.amount_input)

        # 已付金额
        self.paid_input = QLineEdit()
        self.paid_input.setFixedHeight(35)
        self.paid_input.setText(str(self.stage_data.get('paid_amount', '0')))
        form_layout.addRow("已付金额:", self.paid_input)

        # 付款日期
        self.date_input = QDateEdit()
        self.date_input.setCalendarPopup(True)
        self.date_input.setFixedHeight(35)
        payment_date = self.stage_data.get('payment_date')
        if payment_date:
            self.date_input.setDate(QDate.fromString(str(payment_date)[:10], "yyyy-MM-dd"))
        else:
            self.date_input.setDate(QDate.currentDate())
        form_layout.addRow("付款日期:", self.date_input)

        # 备注
        self.remark_input = QTextEdit()
        self.remark_input.setFixedHeight(60)
        self.remark_input.setPlainText(self.stage_data.get('remark', ''))
        form_layout.addRow("备注:", self.remark_input)

        layout.addLayout(form_layout)

        # 按钮
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        save_btn = QPushButton("确定")
        save_btn.clicked.connect(self.accept)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
        """)
        buttons_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
        """)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)
        self.setLayout(layout)

    def get_stage_data(self):
        """获取阶段数据"""
        amount = float(self.amount_input.text() or 0)
        paid = float(self.paid_input.text() or 0)
        
        # 自动计算状态
        status = 1  # 待付
        if paid >= amount and amount > 0:
            status = 3  # 已付清
        elif paid > 0:
            status = 2  # 部分付

        return {
            'id': self.stage_data.get('id'),  # 编辑时保留原id
            'stage_type': self.stage_type,
            'stage_name': self.name_input.text(),
            'amount': amount,
            'paid_amount': paid,
            'status': status,
            'payment_date': self.date_input.date().toString("yyyy-MM-dd"),
            'remark': self.remark_input.toPlainText()
        }


class InventoryDialog(QDialog):
    def __init__(self, api_client, dept_id=None, inventory=None, product_id=None, oe_number=None):
        super().__init__()
        self.api_client = api_client
        self.dept_id = dept_id or "S"
        self.inventory = inventory
        self.preselected_product_id = product_id  # 预选的产品ID
        self.preselected_oe_number = oe_number    # 预选的OE号
        self.is_edit = inventory is not None
        self.products = []
        self.init_ui()
        QTimer.singleShot(0, self.load_products)

    def load_products(self):
        try:
            # 尝试从缓存加载
            self.products = cache_manager.get(CACHE_KEYS['PRODUCTS'], max_age=300)
            if self.products is None:
                self.products = self.api_client.get_products()
                cache_manager.set(CACHE_KEYS['PRODUCTS'], self.products)
            
            self.product_combo.clear()
            self.product_combo.addItem("请选择产品", None)
            
            # 如果有预选的OE号，只显示匹配的产品
            if self.preselected_oe_number:
                for p in self.products:
                    if p.get('oe_number') == self.preselected_oe_number:
                        self.product_combo.addItem(f"{p.get('product_code')} - {p.get('oe_number')}", p)
                        # 自动选中
                        self.product_combo.setCurrentIndex(1)
                        break
            else:
                for p in self.products:
                    if isinstance(p, dict):
                        self.product_combo.addItem(f"{p.get('product_code')} - {p.get('oe_number')}", p)
            
            # 编辑模式：选中当前库存的产品
            if self.inventory:
                target_product_id = self.inventory.get('product_id')
                matched = False
                for i in range(self.product_combo.count()):
                    data = self.product_combo.itemData(i)
                    if isinstance(data, dict) and data.get('id') == target_product_id:
                        self.product_combo.setCurrentIndex(i)
                        matched = True
                        break
                # 2026-06-23：库存产品不在标准列表时（被禁用/跨部门/软删除等），
                # 单独拉一次详情补入下拉框，避免"产品丢失"看不到当前绑定
                if not matched and target_product_id:
                    try:
                        detail = self.api_client.get_product_detail(target_product_id)
                        if isinstance(detail, dict) and detail.get('id'):
                            label = f"{detail.get('product_code', '')} - {detail.get('oe_number', '')} (已停用)"
                            self.product_combo.addItem(label, detail)
                            self.product_combo.setCurrentIndex(self.product_combo.count() - 1)
                    except Exception as ex:
                        print(f"补拉库存产品详情失败: {ex}")
            # 2026-06-23：编辑模式下产品由库存记录自动绑定，下拉框锁定不可改
            if self.inventory and self.inventory.get('product_id'):
                self.product_combo.setEnabled(False)
                self.product_combo.setToolTip("产品由库存记录自动绑定，不可修改")
            # 新建模式：有预选产品ID
            elif self.preselected_product_id:
                for i in range(self.product_combo.count()):
                    data = self.product_combo.itemData(i)
                    if isinstance(data, dict) and data.get('id') == self.preselected_product_id:
                        self.product_combo.setCurrentIndex(i)
                        break
        except Exception as e:
            print(f"加载产品失败: {e}")
    
    def load_suppliers(self):
        """加载供应商列表"""
        try:
            # 尝试从缓存加载
            suppliers = cache_manager.get(CACHE_KEYS['SUPPLIERS'], max_age=300)
            if suppliers is None:
                suppliers = self.api_client.get_suppliers()
                cache_manager.set(CACHE_KEYS['SUPPLIERS'], suppliers)
            
            self.supplier_combo.clear()
            self.supplier_combo.addItem("请选择供应商", None)
            for s in suppliers:
                self.supplier_combo.addItem(f"{s.get('supplier_code')} - {s.get('supplier_name')}", s)
            
            # 如果是编辑，设置选中的供应商
            if self.inventory and self.inventory.get('supplier_id'):
                for i in range(self.supplier_combo.count()):
                    data = self.supplier_combo.itemData(i)
                    if data and data.get('id') == self.inventory.get('supplier_id'):
                        self.supplier_combo.setCurrentIndex(i)
                        break
        except Exception as e:
            print(f"加载供应商失败: {e}")
    
    def load_customers(self):
        """加载客户列表"""
        try:
            # 尝试从缓存加载
            customers = cache_manager.get(CACHE_KEYS['CUSTOMERS'], max_age=300)
            if customers is None:
                customers = self.api_client.get_customers()
                cache_manager.set(CACHE_KEYS['CUSTOMERS'], customers)
            
            self.customer_combo.clear()
            self.customer_combo.addItem("请选择客户", None)
            for c in customers:
                self.customer_combo.addItem(f"{c.get('customer_code')} - {c.get('customer_name')}", c)
            
            # 如果是编辑，设置选中的客户
            if self.inventory and self.inventory.get('customer_id'):
                for i in range(self.customer_combo.count()):
                    data = self.customer_combo.itemData(i)
                    if data and data.get('id') == self.inventory.get('customer_id'):
                        self.customer_combo.setCurrentIndex(i)
                        break
        except Exception as e:
            print(f"加载客户失败: {e}")

    def init_ui(self):
        self.setWindowTitle("编辑库存" if self.is_edit else "新建库存")
        self.setFixedSize(550, 550)
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        form_layout = QFormLayout()
        form_layout.setSpacing(10)

        self.product_combo = QComboBox()
        self.product_combo.setFixedHeight(35)
        form_layout.addRow("产品 *:", self.product_combo)

        # 供应商选择
        self.supplier_combo = QComboBox()
        self.supplier_combo.setFixedHeight(35)
        form_layout.addRow("供应商 *:", self.supplier_combo)
        # 加载供应商列表
        self.load_suppliers()

        # 客户选择
        self.customer_combo = QComboBox()
        self.customer_combo.setFixedHeight(35)
        form_layout.addRow("客户 *:", self.customer_combo)
        # 加载客户列表
        self.load_customers()

        self.stock_type_combo = QComboBox()
        self.stock_type_combo.setFixedHeight(35)
        self.stock_type_combo.addItems(["采购在途", "待入库", "已入库", "历史库存"])
        form_layout.addRow("库存类型:", self.stock_type_combo)

        self.quantity_input = QLineEdit()
        self.quantity_input.setFixedHeight(35)
        form_layout.addRow("数量:", self.quantity_input)

        self.location_input = QLineEdit()
        self.location_input.setFixedHeight(35)
        form_layout.addRow("库位:", self.location_input)

        self.remark_input = QTextEdit()
        self.remark_input.setFixedHeight(80)
        form_layout.addRow("备注:", self.remark_input)

        layout.addLayout(form_layout)

        status_group = QGroupBox("库存状态颜色")
        status_layout = QHBoxLayout()
        status_layout.addWidget(QLabel("● 黄色: 采购在途"))
        status_layout.addWidget(QLabel("● 蓝色: 待入库"))
        status_layout.addWidget(QLabel("● 绿色: 已入库"))
        status_layout.addWidget(QLabel("● 黑色: 历史库存"))
        status_group.setLayout(status_layout)
        layout.addWidget(status_group)

        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        save_btn = QPushButton("保存")
        save_btn.clicked.connect(self.save_inventory)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        buttons_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)
        self.setLayout(layout)

        if self.inventory:
            # 回填数量（使用 total_quantity 或 quantity）
            qty = self.inventory.get('total_quantity') or self.inventory.get('quantity', '')
            self.quantity_input.setText(str(qty))
            # 回填库位
            self.location_input.setText(self.inventory.get('current_location', '') or '')
            # 回填备注
            self.remark_input.setPlainText(self.inventory.get('remark', '') or '')
            # 回填库存类型
            stock_type = self.inventory.get('stock_type', 2)
            # stock_type 转下拉框索引
            index_map = {1: 0, 2: 1, 3: 2, 4: 3}  # 1->采购在途(0), 2->待入库(1), 3->已入库(2), 4->历史库存(3)
            if isinstance(stock_type, str):
                try:
                    stock_type = int(stock_type)
                except ValueError:
                    stock_type = 2
            combo_index = index_map.get(stock_type, 0)
            self.stock_type_combo.setCurrentIndex(combo_index)

    def save_inventory(self):
        product = self.product_combo.currentData()
        if not product:
            QMessageBox.warning(self, "警告", "请选择产品")
            return

        quantity = self.quantity_input.text().strip()
        if not quantity:
            QMessageBox.warning(self, "警告", "请输入数量")
            return

        try:
            # 先转为浮点数，再转为整数（处理后端返回的 123.0 格式）
            quantity = int(float(quantity))
        except ValueError:
            QMessageBox.warning(self, "警告", "数量必须是整数")
            return

        # 库存类型映射：下拉框文本 -> stock_type
        type_map = {"采购在途": 1, "待入库": 2, "已入库": 3, "历史库存": 4}
        stock_type_text = self.stock_type_combo.currentText()
        stock_type = type_map.get(stock_type_text, 1)

        # 获取供应商（从供应商选择下拉框）
        supplier = self.supplier_combo.currentData()
        supplier_id = supplier.get('id') if supplier else None
        
        # 获取客户（从客户选择下拉框）
        customer = self.customer_combo.currentData()
        if not customer:
            QMessageBox.warning(self, "警告", "请选择客户")
            return
        customer_id = customer.get('id')
        
        inventory_data = {
            "dept_id": self.dept_id,
            "product_id": product.get('id'),
            "supplier_id": supplier_id,
            "customer_id": customer_id,
            "quantity": quantity,
            "current_location": self.location_input.text().strip() or 'WAREHOUSE',
            "stock_type": stock_type,
            "remark": self.remark_input.toPlainText().strip()
        }

        try:
            if self.is_edit:
                self.api_client.update_inventory(self.inventory.get('id'), inventory_data)
            else:
                self.api_client.create_inventory(inventory_data)
            # 清除库存缓存
            cache_manager.delete(CACHE_KEYS['INVENTORY_SUMMARY'])
            print("DEBUG - 已清除库存缓存")
            QMessageBox.information(self, "成功", "库存记录已保存")
            self.accept()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存失败: {str(e)}")


class UpdateDialog(QDialog):
    """检查更新对话框"""

    def __init__(self, current_version, latest_version, changelog, force_update, parent=None):
        super().__init__(parent)
        self.current_version = current_version
        self.latest_version = latest_version
        self.force_update = force_update
        self.update_accepted = False
        self.setWindowTitle("发现新版本")
        self.setMinimumSize(450, 300)
        self._setup_ui(changelog)

    def _setup_ui(self, changelog):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # 图标和标题
        icon_label = QLabel("🎉")
        icon_label.setAlignment(Qt.AlignCenter)
        icon_label.setStyleSheet("font-size: 48px;")
        layout.addWidget(icon_label)

        # 版本信息
        version_text = f"当前版本: {self.current_version} → 最新版本: {self.latest_version}"
        version_label = QLabel(version_text)
        version_label.setAlignment(Qt.AlignCenter)
        version_font = QFont()
        version_font.setPointSize(11)
        version_label.setFont(version_font)
        layout.addWidget(version_label)

        # 更新日志
        changelog_label = QLabel("更新内容:")
        layout.addWidget(changelog_label)

        changelog_text = QTextEdit()
        changelog_text.setReadOnly(True)
        changelog_text.setPlainText(changelog or "暂无更新说明")
        changelog_text.setMaximumHeight(120)
        layout.addWidget(changelog_text)

        # 强制更新提示
        if self.force_update:
            warning_label = QLabel("⚠️ 此版本为强制更新，必须安装后才能继续使用")
            warning_label.setStyleSheet("color: #ef4444; font-weight: bold;")
            layout.addWidget(warning_label)

        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        if not self.force_update:
            skip_btn = QPushButton("稍后")
            skip_btn.setDefault(False)
            skip_btn.clicked.connect(self.reject)
            btn_layout.addWidget(skip_btn)

        install_btn = QPushButton("立即更新" if not self.force_update else "安装更新")
        install_btn.setDefault(True)
        install_btn.setStyleSheet("""
            QPushButton {
                background-color: #3b82f6;
                color: white;
                padding: 8px 24px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #2563eb;
            }
        """)
        install_btn.clicked.connect(self._on_install)
        btn_layout.addWidget(install_btn)

        layout.addLayout(btn_layout)

    def _on_install(self):
        self.update_accepted = True
        self.accept()

    def was_update_accepted(self):
        return self.update_accepted


class BugReportDialog(QDialog):
    """Bug 报告提交对话框"""

    def __init__(self, api_client=None, parent=None):
        super().__init__(parent)
        self.api_client = api_client
        self.selected_files = []
        self.setWindowTitle("反馈 Bug")
        self.setMinimumSize(500, 600)
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # 标题
        title_label = QLabel("🐛 反馈问题")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        layout.addWidget(title_label)

        # 表单
        form_layout = QFormLayout()
        form_layout.setSpacing(12)

        # 标题
        self.title_input = QLineEdit()
        self.title_input.setPlaceholderText("简要描述问题（必填）")
        self.title_input.setMaxLength(100)
        form_layout.addRow("问题标题 *:", self.title_input)

        # 描述
        self.desc_input = QTextEdit()
        self.desc_input.setPlaceholderText("详细描述问题，包括复现步骤（必填）")
        self.desc_input.setMaximumHeight(100)
        form_layout.addRow("问题描述 *:", self.desc_input)

        # 版本号
        from config import Config
        self.version_input = QLineEdit()
        self.version_input.setPlaceholderText("如 v1.0.0")
        self.version_input.setText(Config.APP_VERSION)
        form_layout.addRow("版本号:", self.version_input)

        # 联系方式
        self.contact_input = QLineEdit()
        self.contact_input.setPlaceholderText("邮箱或工号（选填）")
        form_layout.addRow("联系方式:", self.contact_input)

        layout.addLayout(form_layout)

        # 截图上传
        screenshot_label = QLabel("截图（可选，最多 5 张）")
        layout.addWidget(screenshot_label)

        self.upload_area = QWidget()
        self.upload_area.setStyleSheet("""
            border: 2px dashed #ccc;
            border-radius: 8px;
            padding: 20px;
            background: #fafafa;
        """)
        upload_layout = QVBoxLayout(self.upload_area)
        upload_layout.setContentsMargins(0, 0, 0, 0)
        upload_layout.setAlignment(Qt.AlignCenter)

        self.upload_label = QLabel("📷 点击或拖拽上传图片\n支持 jpg, png 格式，单张最大 5MB")
        self.upload_label.setAlignment(Qt.AlignCenter)
        self.upload_label.setStyleSheet("color: #666;")
        upload_layout.addWidget(self.upload_label)

        self.upload_area.mousePressEvent = lambda e: self.select_files()
        self.upload_area.setAcceptDrops(True)

        layout.addWidget(self.upload_area)

        # 预览区域
        self.preview_container = QWidget()
        preview_layout = QHBoxLayout(self.preview_container)
        preview_layout.setContentsMargins(0, 0, 0, 0)
        self.preview_container.setMaximumHeight(80)
        self.preview_scroll = QScrollArea()
        self.preview_scroll.setWidgetResizable(True)
        self.preview_scroll.setMaximumHeight(80)
        self.preview_widget = QWidget()
        self.preview_widget_layout = QHBoxLayout(self.preview_widget)
        self.preview_scroll.setWidget(self.preview_widget)
        preview_layout.addWidget(self.preview_scroll)
        layout.addWidget(self.preview_container)

        # 2026-06-25：日志上传
        log_label = QLabel("📋 日志文件（可选）")
        layout.addWidget(log_label)

        log_btn_layout = QHBoxLayout()
        self.log_upload_btn = QPushButton("📁 选择日志文件")
        self.log_upload_btn.clicked.connect(self.select_log_file)
        log_btn_layout.addWidget(self.log_upload_btn)

        # 自动收集日志按钮
        self.collect_log_btn = QPushButton("🔍 自动收集")
        self.collect_log_btn.setToolTip("自动收集客户端运行日志")
        self.collect_log_btn.clicked.connect(self._collect_client_log)
        log_btn_layout.addWidget(self.collect_log_btn)

        log_btn_layout.addStretch()

        self.log_count_label = QLabel("未选择日志")
        self.log_count_label.setStyleSheet("color: #666;")
        log_btn_layout.addWidget(self.log_count_label)
        layout.addLayout(log_btn_layout)

        self.selected_log_files = []  # 存储日志文件路径

        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)

        self.submit_btn = QPushButton("提交反馈")
        self.submit_btn.setStyleSheet("""
            QPushButton {
                background-color: #ef4444;
                color: white;
                padding: 8px 24px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #dc2626;
            }
        """)
        self.submit_btn.clicked.connect(self._on_submit)
        btn_layout.addWidget(self.submit_btn)

        layout.addLayout(btn_layout)

    def select_files(self):
        from PySide6.QtWidgets import QFileDialog
        files, _ = QFileDialog.getOpenFileNames(
            self, "选择截图", "", "图片文件 (*.jpg *.jpeg *.png)"
        )
        if files:
            self.add_files(files)

    def add_files(self, files):
        for file_path in files:
            if len(self.selected_files) >= 5:
                QMessageBox.warning(self, "提示", "最多只能上传 5 张图片")
                break
            # 检查文件大小
            import os
            size = os.path.getsize(file_path)
            if size > 5 * 1024 * 1024:
                QMessageBox.warning(self, "提示", "单张图片最大 5MB")
                continue
            self.selected_files.append(file_path)
        self._update_preview()

    def remove_file(self, index):
        if 0 <= index < len(self.selected_files):
            self.selected_files.pop(index)
            self._update_preview()

    def _update_preview(self):
        # 清除预览
        while self.preview_widget_layout.count():
            item = self.preview_widget_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        for i, file_path in enumerate(self.selected_files):
            preview_item = QWidget()
            preview_layout = QVBoxLayout(preview_item)
            preview_layout.setContentsMargins(5, 5, 5, 5)
            preview_layout.setSpacing(2)

            # 缩略图
            from PySide6.QtGui import QPixmap
            pixmap = QPixmap(file_path).scaled(60, 60, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            img_label = QLabel()
            img_label.setPixmap(pixmap)
            img_label.setAlignment(Qt.AlignCenter)
            preview_layout.addWidget(img_label)

            # 删除按钮
            del_btn = QPushButton("×")
            del_btn.setFixedSize(20, 20)
            del_btn.setStyleSheet("""
                QPushButton {
                    background-color: #ef4444;
                    color: white;
                    border: none;
                    border-radius: 10px;
                    font-size: 14px;
                }
            """)
            del_btn.clicked.connect(lambda _, idx=i: self.remove_file(idx))
            preview_layout.addWidget(del_btn)
            preview_layout.setAlignment(del_btn, Qt.AlignCenter)

            self.preview_widget_layout.addWidget(preview_item)

    def select_log_file(self):
        """选择日志文件"""
        from PySide6.QtWidgets import QFileDialog
        files, _ = QFileDialog.getOpenFileNames(
            self, "选择日志文件", "", "日志文件 (*.log *.txt);;所有文件 (*.*)"
        )
        if files:
            self._add_log_files(files)

    def _add_log_files(self, files):
        """添加日志文件"""
        import os
        for file_path in files:
            if len(self.selected_log_files) >= 3:
                QMessageBox.warning(self, "提示", "最多只能上传 3 个日志文件")
                break
            # 检查文件大小（最大 10MB）
            size = os.path.getsize(file_path)
            if size > 10 * 1024 * 1024:
                QMessageBox.warning(self, "提示", f"日志文件 {os.path.basename(file_path)} 超过 10MB，将被跳过")
                continue
            if file_path not in self.selected_log_files:
                self.selected_log_files.append(file_path)
        self._update_log_count()

    def _collect_client_log(self):
        """自动收集客户端运行日志"""
        import os
        import tempfile
        import subprocess
        from datetime import datetime

        try:
            # 创建临时日志文件
            temp_dir = tempfile.gettempdir()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            log_file = os.path.join(temp_dir, f"pi_manager_log_{timestamp}.log")

            # 收集环境信息
            log_content = []
            log_content.append("=" * 60)
            log_content.append("PI Manager 客户端日志收集")
            log_content.append(f"收集时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            log_content.append("=" * 60)

            # 版本信息
            from config import Config
            log_content.append(f"\n版本信息:")
            log_content.append(f"  APP_VERSION: {Config.APP_VERSION}")
            log_content.append(f"  UPDATE_SERVER: {Config.UPDATE_SERVER_URL}")
            log_content.append(f"  API_BASE_URL: {Config.API_BASE_URL}")

            # 系统信息
            log_content.append(f"\n系统信息:")
            try:
                import platform
                log_content.append(f"  OS: {platform.platform()}")
                log_content.append(f"  Python: {platform.python_version()}")
            except:
                pass

            # 尝试读取应用程序日志
            log_content.append(f"\n最近运行日志:")
            log_content.append("-" * 40)

            # 检查常见日志位置
            log_paths = [
                os.path.expanduser("~/AppData/Local/PI_Manager/logs"),
                os.path.expanduser("~/.pi_manager/logs"),
                os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs"),
            ]

            found_logs = False
            for log_path in log_paths:
                if os.path.exists(log_path):
                    try:
                        for f in os.listdir(log_path):
                            if f.endswith('.log'):
                                fpath = os.path.join(log_path, f)
                                mtime = os.path.getmtime(fpath)
                                log_content.append(f"\n[文件] {fpath} (修改时间: {datetime.fromtimestamp(mtime)})")
                                # 读取最后 100 行
                                with open(fpath, 'r', encoding='utf-8', errors='ignore') as lf:
                                    lines = lf.readlines()
                                    for line in lines[-100:]:
                                        log_content.append(line.rstrip())
                                found_logs = True
                    except Exception as e:
                        log_content.append(f"读取失败: {e}")

            if not found_logs:
                log_content.append("\n未找到应用程序日志文件")

            # 写入临时文件
            with open(log_file, 'w', encoding='utf-8') as f:
                f.write('\n'.join(log_content))

            self._add_log_files([log_file])
            # 使用 QTimer 在主线程中显示消息
            from PySide6.QtCore import QTimer
            QTimer.singleShot(100, lambda: QMessageBox.information(self, "成功", f"已收集日志并添加到附件\n日志文件: {log_file}"))

        except Exception as e:
            from PySide6.QtCore import QTimer
            QTimer.singleShot(100, lambda: QMessageBox.warning(self, "错误", f"收集日志失败: {str(e)}"))

    def _update_log_count(self):
        """更新日志文件计数"""
        count = len(self.selected_log_files)
        if count == 0:
            self.log_count_label.setText("未选择日志")
            self.log_count_label.setStyleSheet("color: #666;")
        else:
            self.log_count_label.setText(f"已选择 {count} 个日志文件")
            self.log_count_label.setStyleSheet("color: #059669; font-weight: bold;")

    def _on_submit(self):
        title = self.title_input.text().strip()
        description = self.desc_input.toPlainText().strip()

        if not title:
            QMessageBox.warning(self, "提示", "请输入问题标题")
            return
        if not description:
            QMessageBox.warning(self, "提示", "请输入问题描述")
            return

        self.submit_btn.setEnabled(False)
        self.submit_btn.setText("提交中...")

        # 直接在主线程中提交（不使用线程以避免跨线程 UI 操作问题）
        self._submit_bug(title, description)

    def _submit_bug(self, title, description):
        """
        🔧 2026-06-29 重写：修复以下问题
        - B1: files 变量类型错误 —— screenshots 用 dict{}，但 logs 用 list.append()，
              dict 没有 append 方法，运行时报 AttributeError
        - B2: MIME 类型写死为 image/jpeg，截图可能是 png
        - B3: 文件句柄关闭逻辑假设 files 是 dict，与 B1 矛盾
        修复后统一用 list 形式（符合 requests 官方 multipart 上传规范），
        requests 会自动为每个 part 设置合适的 Content-Type header。
        """
        try:
            import requests
            import os
            from config import Config

            url = f"{Config.UPDATE_SERVER_URL}/api/bug"
            # 🔧 B1 修复：统一使用 list 格式（field_name, (filename, file_obj, content_type)）
            files = []
            open_files = []  # 🔧 B3 修复：单独跟踪打开的文件句柄，确保 finally 全部关闭

            # 添加截图（最多 5 张）
            for file_path in self.selected_files:
                filename = os.path.basename(file_path)
                ext = filename.rsplit('.', 1)[-1].lower()
                # B2 修复：根据扩展名动态判断 MIME 类型
                if ext in ('jpg', 'jpeg'):
                    content_type = 'image/jpeg'
                elif ext == 'png':
                    content_type = 'image/png'
                else:
                    content_type = 'application/octet-stream'
                fp = open(file_path, 'rb')
                open_files.append(fp)
                files.append(('screenshots', (filename, fp, content_type)))

            # 添加日志文件（最多 3 个）
            for file_path in self.selected_log_files:
                filename = os.path.basename(file_path)
                ext = filename.rsplit('.', 1)[-1].lower()
                if ext == 'log':
                    content_type = 'text/plain'
                elif ext == 'txt':
                    content_type = 'text/plain'
                else:
                    content_type = 'application/octet-stream'
                fp = open(file_path, 'rb')
                open_files.append(fp)
                files.append(('logs', (filename, fp, content_type)))

            # data 部分（非文件字段）
            data = {
                'title': title,
                'description': description,
                'version': self.version_input.text().strip() or Config.APP_VERSION,
                'contact': self.contact_input.text().strip(),
            }

            response = requests.post(url, data=data, files=files, timeout=60)

            # B3 修复：统一在 finally 中关闭所有打开的文件句柄
            for fp in open_files:
                try:
                    fp.close()
                except Exception:
                    pass

            # 响应检查：200 和 201 都算成功
            if response.status_code in (200, 201):
                result = response.json()
                self._show_result(True, f"提交成功！Bug 编号: #{result.get('id', '?')}")
            else:
                self._show_result(False, f"提交失败: HTTP {response.status_code}\n{response.text[:200]}")
        except requests.exceptions.SSLError as e:
            self._show_result(False, f"SSL 错误，更新服务地址可能不可达：\n{str(e)[:200]}")
        except requests.exceptions.ConnectionError as e:
            self._show_result(False, f"连接失败，请检查网络或更新服务地址：\n{str(e)[:200]}")
        except requests.exceptions.Timeout as e:
            self._show_result(False, f"连接超时：\n{str(e)[:200]}")
        except Exception as e:
            self._show_result(False, f"提交失败：\n{str(e)[:200]}")
        finally:
            # 🔧 B3 修复：finally 确保即使请求前抛异常，文件句柄也能关闭
            for fp in open_files:
                try:
                    fp.close()
                except Exception:
                    pass

    def _show_result(self, success, message):
        """显示提交结果"""
        self.submit_btn.setEnabled(True)
        self.submit_btn.setText("提交反馈")
        if success:
            # 先关闭对话框，再显示成功消息
            self.accept()
            QMessageBox.information(None, "成功", "反馈已提交，感谢您的反馈！")
        else:
            QMessageBox.warning(self, "错误", message)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event):
        files = []
        for url in event.mimeData().urls():
            if url.isLocalFile():
                files.append(url.toLocalFile())
        if files:
            self.add_files(files)


def download_and_install_update(download_url, sha256_url, current_exe_path, latest_version):
    """下载并安装更新"""
    import requests
    import hashlib
    import os
    import subprocess
    import tempfile

    print(f"[Update] 开始下载: {download_url}")

    try:
        temp_dir = tempfile.gettempdir()
        exe_filename = f"PI_Manager_Client_{latest_version}.exe"
        temp_exe_path = os.path.join(temp_dir, exe_filename)

        response = requests.get(download_url, stream=True, timeout=60)
        if response.status_code != 200:
            print(f"[Update] 下载失败: HTTP {response.status_code}")
            return False

        with open(temp_exe_path, "wb") as f:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)

        print(f"[Update] 下载完成: {temp_exe_path}")

        # SHA256 校验
        if sha256_url:
            sha256_response = requests.get(sha256_url, timeout=10)
            if sha256_response.status_code == 200:
                expected_sha256 = sha256_response.json().get("sha256", "").strip()
                if expected_sha256:
                    sha256_hash = hashlib.sha256()
                    with open(temp_exe_path, "rb") as f:
                        for chunk in iter(lambda: f.read(8192), b""):
                            sha256_hash.update(chunk)
                    actual_sha256 = sha256_hash.hexdigest()
                    if actual_sha256 != expected_sha256:
                        print(f"[Update] SHA256 校验失败!")
                        os.remove(temp_exe_path)
                        return False

        # 创建安装脚本
        install_script = os.path.join(temp_dir, "update_client.bat")
        script_content = f'''@echo off
chcp 65001 >nul
echo 正在安装更新...
timeout /t 2 /nobreak >nul
del /f /q "{current_exe_path}" 2>nul
move /y "{temp_exe_path}" "{current_exe_path}"
start "" "{current_exe_path}"
'''
        with open(install_script, "w", encoding="utf-8") as f:
            f.write(script_content)

        subprocess.Popen(
            ["cmd.exe", "/c", "start", "cmd", "/c", install_script],
            detached=True,
            creationflags=subprocess.CREATE_NO_WINDOW
        )
        return True

    except Exception as e:
        print(f"[Update] 安装失败: {e}")
        return False


def _strip_version_prefix(version: str) -> str:
    """从 'client/v1.0.0.10' 中提取纯版本号 '1.0.0.10'"""
    for prefix in ('client/', 'server/'):
        if version.startswith(prefix):
            version = version[len(prefix):]
    return version.lstrip('v')


def _normalize_version(version: str) -> list:
    """将版本号字符串拆分为数字段"""
    stripped = _strip_version_prefix(version)
    parts = []
    for part in stripped.split('.'):
        try:
            parts.append(int(part))
        except ValueError:
            parts.append(0)
    return parts


def _compare_versions(v1: str, v2: str) -> int:
    """比较版本号，返回: 1 if v1>v2, -1 if v1<v2, 0 if equal"""
    try:
        p1 = _normalize_version(v1)
        p2 = _normalize_version(v2)
        max_len = max(len(p1), len(p2))
        p1.extend([0] * (max_len - len(p1)))
        p2.extend([0] * (max_len - len(p2)))
        for a, b in zip(p1, p2):
            if a > b:
                return 1
            if a < b:
                return -1
        return 0
    except Exception:
        return 0


def check_for_updates():
    """
    🔧 2026-06-29 修复：
    - A3: latest_version 先去掉 'client/' 和 'v' 前缀再显示
    - A4: 增加 min_compatible 最低版本兼容检查
    - A6: URL 拼接已在返回值层面统一处理
    """
    try:
        import requests
        from config import Config

        current_version = Config.APP_VERSION
        update_server = Config.UPDATE_SERVER_URL

        if not update_server:
            return False, current_version, None, False, None, None, None, False

        url = f"{update_server}/api/version/client/{current_version}"
        response = requests.get(url, timeout=10)

        if response.status_code == 200:
            data = response.json()
            has_update = data.get("has_update", False)

            # A3 修复：去掉 client/ 和 v 前缀后再显示
            latest_raw = data.get("latest", current_version)
            latest_version = _strip_version_prefix(latest_raw)

            # 🔧 2026-06-29 修复：如果最新版本等于当前版本，强制 has_update=False
            # 防止后端比较逻辑错误导致误报更新
            if _compare_versions(latest_version, current_version) <= 0:
                has_update = False

            # A4 修复：检查最低兼容版本
            min_compatible = data.get("min_compatible", "") or ""
            is_blocked = False
            if min_compatible and _compare_versions(current_version, min_compatible) < 0:
                is_blocked = True
                has_update = True  # 强制触发 update dialog 显示阻止消息

            changelog = data.get("changelog", "")
            force_update = data.get("force", False)
            download_url = data.get("download_url", "")
            sha256_url = data.get("sha256_url", "")

            if download_url and not download_url.startswith("http"):
                download_url = f"{update_server}{download_url}"
            if sha256_url and not sha256_url.startswith("http"):
                sha256_url = f"{update_server}{sha256_url}"

            return has_update, latest_version, changelog, force_update, download_url, sha256_url, min_compatible, is_blocked
        else:
            print(f"[Update] 检查更新失败: HTTP {response.status_code}")
            return False, current_version, None, False, None, None, None, False

    except Exception as e:
        print(f"[Update] 检查更新异常: {e}")
        return False, None, None, False, None, None, None, False


def main():
    ctypes.windll.kernel32.SetConsoleOutputCP(65001)
    
    # 启用高DPI缩放（使用推荐方式）
    os.environ["QT_ENABLE_HIGHDPI_SCALING"] = "1"
    os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "0"
    
    app = QApplication(sys.argv)

    # 2026-06-26 修复：全局禁用 QSpinBox/QDoubleSpinBox/QComboBox 的滚轮修改行为
    # 避免用户在滚动页面时不小心改到数值或下拉选项
    class WheelEventFilter(QObject):
        def eventFilter(self, obj, event):
            if event.type() == QEvent.Wheel:
                if isinstance(obj, (QSpinBox, QDoubleSpinBox, QComboBox)):
                    event.ignore()
                    return True
            return super().eventFilter(obj, event)

    app.installEventFilter(WheelEventFilter(app))
    
    # 设置全局字体
    font = QFont()
    font.setFamily("Microsoft YaHei")
    font.setPointSize(10)
    app.setFont(font)

    # 使用带缓存功能的API客户端
    api_client = CachedApiClient()

    try:
        login = LoginWindow(api_client)
        if login.exec() != QDialog.DialogCode.Accepted:
            return

        dept_id = login.get_selected_department() or "S"
        window = MainWindow(api_client, dept_id)
        window.show()

        # 2026-06-25: 异步检查更新（不阻塞界面）
        from config import Config
        if Config.AUTO_CHECK_UPDATE:
            QTimer.singleShot(1000, window._check_update_async)

        sys.exit(app.exec())
    except Exception as e:
        print("Error during application execution:", str(e))
        traceback.print_exc()
        # 显示错误对话框
        error_dialog = QDialog()
        error_dialog.setWindowTitle("启动错误")
        error_dialog.resize(400, 200)
        layout = QVBoxLayout()
        layout.addWidget(QLabel(f"启动失败: {str(e)}"))
        layout.addWidget(QLabel("请检查日志获取详细信息"))
        error_dialog.setLayout(layout)
        error_dialog.exec()
        sys.exit(1)


if __name__ == "__main__":
    main()

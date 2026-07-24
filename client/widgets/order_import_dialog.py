# ============================================================
# 订单导入对话框 - Qt前端实现
# 文件：client/widgets/order_import_dialog.py
# 创建日期：2026-05-29
# 用途：订单导入UI组件
# ============================================================

from typing import List, Dict, Optional, Any
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QTableWidget, QTableWidgetItem, QFileDialog, QMessageBox,
    QProgressBar, QCheckBox, QComboBox, QGroupBox, QScrollArea,
    QWidget, QAbstractItemView, QHeaderView, QLineEdit, QMenu
)
from PySide6.QtCore import Qt, QThread, Signal, Slot, QTimer
from PySide6.QtGui import QColor, QFont

import json
import traceback
import os
import tempfile
from openpyxl import load_workbook, Workbook


class OrderImportDialog(QDialog):
    """
    订单导入对话框
    
    功能：
    - 文件选择和预览
    - 数据预览和字段映射
    - 自动匹配产品
    - 批量导入订单
    """
    
    # 信号定义
    preview_loaded = Signal(list, list, int, int)  # headers, rows, total, columns
    match_completed = Signal(list)  # match results
    import_completed = Signal(bool, int, int, list)  # success, success_count, failed_count, errors
    error_occurred = Signal(str)  # error message
    
    def __init__(self, api_client, parent=None, is_supplement_mode=False, order_id=None):
        super().__init__(parent)
        self.api_client = api_client
        self.preview_data = None
        self.match_results = []
        self.current_mapping = {}
        self._current_temp_pi_no = None  # 存储当前预览的临时PI号 [6.0.2] [6.0.2.1]
        
        # [6.2.1] 补充商品模式参数
        self.is_supplement_mode = is_supplement_mode
        self.order_id = order_id
        self._supplement_order_data = None  # 存储当前订单的完整数据
        # 2026-06-23：客户-产品表缓存（非补充模式也预初始化便于测试）
        self._db_existing_models = {}  # model -> product dict
        self._db_models_loaded = False  # 标记是否已完成加载（避免 race condition）
        # 🔧 2026-06-29：用户右键删除预览行时记录被删除的索引（相对 _raw_preview_rows），
        # 用于在导入时过滤原始 Excel 文件，避免删除的行被导入
        self._deleted_preview_indices = set()

        self.setWindowTitle("补充商品" if is_supplement_mode else "订单导入")
        self.setMinimumSize(1000, 700)
        self.init_ui()

        if is_supplement_mode:
            self._init_supplement_mode()
        else:
            # 自动加载客户列表
            QTimer.singleShot(100, self.load_customers)
    
    def init_ui(self):
        """初始化UI"""
        self.main_layout = QVBoxLayout(self)
        
        # 文件选择区域
        file_group = self._create_file_selection_group()
        self.main_layout.addWidget(file_group)
        
        # 预览区域
        preview_group = self._create_preview_group()
        self.main_layout.addWidget(preview_group)
        
        # PI号预览区域（补充商品模式下隐藏）
        self.pi_preview_group = self._create_pi_preview_group()
        self.main_layout.addWidget(self.pi_preview_group)
        
        # 操作按钮区域
        button_layout = self._create_button_layout()
        self.main_layout.addLayout(button_layout)
    
    def _init_supplement_mode(self):
        """[6.2.1] 初始化补充商品模式"""
        # 隐藏 PI 号预览区域
        if hasattr(self, 'pi_preview_group'):
            self.pi_preview_group.hide()

        # [问题 #27] 补充商品模式：客户应自动读取当前订单，无需选择
        # 隐藏客户选择（继承当前订单客户）
        if hasattr(self, 'customer_combo'):
            self.customer_combo.hide()
        if hasattr(self, 'load_customer_btn'):
            self.load_customer_btn.hide()

        # 隐藏生成 PI 选项
        if hasattr(self, 'generate_pi_checkbox'):
            self.generate_pi_checkbox.hide()

        # 加载当前订单数据（从中获取 customer_id 用于 API 调用）
        self._load_supplement_order_data()

        # 2026-06-23：预加载客户-产品表用于补充商品去重
        # （_db_existing_models / _db_models_loaded 已在 __init__ 中预初始化）
        self._load_db_existing_products()

    def _load_db_existing_products(self, on_loaded=None):
        """[6.2.23] 调 GET /customer-products 拉取该客户所有产品 Model 集合（去重缓存）

        Args:
            on_loaded: 可选回调，DB 加载完后调用（用于在补商品模式下重新过滤预览表）
        """
        customer_id = self._get_current_customer_id()
        if not customer_id:
            return
        try:
            import threading
            def fetch():
                try:
                    resp = self.api_client.get(
                        "/customer-products/",
                        params={"customer_id": customer_id, "page_size": 1000, "page": 1}
                    )
                    items = (resp or {}).get('items') or []
                    self._db_existing_models = {
                        p['customer_model']: p
                        for p in items
                        if p.get('customer_model')
                    }
                    self._db_models_loaded = True
                    print(
                        f"[6.2.23] 补充商品：客户 {customer_id} 共 {len(self._db_existing_models)} 条已存在产品"
                    )
                    if on_loaded:
                        try:
                            on_loaded()
                        except Exception as cb_err:
                            print(f"[6.2.23] on_loaded 回调失败: {cb_err}")
                except Exception as e:
                    print(f"[6.2.23] 预加载客户产品失败: {e}")
                    self._db_existing_models = {}
                    self._db_models_loaded = True  # 失败也标记为已完成（避免永久等待）
                    if on_loaded:
                        try:
                            on_loaded()
                        except Exception:
                            pass
            threading.Thread(target=fetch, daemon=True).start()
        except Exception as e:
            print(f"[6.2.23] 启动预加载失败: {e}")
            self._db_existing_models = {}
            self._db_models_loaded = True

    def _get_current_customer_id(self) -> Optional[int]:
        """读取当前选中的 customer_id（若已选）"""
        # 优先使用补充商品模式下记录的 _target_customer_id
        target = getattr(self, '_target_customer_id', None)
        if target:
            try:
                return int(target)
            except (TypeError, ValueError):
                pass
        cb = getattr(self, 'customer_combo', None)
        if not cb:
            return None
        data = cb.currentData()
        try:
            return int(data) if data else None
        except (TypeError, ValueError):
            return None
    
    def _create_file_selection_group(self) -> QGroupBox:
        """创建文件选择区域"""
        group = QGroupBox("1. 选择文件")
        layout = QVBoxLayout()
        
        # 客户选择行
        customer_layout = QHBoxLayout()
        customer_label = QLabel("客户：")
        customer_layout.addWidget(customer_label)

        # 🔧 2026-07-02 把"加载"按钮改为搜索框，支持按客户名/区域模糊搜索
        self.customer_search = QLineEdit()
        self.customer_search.setPlaceholderText("搜索客户名称/区域")
        self.customer_search.setMinimumWidth(200)
        self.customer_search.setClearButtonEnabled(True)
        self.customer_search.textChanged.connect(self._filter_customers)
        customer_layout.addWidget(self.customer_search)

        self.customer_combo = QComboBox()
        self.customer_combo.setMinimumWidth(200)
        self.customer_combo.currentIndexChanged.connect(self._on_customer_changed)
        customer_layout.addWidget(self.customer_combo)

        # 兼容旧代码：保留 load_customer_btn 引用但隐藏
        self.load_customer_btn = QPushButton("加载")
        self.load_customer_btn.clicked.connect(self.load_customers)
        self.load_customer_btn.hide()

        customer_layout.addStretch()
        layout.addLayout(customer_layout)
        
        # 文件选择行
        file_layout = QHBoxLayout()
        
        self.file_path_label = QLabel("未选择文件")
        self.file_path_label.setStyleSheet("color: gray;")
        file_layout.addWidget(self.file_path_label, 1)
        
        self.browse_btn = QPushButton("浏览...")
        self.browse_btn.clicked.connect(self.browse_file)
        file_layout.addWidget(self.browse_btn)
        
        self.preview_btn = QPushButton("预览")
        self.preview_btn.clicked.connect(self.load_preview)
        self.preview_btn.setEnabled(False)
        file_layout.addWidget(self.preview_btn)
        
        layout.addLayout(file_layout)
        
        group.setLayout(layout)
        return group
    
    def _create_preview_group(self) -> QGroupBox:
        """创建预览区域"""
        group = QGroupBox("2. 数据预览")
        layout = QVBoxLayout()
        
        # 按钮和状态行
        btn_layout = QHBoxLayout()
        
        self.clear_preview_btn = QPushButton("清空")
        self.clear_preview_btn.clicked.connect(self.clear_preview)
        self.clear_preview_btn.setStyleSheet("background-color: #ef4444; color: white; padding: 5px 10px;")
        btn_layout.addWidget(self.clear_preview_btn)
        
        # 2026-06-25：添加"保留不完整行"复选框
        self.keep_incomplete_checkbox = QCheckBox("保留不完整行")
        self.keep_incomplete_checkbox.setToolTip("勾选后，Model 或 Qty 缺失的行也会被导入")
        self.keep_incomplete_checkbox.toggled.connect(self._on_keep_incomplete_toggled)
        btn_layout.addWidget(self.keep_incomplete_checkbox)
        
        btn_layout.addStretch()
        
        # 添加状态标签
        self.preview_status_label = QLabel("请先选择文件并预览")
        self.preview_status_label.setStyleSheet("color: gray; font-style: italic;")
        btn_layout.addWidget(self.preview_status_label)
        
        layout.addLayout(btn_layout)
        
        self.preview_table = QTableWidget()
        self.preview_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.preview_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.preview_table.setMinimumHeight(250)
        # 2026-06-26：启用右键菜单（支持删除行）
        self.preview_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.preview_table.customContextMenuRequested.connect(self._on_preview_context_menu)
        
        layout.addWidget(self.preview_table)
        
        group.setLayout(layout)
        return group
    
    def _create_pi_preview_group(self) -> QGroupBox:
        """创建PI号预览区域"""
        group = QGroupBox("3. PI号设置")
        layout = QVBoxLayout()
        
        # PI号预览行
        pi_layout = QHBoxLayout()
        pi_layout.addWidget(QLabel("预计 PI 号："))
        
        self.pi_preview_input = QLineEdit()
        self.pi_preview_input.setPlaceholderText("选择客户后自动生成")
        self.pi_preview_input.setStyleSheet("background-color: #f3f4f6;")
        pi_layout.addWidget(self.pi_preview_input, 1)
        
        self.refresh_pi_btn = QPushButton("刷新")
        self.refresh_pi_btn.setFixedWidth(60)
        self.refresh_pi_btn.clicked.connect(self._refresh_pi_preview)
        pi_layout.addWidget(self.refresh_pi_btn)
        
        layout.addLayout(pi_layout)
        
        group.setLayout(layout)
        return group
    
    def _refresh_pi_preview(self):
        """刷新PI号预览"""
        customer_id = self.customer_combo.currentData()
        # 从 UserRole+1 获取完整客户信息（用于预览）
        customer = self.customer_combo.currentData(Qt.ItemDataRole.UserRole + 1)
        
        if not customer:
            self.pi_preview_input.clear()
            self.pi_preview_input.setPlaceholderText("请先选择客户")
            return
        
        try:
            customer_code = customer.get('customer_code', 'C001')
            dept_id = getattr(self.parent(), 'dept_id', 'D') if self.parent() else 'D'
            
            from datetime import datetime
            date_part = datetime.now().strftime("%y%m%d")
            preview_pi = f"PI {dept_id}{customer_code[:4].upper().ljust(4, '0')}{date_part}?"
            self.pi_preview_input.setText(preview_pi)
            self._current_temp_pi_no = preview_pi  # [6.0.2] [6.0.2.1] 保存预览值供 start_import() 读取
        except Exception as e:
            print(f"预览PI号失败: {e}")
    
    def _on_customer_changed(self, index):
        """客户选择变更"""
        self._refresh_pi_preview()
        # 更新单条新增按钮状态
        self._update_single_add_button_state()

    def _update_single_add_button_state(self):
        """根据客户选择状态更新单条新增按钮"""
        if not hasattr(self, 'single_add_btn'):
            return

        customer_id = self.customer_combo.currentData()
        has_customer = bool(customer_id)

        self.single_add_btn.setEnabled(has_customer)
        if not has_customer:
            self.single_add_btn.setToolTip("请先选择客户")
            self.single_add_btn.setStyleSheet("""
                QPushButton {
                    background-color: #9ca3af;
                    color: white;
                    font-weight: bold;
                    padding: 10px 20px;
                }
                QPushButton:hover {
                    background-color: #9ca3af;
                }
            """)
        else:
            self.single_add_btn.setToolTip("手动添加单个产品")
            self.single_add_btn.setStyleSheet("""
                QPushButton {
                    background-color: #3b82f6;
                    color: white;
                    font-weight: bold;
                    padding: 10px 20px;
                }
                QPushButton:hover {
                    background-color: #2563eb;
                }
            """)
    
    def clear_preview(self):
        """清空预览表格"""
        self.preview_table.setRowCount(0)
        self.preview_table.setColumnCount(0)
        self.preview_data = None
        self._current_temp_pi_no = None  # [6.0.2] [6.0.2.1] 清空临时PI
        # 🔧 2026-06-29：清空时也清空删除行索引
        self._deleted_preview_indices = set()
        self.preview_status_label.setText("已清空，请重新选择文件")
        self.auto_match_btn.setEnabled(False)
        self.import_btn.setEnabled(False)
        self.create_product_btn.setEnabled(False)
    
    def _create_match_result_group(self) -> QGroupBox:
        """创建匹配结果区域"""
        group = QGroupBox("3. 匹配结果")
        layout = QVBoxLayout()
        
        # 选项区域
        option_layout = QHBoxLayout()
        
        self.delete_block_btn = QPushButton("删除区块")
        self.delete_block_btn.clicked.connect(self.delete_match_block)
        self.delete_block_btn.setStyleSheet("background-color: #ef4444; color: white; padding: 5px 10px;")
        option_layout.addWidget(self.delete_block_btn)
        
        self.auto_match_checkbox = QCheckBox("自动匹配产品")
        self.auto_match_checkbox.setChecked(True)
        option_layout.addWidget(self.auto_match_checkbox)
        
        self.show_unmatched_only_checkbox = QCheckBox("仅显示未匹配项")
        self.show_unmatched_only_checkbox.stateChanged.connect(self.filter_unmatched)
        option_layout.addWidget(self.show_unmatched_only_checkbox)
        
        option_layout.addStretch()
        
        self.match_summary_label = QLabel("待处理: 0 条")
        option_layout.addWidget(self.match_summary_label)
        
        self.create_product_btn = QPushButton("创建产品")
        self.create_product_btn.clicked.connect(self.create_product_for_unmatched)
        self.create_product_btn.setEnabled(False)
        self.create_product_btn.setStyleSheet("background-color: #3b82f6; color: white; padding: 5px 10px;")
        option_layout.addWidget(self.create_product_btn)
        
        layout.addLayout(option_layout)
        
        # 匹配结果表格
        self.match_table = QTableWidget()
        self.match_table.setMinimumHeight(150)
        self.match_table.setColumnCount(5)
        self.match_table.setHorizontalHeaderLabels([
            "行号", "客户产品编号", "OE号", "匹配状态", "产品信息"
        ])
        self.match_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.match_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        
        layout.addWidget(self.match_table)
        
        group.setLayout(layout)
        return group
    
    def delete_match_block(self):
        """删除匹配结果区块"""
        self.match_table.setRowCount(0)
        self.match_results = []
        self.match_summary_label.setText("待处理: 0 条")
        self.import_btn.setEnabled(False)
        self.create_product_btn.setEnabled(False)
        QMessageBox.information(self, "提示", "已清空匹配结果，可重新匹配")
    
    def _create_button_layout(self) -> QHBoxLayout:
        """创建按钮区域"""
        layout = QHBoxLayout()
        
        self.single_add_btn = QPushButton("+ 单条新增")
        self.single_add_btn.clicked.connect(self.open_single_order_dialog)
        self.single_add_btn.setStyleSheet("""
            QPushButton {
                background-color: #3b82f6;
                color: white;
                font-weight: bold;
                padding: 10px 20px;
            }
            QPushButton:hover {
                background-color: #2563eb;
            }
        """)
        layout.addWidget(self.single_add_btn)
        
        layout.addStretch()
        
        self.import_btn = QPushButton("开始导入")
        self.import_btn.clicked.connect(self.start_import)
        self.import_btn.setEnabled(False)
        self.import_btn.setStyleSheet("""
            QPushButton {
                background-color: #2E7D32;
                color: white;
                font-weight: bold;
                padding: 10px 20px;
            }
            QPushButton:disabled {
                background-color: #ccc;
            }
        """)
        layout.addWidget(self.import_btn)
        
        self.close_btn = QPushButton("关闭")
        self.close_btn.clicked.connect(self.close)
        layout.addWidget(self.close_btn)
        
        return layout
    
    def open_single_order_dialog(self):
        """打开单条产品录入对话框"""
        # [问题 #27] 补充商品模式下，从当前订单数据获取 customer_id
        customer_id = None
        if self.is_supplement_mode and self._supplement_order_data:
            customer_id = self._supplement_order_data.get('customer_id')
        else:
            customer_id = self.customer_combo.currentData()

        if not customer_id:
            QMessageBox.warning(self, "提示", "请先选择客户")
            return

        try:
            from widgets.single_order_dialog import SingleOrderDialog
        except ImportError:
            import sys
            sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            from widgets.single_order_dialog import SingleOrderDialog

        # 判断当前模式
        mode = 'supplement' if self.is_supplement_mode else 'import'

        dialog = SingleOrderDialog(
            api_client=self.api_client,
            parent=self,
            customer_id=customer_id,
            mode=mode
        )

        if dialog.exec() == SingleOrderDialog.Accepted:
            product_data = dialog.get_product_data()
            self._add_single_product_to_preview(product_data)

    def _add_single_product_to_preview(self, product_data: dict):
        """将单条产品数据添加到预览表格"""
        if not self.preview_table:
            return

        # [问题 #29] 初始化 preview_data 如果为空（手动添加时没有文件预览数据）
        if self.preview_data is None:
            self.preview_data = {'headers': [], 'rows': [], 'total': 0}

        # 添加到 preview_data 中（用于导入时获取数据）
        product_data['is_temp'] = True  # 标记为临时产品
        self.preview_data['rows'].append(product_data)
        self.preview_data['total'] = len(self.preview_data['rows'])

        # 如果预览表格为空，初始化列
        if self.preview_table.columnCount() == 0:
            headers = ['行号', '客户产品编号', 'OE号', '产品描述', '数量', '单价', '状态']
            self.preview_table.setColumnCount(len(headers))
            self.preview_table.setHorizontalHeaderLabels(headers)

        # 添加新行
        row_idx = self.preview_table.rowCount()
        self.preview_table.insertRow(row_idx)

        # 填充数据
        from PySide6.QtWidgets import QTableWidgetItem
        from PySide6.QtGui import QColor

        self.preview_table.setItem(row_idx, 0, QTableWidgetItem(str(row_idx + 1)))
        self.preview_table.setItem(row_idx, 1, QTableWidgetItem(product_data.get('customer_code', '')))
        self.preview_table.setItem(row_idx, 2, QTableWidgetItem(product_data.get('oe_number', '') or ''))
        self.preview_table.setItem(row_idx, 3, QTableWidgetItem(product_data.get('detail_desc', '') or ''))
        self.preview_table.setItem(row_idx, 4, QTableWidgetItem(str(product_data.get('quantity', 1))))

        price = product_data.get('unit_price')
        price_str = f"${price:.2f}" if price else ''
        self.preview_table.setItem(row_idx, 5, QTableWidgetItem(price_str))

        # 状态标识
        status = '临时' if product_data.get('is_temporary') else '正式'
        status_item = QTableWidgetItem(status)
        if product_data.get('is_temporary'):
            status_item.setBackground(QColor("#fef3c7"))  # 黄色背景标记临时产品
        self.preview_table.setItem(row_idx, 6, status_item)

        # 更新状态标签
        total_rows = self.preview_table.rowCount()
        temp_count = sum(1 for r in range(total_rows)
                        if self.preview_table.item(r, 6) and self.preview_table.item(r, 6).text() == '临时')

        self.preview_status_label.setText(
            f"共 {total_rows} 行数据（手动添加 {total_rows} 条，其中临时产品 {temp_count} 条）"
        )

        # 启用导入按钮
        self.import_btn.setEnabled(True)

        # 高亮新增行
        for col in range(self.preview_table.columnCount()):
            item = self.preview_table.item(row_idx, col)
            if item:
                item.setBackground(QColor("#dbeafe"))  # 浅蓝色背景

        QMessageBox.information(
            self,
            "成功",
            f"产品已添加到预览列表\n"
            f"客户产品编号: {product_data.get('customer_code')}\n"
            f"数量: {product_data.get('quantity')}"
        )
    
    @Slot()
    def load_customers(self):
        """加载客户列表"""
        try:
            response = self.api_client.get("/customers/")
            if response:
                # 🔧 2026-07-02 保存完整客户列表，用于搜索过滤
                self._all_customers = list(response)
                self._filter_customers(self.customer_search.text() if hasattr(self, 'customer_search') else "")

            # 初始化单条新增按钮状态
            QTimer.singleShot(100, self._update_single_add_button_state)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载客户失败: {e}")

    def _filter_customers(self, keyword: str):
        """根据搜索关键词过滤客户列表（按客户名或区域模糊匹配）"""
        if not hasattr(self, '_all_customers'):
            return

        # 阻止信号触发（避免刷新PI预览）
        self.customer_combo.blockSignals(True)
        try:
            self.customer_combo.clear()
            self.customer_combo.addItem("请选择客户", None)

            keyword_lower = keyword.strip().lower()
            for customer in self._all_customers:
                name = (customer.get('customer_name') or '').lower()
                region = (customer.get('region') or customer.get('country') or '').lower()
                code = (customer.get('customer_code') or '').lower()

                # 关键词为空显示全部；否则按客户名/区域/客户代码匹配
                if not keyword_lower or keyword_lower in name or keyword_lower in region or keyword_lower in code:
                    self.customer_combo.addItem(
                        customer.get('customer_name', ''),
                        customer.get('id'))
                    # 存储完整客户信息到 UserRole+1
                    self.customer_combo.setItemData(
                        self.customer_combo.count() - 1,
                        customer,
                        Qt.ItemDataRole.UserRole + 1)
        finally:
            self.customer_combo.blockSignals(False)

    @Slot()
    def browse_file(self):
        """浏览并选择文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择Excel文件",
            "",
            "Excel文件 (*.xlsx *.xls);;所有文件 (*.*)"
        )
        
        if file_path:
            self.file_path_label.setText(file_path)
            self.file_path_label.setStyleSheet("color: black;")
            self.preview_btn.setEnabled(True)
            
            # 自动加载预览
            self.load_preview()
    
    @Slot()
    def load_preview(self):
        """加载文件预览"""
        file_path = self.file_path_label.text()
        if not file_path or file_path == "未选择文件":
            QMessageBox.warning(self, "提示", "请先选择文件")
            return
        
        self.preview_status_label.setText("正在加载预览...")
        self.preview_btn.setEnabled(False)
        
        # 使用工作线程加载预览
        self.preview_worker = PreviewWorker(self.api_client, file_path)
        self.preview_worker.preview_ready.connect(self.on_preview_ready)
        self.preview_worker.error.connect(self.on_preview_error)
        self.preview_worker.finished.connect(lambda: self.preview_btn.setEnabled(True))
        self.preview_worker.start()
    
    @Slot(list, list, int, int)
    def on_preview_ready(self, headers: list, rows: list, total: int, columns: int):
        """预览加载完成 - 显示原始Excel全部内容"""
        # 🔧 2026-06-29：重新加载预览时清空之前删除的行索引
        self._deleted_preview_indices = set()
        # 保存原始（未过滤）数据
        self._raw_preview_rows = rows
        # 2026-06-26：保存初始预览行（用于不完整行计数，勾选过滤后保持不变）
        self._initial_preview_rows = rows
        self.preview_data = {'headers': headers, 'rows': rows, 'total': total}
        self._skipped_indices = set()  # 补充商品模式下被过滤的原始行索引

        # 2026-06-25：查找 Model 和 Qty 列索引
        model_col_idx = self._find_model_column(headers)
        qty_col_idx = self._find_qty_column(headers)

        # [6.2.1/6.2.23] 补充商品模式下过滤已存在的商品
        if self.is_supplement_mode:
            # 若 DB 还在加载，调度一次重过滤；否则直接过滤
            if not self._db_models_loaded:
                self._load_db_existing_products(on_loaded=self._refilter_preview_after_db_load)
            self._skipped_indices = self._compute_skipped_indices(rows, headers)
            display_rows = [r for i, r in enumerate(rows) if i not in self._skipped_indices]
            self.preview_data['rows'] = display_rows
            self.preview_data['total'] = len(display_rows)
            total = len(display_rows)
        else:
            display_rows = rows
            self.preview_data['rows'] = rows

        # 显示所有原始列
        display_headers = ['行号'] + headers
        self.preview_table.setColumnCount(len(display_headers))
        self.preview_table.setHorizontalHeaderLabels(display_headers)

        # 2026-06-25：统计不完整行
        incomplete_count = 0
        incomplete_rows = set()  # 存储不完整行的索引

        self.preview_table.setRowCount(len(display_rows))
        for row_idx, row in enumerate(display_rows):
            self.preview_table.setItem(row_idx, 0, QTableWidgetItem(str(row_idx + 1)))
            for col_idx, value in enumerate(row):
                self.preview_table.setItem(
                    row_idx, col_idx + 1,
                    QTableWidgetItem(str(value) if value else "")
                )

            # 2026-06-25：检查 Model 和 Qty 是否缺失
            is_incomplete = False
            if model_col_idx is not None and model_col_idx < len(row):
                model_val = str(row[model_col_idx]).strip() if row[model_col_idx] else ''
                if not model_val:
                    is_incomplete = True
            else:
                is_incomplete = True  # 列不存在也算缺失

            if qty_col_idx is not None and qty_col_idx < len(row):
                qty_val = str(row[qty_col_idx]).strip() if row[qty_col_idx] else ''
                if not qty_val:
                    is_incomplete = True
            elif qty_col_idx is None:
                is_incomplete = True  # 列不存在也算缺失

            if is_incomplete:
                incomplete_count += 1
                incomplete_rows.add(row_idx)
                # 标红整行
                for col_idx in range(self.preview_table.columnCount()):
                    cell_item = self.preview_table.item(row_idx, col_idx)
                    if cell_item:
                        cell_item.setBackground(QColor("#fee2e2"))  # 浅红色

        self.preview_table.resizeColumnsToContents()
        self.preview_table.setColumnWidth(0, 50)

        # [6.2.1] 补充商品模式下高亮显示
        if self.is_supplement_mode:
            self._highlight_temp_rows()

        suffix = "（已过滤重复商品）" if self.is_supplement_mode else ""
        # 2026-06-23：DB 还在加载时给出提示
        if self.is_supplement_mode and not self._db_models_loaded:
            suffix = "（客户-产品表加载中…）"

        # 2026-06-25：添加不完整行提示
        if incomplete_count > 0:
            incomplete_suffix = f"，{incomplete_count} 行不完整（红色标注）"
            if not self.keep_incomplete_checkbox.isChecked():
                incomplete_suffix += "（将被跳过）"
            else:
                incomplete_suffix += "（已勾选保留）"
        else:
            incomplete_suffix = ""

        self.preview_status_label.setText(f"共 {total} 行数据" + suffix + incomplete_suffix)
        self.preview_status_label.setStyleSheet("color: #dc2626; font-weight: bold;" if incomplete_count > 0 else "color: gray; font-style: italic;")
        self.import_btn.setEnabled(True)
        self.preview_btn.setEnabled(True)

    def _refilter_preview_after_db_load(self):
        """[6.2.23] DB 客户-产品表加载完成后重过滤预览表（仅在补充商品模式生效）"""
        if not self.is_supplement_mode or not hasattr(self, '_raw_preview_rows'):
            return
        try:
            headers = (self.preview_data or {}).get('headers') or []
            raw_rows = self._raw_preview_rows or []
            self._skipped_indices = self._compute_skipped_indices(raw_rows, headers)
            filtered = [r for i, r in enumerate(raw_rows) if i not in self._skipped_indices]
            self.preview_data['rows'] = filtered
            self.preview_data['total'] = len(filtered)

            # 重建表格
            self.preview_table.setRowCount(len(filtered))
            for row_idx, row in enumerate(filtered):
                self.preview_table.setItem(row_idx, 0, QTableWidgetItem(str(row_idx + 1)))
                for col_idx, value in enumerate(row):
                    self.preview_table.setItem(
                        row_idx, col_idx + 1,
                        QTableWidgetItem(str(value) if value else "")
                    )
            self._highlight_temp_rows()
            # 刷新统计（包括不完整行）
            self._refresh_preview_stats()
            print(
                f"[6.2.23] DB 加载完成后重过滤：{len(raw_rows)} → {len(filtered)} 行"
                f"（跳过 {len(self._skipped_indices)}）"
            )
        except Exception as e:
            print(f"[6.2.23] 重过滤失败: {e}")
    
    def _highlight_temp_rows(self):
        """[6.2.1] 高亮临时商品行（黄色背景）"""
        for row_idx in range(self.preview_table.rowCount()):
            # 检查第6列（状态列）是否包含"临时"
            status_item = self.preview_table.item(row_idx, 6)
            is_temp = status_item and '临时' in status_item.text()

            for col_idx in range(self.preview_table.columnCount()):
                cell_item = self.preview_table.item(row_idx, col_idx)
                if cell_item:
                    if is_temp:
                        cell_item.setBackground(QColor("#fef3c7"))
                    else:
                        # 清除背景（用透明色，QColor() 默认是黑色，会变全黑）
                        cell_item.setData(Qt.BackgroundRole, None)

    def _on_preview_context_menu(self, pos):
        """预览表格右键菜单：支持删除行"""
        row = self.preview_table.rowAt(pos.y())
        if row < 0:
            return

        menu = QMenu(self)
        delete_action = menu.addAction("🗑 删除此行")
        delete_action.triggered.connect(lambda: self._delete_preview_row(row))
        menu.exec(self.preview_table.viewport().mapToGlobal(pos))

    def _delete_preview_row(self, row: int):
        """删除预览表格指定行
        🔧 2026-06-29 修复：原实现只删除表格行，但 ImportWorker 重新读取原始 Excel 文件导入，
        导致删除的行仍然被导入。需要同步记录被删除的索引到 _deleted_preview_indices，
        供 start_import 时过滤原始 Excel。
        """
        # 在补充商品模式下，display_rows 是 _raw_preview_rows 的过滤结果（_skipped_indices），
        # preview_table 行号 = _raw_preview_rows 索引中未被 _skipped_indices 过滤的，
        # 所以需要反向映射回 _raw_preview_rows 索引
        raw_index = self._map_display_row_to_raw_index(row)
        if raw_index is not None:
            self._deleted_preview_indices.add(raw_index)
            print(f"[删除行] preview 行 {row} → _raw_preview_rows 索引 {raw_index}")

        self.preview_table.removeRow(row)
        # 重新编号（第一列）
        for r in range(self.preview_table.rowCount()):
            idx_item = self.preview_table.item(r, 0)
            if idx_item:
                idx_item.setText(str(r + 1))
        # 刷新预览统计
        self._refresh_preview_stats()
        # 重新高亮临时行
        self._highlight_temp_rows()

    def _map_display_row_to_raw_index(self, display_row: int) -> Optional[int]:
        """将 preview_table 的显示行号映射到 _raw_preview_rows 中的原始索引
        考虑补充商品模式下的 _skipped_indices 过滤。
        """
        if not hasattr(self, '_raw_preview_rows') or not self._raw_preview_rows:
            return None
        skipped = getattr(self, '_skipped_indices', set()) or set()
        display_to_raw = []
        for raw_idx in range(len(self._raw_preview_rows)):
            if raw_idx not in skipped and raw_idx not in self._deleted_preview_indices:
                display_to_raw.append(raw_idx)
        if 0 <= display_row < len(display_to_raw):
            return display_to_raw[display_row]
        return None
    
    def _on_keep_incomplete_toggled(self, checked: bool):
        """保留不完整行复选框状态变化"""
        self._refresh_preview_stats()
        self._highlight_temp_rows()

    def _refresh_preview_stats(self):
        """刷新预览统计信息"""
        total = self.preview_table.rowCount()
        # 统计临时行
        temp_count = sum(
            1 for r in range(total)
            if self.preview_table.item(r, 6) and '临时' in self.preview_table.item(r, 6).text()
        )
        # 统计不完整行（从原始数据）
        incomplete_count = self._count_incomplete_rows()

        # 文字
        text = f"共 {total} 行数据（{temp_count} 行临时商品）"
        if incomplete_count > 0:
            keep = self.keep_incomplete_checkbox.isChecked()
            if keep:
                text += f"，{incomplete_count} 行不完整（已勾选保留）"
                self.preview_status_label.setStyleSheet("color: #6b7280;")  # 普通色
            else:
                text += f"，{incomplete_count} 行不完整（将被跳过）"
                self.preview_status_label.setStyleSheet("color: #dc2626; font-weight: bold;")  # 红色
        else:
            self.preview_status_label.setStyleSheet("color: #6b7280;")

        self.preview_status_label.setText(text)

    def _count_incomplete_rows(self) -> int:
        """直接从原始预览数据计算不完整行数（不依赖表格行数和过滤后的 rows）"""
        if not hasattr(self, 'preview_data') or not self.preview_data:
            return 0
        # 优先使用保存的初始数据，避免勾选过滤后 rows 被修改导致计数变化
        rows = getattr(self, '_initial_preview_rows', None) or self.preview_data.get('rows') or []
        headers = self.preview_data.get('headers') or []
        if not headers or not rows:
            return 0
        model_col_idx = self._find_model_column(headers)
        qty_col_idx = self._find_qty_column(headers)
        if model_col_idx is None or qty_col_idx is None:
            return 0
        count = 0
        for row in rows:
            if self._is_row_incomplete(row, headers, model_col_idx, qty_col_idx):
                count += 1
        return count

    def _get_row_data(self, row_idx: int) -> list:
        """获取预览表格某行的数据（去除首列行号）"""
        result = []
        for col in range(1, self.preview_table.columnCount()):
            item = self.preview_table.item(row_idx, col)
            result.append(item.text() if item else '')
        return result

    def _is_preview_row_incomplete(self, row_data: list) -> bool:
        """根据当前 preview_data 的列索引判断预览行是否不完整"""
        model_col_idx = self._get_model_col_idx()
        qty_col_idx = self._get_qty_col_idx()
        if model_col_idx is None or qty_col_idx is None:
            return False
        # 列索引：表格第一列是行号，所以实际数据列 = 原始列索引 + 1
        model_col_in_table = model_col_idx + 1
        qty_col_in_table = qty_col_idx + 1
        if model_col_in_table >= len(row_data) and model_col_in_table < self.preview_table.columnCount():
            return True
        if qty_col_in_table >= len(row_data) and qty_col_in_table < self.preview_table.columnCount():
            return True
        # 表格中的列（含行号偏移）
        # 实际数据从 row_data[col-1] 读取
        m = (row_data[model_col_idx] or '').strip() if model_col_idx < len(row_data) else ''
        q = (row_data[qty_col_idx] or '').strip() if qty_col_idx < len(row_data) else ''
        return not m or not q

    def _get_model_col_idx(self) -> Optional[int]:
        """获取 Model 列在原始数据中的索引"""
        if not hasattr(self, 'preview_data') or not self.preview_data:
            return None
        headers = self.preview_data.get('headers') or []
        for i, h in enumerate(headers):
            if h and 'model' in str(h).lower():
                return i
        return None

    def _get_qty_col_idx(self) -> Optional[int]:
        """获取 Qty 列在原始数据中的索引"""
        if not hasattr(self, 'preview_data') or not self.preview_data:
            return None
        headers = self.preview_data.get('headers') or []
        for i, h in enumerate(headers):
            if h and 'qty' in str(h).lower():
                return i
        return None
    
    @Slot(str)
    def on_preview_error(self, error_msg: str):
        """预览加载失败"""
        self.preview_status_label.setText(f"加载失败: {error_msg}")
        QMessageBox.warning(self, "错误", f"加载预览失败: {error_msg}")
    
    @Slot()
    def start_auto_match(self):
        """开始自动匹配"""
        if not self.preview_data:
            QMessageBox.warning(self, "提示", "请先加载预览数据")
            return

        rows = self.preview_data.get('rows', [])
        if not rows:
            QMessageBox.warning(self, "提示", "没有数据可匹配")
            return

        self.auto_match_btn.setEnabled(False)
        self.match_summary_label.setText("匹配中...")

        # 2026-06-23：传 customer_id 给 MatchWorker
        self.match_worker = MatchWorker(
            self.api_client, rows, customer_id=self._get_current_customer_id()
        )
        self.match_worker.match_completed.connect(self.on_match_completed)
        self.match_worker.error.connect(self.on_match_error)
        self.match_worker.finished.connect(lambda: self.auto_match_btn.setEnabled(True))
        self.match_worker.start()

    @Slot(list)
    def on_match_completed(self, results: list):
        """匹配完成（2026-06-23 扩展：matched/created_temp/reused_existing 三色）"""
        self.match_results = results

        matched_count = 0
        created_temp_count = 0
        reused_count = 0

        for idx, result in enumerate(results):
            if idx >= self.match_table.rowCount():
                break

            status = result.get('status', 'unmatched')
            best_match = result.get('best_match')

            if status == "matched" and best_match:
                matched_count += 1
                match_type = best_match.get('match_type', '')
                match_score = best_match.get('match_score', 0)
                if match_type == 'exact_customer_code':
                    status_text = f"✓ 精确匹配 ({match_score}%)"
                    status_color = QColor('#dcfce7')  # 绿
                elif match_type == 'oe_number':
                    status_text = f"~ OE号匹配 ({match_score}%)"
                    status_color = QColor('#fef3c7')  # 黄
                else:
                    status_text = f"~ 名称匹配 ({match_score}%)"
                    status_color = QColor('#fef3c7')  # 黄
            elif status == "created_temp":
                created_temp_count += 1
                product = result.get('product') or {}
                status_text = f"+ 新建临时 (Model={product.get('customer_model', '-')})"
                status_color = QColor('#dbeafe')  # 蓝
            elif status == "reused_existing":
                reused_count += 1
                product = result.get('product') or {}
                mark = "(已转正)" if not product.get('is_temporary') else "(temp)"
                status_text = f"↻ 复用已有 {mark}"
                status_color = QColor('#e0f2fe')  # 青
            else:
                status_text = "✗ 未匹配"
                status_color = QColor('#fee2e2')  # 红

            status_item = QTableWidgetItem(status_text)
            status_item.setBackground(status_color)
            self.match_table.setItem(idx, 3, status_item)

            # 缓存 product_id（导入时直接用）
            if result.get('product_id'):
                if not hasattr(self, '_match_product_ids'):
                    self._match_product_ids = {}
                self._match_product_ids[idx] = result['product_id']

        # 摘要更新
        summary = (
            f"匹配完成：✓ {matched_count} 精确匹配，"
            f"+ {created_temp_count} 新建临时，"
            f"↻ {reused_count} 复用"
        )
        self.match_summary_label.setText(summary)

        # 启用导入按钮
        self.import_btn.setEnabled(True)

        QMessageBox.information(
            self,
            "匹配完成",
            f"匹配完成！\n精确匹配: {matched_count}\n新建临时: {created_temp_count}\n复用已有: {reused_count}",
        )
    
    @Slot(str)
    def on_match_error(self, error_msg: str):
        """匹配失败"""
        self.match_summary_label.setText("匹配失败")
        QMessageBox.warning(self, "错误", f"匹配失败: {error_msg}")
    
    @Slot(int)
    def filter_unmatched(self, state: int):
        """过滤未匹配项"""
        if not self.match_results:
            return
        
        show_only_unmatched = (state == Qt.Checked)
        
        for idx, result in enumerate(self.match_results):
            if idx >= self.match_table.rowCount():
                break
            
            best_match = result.get('best_match')
            is_unmatched = not best_match
            
            # 根据过滤状态显示/隐藏行
            self.match_table.setRowHidden(idx, show_only_unmatched and not is_unmatched)
    
    def create_product_for_unmatched(self):
        """为未匹配项创建产品"""
        if not self.match_results:
            return

        # 收集未匹配的行数据
        unmatched_rows = []
        rows = self.preview_data.get('rows', []) if self.preview_data else []
        for idx, result in enumerate(self.match_results):
            best_match = result.get('best_match')
            if not best_match and idx < len(rows):
                row = rows[idx]
                # 2026-06-23：补充商品模式下行是 list，其他模式可能为 dict；统一按 list 处理
                if isinstance(row, dict):
                    row = row.get('_raw') or []
                unmatched_rows.append({
                    'row_index': idx,
                    'customer_code': row[2] if len(row) > 2 else "",
                    'oe_number': row[3] if len(row) > 3 else "",
                    'product_name': row[4] if len(row) > 4 else "",
                })
        
        if not unmatched_rows:
            QMessageBox.information(self, "提示", "没有未匹配的数据需要创建产品")
            return
        
        # 打开产品创建对话框（使用产品管理模块的CustomerProductDialog）
        try:
            from widgets.customer_product_dialog import CustomerProductDialog
        except ImportError:
            import sys
            sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            from widgets.customer_product_dialog import CustomerProductDialog
        
        dialog = CustomerProductDialog(self.api_client)
        if dialog.exec() == QDialog.Accepted:
            # 产品创建成功，刷新匹配
            self.start_auto_match()
    
    @Slot()
    def start_import(self):
        """开始导入"""
        if not self.preview_data:
            QMessageBox.warning(self, "提示", "请先加载预览数据")
            return
        
        total_rows = self.preview_data.get('total', len(self.preview_data.get('rows', [])))
        
        if total_rows == 0:
            QMessageBox.warning(self, "提示", "没有可导入的商品（已全部存在）")
            return
        
        # [6.2.1] 补充商品模式使用不同的导入逻辑
        if self.is_supplement_mode:
            self._start_supplement_import()
            return
        
        # 2026-06-25：过滤不完整的行
        keep_incomplete = self.keep_incomplete_checkbox.isChecked()
        headers = self.preview_data.get('headers', [])
        model_col_idx = self._find_model_column(headers)
        qty_col_idx = self._find_qty_column(headers)
        rows = self.preview_data.get('rows', [])
        
        filtered_rows = []
        skipped_incomplete = 0
        for row in rows:
            if self._is_row_incomplete(row, headers, model_col_idx, qty_col_idx):
                skipped_incomplete += 1
                if keep_incomplete:
                    filtered_rows.append(row)
            else:
                filtered_rows.append(row)
        
        # 更新预览数据
        self.preview_data['rows'] = filtered_rows
        self.preview_data['total'] = len(filtered_rows)
        total_rows = len(filtered_rows)
        
        if total_rows == 0:
            QMessageBox.warning(self, "提示", f"没有可导入的商品（已过滤 {skipped_incomplete} 行不完整数据）")
            return
        
        # 原订单导入逻辑
        # 确认导入
        warning_msg = f"即将导入{total_rows}个商品"
        if skipped_incomplete > 0 and not keep_incomplete:
            warning_msg += f"\n（将跳过 {skipped_incomplete} 行 Model/Qty 缺失的数据）"
        elif skipped_incomplete > 0:
            warning_msg += f"\n（包含 {skipped_incomplete} 行不完整数据）"
        
        reply = QMessageBox.question(
            self,
            "确认导入",
            warning_msg + "\n是否继续？",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply != QMessageBox.Yes:
            return

        self.import_btn.setEnabled(False)

        # 使用工作线程执行导入
        file_path = self.file_path_label.text()
        customer_id = self.customer_combo.currentData()

        if not customer_id:
            QMessageBox.warning(self, "提示", "请先选择客户")
            self.import_btn.setEnabled(True)
            return

        # 🔧 2026-06-29 修复：如果用户右键删除了某些预览行，
        # 需要在导入前生成一个排除这些行的临时 Excel 文件，
        # 因为后端 /orders/import 接口不支持跳过行参数。
        if self._deleted_preview_indices:
            try:
                temp_file_path = self._build_filtered_excel_for_import(file_path)
            except Exception as e:
                QMessageBox.critical(self, "错误", f"生成临时文件失败: {str(e)}")
                self.import_btn.setEnabled(True)
                return
            # 清理上次的临时文件
            if hasattr(self, '_last_temp_file') and self._last_temp_file and os.path.exists(self._last_temp_file):
                try:
                    os.unlink(self._last_temp_file)
                except OSError:
                    pass
            self._last_temp_file = temp_file_path
            file_path_to_import = temp_file_path
        else:
            file_path_to_import = file_path

        self.import_worker = ImportWorker(self.api_client, file_path_to_import, customer_id)
        self.import_worker.import_completed.connect(self.on_import_completed)
        self.import_worker.error.connect(self.on_import_error)
        self.import_worker.finished.connect(lambda: self.import_btn.setEnabled(True))
        self.import_worker.start()

    def _build_filtered_excel_for_import(self, src_file_path: str) -> str:
        """根据 _deleted_preview_indices 生成过滤后的临时 Excel 文件
        🔧 2026-06-29 新增：用于支持用户删除预览行后正确导入。

        Args:
            src_file_path: 原始 Excel 文件路径

        Returns:
            临时文件路径（调用方负责清理）
        """
        # 原始索引 → Excel 行号（1-based，包含表头）
        # _raw_preview_rows 的索引 0 对应 Excel 第 2 行（第 1 行是表头）
        excel_rows_to_skip = {idx + 2 for idx in self._deleted_preview_indices}

        wb = load_workbook(src_file_path)
        ws = wb.active

        # 从后往前删除，避免索引偏移
        for row_num in sorted(excel_rows_to_skip, reverse=True):
            if row_num <= ws.max_row:
                ws.delete_rows(row_num, 1)

        # 写入临时文件
        fd, temp_path = tempfile.mkstemp(suffix='.xlsx', prefix='pi_import_filtered_')
        os.close(fd)
        wb.save(temp_path)
        wb.close()
        print(f"[过滤导入] 原始文件={src_file_path}, 跳过 {len(excel_rows_to_skip)} 行 → {temp_path}")
        return temp_path
    
    def _start_supplement_import(self):
        """[6.2.1] 补充商品导入"""
        total_rows = self.preview_data.get('total', len(self.preview_data.get('rows', [])))
        
        reply = QMessageBox.question(
            self,
            "确认补充",
            f"即将补充 {total_rows} 个商品到当前订单。\n这些商品将标记为临时商品。\n是否继续？",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply != QMessageBox.Yes:
            return
        
        self.import_btn.setEnabled(False)
        
        try:
            # 获取预览数据（row 可能是 list[Excel] 或 dict[手动添加]）
            rows = self.preview_data.get('rows', [])
            # 2026-06-23：兼容两种行格式
            #  - 列表：Excel 原始行（list[cell_value]）
            #  - 字典：手动添加的产品（含 customer_code/oe_number/detail_desc/...）
            # 统一打包为 dict 列表发送
            payload_items = []
            for row in rows:
                if isinstance(row, dict):
                    payload_items.append(row)
                else:
                    payload_items.append({'raw': list(row), 'is_temp': True})

            # 调用后端 API 追加商品到订单
            response = self.api_client.post(
                f"/orders/{self.order_id}/supplement-items",
                json={'items': payload_items, 'is_temp': True}
            )
            
            if response and response.get('success'):
                # 发送信号通知父窗口刷新
                self.import_completed.emit(True, len(rows), 0, [])
                QMessageBox.information(
                    self,
                    "补充完成",
                    f"成功补充 {len(rows)} 个商品到订单"
                )
                self.close()
            else:
                QMessageBox.warning(
                    self,
                    "补充失败",
                    response.get('message', '未知错误')
                )
                
        except Exception as e:
            QMessageBox.warning(self, "错误", f"补充商品失败: {str(e)}")
        finally:
            self.import_btn.setEnabled(True)
    
    @Slot(bool, int, int, list)
    def on_import_completed(self, success: bool, success_count: int, failed_count: int, errors: list):
        """导入完成"""
        if success:
            QMessageBox.information(
                self,
                "导入完成",
                f"导入完成！\n成功: {success_count} 个产品\n失败: {failed_count} 个产品"
            )
            # 点击确定后自动关闭对话框，父窗口将刷新订单列表
            self.accept()
        else:
            reply = QMessageBox.warning(
                self,
                "导入完成（部分失败）",
                f"导入完成，但有部分失败。\n成功: {success_count} 个产品\n失败: {failed_count} 个产品\n\n点击确定查看错误详情。",
                QMessageBox.Ok | QMessageBox.Cancel
            )
            if reply == QMessageBox.Ok:
                # 显示错误详情
                if errors:
                    error_dialog = ErrorDetailDialog(errors, self)
                    error_dialog.exec()
                # 关闭对话框
                self.accept()
            else:
                return

        # 显示错误详情
        if errors and not success:
            error_dialog = ErrorDetailDialog(errors, self)
            error_dialog.exec()
        
        # [6.0.2] [6.0.2.3] 导入成功后，写入临时PI号到每个已导入订单
        # 注意：is_temp_pi 的实际写入依赖后端 API 支持（任务4/5完成后）
        # 此处仅记录日志，验证 _current_temp_pi_no 是否正确传递
        if success and self._current_temp_pi_no:
            temp_pi_base = self._current_temp_pi_no.rstrip('?')
            print(f"[6.0.2] 导入成功，临时PI基数: {temp_pi_base}，成功 {success_count} 条，is_temp_pi=True")
            # TODO [6.0.2.3]: 后端支持后，遍历成功导入的订单ID，调用 api_client.update_order_temp_pi() 写入正式 temp PI

        # 🔧 2026-06-29：导入完成后清理临时文件
        self._cleanup_temp_file()
    
    @Slot(str)
    def on_import_error(self, error_msg: str):
        """导入失败"""
        # 🔧 2026-06-29：导入失败时也清理临时文件
        self._cleanup_temp_file()
        QMessageBox.warning(self, "错误", f"导入失败: {error_msg}")

    def _cleanup_temp_file(self):
        """清理 _last_temp_file 临时文件（如果存在）"""
        if hasattr(self, '_last_temp_file') and self._last_temp_file:
            try:
                if os.path.exists(self._last_temp_file):
                    os.unlink(self._last_temp_file)
                    print(f"[清理临时文件] 已删除: {self._last_temp_file}")
            except OSError as e:
                print(f"[清理临时文件] 失败: {e}")
            finally:
                self._last_temp_file = None
    
    def _load_supplement_order_data(self):
        """[6.2.1] 加载当前订单数据用于去重"""
        if not self.order_id or not self.parent():
            return
        
        try:
            # 从父窗口获取当前订单数据
            parent_widget = self.parent()
            current_order = None
            
            # 尝试从缓存获取订单数据
            if hasattr(parent_widget, '_current_order_detail'):
                current_order = parent_widget._current_order_detail
            elif hasattr(parent_widget, '_order_summary_cache'):
                # 从缓存中查找当前订单
                cache = parent_widget._order_summary_cache
                orders = cache.get('pi_list', [])
                for order in orders:
                    if order.get('id') == self.order_id or order.get('pi_id') == self.order_id:
                        current_order = order
                        break
            
            if current_order:
                self._supplement_order_data = current_order
                existing_models = {item.get('model') for item in current_order.get('items', []) if item.get('model')}
                print(f"[6.2.1] 加载订单数据完成，现有商品 {len(existing_models)} 个")

                # [问题 #25] 自动获取当前订单客户，预加载客户列表并选中
                order_customer_id = current_order.get('customer_id')
                if order_customer_id:
                    self._target_customer_id = order_customer_id
                    # 先加载客户列表，加载完成后自动选中对应客户
                    QTimer.singleShot(100, lambda: self._select_customer_on_load(order_customer_id))
            else:
                print(f"[6.2.1] 未找到订单数据，order_id={self.order_id}")
                self._supplement_order_data = {'items': []}
                
        except Exception as e:
            print(f"[6.2.1] 加载订单数据失败: {e}")
            self._supplement_order_data = {'items': []}
    
    def _compute_skipped_indices(self, preview_rows: list, headers: list = None) -> set:
        """[6.2.1/6.2.23] 计算应跳过的原始行索引集合

        去重源（任一命中即跳过）：
            1. 当前订单已存在的 items（_supplement_order_data.items[*].model）
            2. 客户-产品表中已存在的产品（_db_existing_models）

        Args:
            preview_rows: Excel 预览原始行（list[list]）
            headers: Excel 表头（list[str]），用于按列名取 Model 字段值

        Returns:
            set[int] 应跳过的原始行索引集合
        """
        if not self._supplement_order_data:
            return set()

        # 1. 当前订单内已有的 Model
        order_models = {
            (item.get('model') or '').strip()
            for item in self._supplement_order_data.get('items', [])
            if (item.get('model') or '').strip()
        }

        # 2. 客户-产品表已存在的 Model（2026-06-23）
        db_models = {
            m.strip() for m in (self._db_existing_models or {}).keys() if m and m.strip()
        }

        # 合并去重集
        existing_models = order_models | db_models
        print(
            f"[6.2.23] 去重源：订单内 {len(order_models)} 条 + 客户-产品表 {len(db_models)} 条"
        )

        # 找到 Model 列的索引
        model_col_idx = self._find_model_column(headers)

        skipped = set()
        skipped_in_order = 0
        skipped_in_db = 0
        for idx, row in enumerate(preview_rows):
            raw_model = ''
            if model_col_idx is not None and model_col_idx < len(row):
                raw_model = (str(row[model_col_idx]) if row[model_col_idx] else '').strip()
            if raw_model and raw_model in existing_models:
                skipped.add(idx)
                if raw_model in order_models:
                    skipped_in_order += 1
                else:
                    skipped_in_db += 1

        if skipped:
            print(
                f"[6.2.23] 补充商品去重完成：跳过 {len(skipped)} 条 "
                f"（订单内 {skipped_in_order} + 客户-产品表 {skipped_in_db}）"
            )
        else:
            print(f"[6.2.23] 补充商品去重完成：无重复，全部进入")
        return skipped

    def _filter_existing_products(self, preview_rows: list, headers: list = None) -> list:
        """[6.2.1/6.2.23] 兼容方法：返回过滤后的行列表（不推荐直接使用，建议用 _compute_skipped_indices）"""
        if not self.is_supplement_mode:
            return preview_rows
        skipped = self._compute_skipped_indices(preview_rows, headers)
        return [r for i, r in enumerate(preview_rows) if i not in skipped]

    def _find_model_column(self, headers) -> Optional[int]:
        """[6.2.23] 在 Excel 表头里定位 Model 列的索引"""
        if not headers:
            return None
        # 优先匹配这些列名（区分优先级）
        candidates = [
            '客户型号(Model)',  # 2026-06-23 新增
            '客户型号',         # 中性别名
            'customer_model',
            'Model',
            'model',
            '客户产品编号',
            'customer_product_no',
        ]
        for cand in candidates:
            for idx, h in enumerate(headers):
                if h and str(h).strip() == cand:
                    return idx
        # 退化匹配：包含「model」字样（不区分大小写）
        for idx, h in enumerate(headers):
            if h and 'model' in str(h).strip().lower():
                return idx
        return None

    def _find_qty_column(self, headers) -> Optional[int]:
        """[6.2.25] 在 Excel 表头里定位 Qty/数量 列的索引"""
        if not headers:
            return None
        candidates = [
            'Qty',
            'qty',
            '数量',
            'quantity',
            'Quantity',
            'num',
            'Num',
        ]
        for cand in candidates:
            for idx, h in enumerate(headers):
                if h and str(h).strip() == cand:
                    return idx
        # 退化匹配：包含「qty」或「数量」字样
        for idx, h in enumerate(headers):
            h_str = str(h).strip().lower()
            if 'qty' in h_str or '数量' in h_str:
                return idx
        return None

    def _is_row_incomplete(self, row, headers, model_col_idx, qty_col_idx) -> bool:
        """[6.2.25] 检查行是否不完整（Model 或 Qty 缺失）"""
        # 检查 Model
        if model_col_idx is not None and model_col_idx < len(row):
            model_val = str(row[model_col_idx]).strip() if row[model_col_idx] else ''
            if not model_val:
                return True
        else:
            return True  # 列不存在也算缺失

        # 检查 Qty
        if qty_col_idx is not None and qty_col_idx < len(row):
            qty_val = str(row[qty_col_idx]).strip() if row[qty_col_idx] else ''
            if not qty_val:
                return True
        elif qty_col_idx is None:
            return True  # 列不存在也算缺失

        return False

    def _select_customer_on_load(self, customer_id):
        """[问题 #25] 客户列表加载完成后自动选中目标客户"""
        if not hasattr(self, '_target_customer_id'):
            return
        
        # 遍历下拉框查找匹配的客户 ID
        for i in range(self.customer_combo.count()):
            cid = self.customer_combo.itemData(i)
            if cid == customer_id:
                self.customer_combo.setCurrentIndex(i)
                print(f"[问题 #25] 自动选中客户 ID={customer_id}, 索引={i}")
                # 清理标记
                try:
                    delattr(self, '_target_customer_id')
                except:
                    pass
                return
        
        print(f"[问题 #25] 未找到客户 ID={customer_id}，当前列表项: {[self.customer_combo.itemData(i) for i in range(self.customer_combo.count())]}")


class PreviewWorker(QThread):
    """预览加载工作线程"""
    preview_ready = Signal(list, list, int, int)
    error = Signal(str)
    
    def __init__(self, api_client, file_path):
        super().__init__()
        self.api_client = api_client
        self.file_path = file_path
    
    def run(self):
        """执行预览加载"""
        try:
            with open(self.file_path, 'rb') as f:
                files = {'file': (self.file_path.split('/')[-1], f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                response = self.api_client.post("/orders/preview", files=files)
            
            if response and response.get('success'):
                self.preview_ready.emit(
                    response.get('headers', []),
                    response.get('preview_rows', []),
                    response.get('total_rows', 0),
                    response.get('column_count', 0)
                )
            else:
                self.error.emit(response.get('error', '未知错误') if response else '无响应')
        except Exception as e:
            self.error.emit(str(e))


class MatchWorker(QThread):
    """自动匹配工作线程（2026-06-23 扩展：支持 auto_create_temporary）"""
    match_completed = Signal(list)
    error = Signal(str)

    def __init__(self, api_client, rows: list, customer_id: int = None):
        super().__init__()
        self.api_client = api_client
        self.rows = rows
        self.customer_id = customer_id

    def run(self):
        """执行自动匹配 + 静默创建 temp"""
        try:
            # 准备匹配项 - 只发送必要的字段
            items = []
            for idx, row in enumerate(self.rows):
                # 假设列顺序：客户产品编号(2), OE号(3), 产品描述(4)
                # 2026-06-23：增加 customer_model 列（Excel 第 8 列）
                customer_code = row[2] if len(row) > 2 else None
                oe_number = row[3] if len(row) > 3 else None
                product_name = row[4] if len(row) > 4 else None
                customer_model = row[7] if len(row) > 7 else None  # 2026-06-23 新增

                # 清理可能的货币符号
                if customer_code and isinstance(customer_code, str):
                    customer_code = customer_code.strip()
                if oe_number and isinstance(oe_number, str):
                    oe_number = oe_number.strip()
                if product_name and isinstance(product_name, str):
                    product_name = product_name.strip()
                if customer_model and isinstance(customer_model, str):
                    customer_model = customer_model.strip()

                item = {
                    'customer_code': customer_code or None,
                    'oe_number': oe_number or None,
                    'product_name': product_name or None,
                    'customer_model': customer_model or None,  # 2026-06-23 新增
                }
                items.append(item)

            # 2026-06-23：发送 auto_create_temporary + customer_id
            payload = {
                'items': items,
                'auto_create_temporary': True,
                'customer_id': self.customer_id,
            }
            response = self.api_client.post("/products/batch-match", data=payload)

            if response and response.get('success'):
                self.match_completed.emit(response.get('results', []))
            else:
                self.error.emit(response.get('error', '未知错误') if response else '无响应')
        except Exception as e:
            self.error.emit(str(e))


class ImportWorker(QThread):
    """导入工作线程"""
    import_completed = Signal(bool, int, int, list)
    error = Signal(str)
    
    def __init__(self, api_client, file_path: str, customer_id: int = None):
        super().__init__()
        self.api_client = api_client
        self.file_path = file_path
        self.customer_id = customer_id
    
    def run(self):
        """执行导入"""
        try:
            endpoint = "/orders/import"
            if self.customer_id:
                endpoint = f"/orders/import?customer_id={self.customer_id}"
            
            with open(self.file_path, 'rb') as f:
                files = {'file': (self.file_path.split('/')[-1], f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                response = self.api_client.post(endpoint, files=files)
            
            if response:
                self.import_completed.emit(
                    response.get('success', False),
                    response.get('success_count', 0),
                    response.get('failed_count', 0),
                    response.get('errors', [])
                )
            else:
                self.error.emit('无响应')
        except Exception as e:
            self.error.emit(str(e))


class ErrorDetailDialog(QDialog):
    """错误详情对话框"""
    
    def __init__(self, errors: list, parent=None):
        super().__init__(parent)
        self.errors = errors
        self.setWindowTitle("导入错误详情")
        self.setMinimumSize(600, 400)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # 错误统计
        stats_label = QLabel(f"共有 {len(self.errors)} 条错误")
        layout.addWidget(stats_label)
        
        # 错误表格
        self.error_table = QTableWidget()
        self.error_table.setColumnCount(3)
        self.error_table.setHorizontalHeaderLabels(["行号", "错误信息", "建议"])
        self.error_table.setRowCount(len(self.errors))
        
        for idx, error in enumerate(self.errors):
            self.error_table.setItem(idx, 0, QTableWidgetItem(str(error.get('row', ''))))
            self.error_table.setItem(idx, 1, QTableWidgetItem(error.get('error', '')))
            
            suggestions = error.get('suggestions', [])
            self.error_table.setItem(idx, 2, QTableWidgetItem(', '.join(suggestions)))
        
        layout.addWidget(self.error_table)
        
        # 关闭按钮
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.close)
        layout.addWidget(close_btn)
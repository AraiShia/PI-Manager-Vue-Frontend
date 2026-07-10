from fastapi import APIRouter, Depends, HTTPException
from typing import List
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.database import get_db
from crud.pi import (
    create_pi_invoice, get_pi_invoice, get_pi_invoices, update_pi_status, get_price_history,
    update_pi_invoice, get_pi_invoice_detail, get_pi_invoices_with_customer,
    get_pi_item, update_pi_item,
    delete_pi_item, inbound_pi_item, inbound_pi_items_batch,  # йңҖжұӮ#40
    delete_pi_invoice,  # 2026-06-15 дҝ®еӨҚпјҡд№ӢеүҚжјҸеҜјпјҢи°ғз”Ёж—¶ NameError
    get_pi_versions, save_pi_snapshot,  # йңҖжұӮ#42
    save_formal_record, load_formal_record, formal_record_exists,  # йңҖжұӮ#42
)
from schemas.pi import PIInvoiceCreate, PIInvoiceResponse, PIInvoiceUpdate
from schemas.pi_detail import PIInvoiceDetailFullResponse
from pydantic import BaseModel
from typing import Optional

router = APIRouter(prefix="/api/pi", tags=["PIз®ЎзҗҶ"])

@router.post("/")
def create_pi_api(pi: PIInvoiceCreate, db: Session = Depends(get_db)):
    try:
        result = create_pi_invoice(db, pi)
        return {
            "id": result.id,
            "dept_id": result.dept_id,
            "pi_no": result.pi_no,
            "customer_id": result.customer_id,
            "total_amount": float(result.total_amount) if result.total_amount else 0,
            "currency": result.currency or "USD",
            "status": result.status or 1,
            "created_at": result.created_at.isoformat() if result.created_at else None
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/", response_model=list[PIInvoiceResponse])
def read_pi_list(skip: int = 0, limit: int = 100, status: int = None, db: Session = Depends(get_db)):
    from models import InvInventory, PiProformaInvoiceItem
    from models import PiPaymentStage as _PS
    from models import PrdCustomerProduct
    results = get_pi_invoices_with_customer(db, skip=skip, limit=limit, status=status)
    # 2026-06-10: дҝ®еӨҚ T5.3 + жҖ»иЎЁеӣһеЎ« вҖ” еҠ дёҠж—Ҙжңҹ/дә§е“Ғж•°/е·Ід»ҳйҮ‘йўқ/еә“еӯҳзӯүеӯ—ж®ө
    for r in results:
        pi_id = r.get("id")
        if not pi_id:
            continue
        # 1) и®ўеҚ•ж—ҘжңҹпјҡеҸ– PI дё»еҚ•зҡ„ created_at
        if not r.get("order_date") and r.get("created_at"):
            r["order_date"] = r["created_at"][:10] if isinstance(r["created_at"], str) else r["created_at"].isoformat()[:10]
        # 2) дә§е“Ғж•°
        item_count = db.query(PiProformaInvoiceItem).filter(
            PiProformaInvoiceItem.pi_id == pi_id
        ).count()
        r["product_count"] = item_count
        # 3) е·Ід»ҳж¬ҫйҮ‘йўқпјҲ2026-06-16 PI йҖҡи·Ҝдҝ®еӨҚпјҡдёҺ update_pi_payment_status дёҖиҮҙпјү
        #    дјҳе…Ҳд»Ҙ ArCustomerPayment ж°ҙеҚ•дёәеҮҶпјҢж— ж°ҙеҚ•ж—¶еӣһйҖҖеҲ°еҲҶжңҹиЎЁе®һд»ҳеҗҲи®Ў
        from models import ArCustomerPayment
        payments_for_pi = db.query(ArCustomerPayment).filter(ArCustomerPayment.pi_id == pi_id).all()
        if payments_for_pi:
            paid_amount_value = float(sum(float(p.actual_amount or p.amount or 0) for p in payments_for_pi))
        else:
            paid_rows = db.query(_PS).filter(
                _PS.pi_id == pi_id,
                _PS.status == 2
            ).all()
            paid_amount_value = float(sum(float(s.amount or 0) for s in paid_rows))
        r["paid_amount"] = paid_amount_value
        # 4) еә“еӯҳеӯ—ж®өпјҲж•°еҖјз»ҹи®Ўпјү
        inv_list = db.query(InvInventory).filter(InvInventory.pi_id == pi_id).all()
        total_qty = sum(float(i.total_quantity or 0) for i in inv_list)
        pending_qty = sum(float(i.pending_quantity or 0) for i in inv_list)
        r["has_inventory"] = total_qty > 0
        r["inventory_quantity"] = total_qty
        r["inventory_pending"] = pending_qty
        r["inventory_count"] = len(inv_list)

        # 4.5) storage_status зҠ¶жҖҒжһҡдёҫпјҲ2026-06-23 ж”¶ж•ӣпјҡдёҺ /pi/detail/{id} еҗҢжәҗпјү
        # жңҹжңӣжҖ»ж•°йҮҸ = жүҖжңү PI item зҡ„ quantity д№Ӣе’Ңпјӣinv_inventory.total_quantity иҒҡеҗҲеҗҺжҜ”иҫғгҖӮ
        from crud.storage_status import StorageStatus
        expected_qty_total = float(
            db.query(func.coalesce(func.sum(PiProformaInvoiceItem.quantity), 0))
            .filter(PiProformaInvoiceItem.pi_id == pi_id).scalar() or 0
        )
        r["storage_status"] = StorageStatus.from_order_inventory(
            db, pi_id=pi_id, expected_total=expected_qty_total
        )
        # 5) 2026-06-11 Phase 7.6: еҮәиҙ§еӣһеЎ«
        from models import ShShipment
        sh_list = db.query(ShShipment).filter(ShShipment.pi_id == pi_id).all()
        r["shipment_count"] = len(sh_list)
        sh_qty = 0.0
        latest_date = None
        for sh in sh_list:
            for shi in (sh.items or []):
                sh_qty += float(shi.quantity or 0)
            if sh.created_at:
                ds = sh.created_at.isoformat()[:10]
                if latest_date is None or ds > latest_date:
                    latest_date = ds
        r["shipped_quantity"] = sh_qty
        r["latest_shipment_date"] = latest_date
        # 7) 2026-06-11 з¬¬ 38 жқЎ: жҙҫз”ҹжңӘд»ҳж¬ҫ/иҝӣеәҰ/зҠ¶жҖҒ
        total = float(r.get("total_amount") or 0)
        paid = float(r.get("paid_amount") or 0)
        unpaid = max(total - paid, 0.0)
        progress = (paid / total * 100.0) if total > 0 else 0.0
        if total <= 0:
            status_text = "жңӘд»ҳж¬ҫ"
        elif progress >= 99.9:
            status_text = "е·Із»“жё…"
        elif progress > 0:
            status_text = "йғЁеҲҶд»ҳж¬ҫ"
        else:
            status_text = "жңӘд»ҳж¬ҫ"
        r["unpaid_amount"] = round(unpaid, 2)
        r["payment_progress"] = round(progress, 1)
        r["payment_status"] = status_text
    print(f"DEBUG - read_pi_list returned {len(results)} items")
    return results

@router.get("/detail/{pi_id}", response_model=PIInvoiceDetailFullResponse)
def read_pi_detail(pi_id: int, db: Session = Depends(get_db)):
    detail = get_pi_invoice_detail(db, pi_id)
    if detail is None:
        raise HTTPException(status_code=404, detail="PIеҚ•дёҚеӯҳеңЁ")
    return detail


# 2026-06-10: дҝ®еӨҚ T6.2 вҖ” д»ҳж¬ҫеҲҶжңҹз«ҜзӮ№
@router.post("/{pi_id}/payment-stages")
def create_payment_stage(pi_id: int, payload: dict, db: Session = Depends(get_db)):
    from models import PiPaymentStage
    pi = get_pi_invoice(db, pi_id)
    if not pi:
        raise HTTPException(status_code=404, detail="PIеҚ•дёҚеӯҳеңЁ")
    stage = PiPaymentStage(
        pi_id=pi_id,
        stage_type=payload.get("stage_type") or payload.get("stage_name") or "deposit",
        stage_no=payload.get("stage_no", 1),
        amount=payload.get("amount", 0),
        percentage=payload.get("percentage"),
        payment_method=payload.get("payment_method"),
        status=1,
    )
    db.add(stage)
    db.commit()
    db.refresh(stage)
    return {
        "id": stage.id,
        "pi_id": stage.pi_id,
        "stage_type": stage.stage_type,
        "stage_no": stage.stage_no,
        "amount": float(stage.amount) if stage.amount else 0,
        "percentage": float(stage.percentage) if stage.percentage else None,
        "status": stage.status,
    }


@router.put("/{pi_id}/payment-stages/{stage_id}")
def update_payment_stage(pi_id: int, stage_id: int, payload: dict, db: Session = Depends(get_db)):
    from models import PiPaymentStage
    stage = db.query(PiPaymentStage).filter(
        PiPaymentStage.id == stage_id,
        PiPaymentStage.pi_id == pi_id
    ).first()
    if not stage:
        raise HTTPException(status_code=404, detail="еҲҶжңҹдёҚеӯҳеңЁ")
    for k in ("stage_type", "amount", "percentage", "status", "paid_date"):
        if k in payload:
            setattr(stage, k, payload[k])
    db.commit()
    db.refresh(stage)
    return {"id": stage.id, "status": stage.status}


@router.delete("/{pi_id}/payment-stages/{stage_id}")
def delete_payment_stage(pi_id: int, stage_id: int, db: Session = Depends(get_db)):
    from models import PiPaymentStage
    stage = db.query(PiPaymentStage).filter(
        PiPaymentStage.id == stage_id,
        PiPaymentStage.pi_id == pi_id
    ).first()
    if not stage:
        raise HTTPException(status_code=404, detail="еҲҶжңҹдёҚеӯҳеңЁ")
    db.delete(stage)
    db.commit()
    return {"deleted": stage_id}


# 2026-06-11 з¬¬ 38 жқЎ: дёҖж¬ЎжҖ§ж·»еҠ д»ҳж¬ҫ(еҲӣе»ә status=2 зҡ„еҲҶжңҹ)
@router.post("/{pi_id}/payments")
def add_payment(pi_id: int, payload: dict, db: Session = Depends(get_db)):
    """
    дёҖж¬ЎжҖ§ж·»еҠ д»ҳж¬ҫ(еҝ«жҚ·е…ҘеҸЈ):
    иҜ·жұӮ: { amount, payment_method?, remark? }
    - amount еҝ…йЎ» > 0
    - иҮӘеҠЁеҸ–дёӢдёҖдёӘ stage_no
    - еҲӣе»ә PiPaymentStage(status=2, paid_date=today)
    - иҝ”еӣһж–°е»әзҡ„еҲҶжңҹ + зҙҜи®Ў paid_amount / unpaid_amount
    """
    from models import PiPaymentStage, PiProformaInvoice, ArCustomerPayment
    from datetime import date, datetime

    pi = db.query(PiProformaInvoice).filter(PiProformaInvoice.id == pi_id).first()
    if not pi:
        raise HTTPException(status_code=404, detail="PIеҚ•дёҚеӯҳеңЁ")

    amount = float(payload.get("amount") or 0)
    if amount <= 0:
        raise HTTPException(status_code=400, detail="amount еҝ…йЎ»еӨ§дәҺ 0")

    # дёӢдёҖдёӘ stage_no
    next_no = (db.query(PiPaymentStage)
                .filter(PiPaymentStage.pi_id == pi_id)
                .count()) + 1

    stage = PiPaymentStage(
        pi_id=pi_id,
        stage_type=payload.get("stage_type") or "deposit",
        stage_no=next_no,
        amount=amount,
        status=2,  # 2=е·Ід»ҳ
        paid_date=date.today(),
    )
    db.add(stage)

    from utils.number_generator import NumberGenerator
    payment_date_value = payload.get("paid_date") or payload.get("payment_date")
    if payment_date_value:
        payment_date = datetime.fromisoformat(str(payment_date_value).replace("Z", "+00:00"))
    else:
        payment_date = datetime.utcnow()

    handling_fee = float(payload.get("handling_fee") or 0)
    actual_amount = max(amount - handling_fee, 0)

    customer_payment = ArCustomerPayment(
        dept_id=pi.dept_id,
        receipt_no=NumberGenerator.generate_receipt_no(db, pi.dept_id),
        pi_id=pi_id,
        customer_id=pi.customer_id,
        amount=amount,
        handling_fee=handling_fee,
        actual_amount=actual_amount,
        is_fully_paid=False,
        payment_date=payment_date,
        payment_method=payload.get("payment_method") or "",
        remark=payload.get("remark") or "",
    )
    db.add(customer_payment)

    # жӣҙж–° PI зҡ„ updated_at
    pi.updated_at = __import__("datetime").datetime.utcnow()
    db.commit()
    db.refresh(stage)
    db.refresh(customer_payment)

    # 2026-06-16 PI йҖҡи·Ҝдҝ®еӨҚпјҡеҶҷеҲҶжңҹеҗҺеҗҢжӯҘ PI дё»иЎЁ status
    from crud.payment import update_pi_payment_status
    update_pi_payment_status(db, pi_id)

    # и®Ўз®—зҙҜи®ЎпјҲ2026-06-16 PI йҖҡи·Ҝдҝ®еӨҚпјҡдёҺ update_pi_payment_status дёҖиҮҙпјү
    total = float(pi.total_amount or 0)
    payments_for_pi = db.query(ArCustomerPayment).filter(ArCustomerPayment.pi_id == pi_id).all()
    if payments_for_pi:
        paid_amount = sum(float(p.actual_amount or p.amount or 0) for p in payments_for_pi)
    else:
        paid_rows = db.query(PiPaymentStage).filter(
            PiPaymentStage.pi_id == pi_id,
            PiPaymentStage.status == 2
        ).all()
        paid_amount = sum(float(s.amount or 0) for s in paid_rows)
    unpaid_amount = max(total - paid_amount, 0.0)
    progress = (paid_amount / total * 100.0) if total > 0 else 0.0

    return {
        "stage": {
            "id": stage.id,
            "pi_id": stage.pi_id,
            "stage_no": stage.stage_no,
            "stage_type": stage.stage_type,
            "amount": float(stage.amount) if stage.amount else 0,
            "status": stage.status,
            "paid_date": stage.paid_date.isoformat() if stage.paid_date else None,
        },
        "receipt": {
            "id": customer_payment.id,
            "receipt_no": customer_payment.receipt_no,
            "actual_amount": float(customer_payment.actual_amount or 0),
            "handling_fee": float(customer_payment.handling_fee or 0),
            "payment_method": customer_payment.payment_method,
            "payment_date": customer_payment.payment_date.isoformat() if customer_payment.payment_date else None,
        },
        "pi_id": pi_id,
        "total_amount": total,
        "paid_amount": round(paid_amount, 2),
        "unpaid_amount": round(unpaid_amount, 2),
        "payment_progress": round(progress, 1),
    }


# жіЁж„Ҹпјҡ/pi/{pi_id} и·Ҝз”ұеңЁдёӢж–№ line 440 йҮҚж–°е®ҡд№үпјҲдҝ®еӨҚеҗҺзүҲжң¬пјү
# 2026-06-22 дҝ®еӨҚпјҡеҺҹ line 274 и·Ҝз”ұпјҲе·ІеҲ йҷӨпјҢеӣ дёәдёҺ line 440 йҮҚеӨҚдјҡеҜјиҮҙPythonеҮҪж•°йҮҚе®ҡд№үй”ҷиҜҜпјү
# й”ҷиҜҜзүҲжң¬зӣҙжҺҘиҝ”еӣһ db_piпјҢдёҚеҢ…еҗ« items
# дҝ®еӨҚзүҲжң¬дҪҝз”Ё get_pi_invoice_detail иҝ”еӣһеҢ…еҗ« items зҡ„е®Ңж•ҙж•°жҚ®

@router.delete("/{pi_id}")
def delete_pi_api(pi_id: int, db: Session = Depends(get_db)):
    try:
        delete_pi_invoice(db, pi_id)
        return {"success": True, "message": "PIе·ІеҲ йҷӨ"}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.post("/batch-delete")
def batch_delete_pi_api(pi_ids: List[int], db: Session = Depends(get_db)):
    """жү№йҮҸеҲ йҷӨPIи®ўеҚ•

    2026-06-15 ејәеҢ–пјҡжҚ•иҺ·жүҖжңү Exception иҖҢйқһд»… ValueErrorпјҢ
    йҳІжӯў SQLAlchemy / зә§иҒ”еҲ йҷӨзӯүејӮеёёеҸҳжҲҗ 500пјҢеүҚз«ҜиғҪж”¶еҲ°еҸҜиҜ»й”ҷиҜҜж¶ҲжҒҜгҖӮ
    """
    deleted = 0
    errors = []
    for pi_id in pi_ids:
        try:
            delete_pi_invoice(db, pi_id)
            deleted += 1
        except Exception as e:
            # дёҡеҠЎй”ҷиҜҜпјҲValueErrorпјү+ еә•еұӮејӮеёёпјҲSQLAlchemyзӯүпјүйғҪеҪ’дёҖдёә errors
            errors.append(f"ID {pi_id}: {type(e).__name__}: {e}")
    return {"deleted": deleted, "total": len(pi_ids), "errors": errors}

@router.get("/export/{pi_id}")
def export_pi_excel(pi_id: int, db: Session = Depends(get_db)):
    from crud.pi import get_pi_invoice_detail
    import io
    from fastapi.responses import StreamingResponse
    
    detail = get_pi_invoice_detail(db, pi_id)
    if detail is None:
        raise HTTPException(status_code=404, detail="PIеҚ•дёҚеӯҳеңЁ")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"PI-{detail['pi_no']}"
        
        # ж ·ејҸе®ҡд№ү
        title_font = Font(name='Arial', size=16, bold=True)
        header_font = Font(name='Arial', size=11, bold=True)
        normal_font = Font(name='Arial', size=10)
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        header_font_white = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # ж Үйўҳ
        ws.merge_cells('A1:H1')
        ws['A1'] = f"PROFORMA INVOICE - {detail['pi_no']}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # еҹәжң¬дҝЎжҒҜ
        status_map = {1: 'иҚүзЁҝ', 2: 'е·ІзЎ®и®Ө', 3: 'е·ІеҸ‘иҙ§', 4: 'е·Іе®ҢжҲҗ'}
        info_data = [
            ('е®ўжҲ·зј–еҸ·', detail.get('customer_code', '')),
            ('е®ўжҲ·еҗҚз§°', detail.get('customer_name', '')),
            ('еёҒз§Қ', detail.get('currency', 'USD')),
            ('жҖ»йҮ‘йўқ', detail.get('total_amount', 0)),
            ('зҠ¶жҖҒ', status_map.get(detail.get('status', 1), 'иҚүзЁҝ')),
            ('еҲӣе»әж—Ҙжңҹ', str(detail.get('created_at', ''))[:10] if detail.get('created_at') else ''),
        ]
        for i, (label, value) in enumerate(info_data, start=3):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = header_font
            ws[f'B{i}'] = value
            ws[f'B{i}'].font = normal_font
        
        # жҳҺз»ҶиЎЁеӨҙ
        row = 10
        headers = ['еәҸеҸ·', 'дә§е“Ғзј–еҸ·', 'OEеҸ·', 'е®ўжҲ·зј–з Ғ', 'жҸҸиҝ°', 'ж•°йҮҸ', 'еҚ•д»·', 'жҖ»д»·', 'еӨҮжіЁ']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # жҳҺз»Ҷж•°жҚ®
        for idx, item in enumerate(detail.get('items', []), 1):
            r = row + idx
            values = [
                idx,
                item.get('product_code', ''),
                item.get('oe_number', ''),
                item.get('customer_code', ''),
                item.get('detail_desc', ''),
                item.get('quantity', 0),
                item.get('unit_price', 0),
                item.get('total_price', 0),
                item.get('remark', '')
            ]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=r, column=col, value=v)
                cell.font = normal_font
                cell.border = thin_border
        
        # еҗҲи®ЎиЎҢ
        total_row = row + len(detail.get('items', [])) + 1
        ws.cell(row=total_row, column=6, value='еҗҲи®Ў').font = header_font
        ws.cell(row=total_row, column=8, value=detail.get('total_amount', 0)).font = header_font
        
        # д»ҳж¬ҫйҳ¶ж®ө
        if detail.get('payment_stages'):
            pay_row = total_row + 2
            ws.cell(row=pay_row, column=1, value='д»ҳж¬ҫйҳ¶ж®ө').font = header_font
            pay_row += 1
            pay_headers = ['йҳ¶ж®өзұ»еһӢ', 'еәҸеҸ·', 'йҮ‘йўқ', 'еә”д»ҳж—Ҙжңҹ', 'зҠ¶жҖҒ']
            for col, h in enumerate(pay_headers, 1):
                cell = ws.cell(row=pay_row, column=col, value=h)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.border = thin_border
            
            for stage in detail['payment_stages']:
                pay_row += 1
                stage_type_map = {'deposit': 'е®ҡйҮ‘', 'balance': 'е°ҫж¬ҫ'}
                stage_status_map = {1: 'еҫ…д»ҳ', 2: 'е·Ід»ҳ'}
                values = [
                    stage_type_map.get(stage['stage_type'], stage['stage_type']),
                    stage.get('stage_no', ''),
                    stage.get('amount', 0),
                    str(stage['due_date'])[:10] if stage.get('due_date') else '',
                    stage_status_map.get(stage['status'], 'еҫ…д»ҳ')
                ]
                for col, v in enumerate(values, 1):
                    cell = ws.cell(row=pay_row, column=col, value=v)
                    cell.font = normal_font
                    cell.border = thin_border
        
        # и°ғж•ҙеҲ—е®Ҫ
        col_widths = [8, 15, 15, 15, 30, 10, 12, 12, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        
        # дҝқеӯҳеҲ°еҶ…еӯҳ
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"PI_{detail['pi_no']}.xlsx"
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except ImportError:
        raise HTTPException(status_code=500, detail="жңҚеҠЎеҷЁзјәе°‘openpyxlеә“пјҢж— жі•еҜјеҮәExcel")

@router.get("/{pi_id}", response_model=PIInvoiceResponse)
def read_pi(pi_id: int, db: Session = Depends(get_db)):
    """рҹ”§ 2026-06-22 дҝ®еӨҚпјҡеҺҹи·Ҝз”ұдёҚиҝ”еӣһitemsеҜјиҮҙеүҚз«ҜеҲ·ж–°жӢҝдёҚеҲ°ж•°жҚ®
    ж”№з”Ё get_pi_invoice_detail иҺ·еҸ–еҢ…еҗ«itemsзҡ„е®Ңж•ҙж•°жҚ®
    """
    from crud.pi import get_pi_invoice_detail
    return get_pi_invoice_detail(db, pi_id)

@router.put("/{pi_id}")
def update_pi_api(pi_id: int, pi_update: PIInvoiceUpdate, db: Session = Depends(get_db)):
    try:
        db_pi = update_pi_invoice(db, pi_id, pi_update)
        if db_pi is None:
            raise HTTPException(status_code=404, detail="PIеҚ•дёҚеӯҳеңЁ")
        return {
            "id": db_pi.id,
            "dept_id": db_pi.dept_id,
            "pi_no": db_pi.pi_no,
            "customer_id": db_pi.customer_id,
            "total_amount": float(db_pi.total_amount) if db_pi.total_amount else 0,
            "currency": db_pi.currency or "USD",
            "status": db_pi.status or 1,
            "created_at": db_pi.created_at.isoformat() if db_pi.created_at else None
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.put("/{pi_id}/status")
def update_pi_status_api(pi_id: int, status_data: dict, db: Session = Depends(get_db)):
    """жӣҙж–°PIеҚ•зҠ¶жҖҒпјҲ0=е·ІеәҹејғпјҢ1=иҚүзЁҝпјҢ2=иҝӣиЎҢдёӯпјҢ3=е·ІеҸ‘иҙ§пјҢ4=е·Іе®ҢжҲҗпјү"""
    try:
        status = status_data.get('status')
        if status not in [0, 1, 2, 3, 4]:
            raise HTTPException(status_code=400, detail="ж— ж•Ҳзҡ„зҠ¶жҖҒеҖј")
        db_pi = update_pi_status(db, pi_id, status)
        return {
            "id": db_pi.id,
            "pi_no": db_pi.pi_no,
            "status": db_pi.status,
            "success": True
        }
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

@router.patch("/{pi_id}/storage-status")
def update_pi_storage_status_api(pi_id: int, data: dict, db: Session = Depends(get_db)):
    """жӣҙж–°PIеҚ•еә“еӯҳзҠ¶жҖҒпјҲж”ҜжҢҒ'зјәиҙ§'ж Үи®°пјү"""
    from models.pi import PiProformaInvoice

    db_pi = db.query(PiProformaInvoice).filter(PiProformaInvoice.id == pi_id).first()
    if not db_pi:
        raise HTTPException(status_code=404, detail="PIеҚ•дёҚеӯҳеңЁ")

    new_status = data.get('storage_status')
    valid_values = {'е·Іе…Ҙеә“', 'йғЁеҲҶе…Ҙеә“', 'жңӘе…Ҙеә“', 'зјәиҙ§', None, ''}
    if new_status not in valid_values:
        raise HTTPException(status_code=400, detail=f"ж— ж•Ҳзҡ„еә“еӯҳзҠ¶жҖҒеҖј: {new_status}")

    db_pi.storage_status = new_status if new_status else None
    db.commit()

    return {"id": pi_id, "storage_status": db_pi.storage_status}

@router.get("/price-history/{customer_id}/{product_id}")
def read_price_history(customer_id: int, product_id: int, db: Session = Depends(get_db)):
    history = get_price_history(db, customer_id, product_id)
    if history is None:
        return {"message": "жҡӮж— еҺҶеҸІд»·ж ји®°еҪ•"}
    return {
        "unit_price": history.unit_price,
        "remark": history.remark,
        "created_at": history.created_at
    }

# 2026-06-15 ж–°еўһпјҡиҺ·еҸ–PIдёӢжүҖжңүи®ўеҚ•йЎ№
@router.get("/{pi_id}/items")
def read_pi_items(pi_id: int, db: Session = Depends(get_db)):
    """иҺ·еҸ–PIдёӢзҡ„жүҖжңүи®ўеҚ•йЎ№"""
    from models import PiProformaInvoiceItem, PiProformaInvoice, CrmCustomer
    # Phase 4: з»ҹдёҖдә§е“Ғи®ҝй—®
    from services.product_lookup import unified_product_lookup

    import logging
    logger = logging.getLogger(__name__)

    logger.info(f"[иҺ·еҸ–и®ўеҚ•йЎ№-API] pi_id={pi_id}")

    items = db.query(PiProformaInvoiceItem).filter(
        PiProformaInvoiceItem.pi_id == pi_id
    ).all()

    logger.info(f"[иҺ·еҸ–и®ўеҚ•йЎ№-API] жҹҘиҜўеҲ° {len(items)} дёӘи®ўеҚ•йЎ№")

    result = []
    for item in items:
        # иҺ·еҸ–е…іиҒ”зҡ„дә§е“ҒдҝЎжҒҜпјҲPhase 4: з»ҹдёҖжҹҘжүҫпјү
        product = None
        product_name = ""
        model = ""
        customer_name = ""
        if item.product_id:
            product = unified_product_lookup(db, item.product_id)
            if product:
                # 2026-06-23 дҝ®еӨҚпјҡдјҳе…ҲдҪҝз”Ё product_nameпјҢдёҺдә§е“Ғз®ЎзҗҶеҲ—иЎЁдҝқжҢҒдёҖиҮҙ
                product_name = product.product_name or product.detail_desc or ""

        # иҺ·еҸ–PIдҝЎжҒҜ
        pi = db.query(PiProformaInvoice).filter(PiProformaInvoice.id == pi_id).first()
        pi_no = pi.pi_no if pi else ""
        if pi and pi.customer_id:
            customer = db.query(CrmCustomer).filter(CrmCustomer.id == pi.customer_id).first()
            if customer:
                customer_name = customer.customer_name or ""
        
        # иҺ·еҸ–е·ІеҮәиҙ§ж•°йҮҸ
        shipped_quantity = 0.0
        # TODO: д»ҺеҮәиҙ§иЎЁи®Ўз®—е·ІеҮәиҙ§ж•°йҮҸ
        
        result.append({
            "id": item.id,
            "pi_id": item.pi_id,
            "pi_no": pi_no,
            "product_id": item.product_id,
            "product_name": product_name or item.detail_desc or "",
            "model": item.customer_code or item.customer_model or "",
            "customer_name": customer_name,
            "quantity": float(item.quantity) if item.quantity else 0,
            "shipped_quantity": shipped_quantity,
            "unit_price": float(item.unit_price) if item.unit_price else 0,
        })
    
    return result


# 2026-06-10 ж–°еўһпјҡPIи®ўеҚ•йЎ№API
@router.get("/items/{item_id}")
def read_pi_item(item_id: int, db: Session = Depends(get_db)):
    """иҺ·еҸ–PIи®ўеҚ•йЎ№иҜҰжғ…"""
    item = get_pi_item(db, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="и®ўеҚ•йЎ№дёҚеӯҳеңЁ")
    return {
        "id": item.id,
        "pi_id": item.pi_id,
        "product_id": item.product_id,
        "oe_number": item.oe_number,
        "customer_code": item.customer_code,
        "detail_desc": item.detail_desc,
        "detail_desc_en": getattr(item, "detail_desc_en", None),
        "product_acquires": getattr(item, "product_acquires", None),
        "product_color": getattr(item, "product_color", None),
        "quantity": float(item.quantity) if item.quantity else 0,
        "unit_price": float(item.unit_price) if item.unit_price else 0,
    }

@router.put("/items/{item_id}")
def update_pi_item_api(item_id: int, update_data: dict, db: Session = Depends(get_db)):
    """жӣҙж–°PIи®ўеҚ•йЎ№"""

    # рҹ”§ 2026-06-22 дҝ®еӨҚпјҡдҪҝз”Ё ASCII йҒҝе…Қ Windows GBK з»Ҳз«Ҝзҡ„ UnicodeEncodeError
    print(f"\n{'='*60}")
    print(f"[DEBUG-API] === update_pi_item_api CALLED ===")
    print(f"[DEBUG-API] item_id: {item_id}")
    print(f"[DEBUG-API] field count: {len(update_data)}")
    print(f"[DEBUG-API] field list: {list(update_data.keys())}")

    # жЈҖжҹҘе…ій”®еӯ—ж®өжҳҜеҗҰеӯҳеңЁ
    key_fields = ['packaging', 'purchase_option_name', 'pack_spec',
                  'customer_prepayment', 'factory_deposit']
    for field in key_fields:
        if field in update_data:
            print(f"[DEBUG-API] [OK] key field '{field}' = {update_data[field]}")
        else:
            print(f"[DEBUG-API] [MISSING] field '{field}'")
    
    print(f"{'='*60}\n")
    
    db_item = update_pi_item(db, item_id, update_data)
    if not db_item:
        raise HTTPException(status_code=404, detail="и®ўеҚ•йЎ№дёҚеӯҳеңЁ")
    
    # рҹ”Қ DEBUG: иҝ”еӣһжӣҙж–°еҗҺзҡ„ж•°жҚ®д»ҘдҫҝеүҚз«ҜйӘҢиҜҒ
    return {
        "success": True,
        "id": db_item.id,
        "product_id": db_item.product_id,
        "message": "и®ўеҚ•йЎ№жӣҙж–°жҲҗеҠҹ",
        # вң… ж–°еўһпјҡиҝ”еӣһе…ій”®еӯ—ж®өеҖјдҫӣйӘҢиҜҒ
        "debug_fields": {
            "detail_desc": getattr(db_item, 'detail_desc', None),
            "product_short_name": getattr(db_item, 'product_short_name', None),
            "product_short_name_en": getattr(db_item, 'product_short_name_en', None),
            "packaging": getattr(db_item, 'packaging', None),
            "purchase_option_name": getattr(db_item, 'purchase_option_name', None),
            "pack_spec": getattr(db_item, 'pack_spec', None),
        }
    }


class ChangeSupplierRequest(BaseModel):
    supplier_name: Optional[str] = None
    factory_short_name: Optional[str] = None
    shop_url: Optional[str] = None
    line_1688_url: Optional[str] = None
    factory_code: Optional[str] = None
    brand: Optional[str] = None
    purchase_price: Optional[float] = None
    factory_price: Optional[float] = None
    factory_deposit: Optional[float] = None
    factory_balance: Optional[float] = None
    invoice_status: Optional[str] = None


@router.put("/items/{item_id}/change-supplier")
def change_supplier_api(item_id: int, body: ChangeSupplierRequest, db: Session = Depends(get_db)):
    """жӣҙжҚў PI item зҡ„дҫӣеә”е•Ҷе№¶йҮҚж–°з”ҹжҲҗйҮҮиҙӯеҚ•"""
    from crud.pi import change_pi_item_supplier
    try:
        result = change_pi_item_supplier(db, item_id, body.model_dump(exclude_unset=True))
        return {"success": True, **result}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


# 2026-06-12 йңҖжұӮ#40пјҡиҪҜеҲ йҷӨ / е…Ҙеә“жҺҘеҸЈ
class InboundSingleRequest(BaseModel):
    quantity: float
    inspector: Optional[str] = None
    remark: Optional[str] = None

class InboundBatchRequest(BaseModel):
    items: list[dict]  # [{"pi_item_id": int, "quantity": float, "remark": str}]
    inspector: Optional[str] = None

@router.delete("/items/{item_id}")
def delete_pi_item_api(item_id: int, db: Session = Depends(get_db)):
    """иҪҜеҲ йҷӨ PI еҚ•е“Ғ"""
    item = delete_pi_item(db, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="е•Ҷе“ҒдёҚеӯҳеңЁжҲ–е·ІеҲ йҷӨ")
    return {"success": True, "item_id": item_id}

@router.post("/items/{item_id}/inbound")
def inbound_pi_item_api(item_id: int, body: InboundSingleRequest, db: Session = Depends(get_db)):
    """еҚ•е“Ғе…Ҙеә“"""
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"[рҹҢҗROUTER] POST /pi/items/{item_id}/inbound  payload: quantity={body.quantity}, inspector={body.inspector}, remark={body.remark!r}")
    try:
        inv = inbound_pi_item(db, item_id, body.quantity, body.inspector, body.remark)
        logger.info(f"[рҹҢҗROUTER] вң… POST /pi/items/{item_id}/inbound  returned inventory_log_id={inv.id}")
        return {"success": True, "inventory_log_id": inv.id}
    except ValueError as e:
        logger.error(f"[рҹҢҗROUTER] вқҢ POST /pi/items/{item_id}/inbound  ValueError: {e}")
        raise HTTPException(status_code=400, detail=str(e))

@router.post("/{pi_id}/inbound-batch")
def inbound_pi_items_batch_api(pi_id: int, body: InboundBatchRequest, db: Session = Depends(get_db)):
    """жү№йҮҸе…Ҙеә“"""
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"[рҹҢҗROUTER] POST /pi/{pi_id}/inbound-batch  items_count={len(body.items)}, inspector={body.inspector}")
    for i, it in enumerate(body.items):
        logger.info(f"[рҹҢҗROUTER]   item[{i}]: {it}")
    result = inbound_pi_items_batch(db, pi_id, body.items, body.inspector)
    logger.info(f"[рҹҢҗROUTER] вң… POST /pi/{pi_id}/inbound-batch  result: {result}")
    return {"success": True, **result}


# ============== 2026-06-12 йңҖжұӮ#42пјҡеҺҶеҸІи®°еҪ• + жӯЈејҸзәӘеҪ• API ==============
class SaveSnapshotRequest(BaseModel):
    change_desc: Optional[str] = ""
    version_no: int

class VersionsResponse(BaseModel):
    id: int
    pi_id: int
    version_no: int
    change_desc: Optional[str]
    created_at: str

@router.get("/{pi_id}/versions", response_model=list[VersionsResponse])
def get_versions(pi_id: int, db: Session = Depends(get_db)):
    """иҺ·еҸ– PI жүҖжңүеҺҶеҸІзүҲжң¬"""
    versions = get_pi_versions(db, pi_id)
    return [
        VersionsResponse(
            id=v.id, pi_id=v.pi_id, version_no=v.version_no,
            change_desc=v.change_desc,
            created_at=v.created_at.isoformat() if v.created_at else "",
        )
        for v in versions
    ]


@router.get("/{pi_id}/versions/{version_id}/snapshot")
def get_version_snapshot(pi_id: int, version_id: int, db: Session = Depends(get_db)):
    """иҺ·еҸ–жҢҮе®ҡзүҲжң¬зҡ„е®Ңж•ҙеҝ«з…§ж•°жҚ®"""
    from models.pi import PiProformaInvoiceVersion
    
    version = db.query(PiProformaInvoiceVersion).filter(
        PiProformaInvoiceVersion.id == version_id,
        PiProformaInvoiceVersion.pi_id == pi_id
    ).first()
    
    if not version:
        raise HTTPException(status_code=404, detail="зүҲжң¬дёҚеӯҳеңЁ")
    
    return {
        "id": version.id,
        "pi_id": version.pi_id,
        "version_no": version.version_no,
        "change_desc": version.change_desc,
        "created_at": version.created_at.isoformat() if version.created_at else "",
        "created_by": version.created_by,
        "snapshot": version.snapshot_data
    }


@router.post("/{pi_id}/versions")
def save_version(pi_id: int, body: SaveSnapshotRequest, db: Session = Depends(get_db)):
    """дҝқеӯҳж–°еҝ«з…§пјҲд№җи§Ӯй”Ғпјү"""
    try:
        v = save_pi_snapshot(db, pi_id, body.change_desc, body.version_no)
        return {"success": True, "version_no": v.version_no}
    except ValueError as e:
        raise HTTPException(status_code=409, detail=str(e))

@router.post("/{pi_id}/formal-record")
def save_formal(pi_id: int, db: Session = Depends(get_db)):
    """дҝқеӯҳжӯЈејҸзәӘеҪ•пјҲJSON ж–Үд»¶пјү"""
    from models.pi import PiProformaInvoice
    pi = db.query(PiProformaInvoice).filter(PiProformaInvoice.id == pi_id).first()
    if not pi:
        raise HTTPException(status_code=404, detail="PI дёҚеӯҳеңЁ")
    try:
        path = save_formal_record(db, pi_id)
        # жӯЈејҸзәӘеҪ•ж–Үд»¶еҶҷе…ҘжҲҗеҠҹеҗҺпјҢеҶҚжҸҗдәӨж•°жҚ®еә“жӯЈејҸзҠ¶жҖҒпјҢйҒҝе…ҚеҚҠжҲҗеҠҹгҖӮ
        if pi.status == 1:
            pi.status = 2
            db.commit()
        return {"success": True, "file_path": path}
    except ValueError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/{pi_id}/formal-record")
def get_formal(pi_id: int, db: Session = Depends(get_db)):
    """иҜ»еҸ–жӯЈејҸзәӘеҪ•"""
    pi = get_pi_invoice(db, pi_id)
    if not pi:
        raise HTTPException(status_code=404, detail="PI дёҚеӯҳеңЁ")
    data = load_formal_record(pi.pi_no)
    if not data:
        raise HTTPException(status_code=404, detail="жӯЈејҸзәӘеҪ•дёҚеӯҳеңЁ")
    return data

@router.get("/{pi_id}/formal-record/exists")
def formal_exists(pi_id: int, db: Session = Depends(get_db)):
    """жЈҖжҹҘжӯЈејҸзәӘеҪ•жҳҜеҗҰеӯҳеңЁпјӣж•°жҚ®еә“жӯЈејҸзҠ¶жҖҒжҳҜи·ЁиҠӮзӮ№зҡ„жқғеЁҒжқҘжәҗгҖӮ"""
    pi = get_pi_invoice(db, pi_id)
    if not pi:
        return {"exists": False}
    return {"exists": pi.status == 2 or formal_record_exists(pi.pi_no)}


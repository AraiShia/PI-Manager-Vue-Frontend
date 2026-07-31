"""
客户回复API路由
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional, Any, cast
from io import BytesIO
from datetime import datetime
from app.database import get_db
from schemas.customer_reply import CustomerReplyCreate, CustomerReplyUpdate, CustomerReplyResponse, BatchRepliesRequest
from crud.customer_reply import (
    get_customer_replies, get_customer_replies_by_pi, get_latest_reply_by_pi,
    get_customer_replies_by_customer, get_customer_reply,
    create_customer_reply, update_customer_reply, delete_customer_reply,
    get_replies_by_items
)
from crud.customer import get_customer as get_customer_by_id
from crud.pi import get_pi_invoice as get_pi, resolve_pi

router = APIRouter(prefix="/api/customer-replies", tags=["客户回复"])


@router.get("", response_model=List[CustomerReplyResponse])
def get_all_replies(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    """获取所有客户回复"""
    return get_customer_replies(db, skip=skip, limit=limit)


@router.get("/pi/{pi_id}", response_model=List[CustomerReplyResponse])
def get_replies_by_pi(pi_id: str, db: Session = Depends(get_db)):
    """获取某PI的所有客户回复（支持PI数字ID或PI编号如PISO9J02607280）"""
    pi_obj = resolve_pi(db, pi_id)
    if not pi_obj:
        return []
    return get_customer_replies_by_pi(db, cast(int, pi_obj.id))


@router.get("/pi/{pi_id}/latest", response_model=Optional[CustomerReplyResponse])
def get_latest_reply(pi_id: str, db: Session = Depends(get_db)):
    """获取某PI的最新客户回复（支持PI数字ID或PI编号如PISO9J02607280）"""
    pi_obj = resolve_pi(db, pi_id)
    if not pi_obj:
        return None
    return get_latest_reply_by_pi(db, cast(int, pi_obj.id))


@router.get("/pi/{pi_id}/list")
def list_replies_with_labels(pi_id: str, db: Session = Depends(get_db)):
    """获取排序后的回复列表（含序号标签，支持PI数字ID或PI编号）"""
    pi = resolve_pi(db, pi_id)
    if not pi:
        raise HTTPException(status_code=404, detail="PI不存在")

    replies = get_customer_replies_by_pi(db, cast(int, pi.id))
    customer = get_customer_by_id(db, cast(int, pi.customer_id)) if pi.customer_id else None

    return {
        "pi_id": cast(int, pi.id),
        "pi_no": pi.pi_no,
        "customer_name": customer.customer_name if customer else "",
        "replies": [
            {
                "id": r.id,
                "reply_type": r.reply_type,
                "sequence_label": r.sequence_label,
                "submitter_name": r.submitter_name,
                "reply_date": r.reply_date.isoformat() if r.reply_date else None,
                "reply_content": r.reply_content,
            }
            for r in replies
        ]
    }


@router.post("/export")
def export_replies(
    pi_id: str,
    customer_name: str = "",
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    selected_ids: Optional[List[int]] = Query(None),
    db: Session = Depends(get_db)
):
    """导出回复记录为 Excel（支持PI数字ID或PI编号）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from models.customer_reply import CustomerReply
    from sqlalchemy import asc

    pi_obj = resolve_pi(db, pi_id)
    target_pi_id = cast(int, pi_obj.id) if pi_obj else (int(pi_id) if str(pi_id).isdigit() else -1)

    query = db.query(CustomerReply).filter(CustomerReply.pi_id == target_pi_id)

    if start_date:
        query = query.filter(CustomerReply.reply_date >= start_date)
    if end_date:
        query = query.filter(CustomerReply.reply_date <= end_date)
    if selected_ids:
        query = query.filter(CustomerReply.id.in_(selected_ids))

    replies = query.order_by(asc(CustomerReply.reply_date)).all()

    wb = Workbook()
    ws = wb.active
    assert ws is not None, "Worksheet initialization failed"
    ws.title = "客户回复记录"

    headers = ["序号", "关联维度/产品", "类型", "提交者", "时间", "内容"]
    header_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    from models import PiProformaInvoiceItem as PiItem
    from crud.product import get_product as get_prod

    for row_idx, r in enumerate(replies, 2):
        prefix = "C" if r.reply_type == "customer" else "R"
        type_text = "客户" if r.reply_type == "customer" else "我方"
        color = "000000" if r.reply_type == "customer" else "1E40AF"
        # 显式转换为 Python 原生类型，避免 openpyxl 类型检查告警
        seq_num = cast(int, r.sequence_num) if r.sequence_num is not None else 0

        # 解析关联维度/单品信息
        target_product_info = "整单通用"
        if r.pi_item_id:
            pi_item = db.query(PiItem).filter(PiItem.id == r.pi_item_id).first()
            if pi_item:
                prod_name = None
                if pi_item.product_id:
                    prod = get_prod(db, cast(int, pi_item.product_id))
                    if prod:
                        prod_name = getattr(prod, 'product_name', None) or getattr(prod, 'name_cn', None)
                target_product_info = f"[单品] {prod_name or getattr(pi_item, 'product_code', None) or f'明细#{r.pi_item_id}'}"

        ws.cell(row=row_idx, column=1, value=f"{prefix}{seq_num}")
        ws.cell(row=row_idx, column=2, value=str(target_product_info))
        ws.cell(row=row_idx, column=3, value=str(type_text))
        ws.cell(row=row_idx, column=4, value=str(r.submitter_name or ""))
        ws.cell(row=row_idx, column=5, value=r.reply_date.strftime("%Y-%m-%d %H:%M:%S") if r.reply_date else "")
        ws.cell(row=row_idx, column=6, value=str(r.reply_content or ""))
        ws.cell(row=row_idx, column=6).font = Font(color=color)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 50

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    date_range = f"{start_date or '开始'}至{end_date or '结束'}"
    export_type = "_选择性导出" if selected_ids else ""
    filename = f"客户回复记录_{customer_name}_{date_range}{export_type}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


@router.post("/batch-by-items")
def batch_get_replies(request: BatchRepliesRequest, db: Session = Depends(get_db)):
    """按商品列表批量获取回复记录"""
    from models import PiProformaInvoiceItem as PiItem
    from crud.product import get_product as get_prod

    raw_replies = get_replies_by_items(db, request.items)

    result = []
    for r in raw_replies:
        item_data: dict[str, Any] = {
            "id": r.id,
            "pi_id": r.pi_id,
            "pi_item_id": r.pi_item_id,
            "product_name": None,
            "pi_no": None,
            "reply_type": r.reply_type,
            "sequence_label": getattr(r, 'sequence_label', ''),
            "submitter_name": r.submitter_name,
            "reply_date": r.reply_date.isoformat() if r.reply_date else None,
            "reply_content": r.reply_content,
            "customer_name": None,
        }

        if r.pi_item_id:
            pi_item = db.query(PiItem).filter(PiItem.id == r.pi_item_id).first()
            if pi_item and pi_item.product_id:
                prod = get_prod(db, cast(int, pi_item.product_id))
                if prod:
                    item_data["product_name"] = getattr(prod, 'product_name', None) or getattr(prod, 'name_cn', None) or ""

        pi_obj = get_pi(db, cast(int, r.pi_id))
        if pi_obj:
            item_data["pi_no"] = pi_obj.pi_no
            cust = get_customer_by_id(db, cast(int, pi_obj.customer_id)) if pi_obj.customer_id else None
            if cust:
                item_data["customer_name"] = cust.customer_name

        result.append(item_data)

    return {"replies": result, "total": len(result)}


@router.post("/export-batch")
def export_batch_replies(
    request: BatchRepliesRequest,
    selected_ids: Optional[List[int]] = Query(None),
    db: Session = Depends(get_db)
):
    """批量导出多商品回复记录为 Excel（单 Sheet）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from models import PiProformaInvoiceItem as PiItem
    from io import BytesIO

    raw_replies = get_replies_by_items(db, request.items)

    if selected_ids:
        raw_replies = [r for r in raw_replies if r.id in selected_ids]

    rows = []
    for r in raw_replies:
        product_name = ""
        if r.pi_item_id:
            pi_item = db.query(PiItem).filter(PiItem.id == r.pi_item_id).first()
            if pi_item and pi_item.product_id:
                from crud.product import get_product as get_prod
                prod = get_prod(db, cast(int, pi_item.product_id))
                if prod:
                    product_name = getattr(prod, 'product_name', None) or getattr(prod, 'name_cn', None) or ""

        pi_obj = get_pi(db, cast(int, r.pi_id))
        pi_no_str = pi_obj.pi_no if pi_obj else ""

        prefix = "C" if r.reply_type in ("customer", "question") else "R"
        seq_label = getattr(r, 'sequence_label', f"{prefix}{r.sequence_num}")

        rows.append({
            "seq": seq_label,
            "product_name": product_name,
            "pi_no": pi_no_str,
            "type_text": "客户提问" if r.reply_type == "question" else ("客户" if r.reply_type == "customer" else "我方回复"),
            "submitter": r.submitter_name or "",
            "date": r.reply_date.strftime("%Y-%m-%d %H:%M:%S") if r.reply_date else "",
            "content": r.reply_content,
            "is_customer": r.reply_type in ("customer", "question"),
        })

    wb = Workbook()
    ws = wb.active
    assert ws is not None, "Worksheet initialization failed"
    ws.title = "客户回复记录"

    headers = ["序号", "商品名", "PI号", "类型", "提交人", "日期", "内容"]
    header_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    customer_font = Font(color="000000")
    reply_font = Font(color="0066CC")

    for row_idx, rd in enumerate(rows, 2):
        font = customer_font if rd["is_customer"] else reply_font
        # 显式转换为 str，消除 openpyxl value 类型告警
        ws.cell(row=row_idx, column=1, value=str(rd["seq"]))
        ws.cell(row=row_idx, column=2, value=str(rd["product_name"]))
        ws.cell(row=row_idx, column=3, value=str(rd["pi_no"]))
        ws.cell(row=row_idx, column=4, value=str(rd["type_text"]))
        ws.cell(row=row_idx, column=5, value=str(rd["submitter"]))
        ws.cell(row=row_idx, column=6, value=str(rd["date"]))
        content_cell = ws.cell(row=row_idx, column=7, value=str(rd["content"] or ""))
        content_cell.font = font

    col_widths = [8, 25, 20, 10, 12, 22, 60]
    for i, w in enumerate(col_widths, 1):
        col_letter = chr(64 + i)
        ws.column_dimensions[col_letter].width = w

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    date_start = str(rows[0]["date"]).split(" ")[0] if rows else ""
    date_end = str(rows[-1]["date"]).split(" ")[0] if rows else ""
    date_range = f"{date_start}_{date_end}" if date_start and date_end and date_start != date_end else (date_start or "")
    export_suffix = "_选择性导出" if selected_ids else ""

    customer_name = ""
    if request.items:
        first_pi_id = request.items[0].get("pi_id")
        if first_pi_id:
            pi_obj = get_pi(db, int(first_pi_id))
            if pi_obj:
                cust = get_customer_by_id(db, cast(int, pi_obj.customer_id)) if pi_obj.customer_id else None
                if cust:
                    customer_name = cust.customer_name

    filename = f"客户回复记录_{customer_name}_{date_range}{export_suffix}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


@router.get("/customer/{customer_id}", response_model=List[CustomerReplyResponse])
def get_replies_by_customer(customer_id: int, db: Session = Depends(get_db)):
    """获取某客户的所有回复"""
    return get_customer_replies_by_customer(db, customer_id)


@router.post("", response_model=CustomerReplyResponse)
def create_reply(reply: CustomerReplyCreate, db: Session = Depends(get_db)):
    """新增客户回复"""
    return create_customer_reply(db, reply)


@router.put("/{reply_id}", response_model=CustomerReplyResponse)
def update_reply(reply_id: int, reply: CustomerReplyUpdate, db: Session = Depends(get_db)):
    """更新客户回复"""
    db_reply = update_customer_reply(db, reply_id, reply)
    if not db_reply:
        raise HTTPException(status_code=404, detail="回复不存在")
    return db_reply


@router.delete("/{reply_id}")
def delete_reply(reply_id: int, db: Session = Depends(get_db)):
    """删除客户回复"""
    success = delete_customer_reply(db, reply_id)
    if not success:
        raise HTTPException(status_code=404, detail="回复不存在")
    return {"message": "删除成功"}

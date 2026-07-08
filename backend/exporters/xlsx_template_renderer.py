"""通用 xlsx 模板渲染器。

加载 xlsx 副本（保留原样式） + 解析 mapping YAML → 根据 data 填充单元格。
"""
import operator
import re
from io import BytesIO
from typing import Any, Dict, List, Optional

import openpyxl
import yaml


class XlsxTemplateRenderer:
    """根据 mapping 渲染 xlsx 模板。"""

    def __init__(self, template_path: str, mapping_path: str):
        self.template_path = template_path
        self.mapping_path = mapping_path
        # 保留原样式（字体/颜色/边框/合并/列宽/行高）
        self.wb = openpyxl.load_workbook(template_path)
        with open(mapping_path, encoding="utf-8") as f:
            self.mapping: Dict[str, Any] = yaml.safe_load(f) or {}

    def render(self, data: Dict[str, Any]) -> bytes:
        """按 mapping 填充数据并返回 xlsx 字节流。"""
        ws = self.wb[self.mapping["template_sheet"]]
        for field in self.mapping.get("fields", []):
            self._apply_field(ws, field, data)
        return self._save_to_bytes(self.wb)

    # ---- field dispatch ----

    def _apply_field(self, ws, field: Dict[str, Any], data: Dict[str, Any]) -> None:
        """分发到各 type 处理器。"""
        ftype = field.get("type")
        if ftype == "static":
            ws[field["cell"]] = field["value"]
        elif ftype == "data":
            value = self._resolve_path(data, field["path"])
            if value is None:
                value = field.get("default", "")
            ws[field["cell"]] = value
        elif ftype == "format":
            template = field["template"]
            sources = field.get("sources", {})
            resolved = {
                key: (self._resolve_path(data, path) or "")
                for key, path in sources.items()
            }
            ws[field["cell"]] = template.format(**resolved)
        elif ftype == "loop":
            self._apply_loop(ws, field, data)
        elif ftype == "calc":
            ws[field["cell"]] = self._eval_formula(field["formula"], {})
        elif ftype == "sum":
            items = self._resolve_path(data, field["data_path"]) or []
            field_name = field["field"]
            total = sum((it.get(field_name) or 0) for it in items)
            ws[field["cell"]] = total
        else:
            raise ValueError(f"unsupported field type: {ftype}")

    # ---- loop ----

    def _apply_loop(self, ws, field: Dict[str, Any], data: Dict[str, Any]) -> None:
        """按起始 cell 向下逐行展开 items。遇 merged 区域时跳过。"""
        items = self._resolve_path(data, field["data_path"]) or []
        # 支持最大行数限制
        max_rows = field.get("max_rows", len(items))
        if max_rows > 0:
            items = items[:max_rows]

        start_cell = field["cell"]
        m = re.match(r"^([A-Z]+)(\d+)$", start_cell)
        if not m:
            raise ValueError(f"invalid start cell: {start_cell}")
        col_letter, row_num = m.group(1), int(m.group(2))
        merged_ranges = list(ws.merged_cells.ranges)

        for i, item in enumerate(items):
            current_row = row_num + i
            for col, expr in field["template_row"].items():
                target_cell = f"{col}{current_row}"
                if self._in_merged_skip_first(merged_ranges, ws, target_cell):
                    continue
                if isinstance(expr, dict):
                    if expr.get("type") == "calc":
                        ws[target_cell] = self._eval_formula(expr["formula"], item)
                    else:
                        ws[target_cell] = str(expr)
                else:
                    # expr 形如 "item.product_name" / "item.quantity"
                    if expr.startswith("item."):
                        key = expr[len("item."):]
                        value = item.get(key)
                        # 处理图片字段（base64 或路径）
                        if key == "photo" and value:
                            # 如果是 base64 图片，尝试写入（openpyxl 不直接支持，留空或写路径）
                            if isinstance(value, str) and len(value) > 1000:
                                value = "[IMAGE]"  # 标记有图片
                        ws[target_cell] = value
                    else:
                        ws[target_cell] = expr

    @staticmethod
    def _in_merged_skip_first(merged_ranges, ws, target_cell: str) -> bool:
        """target_cell 落在 merged 区域且不是该区域左上角 → 跳过。"""
        for mr in merged_ranges:
            if mr.min_row <= ws[target_cell].row <= mr.max_row and \
               mr.min_col <= ws[target_cell].column <= mr.max_col:
                # 左上角 cell
                anchor = ws.cell(row=mr.min_row, column=mr.min_col)
                anchor_addr = anchor.coordinate
                if target_cell != anchor_addr:
                    return True
        return False

    # ---- helpers ----

    @staticmethod
    def _resolve_path(data: Dict[str, Any], path: str):
        """按点路径取值。失败返回 None。"""
        if not path:
            return None
        cur = data
        for part in path.split("."):
            if isinstance(cur, dict) and part in cur:
                cur = cur[part]
            else:
                return None
        return cur

    @staticmethod
    def _eval_formula(formula: str, item: Dict[str, Any]):
        """极简公式求值：'item.X * item.Y' / 'item.X + item.Y' / '10 * 5'。"""
        ops = {
            "*": operator.mul,
            "+": operator.add,
            "-": operator.sub,
            "/": operator.truediv,
        }
        for op_sym, op_func in ops.items():
            if op_sym in formula:
                left_str, right_str = formula.split(op_sym, 1)
                left = XlsxTemplateRenderer._eval_term(left_str.strip(), item)
                right = XlsxTemplateRenderer._eval_term(right_str.strip(), item)
                return op_func(left, right)
        raise ValueError(f"unsupported formula: {formula}")

    @staticmethod
    def _eval_term(term: str, item: Dict[str, Any]):
        """解析单个项：'item.X' / '10'。"""
        if term.startswith("item."):
            return item.get(term[len("item."):], 0)
        try:
            return float(term) if "." in term else int(term)
        except (ValueError, TypeError):
            return 0

    @staticmethod
    def _save_to_bytes(wb) -> bytes:
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

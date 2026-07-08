# ============================================================
# Excel解析服务
# 文件：services/excel_parser.py
# 创建日期：2026-05-29
# 用途：解析Excel文件，支持.xlsx和.xls格式
# ============================================================

from typing import Dict, List, Optional, Tuple, Any
from io import BytesIO
import re

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


class ExcelParser:
    """Excel文件解析器"""
    
    SUPPORTED_EXTENSIONS = ['.xlsx', '.xls']
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
    
    def __init__(self):
        if not openpyxl:
            raise ImportError("请安装 openpyxl: pip install openpyxl")
    
    def parse_preview(
        self, 
        content: bytes, 
        max_rows: int = 10
    ) -> Dict[str, Any]:
        """
        预览Excel文件
        
        Args:
            content: Excel文件内容（bytes）
            max_rows: 预览最大行数
        
        Returns:
            Dict: 包含 headers, preview_rows, total_rows, column_count
        
        🔧 2026-07-02 修复：单行数据异常时跳过该行而不是整个文件失败
        """
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=False)
        sheet = workbook.active
        
        # 获取表头
        headers = []
        try:
            for cell in sheet[1]:
                headers.append(str(cell.value) if cell.value else '')
        except Exception as e:
            print(f"[WARN] 读取表头失败: {e}")
            headers = []
        
        # 获取预览数据
        preview_rows = []
        row_count = min(sheet.max_row - 1, max_rows)  # 减1是因为有表头
        
        for row_idx in range(2, row_count + 2):  # 从第2行开始
            try:
                row_data = []
                for cell in sheet[row_idx]:
                    row_data.append(self._format_cell_value(cell.value))
                # 跳过完全空白的行
                if any(v.strip() if isinstance(v, str) else v for v in row_data):
                    preview_rows.append(row_data)
            except Exception as e:
                # 🔧 2026-07-02 修复：单行数据异常时跳过该行，不影响整个文件
                print(f"[WARN] 跳过异常行 {row_idx}: {e}")
                continue
        
        return {
            'headers': headers,
            'preview_rows': preview_rows,
            'total_rows': sheet.max_row - 1,  # 减去表头行
            'column_count': sheet.max_column
        }
    
    def parse_all(self, content: bytes) -> List[List[str]]:
        """
        解析所有行数据
        
        Args:
            content: Excel文件内容（bytes）
        
        Returns:
            List[List[str]]: 所有行数据
        
        🔧 2026-07-02 修复：单行数据异常时跳过该行
        """
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=False)
        sheet = workbook.active
        
        rows = []
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            try:
                row_data = [self._format_cell_value(cell) for cell in row]
                # 跳过完全空白的行
                if any(v.strip() if isinstance(v, str) else v for v in row_data):
                    rows.append(row_data)
            except Exception as e:
                print(f"[WARN] 跳过异常行 {row_idx}: {e}")
                continue
        
        return rows
    
    def parse_with_mapping(
        self, 
        content: bytes, 
        column_mapping: Dict[int, str]
    ) -> List[Dict[str, Any]]:
        """
        按列映射解析数据
        
        Args:
            content: Excel文件内容（bytes）
            column_mapping: 列索引到字段名的映射 {0: 'pi_no', 1: 'customer_code', ...}
        
        Returns:
            List[Dict]: 映射后的数据列表
        
        🔧 2026-07-02 修复：单行数据异常时跳过该行
        """
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=False)
        sheet = workbook.active
        
        rows = []
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_idx == 1:  # 跳过表头
                continue
            try:
                row_data = {}
                for col_idx, field_name in column_mapping.items():
                    if col_idx < len(row):
                        row_data[field_name] = self._format_cell_value(row[col_idx])
                # 跳过完全空白的行
                if any(v.strip() if isinstance(v, str) else v for v in row_data.values()):
                    rows.append(row_data)
            except Exception as e:
                print(f"[WARN] 跳过异常行 {row_idx}: {e}")
                continue
        
        return rows
    
    def detect_header_mapping(
        self, 
        headers: List[str], 
        field_mapping: Dict[str, str]
    ) -> Dict[int, str]:
        """
        自动检测表头与字段的映射关系
        
        Args:
            headers: Excel表头列表
            field_mapping: 字段名到数据库字段的映射 {'客户产品编号': 'customer_code', ...}
        
        Returns:
            Dict[int, str]: 列索引到数据库字段的映射
        """
        column_mapping = {}
        
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            
            header_clean = header.strip()
            
            # 精确匹配
            if header_clean in field_mapping:
                column_mapping[col_idx] = field_mapping[header_clean]
                continue
            
            # 模糊匹配
            for excel_name, db_field in field_mapping.items():
                if self._is_similar(header_clean, excel_name):
                    column_mapping[col_idx] = db_field
                    break
        
        return column_mapping
    
    def _format_cell_value(self, value: Any) -> str:
        """格式化单元格值"""
        if value is None:
            return ''
        
        if isinstance(value, (int, float)):
            # 处理数字格式
            if isinstance(value, float) and value == int(value):
                return str(int(value))
            return str(value)
        
        if hasattr(value, 'strftime'):  # datetime, date, time
            try:
                if hasattr(value, 'hour'):  # datetime or time
                    return value.strftime('%Y-%m-%d %H:%M:%S')
                else:  # date
                    return value.strftime('%Y-%m-%d')
            except:
                return str(value)
        
        return str(value).strip()
    
    def _is_similar(self, str1: str, str2: str, threshold: float = 0.6) -> bool:
        """判断两个字符串是否相似"""
        str1 = str1.lower().strip()
        str2 = str2.lower().strip()
        
        # 完全包含
        if str1 in str2 or str2 in str1:
            return True
        
        # 编辑距离相似度
        similarity = self._levenshtein_similarity(str1, str2)
        return similarity >= threshold
    
    def _levenshtein_similarity(self, s1: str, s2: str) -> float:
        """计算编辑距离相似度"""
        if not s1 or not s2:
            return 0
        
        len1, len2 = len(s1), len(s2)
        max_len = max(len1, len2)
        
        if max_len == 0:
            return 1
        
        distance = self._levenshtein_distance(s1, s2)
        return 1 - (distance / max_len)
    
    def _levenshtein_distance(self, s1: str, s2: str) -> int:
        """计算编辑距离"""
        len1, len2 = len(s1), len(s2)
        
        # 创建DP表
        dp = [[0] * (len2 + 1) for _ in range(len1 + 1)]
        
        # 初始化
        for i in range(len1 + 1):
            dp[i][0] = i
        for j in range(len2 + 1):
            dp[0][j] = j
        
        # 填充
        for i in range(1, len1 + 1):
            for j in range(1, len2 + 1):
                cost = 0 if s1[i-1] == s2[j-1] else 1
                dp[i][j] = min(
                    dp[i-1][j] + 1,      # 删除
                    dp[i][j-1] + 1,      # 插入
                    dp[i-1][j-1] + cost  # 替换
                )
        
        return dp[len1][len2]
    
    @staticmethod
    def validate_file(filename: str, size: int) -> Tuple[bool, str]:
        """
        验证文件格式和大小
        
        Returns:
            Tuple[bool, str]: (是否有效, 错误信息)
        """
        # 检查扩展名
        valid_ext = False
        for ext in ExcelParser.SUPPORTED_EXTENSIONS:
            if filename.lower().endswith(ext):
                valid_ext = True
                break
        
        if not valid_ext:
            return False, "文件格式错误，请上传 .xlsx 或 .xls 格式的Excel文件"
        
        # 检查文件大小
        if size > ExcelParser.MAX_FILE_SIZE:
            size_mb = size / (1024 * 1024)
            return False, f"文件大小超出限制（最大10MB），当前文件：{size_mb:.2f}MB"
        
        return True, ""


class ExcelTemplate:
    """Excel模板生成器"""
    
    # 标准表头顺序
    STANDARD_HEADERS = [
        '订单日期',
        'ORDER NO.',
        '客户产品编号',
        'OE号',
        '产品描述',
        '数量',
        '单价',
        '货币',
        '金额',
        '折扣',
        '客户ID',
        '客户名称',
        '联系人',
        '联系电话',
        '联系邮箱',
        '供应商ID',
        '供应商名称',
        '采购选项',
        '工厂编号',
        '交期',
        '包装方式',
        '每包数量',
        '纸箱长',
        '纸箱宽',
        '纸箱高',
        '纸箱体积',
        '打包规格',
        '毛重',
        '净重',
        '运输方式',
        '起运港',
        '目的港',
        '目的地',
        '付款条件',
        '交货日期',
        '备注',
        '汇率',
        '预估成本',
        '毛利率',
        '业务员',
        '订单状态',
    ]
    
    @staticmethod
    def create_template(filename: str) -> bool:
        """
        创建标准导入模板
        
        Args:
            filename: 保存路径
        
        Returns:
            bool: 是否成功
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "订单导入"
            
            # 设置表头样式
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # 写入表头
            for col_idx, header in enumerate(ExcelTemplate.STANDARD_HEADERS, start=1):
                cell = sheet.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 设置列宽
            for col_idx in range(1, len(ExcelTemplate.STANDARD_HEADERS) + 1):
                sheet.column_dimensions[get_column_letter(col_idx)].width = 15
            
            # 冻结首行
            sheet.freeze_panes = 'A2'
            
            # 添加示例数据行
            example_data = [
                ['2024-01-15', 'PI20240115001', 'CUST-001', '12345', '发动机曲轴', 100, 50.00, 'USD', 5000.00, '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 7.24, 3000.00, 25, '张三', 'draft'],
            ]
            
            for row_idx, row_data in enumerate(example_data, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)
            
            # 保存
            workbook.save(filename)
            return True
            
        except Exception as e:
            print(f"创建模板失败: {e}")
            return False